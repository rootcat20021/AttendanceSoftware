# -*- coding: utf-8 -*-
def index():
    response.flash = T("Welcome to Attendance Software!")
    return dict(message=T('Attendance Software'))

def MachineCARDNOtoGRNO(CARDNO):
    import re
    CARDNO = re.sub("^5", "PR0012", CARDNO)
    CARDNO = re.sub("^4", "BH0011", CARDNO)
    CARDNO = re.sub("^3", "SS", CARDNO)
    CARDNO = re.sub("^10", "G", CARDNO)
    CARDNO = re.sub("^20", "L", CARDNO)
    return CARDNO

def CARDNOtoGRNO(CARDNO):
    import re
    CARDNO = re.sub("^5", "PR0012", CARDNO)
    CARDNO = re.sub("^4", "BH0011", CARDNO)
    CARDNO = re.sub("^03", "SS", CARDNO)
    CARDNO = re.sub("^010", "G", CARDNO)
    CARDNO = re.sub("^020", "L", CARDNO)
    return CARDNO

def GRNOtoCARDNO(GRNO):
    import re
    GRNO = re.sub("PR0012","5", GRNO)
    GRNO = re.sub("BH0011","4", GRNO)
    GRNO = re.sub("SS","03", GRNO)
    GRNO = re.sub("G","010", GRNO)
    GRNO = re.sub("L","020", GRNO)
    return GRNO

# following function returns the new GRNO. If it is unable to map than the same GRNO is returned
def GRNOtoNewGRNO(GRNO):
    SSCount = db(db.SSAttendanceCount.OldSewadarid == GRNO).select()
    NewGRNO = GRNO
    for Sewadar in SSCount:
        NewGRNO = Sewadar.NewID
        break
    
    return NewGRNO

#@auth.requires_login()
def view_sewadar():
    import os
    from gluon.tools import PluginManager
    SewaSamitiCountRecords = 'No Valid record selected'
    SewaSamitiDatesRecords = 'No Valid record selected'
    MasterRecords = 'No Valid record selected'
    MachineRecords = 'No Valid record selected'
    TextMessage = ''

    GENTS_REQUIRED = 30
    LADIES_REQURED = 36

    response.subtitle = "Sewadar Details"
    from gluon.sqlhtml import form_factory
    form=form_factory(SQLField('SewadarId','string',default="",requires=IS_NOT_EMPTY()),formname='UserDetails')
    if form.accepts(request.vars,session,formname='UserDetails'):
        GRNO = request.vars.SewadarId
        request.flash = GRNO
        datasource  = db((db.SSAttendanceCount.OldSewadarid == GRNO) | (db.SSAttendanceCount.NewID == GRNO)).select()
        for Sewadar in datasource:
#            try:
            if (Sewadar.gender == 'Male') & (Sewadar.Total < GENTS_REQUIRED):
                TextMessage = "Attendance Short by : " + str(GENTS_REQUIRED - Sewadar.Total)
            elif (Sewadar.gender == 'Female') & (Sewadar.Total < LADIES_REQURED):
                TextMessage = "Attendance Short by : " + str(LADIES_REQURED - Sewadar.Total)
            else:
                TextMessage = "Attendance Complete!"
#            except:
#                TextMessage = Sewadar.gender + str(Sewadar.Total)

        columns = ['SSAttendanceCount.NewID','SSAttendanceCount.OldSewadarid','SSAttendanceCount.Name','SSAttendanceCount.Total']
        orderby=columns

        headers = {'SSAttendanceCount.NewID':{'label':T('NewId'),'class':'','width':12,'truncate':12,'selected': False},
                   'SSAttendanceCount.OldSewadarid':{'label':T('OldId'),'class':'','width':12,'truncate':12,'selected': False},
                   'SSAttendanceCount.Name':{'label':T('Name                  '),'class':'','width':30,'truncate':30,'selected': False},
                   'SSAttendanceCount.Total':{'label':T('Total'),'class':'','width':4,'truncate':4,'selected': False}
                  }

        SewaSamitiCountRecords = SQLTABLE(datasource,columns=columns,headers=headers,orderby=orderby,_class='datatable')


        columns = ['SSAttendanceDate.DutyDate','SSAttendanceDate.Duty_Type']
        orderby=columns

        headers = {'SSAttendanceDate.DutyDate':{'label':T('Date'),'class':'','width':12,'truncate':12,'selected': False},
                   'SSAttendanceDate.Duty_Type':{'label':T('Type'),'class':'','width':2,'truncate':10,'selected': False}
                  }

        db.SSAttendanceDate.DutyDate.represent = lambda value, row: value.strftime("%d/%m/%Y")

        datasource = db((db.SSAttendanceDate.OldSewadarID == GRNO) | (db.SSAttendanceDate.SewadarNewID == GRNO)).select(orderby=~db.SSAttendanceDate.DutyDate)
        SewaSamitiDatesRecords = SQLTABLE(datasource,columns=columns,headers=headers,orderby=orderby,_class='datatable')
        #SewaSamitiDatesRecords = plugin_powerTable(db(db.SSAttendanceDate).select(orderby=~db.SSAttendanceDate.DutyDate))
        #records1 = SQLTABLE(db(db.MasterSheet.GR_NO == GRNO).select(),headers='fieldname:capitalize')
        rows = db(db.SSAttendanceCount.NewID == GRNO).select()
        oldgrnumber = GRNO
        for row in rows:
            oldgrnumber = row.OldSewadarid or GRNO

        datasource = db(db.MasterSheet.GR_NO == oldgrnumber).select()

        columns = ['MasterSheet.CANTEEN','MasterSheet.DEV_DTY']
        orderby=columns

        headers = {'MasterSheet.CANTEEN':{'label':T('CANTEEN'),'class':'','width':12,'truncate':12,'selected': False},
                   'MasterSheet.DEV_DTY':{'label':T('JATHA'),'class':'','width':35,'truncate':35,'selected': False}
                  }

        MasterRecords = SQLTABLE(datasource,columns=columns,headers=headers,orderby=orderby,_class='datatable')

#    if auth.user.username == 'admin':
#        MachineRecords = SQLTABLE(db(db.RawData.CARDNO == GRNOtoCARDNO(GRNO)).select('OFFICEPUNCH',orderby=~db.RawData.OFFICEPUNCH),headers='fieldname:capitalize')
#    else:
        MachineRecords = ''



    return dict(TextMessage=TextMessage,form=form,MasterRecords=MasterRecords,MachineRecords=MachineRecords,SewaSamitiDatesRecords=SewaSamitiDatesRecords,SewaSamitiCountRecords=SewaSamitiCountRecords)


@auth.requires(auth.user_id == 3)
def update():

    import os
    response.subtitle = "Upload excel file "
    from gluon.sqlhtml import form_factory
    form2=form_factory(SQLField('Master_xls','upload',uploadfolder='temporary'),formname='MASTER')
    if form2.accepts(request.vars,session,formname='MASTER'):
        request.flash='Received: %s'%request.vars.Master_xls
        path = os.path.join(request.folder,'private','Master_xls.xlsx')
        import shutil
        shutil.copyfileobj(request.vars.Master_xls.file,open(path, 'wb'))
        #Then redirect to the next screen (or do the processing now)
        redirect(URL(r=request, f='uploaddata_master'))

    form3=form_factory(SQLField('SSAttendanceDates_xls','upload',uploadfolder='temporary'),formname='SSATTENDANCE')
    if form3.accepts(request.vars,session,formname='SSATTENDANCE'):
        request.flash='Received: %s'%request.vars.SSAttendanceDates_xls
        path = os.path.join(request.folder,'private','SSAttendanceDates_xls.xlsx')
        import shutil
        shutil.copyfileobj(request.vars.SSAttendanceDates_xls.file,open(path, 'wb'))
        #Then redirect to the next screen (or do the processing now)
        redirect(URL(r=request, f='uploaddata_SSAttendance'))

    form4=form_factory(SQLField('SSAttendanceCount_xls','upload',uploadfolder='temporary'),formname='SSATTENDANCECOUNT')
    if form4.accepts(request.vars,session,formname='SSATTENDANCECOUNT'):
        request.flash='Received: %s'%request.vars.SSAttendanceCount_xls
        path = os.path.join(request.folder,'private','SSAttendanceCount_xls.xlsx')
        import shutil
        shutil.copyfileobj(request.vars.SSAttendanceCount_xls.file,open(path, 'wb'))
        #Then redirect to the next screen (or do the processing now)
        redirect(URL(r=request, f='uploaddata_SSAttendanceCount'))

    form5=form_factory(SQLField('CardList_xls','upload',uploadfolder='temporary'),formname='CARDLIST')
    if form5.accepts(request.vars,session,formname='CARDLIST'):
        request.flash='Received: %s'%request.vars.CardList_xls
        path = os.path.join(request.folder,'private','CardList_xls.xlsx')
        import shutil
        shutil.copyfileobj(request.vars.CardList_xls.file,open(path, 'wb'))
        #Then redirect to the next screen (or do the processing now)
        redirect(URL(r=request, f='uploaddata_CardList'))

    form6=form_factory(SQLField('STARLINK_zip','upload',uploadfolder='temporary'),formname='STARLINK')
    if form6.accepts(request.vars,session,formname='STARLINK'):
        request.flash='Received: %s'%request.vars.STARLINK_zip
        path = os.path.join(request.folder,'private','STARLINK.zip')
        import shutil
        shutil.copyfileobj(request.vars.STARLINK_zip.file,open(path, 'wb'))
        #Then redirect to the next screen (or do the processing now)
        redirect(URL(r=request, f='uploaddata_MachineAttendance'))

    
    form7=form_factory(SQLField('MACHINE_xlsx','upload',uploadfolder='temporary'),formname='MachineManual')
    if form7.accepts(request.vars,session,formname='MachineManual'):
        request.flash='Received: %s'%request.vars.MACHINE_xlsx
        path = os.path.join(request.folder,'private','MachineAttendance.xlsx')
        import shutil
        shutil.copyfileobj(request.vars.MACHINE_xlsx.file,open(path, 'wb'))
        #Then redirect to the next screen (or do the processing now)
        redirect(URL(r=request, f='uploaddata_MachineManualAttendance'))

    form8=form_factory(SQLField('Initiation_xlsx','upload',uploadfolder='temporary'),formname='InitiationList')
    if form8.accepts(request.vars,session,formname='InitiatedList'):
        request.flash='Received: %s'%request.vars.Initiation_xlsx
        path = os.path.join(request.folder,'private','InitiatedList.xlsx')
        import shutil
        shutil.copyfileobj(request.vars.Initiation_xlsx.file,open(path, 'wb'))
        #Then redirect to the next screen (or do the processing now)
        redirect(URL(r=request, f='uploaddata_InitiationList'))

    form9=form_factory(SQLField('SSParshadStatus_xlsx','upload',uploadfolder='temporary'),formname='SSParshadStatus')
    if form9.accepts(request.vars,session,formname='SSParshadStatus'):
        request.flash='Received: %s'%request.vars.SSParshadStatus_xlsx
        path = os.path.join(request.folder,'private','SSParshadStatus.xlsx')
        import shutil
        shutil.copyfileobj(request.vars.SSParshadStatus_xlsx.file,open(path, 'wb'))
        #Then redirect to the next screen (or do the processing now)
        redirect(URL(r=request, f='uploaddata_SSParshadStatus'))
    return dict(form2=form2,form3=form3,form4=form4,form5=form5,form6=form6,form7=form7,form8=form8,form9=form9)


def uploaddata_InitiationList():
    import xlrd
    import os
    path = os.path.join(request.folder,'private','InitiatedList.xlsx')
    workbook = xlrd.open_workbook(path)

    worksheet = workbook.sheet_names()
    #Attempt with any worksheet found
    worksheet = workbook.sheet_by_name(worksheet[0])
    header_row = 0
    import re, time

    #my headers are the headers in DB
    myheaders = ['NewGRNO','Status']

    #headers are the name used in XLS
    headers = ['NewGRNO','Status']
    header_cells_in_xls = [worksheet.cell_value(0,idx) for idx in range(worksheet.ncols)]

    try:
        db(db.InitiatedList.id > 0).delete()
    except:
        pass


    indexcol = [];
    for header_cell in headers:
            try:
                indexcol.append(header_cells_in_xls.index(header_cell))
            except Exception,e :
                print 'Header not found in xls: %s' %e
            else:
                print 'OK'

    print "INDEXCOL = ", indexcol
    row_dict_list = []

    throttle = 1
    for row in xrange(header_row+1, worksheet.nrows):
       i=0
       row_dict = {}

       for col in indexcol:
           cell_type = worksheet.cell_type(row, col)
           if cell_type == xlrd.XL_CELL_EMPTY:
               value = None
           elif cell_type == xlrd.XL_CELL_TEXT:
              value = worksheet.cell_value(row, col)
           elif cell_type == xlrd.XL_CELL_NUMBER:
              value = float(worksheet.cell_value(row,col))
           elif cell_type == xlrd.XL_CELL_DATE:
              from datetime import datetime
              value = datetime(*xlrd.xldate_as_tuple(worksheet.cell_value(row, col), workbook.datemode))
           elif cell_type == xlrd.XL_CELL_BOOLEAN:
              value = bool(worksheet.cell_value(row, col))
           else:
              value = worksheet.cell_value(row, col)

           if value == None:
               row_dict[myheaders[i]] = "YES"
           else:
               row_dict[myheaders[i]] = value

           i=i+1
       try:
           db.InitiatedList.insert(**row_dict)
       except:
           duplicate = 1

    response.flash = T("Entry Successful!")

    return locals()

def uploaddata_SSParshadStatus():
    import xlrd
    import os
    path = os.path.join(request.folder,'private','SSParshadStatus.xlsx')
    workbook = xlrd.open_workbook(path)

    worksheet = workbook.sheet_names()
    #Attempt with any worksheet found
    worksheet = workbook.sheet_by_name(worksheet[0])
    header_row = 0
    import re, time

    #my headers are the headers in DB
    myheaders = ['NewGRNO','Status']

    #headers are the name used in XLS
    headers = ['NewGRNO','Status']
    header_cells_in_xls = [worksheet.cell_value(0,idx) for idx in range(worksheet.ncols)]


    try:
        db(db.SSTentativeParshadList.id > 0).delete()
    except:
        pass

    indexcol = [];
    for header_cell in headers:
            try:
                indexcol.append(header_cells_in_xls.index(header_cell))
            except Exception,e :
                print 'Header not found in xls: %s' %e
            else:
                print 'OK'

    print "INDEXCOL = ", indexcol
    row_dict_list = []

    throttle = 1
    for row in xrange(header_row+1, worksheet.nrows):
       i=0
       row_dict = {}

       for col in indexcol:
           cell_type = worksheet.cell_type(row, col)
           if cell_type == xlrd.XL_CELL_EMPTY:
               value = None
           elif cell_type == xlrd.XL_CELL_TEXT:
              value = worksheet.cell_value(row, col)
           elif cell_type == xlrd.XL_CELL_NUMBER:
              value = float(worksheet.cell_value(row,col))
           elif cell_type == xlrd.XL_CELL_DATE:
              from datetime import datetime
              value = datetime(*xlrd.xldate_as_tuple(worksheet.cell_value(row, col), workbook.datemode))
           elif cell_type == xlrd.XL_CELL_BOOLEAN:
              value = bool(worksheet.cell_value(row, col))
           else:
              value = worksheet.cell_value(row, col)
           row_dict[myheaders[i]] = value
           i=i+1
       try:
           db.SSTentativeParshadList.insert(**row_dict)
       except:
           duplicate = 1

    response.flash = T("Entry Successful!")

    return locals()



def uploaddata_MachineAttendance():
    import os
    path = os.path.join(request.folder,'private','STARLINK.zip')
    os.system("rm -rf " + os.path.join(request.folder,'private','STARLINK') + ";unzip " + path + " -d " + os.path.join(request.folder,'private'));
    AddAttendanceFileToMachineAttendanceDatabase(os.path.join(request.folder,'private','STARLINK'))

    #Translate to new GRNO
    SSCount = db(db.SSAttendanceCount).select()
    for Sewadar in SSCount:
        if Sewadar.OldSewadarid == '':
            pass
        else:
            db(db.MachineAttendance.GRNO == Sewadar.OldSewadarid).update(NewGRNO = Sewadar.NewID)

    #Translate to new GNRO using Date database too as they sometime seem to have additional info
    SSDates = db(db.SSAttendanceDate).select(db.SSAttendanceDate.OldSewadarID,db.SSAttendanceDate.SewadarNewID,distinct=True)
    for Sewadar in SSDates:
        if Sewadar.OldSewadarID == '':
            pass
        else:
            db(db.MachineAttendance.GRNO == Sewadar.OldSewadarID).update(NewGRNO = Sewadar.SewadarNewID)

    
    return "Uploaded STARLINK FOLDER"

def AddAttendanceFileToMachineAttendanceDatabase(walk_dir):
    import os
    import sys
    import re

    a = re.compile(".*.TXT")
    t = re.compile(".*:.*:.*")
    file_list = []



    for root, subdirs, files in os.walk(walk_dir):
        file_list[:] = [filename for filename in files if a.match(filename)]
        for filename in file_list:
            file_path = os.path.join(root, filename)
            with open(file_path, 'rb') as fp:
                row_dict = {}
                for line in fp:
                    try:
                        dummy_0, CARDNO, TIME, IO ,dummy_1 = re.split(r'\s+',line)
                        row_dict['GRNO'] = MachineCARDNOtoGRNO(CARDNO)
                        d1, d2, m1, m2, y1, y2 = list(filename.replace(".TXT","").replace("SL",""))
                        if t.match(TIME):
                            row_dict['DATETIME'] = '20' + y1 + y2 + '-' + m1 + m2 + '-' + d1 + d2 + ' ' + TIME
                        else:
                            row_dict['DATETIME'] = '20' + y1 + y2 + '-' + m1 + m2 + '-' + d1 + d2 + ' ' + TIME + ':00'

                        row_dict['TYPE'] = 'DMACHINE'
                        row_dict['NewGRNO'] = row_dict['GRNO']
                        row_dict['IO'] = IO
                        try:
                            db.MachineAttendance.insert(**row_dict)
                        except:
                            duplicate = 1
                    except:
                        pass

        for subdir in subdirs:
            AddAttendanceFileToMachineAttendanceDatabase(subdir)

    return 0


def uploaddata_CardList():
    import xlrd
    import os
    path = os.path.join(request.folder,'private','CardList_xls.xlsx')
    workbook = xlrd.open_workbook(path)
    worksheet = workbook.sheet_names()
    #Attempt with any worksheet found
    worksheet = workbook.sheet_by_name(worksheet[0])
    header_row = 0
    import re

    #my headers are the headers in DB
    myheaders = ['SewadarOldID','SewadarNewID','PROXIMITY_CARDLIST']

    #headers are the name used in XLS
    headers = ['GR_NO','A_CARD']
    header_cells_in_xls = [worksheet.cell_value(0,idx) for idx in range(worksheet.ncols)]



    indexcol = [];
    for header_cell in headers:
            try:
                indexcol.append(header_cells_in_xls.index(header_cell))
            except Exception,e :
                print 'Header not found in xls: %s' %e
            else:
                print 'OK'

    print "INDEXCOL = ", indexcol
    row_dict_list = []

    try:
        db(db.MasterSheet).delete()
    except:
        print "Created Master Table"

    for row in xrange(header_row+1, worksheet.nrows):
       i=0
       row_dict = {}
       for col in indexcol:
           cell_type = worksheet.cell_type(row, col)
           if cell_type == xlrd.XL_CELL_EMPTY:
               value = None
           elif cell_type == xlrd.XL_CELL_TEXT:
              value = worksheet.cell_value(row, col)
           elif cell_type == xlrd.XL_CELL_NUMBER:
              value = float(worksheet.cell_value(row,col))
           elif cell_type == xlrd.XL_CELL_DATE:
              from datetime import datetime
              value = datetime(*xlrd.xldate_as_tuple(worksheet.cell_value(row, col), workbook.datemode))
           elif cell_type == xlrd.XL_CELL_BOOLEAN:
              value = bool(worksheet.cell_value(row, col))
           else:
              value = worksheet.cell_value(row, col)
           row_dict[myheaders[i]] = value
           i=i+1
       db.MasterSheet.insert(**row_dict)

    response.flash = T("Entry Successful!")
    return locals()


def uploaddata_master():
    import xlrd
    import os
    path = os.path.join(request.folder,'private','Master_xls.xlsx')
    workbook = xlrd.open_workbook(path)
    worksheet = workbook.sheet_names()
    #Attempt with any worksheet found
    worksheet = workbook.sheet_by_name(worksheet[0])
    header_row = 0
    import re

    #my headers are the headers in DB
    myheaders = ['GR_NO','SewadarNewID','FH_NAME','CANTEEN','DEV_DTY','DUTY_AT','RES_ADDR1','RES_ADDR2','RES_PHONE1','RES_PHONE2','INITIATED_SS','RES_PHONE3']

    #headers are the name used in XLS
    headers = ['GR_NO','SewadarNewID','FH_NAME','CANTEEN','DEV_DTY','DUTY_AT','RES_ADDR1','RES_ADDR2','RES_PHONE1','RES_PHONE2','DOI_YN','RES_PHONE3']
    header_cells_in_xls = [worksheet.cell_value(0,idx) for idx in range(worksheet.ncols)]



    indexcol = [];
    for header_cell in headers:
            try:
                indexcol.append(header_cells_in_xls.index(header_cell))
            except Exception,e :
                print 'Header not found in xls: %s' %e
            else:
                print 'OK'

    print "INDEXCOL = ", indexcol
    row_dict_list = []

    try:
        db(db.MasterSheet).delete()
    except:
        print "Created Master Table"

    for row in xrange(header_row+1, worksheet.nrows):
       i=0
       row_dict = {}
       for col in indexcol:
           cell_type = worksheet.cell_type(row, col)
           if cell_type == xlrd.XL_CELL_EMPTY:
               value = None
           elif cell_type == xlrd.XL_CELL_TEXT:
              value = worksheet.cell_value(row, col)
           elif cell_type == xlrd.XL_CELL_NUMBER:
              value = float(worksheet.cell_value(row,col))
           elif cell_type == xlrd.XL_CELL_DATE:
              from datetime import datetime
              value = datetime(*xlrd.xldate_as_tuple(worksheet.cell_value(row, col), workbook.datemode))
           elif cell_type == xlrd.XL_CELL_BOOLEAN:
              value = bool(worksheet.cell_value(row, col))
           else:
              value = worksheet.cell_value(row, col)
           row_dict[myheaders[i]] = value
           i=i+1
       db.MasterSheet.insert(**row_dict)

    response.flash = T("Entry Successful!")
    return locals()


def uploaddata_MachineManualAttendance():
    import xlrd
    import os
    path = os.path.join(request.folder,'private','MachineAttendance.xlsx')
    workbook = xlrd.open_workbook(path)

    worksheet = workbook.sheet_names()
#Attempt with any worksheet found
    worksheet = workbook.sheet_by_name(worksheet[0])
    header_row = 0
    import re, time

#my headers are the headers in DB
    myheaders = ['NewGRNO','GRNO','DATETIME','TYPE','IO']

#headers are the name used in XLS
    headers = ['NewGRNO','GRNO','DATETIME','TYPE','IO']
    header_cells_in_xls = [worksheet.cell_value(0,idx) for idx in range(worksheet.ncols)]



    indexcol = [];
    for header_cell in headers:
        try:
            indexcol.append(header_cells_in_xls.index(header_cell))
        except Exception,e :
            print 'Header not found in xls: %s' %e
        else:
            print 'OK'

    row_dict_list = []

    throttle = 1
    for row in xrange(header_row+1, worksheet.nrows):
        i=0
        row_dict = {}
        throttle = throttle + 1
        if throttle % 1000 == 0:
            time.sleep (50.0 / 1000.0);

        for col in indexcol:
            cell_type = worksheet.cell_type(row, col)
            if cell_type == xlrd.XL_CELL_EMPTY:
                value = None
            elif cell_type == xlrd.XL_CELL_TEXT:
                value = worksheet.cell_value(row, col)
            elif cell_type == xlrd.XL_CELL_NUMBER:
                value = float(worksheet.cell_value(row,col))
            elif cell_type == xlrd.XL_CELL_DATE:
                from datetime import datetime
                value = datetime(*xlrd.xldate_as_tuple(worksheet.cell_value(row, col), workbook.datemode))
            elif cell_type == xlrd.XL_CELL_BOOLEAN:
                value = bool(worksheet.cell_value(row, col))
            else:
                value = worksheet.cell_value(row, col)
            row_dict[myheaders[i]] = value
            i=i+1

        try:
            db.MachineAttendance.insert(**row_dict)
        except:
            a = 1

    return locals()

def uploaddata_SSAttendance():
    import xlrd
    import os
    path = os.path.join(request.folder,'private','SSAttendanceDates_xls.xlsx')
    workbook = xlrd.open_workbook(path)

    worksheet = workbook.sheet_names()
    #Attempt with any worksheet found
    worksheet = workbook.sheet_by_name(worksheet[0])
    header_row = 0
    import re, time

    #my headers are the headers in DB
    myheaders = ['OldSewadarID','SewadarNewID','Name','Gender','DepartmentID','DutyDate','Duty_Type']

    #headers are the name used in XLS
    headers = ['OldSewadarID','SewadarNewID','Name','Gender','DepartmentID','DutyDate','Duty Type']
    header_cells_in_xls = [worksheet.cell_value(0,idx) for idx in range(worksheet.ncols)]



    indexcol = [];
    for header_cell in headers:
            try:
                indexcol.append(header_cells_in_xls.index(header_cell))
            except Exception,e :
                print 'Header not found in xls: %s' %e
            else:
                print 'OK'

    print "INDEXCOL = ", indexcol
    row_dict_list = []

    throttle = 1
#    db(db.SSAttendanceDate).delete()
    for row in xrange(header_row+1, worksheet.nrows):
       i=0
       row_dict = {}
#       throttle = throttle + 1
#       if throttle % 1000 == 0:
#            time.sleep (20.0 / 1000.0);

       for col in indexcol:
           cell_type = worksheet.cell_type(row, col)
           if cell_type == xlrd.XL_CELL_EMPTY:
               value = None
           elif cell_type == xlrd.XL_CELL_TEXT:
              value = worksheet.cell_value(row, col)
           elif cell_type == xlrd.XL_CELL_NUMBER:
              value = float(worksheet.cell_value(row,col))
           elif cell_type == xlrd.XL_CELL_DATE:
              from datetime import datetime
              value = datetime(*xlrd.xldate_as_tuple(worksheet.cell_value(row, col), workbook.datemode))
           elif cell_type == xlrd.XL_CELL_BOOLEAN:
              value = bool(worksheet.cell_value(row, col))
           else:
              value = worksheet.cell_value(row, col)
           row_dict[myheaders[i]] = value
           i=i+1
       try:
           db.SSAttendanceDate.insert(**row_dict)
       except:
           duplicate = 1
       #row_dict_list.append(row_dict)

    #
    #for row in row_dict_list:
    #   db.SSAttendanceDate.insert(**row)

    response.flash = T("Entry Successful!")

    return locals()

def uploaddata_SSAttendanceCount():
    import xlrd
    import os
    path = os.path.join(request.folder,'private','SSAttendanceCount_xls.xlsx')
    workbook = xlrd.open_workbook(path)
    worksheet = workbook.sheet_names()
    #Attempt with any worksheet found
    worksheet = workbook.sheet_by_name(worksheet[0])
    header_row = 0
    import re


    #my headers are the headers in DB
    myheaders = ['NewID','OldSewadarid','Name','Father_Husband_Name','status','gender','B','w','V1','V2','V3','V4','Total','areaname']

    #headers are the name used in XLS
    headers = ['NewID','OldSewadarid','Name','Father_Husband_Name','status','gender','B','w','V1','V2','V3','V4','Total','areaname']
    header_cells_in_xls = [worksheet.cell_value(0,idx) for idx in range(worksheet.ncols)]



    indexcol = [];
    for header_cell in headers:
            try:
                indexcol.append(header_cells_in_xls.index(header_cell))
            except Exception,e :
                print 'Header not found in xls: %s' %e
            else:
                print 'OK'

    print "INDEXCOL = ", indexcol
    row_dict_list = []

    db(db.SSAttendanceCount).delete()
    for row in xrange(header_row+1, worksheet.nrows):
       i=0
       row_dict = {}
       for col in indexcol:
           cell_type = worksheet.cell_type(row, col)
           if cell_type == xlrd.XL_CELL_EMPTY:
               value = None
           elif cell_type == xlrd.XL_CELL_TEXT:
              value = worksheet.cell_value(row, col)
           elif cell_type == xlrd.XL_CELL_NUMBER:
              value = float(worksheet.cell_value(row,col))
           elif cell_type == xlrd.XL_CELL_DATE:
              from datetime import datetime
              value = datetime(*xlrd.xldate_as_tuple(worksheet.cell_value(row,col), workbook.datemode))
           elif cell_type == xlrd.XL_CELL_BOOLEAN:
              value = bool(worksheet.cell_value(row, col))
           else:
              value = worksheet.cell_value(row, col)
           row_dict[myheaders[i]] = value
           i=i+1
       db.SSAttendanceCount.insert(**row_dict)



    return locals()


def xlsdiff():
    import xlrd
    import os
    path1 = os.path.join(request.folder,'private','XL1.xlsx')
    path2 = os.path.join(request.folder,'private','XL2.xlsx')
    workbook1 = xlrd.open_workbook(path1)
    worksheet1 = workbook1.sheet_names()
    workbook2 = xlrd.open_workbook(path2)
    worksheet2 = workbook2.sheet_names()
    #Attempt with any worksheet found
    worksheet1 = workbook1.sheet_by_name(worksheet1[0])
    worksheet2 = workbook2.sheet_by_name(worksheet2[0])
    header_row = 0
    import re

    #Now define the table
    db.define_table('XLS1',
                    Field('GRNO','string'),
                    Field('DutyDate','string'),
                    Field('DutyDate','string'),
                    migrate=True,
                    redefine=True,
                    format='%(SewadarNewID)s')

    db(db.XLS1.id > 0).delete()

    db.define_table('XLS2',
                    Field('GRNO','string'),
                    Field('DutyDate','string'),
                    Field('DutyDate','string'),
                    migrate=True,
                    redefine=True,
                    format='%(SewadarNewID)s')

    db(db.XLS2.id > 0).delete()

    #my headers are the headers in DB
    myheaders = ['GRNO','DutyDate','DutyType']

    #headers are the name used in XLS
    headers = ['GRNO','DutyDate','DutyType']
    header_cells_in_xls1 = [worksheet1.cell_value(0,idx) for idx in range(worksheet1.ncols)]
    header_cells_in_xls1 = [worksheet2.cell_value(0,idx) for idx in range(worksheet2.ncols)]



    indexcol = [];
    for header_cell in headers:
            try:
                indexcol.append(header_cells_in_xls1.index(header_cell))
            except Exception,e :
                print 'Header not found in xls: %s' %e
            else:
                print 'OK'

    print "INDEXCOL = ", indexcol
    row_dict_list = []

    db(db.XLS1).delete()
    for row in xrange(header_row+1, worksheet1.nrows):
       i=0
       row_dict = {}
       for col in indexcol:
           cell_type = worksheet1.cell_type(row, col)
           if cell_type == xlrd.XL_CELL_EMPTY:
               value = None
           elif cell_type == xlrd.XL_CELL_TEXT:
              value = worksheet1.cell_value(row, col)
           elif cell_type == xlrd.XL_CELL_NUMBER:
              value = float(worksheet1.cell_value(row,col))
           elif cell_type == xlrd.XL_CELL_DATE:
              from datetime import datetime
              value = datetime(*xlrd.xldate_as_tuple(worksheet1.cell_value(row,col), workbook1.datemode))
           elif cell_type == xlrd.XL_CELL_BOOLEAN:
              value = bool(worksheet1.cell_value(row, col))
           else:
              value = worksheet1.cell_value(row, col)
           row_dict[myheaders[i]] = value
           i=i+1
       db.XlS1.insert(**row_dict)

    for row in xrange(header_row+1, worksheet1.nrows):
       i=0
       row_dict = {}
       for col in indexcol:
           cell_type = worksheet2.cell_type(row, col)
           if cell_type == xlrd.XL_CELL_EMPTY:
               value = None
           elif cell_type == xlrd.XL_CELL_TEXT:
              value = worksheet2.cell_value(row, col)
           elif cell_type == xlrd.XL_CELL_NUMBER:
              value = float(worksheet2.cell_value(row,col))
           elif cell_type == xlrd.XL_CELL_DATE:
              from datetime import datetime
              value = datetime(*xlrd.xldate_as_tuple(worksheet2.cell_value(row,col), workbook2.datemode))
           elif cell_type == xlrd.XL_CELL_BOOLEAN:
              value = bool(worksheet2.cell_value(row, col))
           else:
              value = worksheet2.cell_value(row, col)
           row_dict[myheaders[i]] = value
           i=i+1
       db.XlS2.insert(**row_dict)





def view_date():
    from gluon.sqlhtml import form_factory
    import datetime

    MachineRecords = 'No Date Selected'

    form=form_factory(SQLField('DateStart','date',default=datetime.datetime.today()),SQLField('DateEnd','date',default=datetime.datetime.today()+datetime.timedelta(days=1)),formname='DateSelect')
    if form.accepts(request.vars,session,formname='DateSelect'):
        DateSelectedStart = request.vars.DateStart
        DateSelectedEnd = request.vars.DateEnd

        extracolumns = [{'label':'GRNO',
                         'class':'string',
                         'width':'10', #width in pixels or %
                         'content':lambda row, rc: A(CARDNOtoGRNO(row.CARDNO),_href='edit/%s'%row.CARDNO),
                         'selected': False #agregate class selected to this column
                         }]


        rows = db((db.RawData.OFFICEPUNCH > DateSelectedStart) & (db.RawData.OFFICEPUNCH < DateSelectedEnd) ).select(db.RawData.CARDNO,db.RawData.OFFICEPUNCH,orderby=db.RawData.CARDNO)

        for row in rows:
            print row.GRNO,",",row.CARDNO
        MachineRecords = SQLTABLE(rows,headers='fieldname:capitalize',extracolumns=extracolumns)

    return dict(form=form,MachineRecords=MachineRecords)

def Roundoffdate(DATETIME, CUTOFF):
    import datetime
    RoundedDate = DATETIME
    if (DATETIME.hour < int(CUTOFF)) or (DATETIME.hour == int(CUTOFF) and DATETIME.minute == 0):
       RoundedDate = RoundedDate.replace(hour=0, minute=0, second=0, microsecond=0)
    else:
       RoundedDate = (RoundedDate + datetime.timedelta(days=1)).replace(hour=0, minute=0, second=0, microsecond=0) 
    return RoundedDate

def ParshadList():
    import os
    from openpyxl import Workbook
    response.headers['Connection'] =  'keep-alive'
    dworkbook = Workbook()
    dpath = os.path.join(request.folder,'private','TentativeParshadList.xlsx')
    try:
        os.remove(dpath)
    except:
        pass

    
    from gluon.sqlhtml import form_factory
    import datetime
    import time

    DateSelectedStart = datetime.datetime.today()
    DateSelectedEnd = datetime.datetime.today()
    DAY_END_TIME = 19


    CURRENT_VISIT = {}
    CURRENT_VISIT['COUNT'] = 5
    CURRENT_VISIT['D0'] = '11-March-2015'
    CURRENT_VISIT['D1'] = '12-March-2015'
    CURRENT_VISIT['D2'] = '13-March-2015'
    CURRENT_VISIT['D3'] = '14-March-2015'
    CURRENT_VISIT['D4'] = '15-March-2015'
    CURRENT_VISIT['D5'] = '16-March-2015'


    VISIT_DATES = {}
    TOTAL_VISIT = 3
    VISIT_DATES['V0','COUNT'] = 5
    VISIT_DATES['V0','D0'] = '05-October-2011 00:00:00'
    VISIT_DATES['V0','D1'] = '06-October-2011 00:00:00'
    VISIT_DATES['V0','D2'] = '07-October-2011 00:00:00'
    VISIT_DATES['V0','D3'] = '08-October-2011 00:00:00'
    VISIT_DATES['V0','D4'] = '09-October-2011 00:00:00'
    VISIT_DATES['V0','D5'] = '10-October-2011 00:00:00'

    VISIT_DATES['V1','COUNT'] = 5
    VISIT_DATES['V1','D0'] = '23-November-2011 00:00:00'
    VISIT_DATES['V1','D1'] = '24-November-2011 00:00:00'
    VISIT_DATES['V1','D2'] = '25-November-2011 00:00:00'
    VISIT_DATES['V1','D3'] = '26-November-2011 00:00:00'
    VISIT_DATES['V1','D4'] = '27-November-2011 00:00:00'
    VISIT_DATES['V1','D5'] = '28-November-2011 00:00:00'

    VISIT_DATES['V2','COUNT'] = 5
    VISIT_DATES['V2','D0'] = '07-March-2012 00:00:00'
    VISIT_DATES['V2','D1'] = '08-March-2012 00:00:00'
    VISIT_DATES['V2','D2'] = '09-March-2012 00:00:00'
    VISIT_DATES['V2','D3'] = '10-March-2012 00:00:00'
    VISIT_DATES['V2','D4'] = '11-March-2012 00:00:00'
    VISIT_DATES['V2','D5'] = '12-March-2012 00:00:00'

    VISIT_DATES['V3','COUNT'] = 5
    VISIT_DATES['V3','D0'] = '10-October-2012 00:00:00'
    VISIT_DATES['V3','D1'] = '11-October-2012 00:00:00'
    VISIT_DATES['V3','D2'] = '12-October-2012 00:00:00'
    VISIT_DATES['V3','D3'] = '13-October-2012 00:00:00'
    VISIT_DATES['V3','D4'] = '14-October-2012 00:00:00'
    VISIT_DATES['V3','D5'] = '15-October-2012 00:00:00'

    VISIT_DATES['V4','COUNT'] = 5
    VISIT_DATES['V4','D0'] = '28-November-2012 00:00:00'
    VISIT_DATES['V4','D1'] = '29-November-2012 00:00:00'
    VISIT_DATES['V4','D2'] = '30-November-2012 00:00:00'
    VISIT_DATES['V4','D3'] = '01-December-2012 00:00:00'
    VISIT_DATES['V4','D4'] = '02-December-2012 00:00:00'
    VISIT_DATES['V4','D5'] = '03-December-2012 00:00:00'

    VISIT_DATES['V5','COUNT'] = 5
    VISIT_DATES['V5','D0'] = '13-March-2013 00:00:00'
    VISIT_DATES['V5','D1'] = '14-March-2013 00:00:00'
    VISIT_DATES['V5','D2'] = '15-March-2013 00:00:00'
    VISIT_DATES['V5','D3'] = '16-March-2013 00:00:00'
    VISIT_DATES['V5','D4'] = '17-March-2013 00:00:00'
    VISIT_DATES['V5','D5'] = '18-March-2013 00:00:00'

    VISIT_DATES['V6','COUNT'] = 5
    VISIT_DATES['V6','D0'] = '09-October-2013 00:00:00'
    VISIT_DATES['V6','D1'] = '10-October-2013 00:00:00'
    VISIT_DATES['V6','D2'] = '11-October-2013 00:00:00'
    VISIT_DATES['V6','D3'] = '12-October-2013 00:00:00'
    VISIT_DATES['V6','D4'] = '13-October-2013 00:00:00'
    VISIT_DATES['V6','D5'] = '14-October-2013 00:00:00'

    VISIT_DATES['V7','COUNT'] = 3
    VISIT_DATES['V7','D0'] = '10-January-2014 00:00:00'
    VISIT_DATES['V7','D1'] = '11-January-2014 00:00:00'
    VISIT_DATES['V7','D2'] = '12-January-2014 00:00:00'
    VISIT_DATES['V7','D3'] = '13-January-2014 00:00:00'
    VISIT_DATES['V7','D4'] = '14-January-2014 00:00:00'

    VISIT_DATES['V8','COUNT'] = 5
    VISIT_DATES['V8','D0'] = '01-October-2014 00:00:00'
    VISIT_DATES['V8','D1'] = '02-October-2014 00:00:00'
    VISIT_DATES['V8','D2'] = '03-October-2014 00:00:00'
    VISIT_DATES['V8','D3'] = '04-October-2014 00:00:00'
    VISIT_DATES['V8','D4'] = '05-October-2014 00:00:00'
    VISIT_DATES['V8','D5'] = '06-October-2014 00:00:00'

    VISIT_DATES['V9','COUNT'] = 5
    VISIT_DATES['V9','D0'] = '26-October-2014 00:00:00'
    VISIT_DATES['V9','D1'] = '27-October-2014 00:00:00'
    VISIT_DATES['V9','D2'] = '28-October-2014 00:00:00'
    VISIT_DATES['V9','D3'] = '29-October-2014 00:00:00'
    VISIT_DATES['V9','D4'] = '30-October-2014 00:00:00'
    VISIT_DATES['V9','D5'] = '31-October-2014 00:00:00'


    message = "ALL OK"
    ParshadList = {}

    form=form_factory(SQLField('DAY_END_TIME','string',default=19,requires=IS_IN_SET(range(0,23,1))),SQLField('DateStart','date',default=datetime.datetime.today()-datetime.timedelta(days=31)),SQLField('DateEnd','date',default=datetime.datetime.today()),formname='DateSelect')

    if form.accepts(request.vars,session,formname='DateSelect'):
        DateSelectedStart = request.vars.DateStart
        DateSelectedEnd = request.vars.DateEnd
        DAY_END_TIME = request.vars.DAY_END_TIME
        SSDate = db((db.SSAttendanceDate.DutyDate >= (datetime.datetime.strptime(DateSelectedStart,'%Y-%m-%d').replace(hour=0, minute=0, second=0, microsecond=0))) & (db.SSAttendanceDate.DutyDate <= (datetime.datetime.strptime(DateSelectedEnd,'%Y-%m-%d').replace(hour=23, minute=59, second=59, microsecond=999)))).select('SewadarNewID','DutyDate','Duty_Type')
        MachineDate = db((db.MachineAttendance.DATETIME >= (datetime.datetime.strptime(DateSelectedStart,'%Y-%m-%d') - datetime.timedelta(hours=((24 - int(DAY_END_TIME)) % 24)))) & (db.MachineAttendance.DATETIME <= (datetime.datetime.strptime(DateSelectedEnd,'%Y-%m-%d') + datetime.timedelta(hours=int(DAY_END_TIME))))).select('GRNO','NewGRNO','DATETIME','TYPE')

        ParshadList['SEWADARS'] = []
        SSAttendanceDictionary = {}
        for SSEntry in SSDate:
            ParshadList['SEWADARS'].append(SSEntry.SewadarNewID)
            if SSEntry.Duty_Type == 'W':
                SSAttendanceDictionary[SSEntry.SewadarNewID,Roundoffdate(SSEntry.DutyDate,23),'W'] = SSEntry.DutyDate
                try:
                    ParshadList[SSEntry.SewadarNewID,'SSCount'] = ParshadList[SSEntry.SewadarNewID,'SSCount'] + 2
                except:
                    ParshadList[SSEntry.SewadarNewID,'SSCount'] = 2
            else:
                SSAttendanceDictionary[SSEntry.SewadarNewID,Roundoffdate(SSEntry.DutyDate,23),'D'] = SSEntry.DutyDate
                try:
                    ParshadList[SSEntry.SewadarNewID,'SSCount'] = ParshadList[SSEntry.SewadarNewID,'SSCount'] + 1
                except:
                    ParshadList[SSEntry.SewadarNewID,'SSCount'] = 1


        MachineAttendanceAdditional = {}


        try:
            db.tempMachineAttendanceAdditional.drop()
        except:
            pass

        #Now define the table
        db.define_table('tempMachineAttendanceAdditional',
                Field('GRNO','string'),
                Field('NewGRNO','string'),
                Field('Duty_Type','string'),
                Field('DutyDate','datetime'),
                Field('DutyDateList','list:string'),
                migrate=True,
                redefine=True,
                format='%(NewGRNO)s')

        db(db.tempMachineAttendanceAdditional.id > 0).delete()

        for MEntry in MachineDate:
            #Select earliest entry but give preference to WW attendance
            if MEntry.TYPE == 'WMANUAL':
                try:
                    a = SSAttendanceDictionary[MEntry.NewGRNO,Roundoffdate(MEntry.DATETIME,DAY_END_TIME),'W']
                except:
                    try:
                        MachineAttendanceAdditional[MEntry.NewGRNO,MEntry.GRNO,Roundoffdate(MEntry.DATETIME,DAY_END_TIME),'W'].append(MEntry.DATETIME) 
                    except:
                        MachineAttendanceAdditional[MEntry.NewGRNO,MEntry.GRNO,Roundoffdate(MEntry.DATETIME,DAY_END_TIME),'W'] = [MEntry.DATETIME]


            else:
                try:
                    a = SSAttendanceDictionary[MEntry.NewGRNO,Roundoffdate(MEntry.DATETIME,DAY_END_TIME),'D']
                except:
                    try:
                        a = SSAttendanceDictionary[MEntry.NewGRNO,Roundoffdate(MEntry.DATETIME,DAY_END_TIME),'W']
                    except:
                        try:
                            MachineAttendanceAdditional[MEntry.NewGRNO,MEntry.GRNO,Roundoffdate(MEntry.DATETIME,DAY_END_TIME),'D'].append(MEntry.DATETIME) 
                        except:
                            MachineAttendanceAdditional[MEntry.NewGRNO,MEntry.GRNO,Roundoffdate(MEntry.DATETIME,DAY_END_TIME),'D'] = [MEntry.DATETIME]


        del SSAttendanceDictionary


        for key, value in MachineAttendanceAdditional.iteritems():
            NewGRNO, GRNO, DutyDate , DutyType= key
            ParshadList['SEWADARS'].append(NewGRNO)
            if DutyType == 'W':
                try:
                    MachineAttendanceAdditional.pop([NewGRNO,GRNO,DutyDate,'D'], None)
                except:
                    pass

                try:
                    ParshadList[NewGRNO,'SSCount'] = ParshadList[NewGRNO,'SSCount'] + 2
                except:
                    ParshadList[NewGRNO,'SSCount'] = 2
            else:
                try:
                    MachineAttendanceAdditional.pop([NewGRNO,GRNO,DutyDate,'W'], None)
                    try:
                        ParshadList[NewGRNO,'SSCount'] = ParshadList[NewGRNO,'SSCount'] + 2
                    except:
                        ParshadList[NewGRNO,'SSCount'] = 2

                except:
                    try:
                        ParshadList[NewGRNO,'SSCount'] = ParshadList[NewGRNO,'SSCount'] + 1
                    except:
                        ParshadList[NewGRNO,'SSCount'] = 1


        for key, value in MachineAttendanceAdditional.iteritems():
            MachineAttendanceAdditionalDictionary = {}
            NewGRNO, GRNO, DutyDate , DutyType= key
            ParshadList['SEWADARS'].append(NewGRNO)
            MachineAttendanceAdditionalDictionary['NewGRNO'] = NewGRNO 
            MachineAttendanceAdditionalDictionary['GRNO'] = GRNO
            MachineAttendanceAdditionalDictionary['Duty_Type'] = DutyType
            MachineAttendanceAdditionalDictionary['DutyDate'] = DutyDate
            MachineAttendanceAdditionalDictionary['DutyDateList'] = value
            try:
                ParshadList[NewGRNO,'Before Visit Machine Additional'].append(DutyDate)
            except:
                ParshadList[NewGRNO,'Before Visit Machine Additional'] = [DutyDate]
            db.tempMachineAttendanceAdditional.insert(**MachineAttendanceAdditionalDictionary)


        del MachineAttendanceAdditional


        #dSSdate = dworkbook.create_sheet(0)
        #dSSdate.title = "SSDate"
        #dSSdate.append(['SewadarNewID','DutyDate','Duty_Type'])
        #for row in SSDate:
        #    dSSdate.append([row.SewadarNewID , row.DutyDate , row.Duty_Type])


        #dMachineAttendance = dworkbook.create_sheet(0)
        #dMachineAttendance.title = "MachineAttendance"
        #dMachineAttendance.append(['GRNO','NewGRNO','DATETIME','TYPE'])
        #for row in MachineDate:
        #    dMachineAttendance.append([row.GRNO,row.NewGRNO,row.DATETIME,row.TYPE])

        MachineDifference = db(db.tempMachineAttendanceAdditional.id > 0).select()

        #dMachineDifference = dworkbook.create_sheet(0)
        #dMachineDifference.title = "MachineDifference"
        #dMachineDifference.append(['NewGRNO','GRNO','Duty_Type','DutyDate','DutyDateList'])
        #for row in MachineDifference:
        #    DutyDateList = ", ".join(map(str, row.DutyDateList))
        #    dMachineDifference.append([row.NewGRNO,row.GRNO,row.Duty_Type,row.DutyDate,DutyDateList])


        
        ParshadList['SEWADARS'] = set(ParshadList['SEWADARS'])

        for visit in xrange(0,TOTAL_VISIT):
            VCOUNT = 0
            VCOUNT = VISIT_DATES['V'+str(visit),'COUNT']

            vc = 'V'+str(visit)
            dc = 'D'+str(day)
            dc1 = 'D'+str(day+VCOUNT)
            SSDate = db((db.SSAttendanceDate.DutyDate >= VISIT_DATES[vc,dc]) & (db.SSAttendanceDate.DutyDate <= VISIT_DATES[vc,dc1])).select('SewadarNewID')
            MachineDate = db((db.MachineAttendance.DATETIME >= VISIT_DATES[vc,dc]) & (db.SSAttendanceDate.DutyDate <= VISIT_DATES[vc,dc1])).select('NewGRNO')
            
            for day in xrange(0,VCOUNT):
                for row in SSDate:
                    ParshadList['SEWADARS'].append(row.SewadarNewID)
                    try:
                        a = ParshadList[row.SewadarNewID,vc]
                    except:
                        ParshadList[row.SewadarNewID,vc] = VCOUNT - day

                for row in MachineDate:
                    ParshadList['SEWADARS'].append(row.NewGRNO)
                    try:
                        if ParshadList[row.NewGRNO,vc] < VCOUNT - day :
                            ParshadList[row.NewGRNO,vc] = VCOUNT - day
                            try:
                                ParshadList[row.NewGRNO,vc,'Additional Machine Visit Days'].append(vc + dc + ' onwards')
                            except:
                                ParshadList[row.NewGRNO,vc,'Additional Machine Visit Days'] = [vc + dc + ' onwards']
                    except:
                        ParshadList[row.NewGRNO,vc] = (VCOUNT - day)
                        try:
                            ParshadList[row.NewGRNO,vc,'Additional Machine Visit Days'].append(vc + dc + ' onwards')
                        except:
                            ParshadList[row.NewGRNO,vc,'Additional Machine Visit Days'] = [vc + dc + ' onwards']
                del SSDate
                del MachineDate



        redirect(URL(r=request, f='testredirect'))

        InitiatedList = db(db.InitiatedList.id > 0).select('NewGRNO','Status')

        for row in InitiatedList:
            ParshadList['SEWADARS'].append(row.NewGRNO)
            ParshadList[row.NewGRNO,'Initiation Status'] = row.Status

        
        for day in xrange(0,CURRENT_VISIT['COUNT']):
            SSDate = db((db.SSAttendanceDate.DutyDate >= CURRENT_VISIT['D'+str(day)]) & (db.SSAttendanceDate.DutyDate <= CURRENT_VISIT['D'+str(day+1)])).select('SewadarNewID','DutyDate','Duty_Type')
            MachineDate = db((db.MachineAttendance.DATETIME >= CURRENT_VISIT['D'+str(day)]) & (db.SSAttendanceDate.DutyDate <= CURRENT_VISIT['D'+str(day+1)])).select('GRNO','NewGRNO','DATETIME','TYPE')
            for row in SSDate:
                ParshadList['SEWADARS'].append(row.SewadarNewID)
                try:
                    a = ParshadList[row.SewadarNewID,'CV']
                except:
                    ParshadList[row.SewadarNewID,'CV'] = CURRENT_VISIT['COUNT'] - day

            for row in MachineDate:
                ParshadList['SEWADARS'].append(row.SewadarNewID)
                try:
                    if ParshadList[row.SewadarNewID,'CV'] < CURRENT_VISIT['COUNT'] - day :
                        ParshadList[row.SewadarNewID,'CV'] = CURRENT_VISIT['COUNT'] - day
                        try:
                            ParshadList[row.SewadarNewID,'CV','Additional Machine Current Visit Days'].append('D' +day+ ' onwards')
                        except:
                            ParshadList[row.SewadarNewID,'CV','Additional Machine Current Visit Days'] = ['D' +day+ ' onwards']
                except:
                    ParshadList[row.SewadarNewID,'CV'] = (CURRENT_VISIT['COUNT'] - day)
                    try:
                        ParshadList[row.SewadarNewID,'CV','Additional Machine Current Visit Days'].append('D' +day+ ' onwards')
                    except:
                        ParshadList[row.SewadarNewID,'CV','Additional Machine Current Visit Days'] = ['D' +day+ ' onwards']


        ParshadList['SEWADARS'] = set(ParshadList['SEWADARS'])


        dParshadList = dworkbook.create_sheet(0)
        dParshadList.title = "ParshadList"
        dParshadList.append(['NewGRNO','SSCount','CVCount','MachineBeforeVisitAdditional','MachineCVAdditional','MachinePreviousVisitAddition','InitiationStatus'])
        for Sewadar in ParshadList['SEWADARS']:
            try:
                SSCount = ParshadList[Sewadar,'SSCount'] 
            except:
                SSCount = "Not Found"
            try:
                CVCount = ParshadList[Sewadar,'CV'] 
            except:
                CVCount = "Not Found"
            try:
                MachineBeforeVisitAdditional = ", ".join(map(str, ParshadList[Sewadar,'Before Visit Machine Additional']))
            except:
                MachineBeforeVisitAdditional = "Not Found"
            try:
                MachinePreviousVisitAddition = ", ".join(map(str, ParshadList[Sewadar,'Additional Machine Visit Days']))
            except:
                MachinePreviousVisitAddition = "Not Found"
            try:
                MachingCVAdditional = ParshadList[Sewadar,'CV','Additional Machine Current Visit Days']
            except:
                MachingCVAdditional = "Not Found"
            try:
                InitiationStatus = ParshadList[Sewadar,'Initiation Status']
            except:
                InitiationStatus = "Not Found"
            dParshadList.append([Sewadar,SSCount,CVCount,MachineBeforeVisitAdditional,MachingCVAdditional,MachinePreviousVisitAddition,InitiationStatus])

        del ParshadList
        try:
            dworkbook.save(dpath)
        except:
            message = "File too big probably"

    return dict(form=form,message=message)

def testredirect():
    return locals()


#@auth.requires_login()
def AttendanceRegister():
    from gluon.sqlhtml import form_factory
    import datetime
    import time

    DateSelectedStart = datetime.datetime.today()
    DateSelectedEnd = datetime.datetime.today()
    Register = "EMPTY"
    DAttendanceRegisterTable = "EMPTY"
    datasource = {}
    columns = []
    orderby = []
    headers = {}
    DEV_DTY = ""
    TextMessage = "Jatha Wise Attendance Report"
    ReportDate =  0
    GENTS_REQUIRED = 30
    LADIES_REQURED = 36

    try:
        db.tempAttendanceRegisterTable.drop()
    except:
        pass

    SewadarDetails = {'Sewadars':[]}
#    form=form_factory(SQLField('JATHA','string',requires=IS_IN_DB(db,'MasterSheet.DEV_DTY','%(DEV_DTY)s')),SQLField('DateStart','date',default=datetime.datetime.today()-datetime.timedelta(days=31)),SQLField('DateEnd','date',default=datetime.datetime.today()),formname='DateSelect')
    my_list = db().select(db.MasterSheet.DEV_DTY, distinct=True).as_list()
    myjathalist = []
    for mydict in my_list:
        myjathalist.append(mydict['DEV_DTY'])

    myjathalist.sort()
    form=form_factory(SQLField('JATHA','string',requires=IS_IN_SET(myjathalist)),SQLField('DateStart','date',default=datetime.datetime.today()-datetime.timedelta(days=31)),SQLField('DateEnd','date',default=datetime.datetime.today()),formname='DateSelect')
    if form.accepts(request.vars,session,formname='DateSelect'):
        DateSelectedStart = request.vars.DateStart
        DateSelectedEnd = request.vars.DateEnd
        ReportDate = datetime.datetime.today()-datetime.timedelta(days=3001)

        SSDate = db((db.SSAttendanceDate.DutyDate >= DateSelectedStart) & (db.SSAttendanceDate.DutyDate <= DateSelectedEnd)).select()
        for Sewadar in SSDate:
            delta = Sewadar.DutyDate - datetime.datetime.strptime(DateSelectedStart,'%Y-%m-%d')
            SewadarDetails[Sewadar.SewadarNewID,'DAYS',delta.days] = Sewadar.DutyDate
            SewadarDetails[Sewadar.SewadarNewID,'TYPE',delta.days] = Sewadar.Duty_Type
            if ReportDate < Sewadar.DutyDate:
                ReportDate = Sewadar.DutyDate

        SSCount = db(db.SSAttendanceCount).select()
        for Sewadar in SSCount:
            SewadarDetails[Sewadar.NewID,'Gender'] = Sewadar.gender
            SewadarDetails[Sewadar.NewID,'TotalCount'] = Sewadar.Total
            SewadarDetails[Sewadar.NewID,'NAME'] = Sewadar.Name


        MasterTable = db(db.MasterSheet.GR_NO == 'SS015298').select()
        for Sewadar in MasterTable:
                 SewadarDetails[Sewadar.GR_NO,'SewadarNewID'] = Sewadar.SewadarNewID
                 SewadarDetails[Sewadar.SewadarNewID,'DEV_DTY'] = Sewadar.DEV_DTY

        DEV_DTY = request.vars.JATHA
#        if auth.user.username == 'admin':
#            DEV_DTY = request.vars.JATHA
#        else:
#            try:
#                DEV_DTY = SewadarDetails[auth.user.username,'DEV_DTY']
#            except:
#                #try:
#                    DEV_DTY = SewadarDetails[SewadarDetails[auth.user.username,'SewadarNewID'],'DEV_DTY']
#                #except:
#                #    TextMessage = 'This ID is not allocated to any Jatha'


        MasterTable = db(db.MasterSheet.DEV_DTY == DEV_DTY).select()
        for Sewadar in MasterTable:
            try:
                #Check if this ID was in SewaSamiti list
                SewadarDetails[Sewadar.SewadarNewID,'NAME']
                SewadarDetails['Sewadars'].append(Sewadar.SewadarNewID)
                SewadarDetails[Sewadar.SewadarNewID,'CANTEEN'] = Sewadar.CANTEEN
                SewadarDetails[Sewadar.SewadarNewID,'DEV_DTY'] = Sewadar.DEV_DTY
                SewadarDetails[Sewadar.SewadarNewID,'OLD_ID'] = Sewadar.GR_NO
            except:
                TextMessage = 'Sewadar ID not in list'


        Fields = [Field('D%02i' %i,'string') for i in range((datetime.datetime.strptime(DateSelectedEnd,'%Y-%m-%d') - datetime.datetime.strptime(DateSelectedStart,'%Y-%m-%d')).days)]


        #Now define the table
        db.define_table('tempAttendanceRegisterTable',
                Field('GR_NO','string'),
                Field('SewadarNewID','string'),
                Field('NAME','string'),
                Field('DEV_DTY','string'),
                Field('CANTEEN','string'),
                Field('TOTAL','integer'),
                Field('STATUS','string'),
                Field('REQD','integer'),
                Field('GENDER','string'),
                *Fields,
                migrate=True,
                redefine=True,
                format='%(SewadarNewID)s')

        db(db.tempAttendanceRegisterTable.id > 0).delete()

        mydict = {}
        mydict.clear()

        for Sewadar in SewadarDetails['Sewadars']:
            mydict['SewadarNewID'] = Sewadar
            mydict['GR_NO'] = SewadarDetails[Sewadar,'OLD_ID']
            mydict['NAME'] = SewadarDetails[Sewadar,'NAME']
            mydict['DEV_DTY'] = SewadarDetails[Sewadar,'DEV_DTY']
            mydict['CANTEEN'] = SewadarDetails[Sewadar,'CANTEEN']
            mydict['TOTAL'] = SewadarDetails[Sewadar,'TotalCount']
            mydict['GENDER'] = SewadarDetails[Sewadar,'Gender']
            mydict['STATUS'] = ''
            mydict['REQD'] = (GENTS_REQUIRED - SewadarDetails[Sewadar,'TotalCount']) if SewadarDetails[Sewadar,'Gender'] == 'Male' else (LADIES_REQURED - SewadarDetails[Sewadar,'TotalCount'])
            if mydict['REQD'] < 1:
                mydict['REQD'] = 0

            columns = ['tempAttendanceRegisterTable.SewadarNewID','tempAttendanceRegisterTable.GR_NO','tempAttendanceRegisterTable.NAME','tempAttendanceRegisterTable.DEV_DTY','tempAttendanceRegisterTable.CANTEEN','tempAttendanceRegisterTable.TOTAL','tempAttendanceRegisterTable.STATUS','tempAttendanceRegisterTable.REQD']
            orderby=columns

            headers = {'tempAttendanceRegisterTable.SewadarNewID':{'label':T('SewadarNewID'),'class':'','width':12,'truncate':12,'selected': False},
               'tempAttendanceRegisterTable.GR_NO':{'label':T('GR_NO'),'class':'','width':10,'truncate':10,'selected': False},
               'tempAttendanceRegisterTable.NAME':{'label':T('NAME'),'class':'','width':10,'truncate': 10,'selected': False},
               'tempAttendanceRegisterTable.DEV_DTY':{'label':T('DEV_DTY'),'class':'','width':11,'truncate': 11,'selected': False},
               'tempAttendanceRegisterTable.CANTEEN':{'label':T('CANTEEN'),'class':'','width':13,'truncate': 13,'selected': False},
               'tempAttendanceRegisterTable.TOTAL':{'label':T('TOTAL'),'class':'','width':3,'truncate': 3,'selected': False},
               'tempAttendanceRegisterTable.STATUS':{'label':T('STATUS'),'class':'','width':3,'truncate': 3,'selected': False},
               'tempAttendanceRegisterTable.REQD':{'label':T('REQD'),'class':'','width':4,'truncate': 4,'selected': False}
               }


            for i in range((datetime.datetime.strptime(DateSelectedEnd,'%Y-%m-%d') - datetime.datetime.strptime(DateSelectedStart,'%Y-%m-%d')).days):
                dateindex = datetime.datetime.strptime(DateSelectedStart,'%Y-%m-%d') + datetime.timedelta(i)
                columns.append('tempAttendanceRegisterTable.D%02i' %i)
                headers['tempAttendanceRegisterTable.D%02i' %i] = {'label':datetime.date.strftime(dateindex,'%d\n%m'),'class':'','width':'3','truncate': 3,'selected': False}
                try:
                    mydict['D%02i' %i] = SewadarDetails[Sewadar,'TYPE',i]
                except:
                    mydict['D%02i' %i] = ''

            db.tempAttendanceRegisterTable.insert(**mydict)

        Register = SQLTABLE(SSDate,headers='fieldname:capitalize')

        datasource = db(db.tempAttendanceRegisterTable).select(orderby=db.tempAttendanceRegisterTable.GR_NO)

        DAttendanceRegisterTable = SQLTABLE(datasource,columns=columns,headers=headers,orderby='GR_NO',_class='datatable')
        TextMessage = 'Attendance Report for ' + DEV_DTY +' from ' + datetime.date.strftime(datetime.datetime.strptime(DateSelectedStart, "%Y-%m-%d"),"%d-%b-%Y") + ' to ' + datetime.date.strftime(datetime.datetime.strptime(DateSelectedEnd, "%Y-%m-%d"),"%d-%b-%Y") + ' '
        ReportDate = datetime.date.strftime(ReportDate,"%d-%m-%Y")
        #tables = plugins.powerTable
        #tables.datasource = datasource
        #tables.uitheme = 'cupertino'
        #tables.dtfeatures['sPaginationType'] = 'two button'
        #tables.keycolumn = 'tempAttendanceRegisterTable.id'
        #tables.showkeycolumn = False
        #tables.columns = ['tempAttendanceRegisterTable.GR_NO','tempAttendanceRegisterTable.DEV_DTY','tempAttendanceRegisterTable.CANTEEN']
        #tables.hiddencolumns = ['tempAttendanceRegisterTable.GR_NO','tempAttendanceRegisterTable.DEV_DTY','tempAttendanceRegisterTable.CANTEEN']
        #tables.headers = 'labels'
        #created_table = tables.create()
        #DAttendanceRegisterTable = plugin_powerTable(datasource)


    return dict(DAttendanceRegisterTable=DAttendanceRegisterTable,form=form,TextMessage=TextMessage,ReportDate=ReportDate,SewadarDetails=SewadarDetails)

def user():
    """
    exposes:
    http://..../[app]/default/user/login
    http://..../[app]/default/user/logout
    http://..../[app]/default/user/register
    http://..../[app]/default/user/profile
    http://..../[app]/default/user/retrieve_password
    http://..../[app]/default/user/change_password
    http://..../[app]/default/user/manage_users (requires membership in
    use @auth.requires_login()
        @auth.requires_membership('group name')
        @auth.requires_permission('read','table name',record_id)
    to decorate functions that need access control
    """
    return dict(form=auth())

@cache.action()
def download():
    """
    allows downloading of uploaded files
    http://..../[app]/default/download/[filename]
    """
    return response.download(request, db)


def call():
    """
    exposes services. for example:
    http://..../[app]/default/call/jsonrpc
    decorate with @services.jsonrpc the functions to expose
    supports xml, json, xmlrpc, jsonrpc, amfrpc, rss, csv
    """
    return service()



def data():
    """
    http://..../[app]/default/data/tables
    http://..../[app]/default/data/create/[table]
    http://..../[app]/default/data/read/[table]/[id]
    http://..../[app]/default/data/update/[table]/[id]
    http://..../[app]/default/data/delete/[table]/[id]
    http://..../[app]/default/data/select/[table]
    http://..../[app]/default/data/search/[table]
    but URLs must be signed, i.e. linked with
      A('table',_href=URL('data/tables',user_signature=True))
    or with the signed load operator
      LOAD('default','data.load',args='tables',ajax=True,user_signature=True)
    """
    return dict(form=crud())

def MachineSewaSamitiDifference():
    from gluon.sqlhtml import form_factory
    import datetime
    import time

    DateSelectedStart = datetime.datetime.today()
    DateSelectedEnd = datetime.datetime.today()
    DAY_END_TIME = 19

    SSAttendance = 'Not queried'
    MachineAttendance = 'Not queried'
    MachineDifference = 'Not queried'

    form=form_factory(SQLField('DAY_END_TIME','string',default=19,requires=IS_IN_SET(range(0,23,1))),SQLField('DateStart','date',default=datetime.datetime.today()-datetime.timedelta(days=31)),SQLField('DateEnd','date',default=datetime.datetime.today()),formname='DateSelect')

    if form.accepts(request.vars,session,formname='DateSelect'):
        DateSelectedStart = request.vars.DateStart
        DateSelectedEnd = request.vars.DateEnd
        DAY_END_TIME = request.vars.DAY_END_TIME
        SSDate = db((db.SSAttendanceDate.DutyDate >= (datetime.datetime.strptime(DateSelectedStart,'%Y-%m-%d').replace(hour=0, minute=0, second=0, microsecond=0))) & (db.SSAttendanceDate.DutyDate <= (datetime.datetime.strptime(DateSelectedEnd,'%Y-%m-%d').replace(hour=23, minute=59, second=59, microsecond=999)))).select('SewadarNewID','DutyDate')
        MachineDate = db((db.MachineAttendance.DATETIME >= (datetime.datetime.strptime(DateSelectedStart,'%Y-%m-%d') - datetime.timedelta(hours=((24 - int(DAY_END_TIME)) % 24)))) & (db.MachineAttendance.DATETIME <= (datetime.datetime.strptime(DateSelectedEnd,'%Y-%m-%d') + datetime.timedelta(hours=int(DAY_END_TIME))))).select('GRNO','NewGRNO','DATETIME')

        SSAttendanceDictionary = {}
        for SSEntry in SSDate:
            SSAttendanceDictionary[SSEntry.SewadarNewID,Roundoffdate(SSEntry.DutyDate,23)] = SSEntry.DutyDate

        MachineAttendanceDictionary = {}
        MachineAttendanceAdditional = {}


        try:
            db.tempMachineAttendanceAdditional.drop()
        except:
            pass

        #Now define the table
        db.define_table('tempMachineAttendanceAdditional',
                Field('GRNO','string'),
                Field('NewGRNO','string'),
                Field('DutyDate','datetime'),
                Field('DutyDateList','list:string'),
                migrate=True,
                redefine=True,
                format='%(NewGRNO)s')

        db(db.tempMachineAttendanceAdditional.id > 0).delete()

        for MEntry in MachineDate:
            #Select earliest entry
            try:
                if MachineAttendanceDictionary[MEntry.NewGRNO,Roundoffdate(MEntry.DATETIME,DAY_END_TIME)] > MEntry.DATETIME:
                    MachineAttendanceDictionary[MEntry.NewGRNO,Roundoffdate(MEntry.DATETIME,DAY_END_TIME)] = MEntry.DATETIME
            except:
                MachineAttendanceDictionary[MEntry.NewGRNO,Roundoffdate(MEntry.DATETIME,DAY_END_TIME)] = MEntry.DATETIME

            try:
                MachineAttendanceDictionary[MEntry.NewGRNO,'LIST',Roundoffdate(MEntry.DATETIME,DAY_END_TIME)].append(MEntry.DATETIME)
            except:
                MachineAttendanceDictionary[MEntry.NewGRNO,'LIST',Roundoffdate(MEntry.DATETIME,DAY_END_TIME)] = [MEntry.DATETIME]

            #Find difference
            try:
                a = SSAttendanceDictionary[MEntry.NewGRNO,Roundoffdate(MEntry.DATETIME,DAY_END_TIME)]
            except:
                try:
                    MachineAttendanceAdditional[MEntry.NewGRNO,MEntry.GRNO,Roundoffdate(MEntry.DATETIME,DAY_END_TIME)].append(MEntry.DATETIME) 
                except:
                    MachineAttendanceAdditional[MEntry.NewGRNO,MEntry.GRNO,Roundoffdate(MEntry.DATETIME,DAY_END_TIME)] = [MEntry.DATETIME]



        for key, value in MachineAttendanceAdditional.iteritems():
            MachineAttendanceAdditionalDictionary = {}
            NewGRNO, GRNO, DutyDate = key
            MachineAttendanceAdditionalDictionary['NewGRNO'] = NewGRNO 
            MachineAttendanceAdditionalDictionary['GRNO'] = GRNO
            MachineAttendanceAdditionalDictionary['DutyDate'] = DutyDate
            MachineAttendanceAdditionalDictionary['DutyDateList'] = value
            db.tempMachineAttendanceAdditional.insert(**MachineAttendanceAdditionalDictionary)


        columns = ['SSAttendanceDate.SewadarNewID','SSAttendanceDate.DutyDate']
        orderby=columns

        headers = {'SSAttendanceDate.SewadarNewID':{'label':T('NewId'),'class':'','width':12,'truncate':12,'selected': False},
                   'SSAttendanceDate.DutyDate':{'label':T('Duty Date'),'class':'','width':12,'truncate':4,'selected': False}
                  }
        SSAttendance = SQLTABLE(SSDate)
        #SSAttendance = SQLTABLE(SSDate,columns=columns,headers=headers,orderby=orderby,_class='datatable')

        columns = ['MachineAttendance.GRNO','MachineAttendance.DUTY_DATE']
        orderby=columns

        headers = {'MachineAttendance.NewGRNO':{'label':T('NewId'),'class':'','width':12,'truncate':12,'selected': False},
                   'MachineAttendance.DATETIME':{'label':T('Duty Date'),'class':'','width':12,'truncate':4,'selected': False}
                  }

        MachineAttendance = SQLTABLE(MachineDate)
        #MachineAttendance = SQLTABLE(MachineDate,columns=columns,headers=headers,orderby=orderby,_class='datatable')

        #Machine Extra Attendance
        MachineDifference = SQLTABLE(db(db.tempMachineAttendanceAdditional.id > 0).select())

    return dict(form=form,MachineAttendance=MachineAttendance,SSAttendance=SSAttendance,MachineDifference=MachineDifference)



#@auth.requires_login()
def AttendanceRegister():
    from gluon.sqlhtml import form_factory
    import datetime
    import time

    DateSelectedStart = datetime.datetime.today()
    DateSelectedEnd = datetime.datetime.today()
    Register = "EMPTY"
    DAttendanceRegisterTable = "EMPTY"
    datasource = {}
    columns = []
    orderby = []
    headers = {}
    DEV_DTY = ""
    TextMessage = "Jatha Wise Attendance Report"
    ReportDate =  0
    GENTS_REQUIRED = 30
    LADIES_REQURED = 36

    try:
        db.tempAttendanceRegisterTable.drop()
    except:
        pass

    SewadarDetails = {'Sewadars':[]}
#    form=form_factory(SQLField('JATHA','string',requires=IS_IN_DB(db,'MasterSheet.DEV_DTY','%(DEV_DTY)s')),SQLField('DateStart','date',default=datetime.datetime.today()-datetime.timedelta(days=31)),SQLField('DateEnd','date',default=datetime.datetime.today()),formname='DateSelect')
    my_list = db().select(db.MasterSheet.DEV_DTY, distinct=True).as_list()
    myjathalist = []
    for mydict in my_list:
        myjathalist.append(mydict['DEV_DTY'])

    myjathalist.sort()
    form=form_factory(SQLField('JATHA','string',requires=IS_IN_SET(myjathalist)),SQLField('DateStart','date',default=datetime.datetime.today()-datetime.timedelta(days=31)),SQLField('DateEnd','date',default=datetime.datetime.today()),formname='DateSelect')
    if form.accepts(request.vars,session,formname='DateSelect'):
        DateSelectedStart = request.vars.DateStart
        DateSelectedEnd = request.vars.DateEnd
        ReportDate = datetime.datetime.today()-datetime.timedelta(days=3001)

        SSDate = db((db.SSAttendanceDate.DutyDate >= DateSelectedStart) & (db.SSAttendanceDate.DutyDate <= DateSelectedEnd)).select()
        for Sewadar in SSDate:
            delta = Sewadar.DutyDate - datetime.datetime.strptime(DateSelectedStart,'%Y-%m-%d')
            SewadarDetails[Sewadar.SewadarNewID,'DAYS',delta.days] = Sewadar.DutyDate
            SewadarDetails[Sewadar.SewadarNewID,'TYPE',delta.days] = Sewadar.Duty_Type
            if ReportDate < Sewadar.DutyDate:
                ReportDate = Sewadar.DutyDate

        SSCount = db(db.SSAttendanceCount).select()
        for Sewadar in SSCount:
            SewadarDetails[Sewadar.NewID,'Gender'] = Sewadar.gender
            SewadarDetails[Sewadar.NewID,'TotalCount'] = Sewadar.Total
            SewadarDetails[Sewadar.NewID,'NAME'] = Sewadar.Name


        MasterTable = db(db.MasterSheet.GR_NO == 'SS015298').select()
        for Sewadar in MasterTable:
                 SewadarDetails[Sewadar.GR_NO,'SewadarNewID'] = Sewadar.SewadarNewID
                 SewadarDetails[Sewadar.SewadarNewID,'DEV_DTY'] = Sewadar.DEV_DTY

        DEV_DTY = request.vars.JATHA
#        if auth.user.username == 'admin':
#            DEV_DTY = request.vars.JATHA
#        else:
#            try:
#                DEV_DTY = SewadarDetails[auth.user.username,'DEV_DTY']
#            except:
#                #try:
#                    DEV_DTY = SewadarDetails[SewadarDetails[auth.user.username,'SewadarNewID'],'DEV_DTY']
#                #except:
#                #    TextMessage = 'This ID is not allocated to any Jatha'


        MasterTable = db(db.MasterSheet.DEV_DTY == DEV_DTY).select()
        for Sewadar in MasterTable:
            try:
                #Check if this ID was in SewaSamiti list
                SewadarDetails[Sewadar.SewadarNewID,'NAME']
                SewadarDetails['Sewadars'].append(Sewadar.SewadarNewID)
                SewadarDetails[Sewadar.SewadarNewID,'CANTEEN'] = Sewadar.CANTEEN
                SewadarDetails[Sewadar.SewadarNewID,'DEV_DTY'] = Sewadar.DEV_DTY
                SewadarDetails[Sewadar.SewadarNewID,'OLD_ID'] = Sewadar.GR_NO
            except:
                TextMessage = 'Sewadar ID not in list'


        Fields = [Field('D%02i' %i,'string') for i in range((datetime.datetime.strptime(DateSelectedEnd,'%Y-%m-%d') - datetime.datetime.strptime(DateSelectedStart,'%Y-%m-%d')).days)]


        #Now define the table
        db.define_table('tempAttendanceRegisterTable',
                Field('GR_NO','string'),
                Field('SewadarNewID','string'),
                Field('NAME','string'),
                Field('DEV_DTY','string'),
                Field('CANTEEN','string'),
                Field('TOTAL','integer'),
                Field('STATUS','string'),
                Field('REQD','integer'),
                Field('GENDER','string'),
                *Fields,
                migrate=True,
                redefine=True,
                format='%(SewadarNewID)s')

        db(db.tempAttendanceRegisterTable.id > 0).delete()

        mydict = {}
        mydict.clear()

        for Sewadar in SewadarDetails['Sewadars']:
            mydict['SewadarNewID'] = Sewadar
            mydict['GR_NO'] = SewadarDetails[Sewadar,'OLD_ID']
            mydict['NAME'] = SewadarDetails[Sewadar,'NAME']
            mydict['DEV_DTY'] = SewadarDetails[Sewadar,'DEV_DTY']
            mydict['CANTEEN'] = SewadarDetails[Sewadar,'CANTEEN']
            mydict['TOTAL'] = SewadarDetails[Sewadar,'TotalCount']
            mydict['GENDER'] = SewadarDetails[Sewadar,'Gender']
            mydict['STATUS'] = ''
            mydict['REQD'] = (GENTS_REQUIRED - SewadarDetails[Sewadar,'TotalCount']) if SewadarDetails[Sewadar,'Gender'] == 'Male' else (LADIES_REQURED - SewadarDetails[Sewadar,'TotalCount'])
            if mydict['REQD'] < 1:
                mydict['REQD'] = 0

            columns = ['tempAttendanceRegisterTable.SewadarNewID','tempAttendanceRegisterTable.GR_NO','tempAttendanceRegisterTable.NAME','tempAttendanceRegisterTable.DEV_DTY','tempAttendanceRegisterTable.CANTEEN','tempAttendanceRegisterTable.TOTAL','tempAttendanceRegisterTable.STATUS','tempAttendanceRegisterTable.REQD']
            orderby=columns

            headers = {'tempAttendanceRegisterTable.SewadarNewID':{'label':T('SewadarNewID'),'class':'','width':12,'truncate':12,'selected': False},
               'tempAttendanceRegisterTable.GR_NO':{'label':T('GR_NO'),'class':'','width':10,'truncate':10,'selected': False},
               'tempAttendanceRegisterTable.NAME':{'label':T('NAME'),'class':'','width':10,'truncate': 10,'selected': False},
               'tempAttendanceRegisterTable.DEV_DTY':{'label':T('DEV_DTY'),'class':'','width':11,'truncate': 11,'selected': False},
               'tempAttendanceRegisterTable.CANTEEN':{'label':T('CANTEEN'),'class':'','width':13,'truncate': 13,'selected': False},
               'tempAttendanceRegisterTable.TOTAL':{'label':T('TOTAL'),'class':'','width':3,'truncate': 3,'selected': False},
               'tempAttendanceRegisterTable.STATUS':{'label':T('STATUS'),'class':'','width':3,'truncate': 3,'selected': False},
               'tempAttendanceRegisterTable.REQD':{'label':T('REQD'),'class':'','width':4,'truncate': 4,'selected': False}
               }


            for i in range((datetime.datetime.strptime(DateSelectedEnd,'%Y-%m-%d') - datetime.datetime.strptime(DateSelectedStart,'%Y-%m-%d')).days):
                dateindex = datetime.datetime.strptime(DateSelectedStart,'%Y-%m-%d') + datetime.timedelta(i)
                columns.append('tempAttendanceRegisterTable.D%02i' %i)
                headers['tempAttendanceRegisterTable.D%02i' %i] = {'label':datetime.date.strftime(dateindex,'%d\n%m'),'class':'','width':'3','truncate': 3,'selected': False}
                try:
                    mydict['D%02i' %i] = SewadarDetails[Sewadar,'TYPE',i]
                except:
                    mydict['D%02i' %i] = ''

            db.tempAttendanceRegisterTable.insert(**mydict)

        Register = SQLTABLE(SSDate,headers='fieldname:capitalize')

        datasource = db(db.tempAttendanceRegisterTable).select(orderby=db.tempAttendanceRegisterTable.GR_NO)

        DAttendanceRegisterTable = SQLTABLE(datasource,columns=columns,headers=headers,orderby='GR_NO',_class='datatable')
        TextMessage = 'Attendance Report for ' + DEV_DTY +' from ' + datetime.date.strftime(datetime.datetime.strptime(DateSelectedStart, "%Y-%m-%d"),"%d-%b-%Y") + ' to ' + datetime.date.strftime(datetime.datetime.strptime(DateSelectedEnd, "%Y-%m-%d"),"%d-%b-%Y") + ' '
        ReportDate = datetime.date.strftime(ReportDate,"%d-%m-%Y")
        #tables = plugins.powerTable
        #tables.datasource = datasource
        #tables.uitheme = 'cupertino'
        #tables.dtfeatures['sPaginationType'] = 'two button'
        #tables.keycolumn = 'tempAttendanceRegisterTable.id'
        #tables.showkeycolumn = False
        #tables.columns = ['tempAttendanceRegisterTable.GR_NO','tempAttendanceRegisterTable.DEV_DTY','tempAttendanceRegisterTable.CANTEEN']
        #tables.hiddencolumns = ['tempAttendanceRegisterTable.GR_NO','tempAttendanceRegisterTable.DEV_DTY','tempAttendanceRegisterTable.CANTEEN']
        #tables.headers = 'labels'
        #created_table = tables.create()
        #DAttendanceRegisterTable = plugin_powerTable(datasource)


    return dict(DAttendanceRegisterTable=DAttendanceRegisterTable,form=form,TextMessage=TextMessage,ReportDate=ReportDate,SewadarDetails=SewadarDetails)

def user():
    """
    exposes:
    http://..../[app]/default/user/login
    http://..../[app]/default/user/logout
    http://..../[app]/default/user/register
    http://..../[app]/default/user/profile
    http://..../[app]/default/user/retrieve_password
    http://..../[app]/default/user/change_password
    http://..../[app]/default/user/manage_users (requires membership in
    use @auth.requires_login()
        @auth.requires_membership('group name')
        @auth.requires_permission('read','table name',record_id)
    to decorate functions that need access control
    """
    return dict(form=auth())

@cache.action()
def download():
    """
    allows downloading of uploaded files
    http://..../[app]/default/download/[filename]
    """
    return response.download(request, db)


def call():
    """
    exposes services. for example:
    http://..../[app]/default/call/jsonrpc
    decorate with @services.jsonrpc the functions to expose
    supports xml, json, xmlrpc, jsonrpc, amfrpc, rss, csv
    """
    return service()



def data():
    """
    http://..../[app]/default/data/tables
    http://..../[app]/default/data/create/[table]
    http://..../[app]/default/data/read/[table]/[id]
    http://..../[app]/default/data/update/[table]/[id]
    http://..../[app]/default/data/delete/[table]/[id]
    http://..../[app]/default/data/select/[table]
    http://..../[app]/default/data/search/[table]
    but URLs must be signed, i.e. linked with
      A('table',_href=URL('data/tables',user_signature=True))
    or with the signed load operator
      LOAD('default','data.load',args='tables',ajax=True,user_signature=True)
    """
    return dict(form=crud())
