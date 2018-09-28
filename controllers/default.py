# -*- coding: utf-8 -*-
import CommonFunctions

from gluon.tools import Mail
mail = Mail()
mail.settings.server = 'smtp.gmail.com:25'
mail.settings.sender = 'softwareattendance@gmail.com'
mail.settings.login = 'softwareattendance:hajari123'

@auth.requires_login()
def index():
    response.flash = T("Welcome to Attendance Software!")
    return dict(message=T('Attendance Software'))

def Roundoffdate(DATETIME, CUTOFF):
    import datetime
    RoundedDate = DATETIME
    if (DATETIME.hour < int(CUTOFF)) or (DATETIME.hour == int(CUTOFF) and DATETIME.minute == 0):
       RoundedDate = RoundedDate.replace(hour=0, minute=0, second=0, microsecond=0)
    else:
       RoundedDate = (RoundedDate + datetime.timedelta(days=1)).replace(hour=0, minute=0, second=0, microsecond=0)
    return RoundedDate

def MachineCARDNOtoGRNO(CARDNO):
    import re
    CARDNO = re.sub("^5", "PR0012", CARDNO)
    CARDNO = re.sub("^4", "BH0011", CARDNO)
    CARDNO = re.sub("^3", "SS", CARDNO)
    CARDNO = re.sub("^10", "G", CARDNO)
    CARDNO = re.sub("^20", "L", CARDNO)
    return CARDNO

def CARDNOtoGRNO(CARDNO):
    import re
    CARDNO = re.sub("^5", "PR0012", CARDNO)
    CARDNO = re.sub("^4", "BH0011", CARDNO)
    CARDNO = re.sub("^03", "SS", CARDNO)
    CARDNO = re.sub("^010", "G", CARDNO)
    CARDNO = re.sub("^020", "L", CARDNO)
    return CARDNO

def GRNOtoCARDNO(GRNO):
    import re
    prog = re.compile('^PR0012')
    if prog.match(GRNO):
        GRNO = re.sub("^PR0012","5", GRNO)

    prog = re.compile('^BH0011')
    if prog.match(GRNO):
        GRNO = re.sub("^BH0011","4", GRNO)

    prog = re.compile('^SS')
    if prog.match(GRNO):
        GRNO = re.sub("^SS","3", GRNO)

    prog = re.compile('^G')
    if prog.match(GRNO):
        GRNO = re.sub("^G","10", GRNO)

    prog = re.compile('^L')
    if prog.match(GRNO):
        GRNO = re.sub("^L","20", GRNO)

    return GRNO


@auth.requires_login()
def view_sewadar():
    import cgi
    import pandas as pd
    import os
    from gluon.tools import PluginManager
    import datetime
    SewaSamitiCountRecords = 'No Valid record selected'
    SewaSamitiDatesRecords = 'No Valid record selected'
    MasterRecords = 'No Valid record selected'
    MachineRecords = 'No Valid record selected'
    TextMessage = ''

    ParshadStatus = 'Not fetched'
    ParshadRemarks = 'Not fetched'

    GENTS_REQUIRED = 30
    LADIES_REQUIRED = 36
    SS_GENTS_REQUIRED = 30
    SS_LADIES_REQUIRED = 36
    DATE_PRINT = datetime.datetime.strptime('2016-11-28', "%Y-%m-%d")

    LastUpdated = datetime.datetime(2000, 1, 1)
    datasource  = db(db.LocalVariables.id > 0).select()
    for data in datasource:
        LastUpdated = data['LastUpdated']
    updation_message = "Last Updated On: " + datetime.datetime.strftime(LastUpdated,'%d-%b-%Y')

    WWSchedule = {}


    response.subtitle = "Sewadar Details"
    from gluon.sqlhtml import form_factory
    form=form_factory(SQLField('SewadarId','string',default="",requires=IS_NOT_EMPTY()),formname='UserDetails')
    if form.accepts(request.vars,session,formname='UserDetails'):
        GRNO = '%' + (request.vars.SewadarId).upper() + '%'
        request.flash = GRNO
        datasource  = db((db.SSAttendanceCount.OldSewadarid.like(GRNO)) | (db.SSAttendanceCount.NewID.like(GRNO))).select()
        print "SSCount read"
        SewadarNewID = ''
        for Sewadar in datasource:
            SewadarNewID = Sewadar.NewID
            if ((Sewadar.gender == 'Male')) & (Sewadar.Total < GENTS_REQUIRED):
                if GENTS_REQUIRED - Sewadar.Total > SS_GENTS_REQUIRED:
                    TextMessage = "Attendance Short by : " + str(SS_GENTS_REQUIRED)
                else:
                    TextMessage = "Attendance Short by : " + str(GENTS_REQUIRED - Sewadar.Total)
                    TextMessage = HTML(BODY(H2(T(TextMessage),_style="color: red;")))
            elif ((Sewadar.gender == 'Female')) & (Sewadar.Total < LADIES_REQUIRED):
                if LADIES_REQUIRED - Sewadar.Total > SS_LADIES_REQUIRED:
                    TextMessage = "Attendance Short by : " + str(SS_LADIES_REQUIRED)
                else:
                    TextMessage = "Attendance Short by : " + str(LADIES_REQUIRED - Sewadar.Total)
                    TextMessage = HTML(BODY(H2(T(TextMessage),_style="color: red;")))
            elif ((Sewadar.gender == 'Male') | (Sewadar.gender == 'Female')):
                TextMessage = "Attendance Complete"
                TextMessage = HTML(BODY(H2(T(TextMessage),_style="color: green;")))
            else:
                TextMessage = "Error in Field"

        columns = ['SSAttendanceCount.NewID','SSAttendanceCount.OldSewadarid','SSAttendanceCount.Name']
        orderby=columns

        headers = {'SSAttendanceCount.NewID':{'label':T('NewId'),'class':'','width':12,'truncate':12,'selected': False},
                   'SSAttendanceCount.OldSewadarid':{'label':T('OldId'),'class':'','width':12,'truncate':12,'selected': False},
                   'SSAttendanceCount.Name':{'label':T('Name                  '),'class':'','width':30,'truncate':30,'selected': False}
                  }

        SewaSamitiCountRecords = SQLTABLE(datasource,columns=columns,headers=headers,orderby=orderby,_class='datatable')


        columns = ['SSAttendanceDate.DutyDate','SSAttendanceDate.Duty_Type']
        orderby=columns

        headers = {'SSAttendanceDate.DutyDate':{'label':T('Date'),'class':'','width':12,'truncate':12,'selected': False},
                   'SSAttendanceDate.Duty_Type':{'label':T('Type'),'class':'','width':2,'truncate':10,'selected': False}
                  }

        db.SSAttendanceDate.DutyDate.represent = lambda value, row: value.strftime("%d/%m/%Y")


        datasource = db((db.SSAttendanceDate.SewadarNewID.like(SewadarNewID)) & (db.SSAttendanceDate.DutyDate > DATE_PRINT )).select(orderby=~db.SSAttendanceDate.DutyDate)
        print "SSDate read"

        #datasource1 = db(db.SSAttendanceDate.OldSewadarID == "IAMNONEXISTENT").select()
        #for row in datasource:
        #    if row.DutyDate > DATE_PRINT:
        #        datasource1.records.append(row)



        SewaSamitiDatesRecords = SQLTABLE(datasource,columns=columns,headers=headers,orderby=orderby,_class='datatable')
        rows = db(db.SSAttendanceCount.NewID.like(SewadarNewID)).select()

        datasource = db(db.MasterSheet.SewadarNewID.like(SewadarNewID)).select()
        print "MastyerrSheet read"

        columns = ['MasterSheet.CANTEEN','MasterSheet.DEV_DTY']
        orderby=columns

        headers = {'MasterSheet.CANTEEN':{'label':T('CANTEEN'),'class':'','width':12,'truncate':12,'selected': False},
                   'MasterSheet.DEV_DTY':{'label':T('JATHA'),'class':'','width':35,'truncate':35,'selected': False}
                  }

        MasterRecords = SQLTABLE(datasource,columns=columns,headers=headers,orderby=orderby,_class='datatable')

#    if auth.user.username == 'admin':
#        MachineRecords = SQLTABLE(db(db.RawData.CARDNO == GRNOtoCARDNO(GRNO)).select('OFFICEPUNCH',orderby=~db.RawData.OFFICEPUNCH),headers='fieldname:capitalize')
#    else:
        MachineRecords = ''
        for row in datasource:
            Jatha = row.DEV_DTY
            if row.Gender == 'M':
                WWSchedule = db(db.WWSchedule.Jatha == Jatha).select()
            else:
                WWSchedule = db(db.WWScheduleLadies.Jatha == Jatha).select()


        df = pd.read_excel('/home/rootcat/new_web2py/web2py/applications/AttendanceSoftware/private/df_ParshadStatus.xlsx',sheetname='Sheet1',index_col=0)
        try:
            ParshadStatus = str(df.at[SewadarNewID.replace('BH0011',''),'ParshadStatus'])
        except:
            ParshadStatus = "Not Found"

        if (ParshadStatus == "Waiting") or (ParshadStatus == "OK"):
            ParshadStatus = HTML(BODY(H2(T('OK'),_style="color: green;")))
        else:
            ParshadStatus = HTML(BODY(H2(T(ParshadStatus),_style="color: red;")))

        try:
            ParshadRemarks = df.at[SewadarNewID.replace('BH0011',''),'ParshadRemark'].split('.')
        except Exception,e:
            ParshadRemarks = ["Not Found" + str(e)]
        i = 0
        for each in ParshadRemarks:
            if (ParshadRemarks[i] == "OK") | (ParshadRemarks[i] == "CORE TEAM"):
                ParshadRemarks[i] =HTML(BODY(H3(T(ParshadRemarks[i]),_style="color: green;")))
            else:
                ParshadRemarks[i] =HTML(BODY(H3(T(ParshadRemarks[i]),_style="color: red;")))
            i = i + 1

    return dict(updation_message=updation_message,TextMessage=TextMessage,form=form,MasterRecords=MasterRecords,MachineRecords=MachineRecords,SewaSamitiDatesRecords=SewaSamitiDatesRecords,SewaSamitiCountRecords=SewaSamitiCountRecords,WWSchedule=WWSchedule,ParshadStatus=ParshadStatus,ParshadRemarks=ParshadRemarks)


@auth.requires(auth.user_id == 3)
def update():

    import os
    response.subtitle = "Upload excel file "
    from gluon.sqlhtml import form_factory
    form2=form_factory(SQLField('Master_xls','upload',uploadfolder='temporary'),formname='MASTER')
    if form2.accepts(request.vars,session,formname='MASTER'):
        request.flash='Received: %s'%request.vars.Master_xls
        path = os.path.join(request.folder,'private','Master_xls.xlsx')
        import shutil
        shutil.copyfileobj(request.vars.Master_xls.file,open(path, 'wb'))
        #Then redirect to the next screen (or do the processing now)
        redirect(URL(r=request, f='uploaddata_master'))


    form3=form_factory(SQLField('SSAttendanceDates_xls','upload',uploadfolder='temporary'),SQLField('RunNow','string',default='No',requires=IS_IN_SET(['Yes','No'])),formname='SSATTENDANCE')
    if form3.accepts(request.vars,session,formname='SSATTENDANCE'):
        request.flash='Received: %s' % request.vars.SSAttendanceDates_xls
        path = os.path.join(request.folder,'private','SSAttendanceDates_xls.xlsx')
        import shutil
        shutil.copyfileobj(request.vars.SSAttendanceDates_xls.file,open(path, 'wb'))
        from datetime import timedelta as timed
        scheduler.queue_task('uploaddata_SSAttendance',
            start_time=request.now + timed(seconds=30),
            timeout = 6000)

        #Then redirect to the next screen (or do the processing now)
        #if request.vars.RunNow == 'Yes':
        #    redirect(URL(r=request, f='uploaddata_SSAttendance'))

    form4=form_factory(SQLField('SSAttendanceCount_xls','upload',uploadfolder='temporary'),formname='SSATTENDANCECOUNT')
    if form4.accepts(request.vars,session,formname='SSATTENDANCECOUNT'):
        request.flash='Received: %s'%request.vars.SSAttendanceCount_xls
        path = os.path.join(request.folder,'private','SSAttendanceCount_xls.xlsx')
        import shutil
        shutil.copyfileobj(request.vars.SSAttendanceCount_xls.file,open(path, 'wb'))
        #Then redirect to the next screen (or do the processing now)
        redirect(URL(r=request, f='uploaddata_SSAttendanceCount'))

    form5=form_factory(SQLField('CardList_xls','upload',uploadfolder='temporary'),formname='CARDLIST')
    if form5.accepts(request.vars,session,formname='CARDLIST'):
        request.flash='Received: %s'%request.vars.CardList_xls
        path = os.path.join(request.folder,'private','CardList_xls.xlsx')
        import shutil
        shutil.copyfileobj(request.vars.CardList_xls.file,open(path, 'wb'))
        #Then redirect to the next screen (or do the processing now)
        redirect(URL(r=request, f='uploaddata_CardList'))

    form6=form_factory(SQLField('STARLINK_zip','upload',uploadfolder='temporary'),formname='STARLINK')
    if form6.accepts(request.vars,session,formname='STARLINK'):
        request.flash='Received: %s'%request.vars.STARLINK_zip
        path = os.path.join(request.folder,'private','STARLINK.zip')
        import shutil
        shutil.copyfileobj(request.vars.STARLINK_zip.file,open(path, 'wb'))
        #Then redirect to the next screen (or do the processing now)
        redirect(URL(r=request, f='uploaddata_MachineAttendance'))


    form7=form_factory(SQLField('MACHINE_xlsx','upload',uploadfolder='temporary'),formname='MachineManual')
    if form7.accepts(request.vars,session,formname='MachineManual'):
        request.flash='Received: %s'%request.vars.MACHINE_xlsx
        path = os.path.join(request.folder,'private','MachineAttendance.xlsx')
        import shutil
        shutil.copyfileobj(request.vars.MACHINE_xlsx.file,open(path, 'wb'))
        #Then redirect to the next screen (or do the processing now)
        redirect(URL(r=request, f='uploaddata_MachineManualAttendance'))

    form8=form_factory(SQLField('Initiation_xlsx','upload',uploadfolder='temporary'),formname='InitiationList')
    if form8.accepts(request.vars,session,formname='InitiatedList'):
        request.flash='Received: %s'%request.vars.Initiation_xlsx
        path = os.path.join(request.folder,'private','InitiatedList.xlsx')
        import shutil
        shutil.copyfileobj(request.vars.Initiation_xlsx.file,open(path, 'wb'))
        #Then redirect to the next screen (or do the processing now)
        redirect(URL(r=request, f='uploaddata_InitiationList'))

    form9=form_factory(SQLField('SSParshadStatus_xlsx','upload',uploadfolder='temporary'),formname='SSParshadStatus')
    if form9.accepts(request.vars,session,formname='SSParshadStatus'):
        request.flash='Received: %s'%request.vars.SSParshadStatus_xlsx
        path = os.path.join(request.folder,'private','SSParshadStatus.xlsx')
        import shutil
        shutil.copyfileobj(request.vars.SSParshadStatus_xlsx.file,open(path, 'wb'))
        #Then redirect to the next screen (or do the processing now)
        redirect(URL(r=request, f='uploaddata_SSParshadStatus'))

    form10=form_factory(SQLField('PreviousParshadList_xlsx','upload',uploadfolder='temporary'),formname='PreviousParshadStatus')
    if form10.accepts(request.vars,session,formname='PreviousParshadStatus'):
        request.flash='Received: %s'%request.vars.PreviousParshadList_xlsx
        path = os.path.join(request.folder,'private','PreviousParshadList.xlsx')
        import shutil
        shutil.copyfileobj(request.vars.PreviousParshadList_xlsx.file,open(path, 'wb'))
        #Then redirect to the next screen (or do the processing now)
        redirect(URL(r=request, f='uploaddata_PreviousParshadList'))

    form11=form_factory(SQLField('AllCardList_xls','upload',uploadfolder='temporary'),formname='ALLCARDLIST')
    if form11.accepts(request.vars,session,formname='ALLCARDLIST'):
        request.flash='Received: %s'%request.vars.AllCardList_xls
        path = os.path.join(request.folder,'private','AllCardList_xls.xlsx')
        import shutil
        shutil.copyfileobj(request.vars.AllCardList_xls.file,open(path, 'wb'))
        #Then redirect to the next screen (or do the processing now)
        redirect(URL(r=request, f='uploaddata_AllCardList'))

    form12=form_factory(SQLField('MailException_xls','upload',uploadfolder='temporary'),formname='MailException')
    if form12.accepts(request.vars,session,formname='MailException'):
        request.flash='Received: %s'%request.vars.MailException_xls
        path = os.path.join(request.folder,'private','MailException_xls.xlsx')
        import shutil
        shutil.copyfileobj(request.vars.MailException_xls.file,open(path, 'wb'))
        #Then redirect to the next screen (or do the processing now)
        redirect(URL(r=request, f='uploaddata_MailException'))

    form13=form_factory(SQLField('WWLadies','upload',uploadfolder='temporary'),formname='WWDuty')
    if form13.accepts(request.vars,session,formname='WWLadies'):
        request.flash='Received: %s'%request.vars.WWLadies
        path = os.path.join(request.folder,'private','LadiesWW.xlsx')
        import shutil
        shutil.copyfileobj(request.vars.WWLadies.file,open(path, 'wb'))
        #Then redirect to the next screen (or do the processing now)
        redirect(URL(r=request, f='uploaddata_WWScheduleLadies'))

    return dict(form2=form2,form3=form3,form4=form4,form5=form5,form6=form6,form7=form7,form8=form8,form9=form9,form10=form10,form11=form11,form12=form12,form13=form13)


def uploaddata_MailException():
    import xlrd
    import os
    path = os.path.join(request.folder,'private','MailException_xls.xlsx')
    workbook = xlrd.open_workbook(path)

    worksheet = workbook.sheet_names()
    #Attempt with any worksheet found
    worksheet = workbook.sheet_by_name(worksheet[0])
    header_row = 0
    import re, time

    #my headers are the headers in DB
    myheaders = ['NewGRNO','Status']

    #headers are the name used in XLS
    headers = ['NewGRNO','Status']
    header_cells_in_xls = [worksheet.cell_value(0,idx) for idx in range(worksheet.ncols)]

    try:
        db(db.ParshadMailException.id > 0).delete()
    except:
        pass


    indexcol = [];
    for header_cell in headers:
            try:
                indexcol.append(header_cells_in_xls.index(header_cell))
            except Exception,e :
                print 'Header not found in xls: %s' %e
            else:
                print 'OK'

    print "INDEXCOL = ", indexcol
    row_dict_list = []

    throttle = 1
    for row in xrange(header_row+1, worksheet.nrows):
       i=0
       row_dict = {}

       for col in indexcol:
           cell_type = worksheet.cell_type(row, col)
           if cell_type == xlrd.XL_CELL_EMPTY:
               value = None
           elif cell_type == xlrd.XL_CELL_TEXT:
              value = worksheet.cell_value(row, col)
           elif cell_type == xlrd.XL_CELL_NUMBER:
              value = float(worksheet.cell_value(row,col))
           elif cell_type == xlrd.XL_CELL_DATE:
              from datetime import datetime
              value = datetime(*xlrd.xldate_as_tuple(worksheet.cell_value(row, col), workbook.datemode))
           elif cell_type == xlrd.XL_CELL_BOOLEAN:
              value = bool(worksheet.cell_value(row, col))
           else:
              value = worksheet.cell_value(row, col)

           if i<1:
               value = re.sub("\s*","",value)

           if value == None:
               row_dict[myheaders[i]] = "YES"
           else:
               row_dict[myheaders[i]] = value

           i=i+1
       try:
           db.ParshadMailException.insert(**row_dict)
       except:
           duplicate = 1

    response.flash = T("Entry Successful!")

    return locals()

def uploaddata_PreviousParshadList():
    import xlrd
    import os
    path = os.path.join(request.folder,'private','PreviousParshadList.xlsx')
    workbook = xlrd.open_workbook(path)

    worksheet = workbook.sheet_names()
    #Attempt with any worksheet found
    worksheet = workbook.sheet_by_name(worksheet[0])
    header_row = 0
    import re, time

    #my headers are the headers in DB
    myheaders = ['NewGRNO','Status']

    #headers are the name used in XLS
    headers = ['NewGRNO','Status']
    header_cells_in_xls = [worksheet.cell_value(0,idx) for idx in range(worksheet.ncols)]

    try:
        db(db.PreviousParshadList.id > 0).delete()
    except:
        pass


    indexcol = [];
    for header_cell in headers:
            try:
                indexcol.append(header_cells_in_xls.index(header_cell))
            except Exception,e :
                print 'Header not found in xls: %s' %e
            else:
                print 'OK'

    print "INDEXCOL = ", indexcol
    row_dict_list = []

    throttle = 1
    for row in xrange(header_row+1, worksheet.nrows):
       i=0
       row_dict = {}

       for col in indexcol:
           cell_type = worksheet.cell_type(row, col)
           if cell_type == xlrd.XL_CELL_EMPTY:
               value = None
           elif cell_type == xlrd.XL_CELL_TEXT:
              value = worksheet.cell_value(row, col)
           elif cell_type == xlrd.XL_CELL_NUMBER:
              value = float(worksheet.cell_value(row,col))
           elif cell_type == xlrd.XL_CELL_DATE:
              from datetime import datetime
              value = datetime(*xlrd.xldate_as_tuple(worksheet.cell_value(row, col), workbook.datemode))
           elif cell_type == xlrd.XL_CELL_BOOLEAN:
              value = bool(worksheet.cell_value(row, col))
           else:
              value = worksheet.cell_value(row, col)

           if i<1:
               value = re.sub("\s*","",value)

           if value == None:
               row_dict[myheaders[i]] = "YES"
           else:
               row_dict[myheaders[i]] = value

           i=i+1
       try:
           db.PreviousParshadList.insert(**row_dict)
       except:
           duplicate = 1

    response.flash = T("Entry Successful!")

    return locals()

def uploaddata_InitiationList():
    import xlrd
    import os
    path = os.path.join(request.folder,'private','InitiatedList.xlsx')
    workbook = xlrd.open_workbook(path)

    worksheet = workbook.sheet_names()
    #Attempt with any worksheet found
    worksheet = workbook.sheet_by_name(worksheet[0])
    header_row = 0
    import re, time

    #my headers are the headers in DB
    myheaders = ['NewGRNO','Status']

    #headers are the name used in XLS
    headers = ['NewGRNO','Status']
    header_cells_in_xls = [worksheet.cell_value(0,idx) for idx in range(worksheet.ncols)]

    try:
        db(db.InitiatedList.id > 0).delete()
    except:
        pass


    indexcol = [];
    for header_cell in headers:
            try:
                indexcol.append(header_cells_in_xls.index(header_cell))
            except Exception,e :
                print 'Header not found in xls: %s' %e
            else:
                print 'OK'

    print "INDEXCOL = ", indexcol
    row_dict_list = []

    throttle = 1
    for row in xrange(header_row+1, worksheet.nrows):
       i=0
       row_dict = {}

       for col in indexcol:
           cell_type = worksheet.cell_type(row, col)
           if cell_type == xlrd.XL_CELL_EMPTY:
               value = None
           elif cell_type == xlrd.XL_CELL_TEXT:
              value = worksheet.cell_value(row, col)
           elif cell_type == xlrd.XL_CELL_NUMBER:
              value = float(worksheet.cell_value(row,col))
           elif cell_type == xlrd.XL_CELL_DATE:
              from datetime import datetime
              value = datetime(*xlrd.xldate_as_tuple(worksheet.cell_value(row, col), workbook.datemode))
           elif cell_type == xlrd.XL_CELL_BOOLEAN:
              value = bool(worksheet.cell_value(row, col))
           else:
              value = worksheet.cell_value(row, col)

           if i<1:
               value = re.sub("\s*","",value)

           if value == None:
               row_dict[myheaders[i]] = "YES"
           else:
               row_dict[myheaders[i]] = value

           i=i+1
       try:
           db.InitiatedList.insert(**row_dict)
       except:
           duplicate = 1

    response.flash = T("Entry Successful!")

    return locals()

def uploaddata_SSParshadStatus():
    import xlrd
    import os
    path = os.path.join(request.folder,'private','SSParshadStatus.xlsx')
    workbook = xlrd.open_workbook(path)

    worksheet = workbook.sheet_names()
    #Attempt with any worksheet found
    worksheet = workbook.sheet_by_name(worksheet[0])
    header_row = 0
    import re, time

    #my headers are the headers in DB
    myheaders = ['NewGRNO','Status']

    #headers are the name used in XLS
    headers = ['NewGRNO','Status']
    header_cells_in_xls = [worksheet.cell_value(0,idx) for idx in range(worksheet.ncols)]


    try:
        db(db.SSTentativeParshadList.id > 0).delete()
    except:
        pass

    indexcol = [];
    for header_cell in headers:
            try:
                indexcol.append(header_cells_in_xls.index(header_cell))
            except Exception,e :
                print 'Header not found in xls: %s' %e
            else:
                print 'OK'

    print "INDEXCOL = ", indexcol
    row_dict_list = []

    throttle = 1
    for row in xrange(header_row+1, worksheet.nrows):
       i=0
       row_dict = {}

       for col in indexcol:
           cell_type = worksheet.cell_type(row, col)
           if cell_type == xlrd.XL_CELL_EMPTY:
               value = None
           elif cell_type == xlrd.XL_CELL_TEXT:
              value = worksheet.cell_value(row, col)
           elif cell_type == xlrd.XL_CELL_NUMBER:
              value = float(worksheet.cell_value(row,col))
           elif cell_type == xlrd.XL_CELL_DATE:
              from datetime import datetime
              value = datetime(*xlrd.xldate_as_tuple(worksheet.cell_value(row, col), workbook.datemode))
           elif cell_type == xlrd.XL_CELL_BOOLEAN:
              value = bool(worksheet.cell_value(row, col))
           else:
              value = worksheet.cell_value(row, col)


           if i<1:
               value = re.sub("\s*","",value)

           row_dict[myheaders[i]] = value
           i=i+1
       try:
           db.SSTentativeParshadList.insert(**row_dict)
       except:
           duplicate = 1

    response.flash = T("Entry Successful!")

    return locals()



def uploaddata_MachineAttendance():
    import os
    import datetime
    path = os.path.join(request.folder,'private','STARLINK.zip')
    os.system("rm -rf " + os.path.join(request.folder,'private','STARLINK') + ";unzip " + path + " -d " + os.path.join(request.folder,'private','STARLINK'));

    #Translate to new GRNO
    SSCount = db(db.SSAttendanceCount).select()
    GROldToNew = {}
    for Sewadar in SSCount:
        if Sewadar.OldSewadarid == '':
            pass
        else:
            GROldToNew[str(Sewadar.OldSewadarid)] = str(Sewadar.NewID)


    #Translate to new GNRO using Date database too as they sometime seem to have additional info
    SSDates = db(db.SSAttendanceDate).select(db.SSAttendanceDate.OldSewadarID,db.SSAttendanceDate.SewadarNewID,distinct=True)
    TotalTranslations = len(SSDates)
    for Sewadar in SSDates:
        if Sewadar.OldSewadarID == '':
            pass
        else:
            GROldToNew[str(Sewadar.OldSewadarID)] = str(Sewadar.SewadarNewID)

    AddAttendanceFileToMachineAttendanceDatabase(os.path.join(request.folder,'private','STARLINK'),**GROldToNew)
    return "Uploaded STARLINK FOLDER"

def AddAttendanceFileToMachineAttendanceDatabase(walk_dir,**GROldToNew):
    import os
    import sys
    import re
    import datetime

    a = re.compile(".*.TXT")
    t = re.compile(".*:.*:.*")
    file_list = []


    for root, subdirs, files in os.walk(walk_dir):
        file_list[:] = [filename for filename in files if a.match(filename)]
        for filename in file_list:
            file_path = os.path.join(root, filename)
            print "Processing filename " + str(file_path)
            with open(file_path, 'rb') as fp:
                row_dict = {}
                for line in fp:
                    try:
                        dummy_0, CARDNO, TIME, IO ,dummy_1 = re.split(r'\s+',line)
                        row_dict['GRNO'] = MachineCARDNOtoGRNO(CARDNO)
                        d1, d2, m1, m2, y1, y2 = list(filename.replace(".TXT","").replace("SL",""))
                        if t.match(TIME):
                            row_dict['DATETIME'] = '20' + y1 + y2 + '-' + m1 + m2 + '-' + d1 + d2 + ' ' + TIME
                        else:
                            row_dict['DATETIME'] = '20' + y1 + y2 + '-' + m1 + m2 + '-' + d1 + d2 + ' ' + TIME + ':00'

                        row_dict['TYPE'] = 'DMACHINE'
                        try:
                            row_dict['NewGRNO'] =  GROldToNew[row_dict['GRNO']]
                        except:
                            row_dict['NewGRNO'] = row_dict['GRNO']

                        row_dict['IO'] = IO
                        now = datetime.datetime.now() + datetime.timedelta(hours=12)
                        try:
                            if datetime.datetime.strptime(row_dict['DATETIME'], "%Y-%m-%d %H:%M:%S") < now:
                                try:
                                    db.MachineAttendance.insert(**row_dict)
                                except:
                                    print "Maybe entry already exists"
                                    print "CARDNO = " + CARDNO + " GRNO = " + row_dict['NewGRNO'] + " TIME = " + row_dict['DATETIME']
                        except:
                            print "malformed or future date " + row_dict['DATETIME'] + " check"



                    except:
                        pass

        for subdir in subdirs:
            AddAttendanceFileToMachineAttendanceDatabase(subdir,**GROldToNew)

    return 0




def uploaddata_master():
    import xlrd
    import os
    path = os.path.join(request.folder,'private','Master_xls.xlsx')
    pathlog = os.path.join(request.folder,'private','log_uploadMaster')
    logf = open(pathlog,'w')
    logf.close()
    workbook = xlrd.open_workbook(path)
    worksheet = workbook.sheet_names()
    #Attempt with any worksheet found
    worksheet = workbook.sheet_by_name(worksheet[0])
    header_row = 0
    import re

    #my headers are the headers in DB
    myheaders = ['SewadarNewID','GR_NO','NAME','CANTEEN','DEV_DTY','INITIATED_SS','Gender','Age']

    #headers are the name used in XLS
    headers = ['GR ID','OldID','NAME','Canteen No','JATHA','INITIATED_SS','gender','AGE']
    header_cells_in_xls = [worksheet.cell_value(0,idx) for idx in range(worksheet.ncols)]


    print "Master Sheet updating"

    indexcol = [];
    for header_cell in headers:
            logf = open(pathlog,'a')
            try:
                indexcol.append(header_cells_in_xls.index(header_cell))
                logf.write(header_cell + " :header found in xls = " + str(header_cells_in_xls.index(header_cell)) + "\n")
            except Exception,e :
                logf.write("header not found in xls = " + str(e) + "\n")
            else:
                print 'OK'
            logf.close()

    print "INDEXCOL = ", indexcol
    logf = open(pathlog,'a')
    logf.write("indexcol = " + str(indexcol) + "\n")
    logf.close()
    row_dict_list = []

    db.MasterSheet.drop()
    db.define_table('MasterSheet',
                    Field('GR_NO','string'),
                    Field('SewadarNewID','string'),
                    Field('NAME','string'),
                    Field('INITIATED_SS','string'),
                    Field('CANTEEN','string'),
                    Field('DEV_DTY','string'),
                    Field('Gender','string'),
                    Field('Age','integer'),
                    migrate=True,
                    redefine=True,
                    format='%(GR_NO)s')

    for row in xrange(header_row+1, worksheet.nrows):
       i=0
       row_dict = {}
       for col in indexcol:
           cell_type = worksheet.cell_type(row, col)
           if cell_type == xlrd.XL_CELL_EMPTY:
               value = None
           elif cell_type == xlrd.XL_CELL_TEXT:
              value = worksheet.cell_value(row, col)
           elif cell_type == xlrd.XL_CELL_NUMBER:
              value = float(worksheet.cell_value(row,col))
           elif cell_type == xlrd.XL_CELL_DATE:
              from datetime import datetime
              value = datetime(*xlrd.xldate_as_tuple(worksheet.cell_value(row, col), workbook.datemode))
           elif cell_type == xlrd.XL_CELL_BOOLEAN:
              value = bool(worksheet.cell_value(row, col))
           else:
              value = worksheet.cell_value(row, col)

           #Uncomment for debug
           #logf = open(pathlog,'a')
           #logf.write("i = " + str(i) + "\n")
           #logf.write("value = " + str(value) + "\n")
           #logf.close()
           if i<=1:
              value = re.sub("\s*","",value)

           try:
               row_dict[myheaders[i]] = value.upper()
           except:
               row_dict[myheaders[i]] = value
           i=i+1
       db.MasterSheet.insert(**row_dict)

    response.flash = T("Entry Successful!")
    return locals()



def hmm():
    from datetime import timedelta as timed
    import datetime
    db.SSAttendanceDate.drop()
    scheduler.queue_task('uploaddata_SSAttendance',
        start_time=datetime.datetime.now() + timed(seconds=3),
        timeout = 6000)
    return 0

def uploaddata_MachineManualAttendance():
    import xlrd
    import os
    db.commit()
    path = os.path.join(request.folder,'private','MachineAttendance.xlsx')
    workbook = xlrd.open_workbook(path)

    worksheet = workbook.sheet_names()
#Attempt with any worksheet found
    worksheet = workbook.sheet_by_name(worksheet[0])
    header_row = 0
    import re, time

#my headers are the headers in DB
    myheaders = ['NewGRNO','GRNO','DATETIME','TYPE','IO']

#headers are the name used in XLS
    headers = ['NewGRNO','GRNO','DATETIME','TYPE','IO']
    header_cells_in_xls = [worksheet.cell_value(0,idx) for idx in range(worksheet.ncols)]



    indexcol = [];
    for header_cell in headers:
        try:
            indexcol.append(header_cells_in_xls.index(header_cell))
        except Exception,e :
            print 'Header not found in xls: %s' %e
            pass
        else:
            print 'OK'

    row_dict_list = []
    failed = 0
    passed = 0


    for row in xrange(header_row+1, worksheet.nrows):
        i=0
        row_dict = {}

        for col in indexcol:
            cell_type = worksheet.cell_type(row, col)
            if cell_type == xlrd.XL_CELL_EMPTY:
                value = None
            elif cell_type == xlrd.XL_CELL_TEXT:
                value = worksheet.cell_value(row, col)
            elif cell_type == xlrd.XL_CELL_NUMBER:
                value = float(worksheet.cell_value(row,col))
            elif cell_type == xlrd.XL_CELL_DATE:
                from datetime import datetime
                value = datetime(*xlrd.xldate_as_tuple(worksheet.cell_value(row, col), workbook.datemode))
            elif cell_type == xlrd.XL_CELL_BOOLEAN:
                value = bool(worksheet.cell_value(row, col))
            else:
                value = worksheet.cell_value(row, col)
            row_dict[myheaders[i]] = value
            i=i+1

        try:
            db.MachineAttendance.insert(**row_dict)
            passed = passed + 1
        except:
            #print "failed: " + row_dict['NewGRNO'] + " " + datetime.strftime(row_dict['DATETIME'],"%d-%b-%Y %h:%m:%s")
            failed = failed + 1
            a = 1

    db.commit()
    return locals()


def uploaddata_SSAttendanceCount():
    import xlrd
    import os
    path = os.path.join(request.folder,'private','SSAttendanceCount_xls.xlsx')
    workbook = xlrd.open_workbook(path)
    worksheet = workbook.sheet_names()
    #Attempt with any worksheet found
    worksheet = workbook.sheet_by_name(worksheet[0])
    header_row = 0
    import re


    #my headers are the headers in DB
    myheaders = ['NewID','OldSewadarid','Name','Father_Husband_Name','status','gender','B','w','V1','V2','V3','V4','Initiated_Status','TotalVisit','Total','areaname']

    #headers are the name used in XLS
    headers = ['NewID','OldSewadarid','Name','Father_Husband_Name','status','Gender','B','w','V1','V2','V3','V4','Initiated_Status','TotalVisit','Total','areaname']
    header_cells_in_xls = [worksheet.cell_value(0,idx) for idx in range(worksheet.ncols)]



    indexcol = [];
    for header_cell in headers:
            try:
                indexcol.append(header_cells_in_xls.index(header_cell))
            except Exception,e :
                print 'Header not found in xls: %s' %e
            else:
                print 'OK'
    value = ""
    print "INDEXCOL = ", indexcol
    row_dict_list = []

    db(db.SSAttendanceCount.id > 0).delete()
    for row in xrange(header_row+1, worksheet.nrows):
       i=0
       row_dict = {}
       for col in indexcol:
           cell_type = worksheet.cell_type(row, col)
           if cell_type == xlrd.XL_CELL_EMPTY:
              value = None
           elif cell_type == xlrd.XL_CELL_TEXT:
              value = worksheet.cell_value(row, col)
           elif cell_type == xlrd.XL_CELL_NUMBER:
              value = float(worksheet.cell_value(row,col))
           elif cell_type == xlrd.XL_CELL_DATE:
              from datetime import datetime
              value = datetime(*xlrd.xldate_as_tuple(worksheet.cell_value(row,col), workbook.datemode))
           elif cell_type == xlrd.XL_CELL_BOOLEAN:
              value = bool(worksheet.cell_value(row, col))
           else:
              value = worksheet.cell_value(row, col)

           if myheaders[i].upper() == 'OLDSEWADARID':
               if (value == 'NA') | (value == ''):
                   value = worksheet.cell_value(row, header_cells_in_xls.index('NewID'))


           if myheaders[i].upper() == 'GENDER':
               if (value.find('f') != -1) |( value.find('F') != -1) | (value.find('la') != -1) | (value.find('La') != -1):
                   value = "Female"
               elif (value.find('m') != -1) | (value.find('M') != -1) | (value.find('g') != -1) | (value.find('G') != -1):
                   value = "Male"
               else:
                   raise ValueError('Gender not in required format')

           #if myheaders[i].upper() == 'STATUS':
           #    if ((value.upper() == 'ELDERLY') |(value.upper() == 'RETIRED') | (value.upper() == 'PERMANENT') | (value == 'permanent') | (value == 'PERMANENT')):
           #        value = 'Permanent'
           #    else:
           #        value = 'Slip'

           if i<=1:
               value = re.sub("\s*","",value)

           row_dict[myheaders[i]] = value
           i=i+1
       db.SSAttendanceCount.insert(**row_dict)



    return locals()

@auth.requires(auth.user_id == 3)
def ParshadList():
    import cv2
    print cv2.__version__
    import os
    from openpyxl import Workbook
    from openpyxl.cell import get_column_letter
    response.headers['Connection'] =  'keep-alive'


    from gluon.sqlhtml import form_factory
    import datetime
    import time



    db.commit()
    db2.commit()
    message = 'Schedular based'

    #form=form_factory(SQLField('DAY_END_TIME','string',default=19,requires=IS_IN_SET(range(0,25,1))),SQLField('DateStart','date',default=datetime.datetime.today()-datetime.timedelta(days=31)),SQLField('DateEnd','date',default=datetime.datetime.today()),SQLField('LALastLadiesNewGRNO','integer',default=9999),SQLField('LBLastLadiesNewGRNO','integer',default=9999),SQLField('GALastGentsNewGRNO','integer',default=9999),SQLField('GBLastGentsNewGRNO','integer',default=9999),SQLField('LastOSS','integer',default=40000),SQLField('SSCountCutOffLadies','integer',default=36),SQLField('SSCountCutOffGents','integer',default=30),SQLField('VisitCountCutOff','integer',default=9),SQLField('CVCutOff','integer',default=4),SQLField('WWCutOff','integer',default=2),SQLField('DumpMachineAttendance','string',requires=IS_IN_SET(['YES','NO']),default='NO'),SQLField('DumpSSAttendance','string',requires=IS_IN_SET(['YES','NO']),default='NO'),SQLField('CANTEENWISE_REPORT','string',requires=IS_IN_SET(['YES','NO','FLAT']),default='NO'),formname='DateSelect')
    form=form_factory(SQLField('DAY_END_TIME','string',default=19,requires=IS_IN_SET(range(0,25,1))),SQLField('DateStart','date',default=datetime.datetime.strptime('15-November-2017 00:00:00','%d-%B-%Y %H:%M:%S')),SQLField('DateEnd','date',default=datetime.datetime.strptime('14-November-2018 23:59:59','%d-%B-%Y %H:%M:%S')),SQLField('MandatoryDaysDateStart','date',default=datetime.datetime.strptime('08-November-2018 00:00:00','%d-%B-%Y %H:%M:%S')),SQLField('MandatoryDaysDateEnd','date',default=datetime.datetime.strptime('11-November-2018 23:59:59','%d-%B-%Y %H:%M:%S')),SQLField('MandatoryDaysCountCutoff','integer',default=0),SQLField('SSCountCutOffLadies','integer',default=36),SQLField('SSCountCutOffGents','integer',default=30),SQLField('VisitCountCutOff','integer',default=9),SQLField('CVCutOff','integer',default=0),SQLField('WWCutOff','integer',default=4),SQLField('WWWaiver','integer',default=367),SQLField('WWAgeWaiver','integer',default=65),SQLField('MailSubject','string',default='Parshad Status'),formname='DateSelect')

    if form.accepts(request.vars,session,formname='DateSelect'):
        DateSelectedStart = request.vars.DateStart
        DateSelectedEnd = request.vars.DateEnd
        MandatoryDaysDateSelectedStart = request.vars.MandatoryDaysDateStart
        MandatoryDaysDateSelectedEnd = request.vars.MandatoryDaysDateEnd
        MandatoryDaysCountCutoff = request.vars.MandatoryDaysCountCutoff
        SSCountCutOffGents = request.vars.SSCountCutOffGents
        SSCountCutOffLadies = request.vars.SSCountCutOffLadies
        CVCutOff = request.vars.CVCutOff
        VisitCountCutOff = request.vars.VisitCountCutOff
        WWCutOff = request.vars.WWCutOff
        WWWaiver = request.vars.WWWaiver
        WWAgeWaiver = request.vars.WWAgeWaiver
        DAY_END_TIME = request.vars.DAY_END_TIME
        MailSubject = request.vars.MailSubject

        from datetime import timedelta as timed
        #scheduler.queue_task('DevParshadListScheduled', pvars={'DateSelectedStart':DateSelectedStart, 'DateSelectedEnd':DateSelectedEnd, 'LALastLadiesNewGRNO':LALastLadiesNewGRNO, 'LBLastLadiesNewGRNO':LBLastLadiesNewGRNO, 'GALastGentsNewGRNO':GALastGentsNewGRNO, 'GBLastGentsNewGRNO':GBLastGentsNewGRNO, 'LastOSS':LastOSS, 'SSCountCutOffGents':SSCountCutOffGents, 'SSCountCutOffLadies':SSCountCutOffLadies, 'CVCutOff':CVCutOff, 'VisitCountCutOff':VisitCountCutOff,'WWCutOff':WWCutOff, 'DAY_END_TIME':DAY_END_TIME, 'DumpMachineAttendance':DumpMachineAttendance, 'DumpSSAttendance':DumpSSAttendance, 'CANTEENWISE_REPORT':CANTEENWISE_REPORT},
        #    start_time=request.now + timed(seconds=1),
        #    timeout = 6000)
        scheduler.queue_task('ParshadListScheduled', pvars={'DateSelectedStart':DateSelectedStart, 'DateSelectedEnd':DateSelectedEnd,'MandatoryDaysDateStart':MandatoryDaysDateSelectedStart,'MandatoryDaysDateEnd':MandatoryDaysDateSelectedEnd,'MandatoryDaysCountCutoff':MandatoryDaysCountCutoff,'SSCountCutOffGents':SSCountCutOffGents, 'SSCountCutOffLadies':SSCountCutOffLadies, 'CVCutOff':CVCutOff, 'VisitCountCutOff':VisitCountCutOff,'WWCutOff':WWCutOff,'WWWaiver':WWWaiver, 'WWAgeWaiver':WWAgeWaiver, 'DAY_END_TIME':DAY_END_TIME,'MailSubject':MailSubject},
            start_time=request.now + timed(seconds=0),
            timeout = 6000)

    return dict(form=form,message=message)

def download_ParshadList():
    import os
    dpath = os.path.join(request.folder,'private','TentativeParshadList.xlsx')
    return response.stream(open(dpath,'rb'), chunk_size=10**6)



#This page dumps out an attendance register. Each cel contains Machine as well as sewa Samiti
#date-time. This can be useful for debugging.
#Additional tabs are added for monthly printout
@auth.requires_login()
def AttendanceRegisterDetailed():
    import os
    import dpath.util
    from openpyxl import Workbook
    from openpyxl.cell import get_column_letter
    from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
    response.headers['Connection'] =  'keep-alive'
    dworkbook = Workbook()
    dpath = os.path.join(request.folder,'private','AttendanceRegisterDetailed.xlsx')

    from gluon.sqlhtml import form_factory
    import datetime
    import time

    SS_GENTS_REQUIRED = 30
    SS_LADIES_REQUIRED = 36
    DateSelectedStart = datetime.datetime.today()
    DateSelectedEnd = datetime.datetime.today()
    DAY_END_TIME = 19

    message = "ALL OK "
    Note3 = 'NON INITIATED SEWADARS ARE NOT ELIGIBLE FOR PARSHAD SLIP.'
    Note4 = 'ALL DISCRIPENCIES TO BE INTIMATED WITHIN ONE WEEK. ALL ATTENDANCE TO BE MARKED IN ATTENDANCE MACHINE.'
    Note5 = 'PLEASE INFORM BADGE OFFICE IN CASE OF CHANGE IN PHONE NUMBERS.'
    Note6 = 'ENDNOTE'
    Note7 = 'ENDNOTE'
    NoteCount = 0
    AttendanceRegister = {}

    #Prepare Jatha List
    my_list = db().select(db.MasterSheet.DEV_DTY, orderby=db.MasterSheet.DEV_DTY, distinct=True).as_list()

    myjathalist = []
    for mydict in my_list:
        myjathalist.append(mydict['DEV_DTY'])

    myjathalist.sort()

    #Add Gender and Name
    SSCount = db(db.SSAttendanceCount).select()
    for Sewadar in SSCount:
        AttendanceRegister[Sewadar.NewID,'Gender'] = Sewadar.gender
        AttendanceRegister[Sewadar.NewID,'TotalCount'] = Sewadar.Total
        AttendanceRegister[Sewadar.NewID,'NAME'] = Sewadar.Name
        AttendanceRegister[Sewadar.NewID,'OldSewadarid'] = Sewadar.OldSewadarid

    form=form_factory(SQLField('DAY_END_TIME','string',default=19,requires=IS_IN_SET(range(0,25,1))),SQLField('DateStart','date',default=datetime.datetime.today()-datetime.timedelta(days=31)),SQLField('DateEnd','date',default=datetime.datetime.today()),SQLField('SSCountCutOffLadies','integer',default=36),SQLField('SSCountCutOffGents','integer',default=30),SQLField('CompletionDate','date',default=datetime.datetime.strptime('29-09-2015','%d-%m-%Y')),SQLField('Note3','str',default=Note3),SQLField('Note4','str',default=Note4),SQLField('Note5','str',default=Note5),SQLField('Note6','str',default=Note6),SQLField('Note7','str',default=Note7),formname='DateSelect')
    if form.accepts(request.vars,session,formname='DateSelect'):
        try:
            os.remove(dpath)
        except:
            pass

        DateSelectedStart = request.vars.DateStart
        DateSelectedEnd = request.vars.DateEnd
        GENTS_REQUIRED = int(request.vars.SSCountCutOffGents)
        LADIES_REQUIRED = int(request.vars.SSCountCutOffLadies)
        DAY_END_TIME = request.vars.DAY_END_TIME
        COMPLETION_DATE = datetime.datetime.strftime(datetime.datetime.strptime(request.vars.CompletionDate,'%Y-%m-%d'),'%d-%b-%Y')
        SSDate = db((db.SSAttendanceDate.DutyDate >= (datetime.datetime.strptime(DateSelectedStart,'%Y-%m-%d').replace(hour=0, minute=0, second=0, microsecond=0))) & (db.SSAttendanceDate.DutyDate <= (datetime.datetime.strptime(DateSelectedEnd,'%Y-%m-%d').replace(hour=23, minute=59, second=59, microsecond=999)))).select('SewadarNewID','DutyDate','Duty_Type')
        MachineDate = db((db.MachineAttendance.DATETIME >= (datetime.datetime.strptime(DateSelectedStart,'%Y-%m-%d') - datetime.timedelta(hours=((24 - int(DAY_END_TIME)) % 24)))) & (db.MachineAttendance.DATETIME <= (datetime.datetime.strptime(DateSelectedEnd,'%Y-%m-%d') + datetime.timedelta(hours=int(DAY_END_TIME))))).select('GRNO','NewGRNO','DATETIME','TYPE')
        print "Machine Attendance between date"
        print "length of MachineDate =" + str(len(MachineDate))


        print "Collecting SSDate"
        for SSEntry in SSDate:
            try:
                AttendanceRegister['SEWADARS'].append(SSEntry.SewadarNewID)
            except:
                AttendanceRegister['SEWADARS'] = [SSEntry.SewadarNewID]

            if SSEntry.Duty_Type == 'W':
                AttendanceRegister[SSEntry.SewadarNewID,Roundoffdate(SSEntry.DutyDate,23)] = [datetime.date.strftime(SSEntry.DutyDate,"%d-%b-%Y") + "(SSW)"]
                AttendanceRegister[SSEntry.SewadarNewID,Roundoffdate(SSEntry.DutyDate,23),'SSTYPE'] = 'W'
                try:
                    AttendanceRegister[SSEntry.SewadarNewID,'SSCount'] = AttendanceRegister[SSEntry.SewadarNewID,'SSCount'] + 2
                except:
                    AttendanceRegister[SSEntry.SewadarNewID,'SSCount'] = 2
            else:
                AttendanceRegister[SSEntry.SewadarNewID,Roundoffdate(SSEntry.DutyDate,23)] = [datetime.date.strftime(SSEntry.DutyDate,"%d-%b-%Y") + "(SSD)"]
                AttendanceRegister[SSEntry.SewadarNewID,Roundoffdate(SSEntry.DutyDate,23),'SSTYPE'] = 'D'
                try:
                    AttendanceRegister[SSEntry.SewadarNewID,'SSCount'] = AttendanceRegister[SSEntry.SewadarNewID,'SSCount'] + 1
                except:
                    AttendanceRegister[SSEntry.SewadarNewID,'SSCount'] = 1


        print "Collecting Machine Date"

        for MEntry in MachineDate:
            try:
                AttendanceRegister['SEWADARS'].append(MEntry.NewGRNO)
            except:
                AttendanceRegister['SEWADARS'] = [MEntry.NewGRNO]

            if MEntry.TYPE == 'WMANUAL':
                try:
                    AttendanceRegister[MEntry.NewGRNO,Roundoffdate(MEntry.DATETIME,DAY_END_TIME)].append(datetime.date.strftime(MEntry.DATETIME,"%d-%b-%Y %H:%M:%S") + "(MMW)")
                except:
                    AttendanceRegister[MEntry.NewGRNO,Roundoffdate(MEntry.DATETIME,DAY_END_TIME)] = [datetime.date.strftime(MEntry.DATETIME,"%d-%b-%Y %H:%M:%S") + "(MMW)"]
            elif MEntry.TYPE == 'DMACHINE':
                try:
                    AttendanceRegister[MEntry.NewGRNO,Roundoffdate(MEntry.DATETIME,DAY_END_TIME)].append(datetime.date.strftime(MEntry.DATETIME,"%d-%b-%Y %H:%M:%S") + "(MD)")
                except:
                    AttendanceRegister[MEntry.NewGRNO,Roundoffdate(MEntry.DATETIME,DAY_END_TIME)] = [datetime.date.strftime(MEntry.DATETIME,"%d-%b-%Y %H:%M:%S") + "(MD)"]
            else:
                try:
                    AttendanceRegister[MEntry.NewGRNO,Roundoffdate(MEntry.DATETIME,DAY_END_TIME)].append(datetime.date.strftime(Roundoffdate(MEntry.DATETIME,DAY_END_TIME),"%d-%b-%Y") + "(MMD)")
                except:
                    AttendanceRegister[MEntry.NewGRNO,Roundoffdate(MEntry.DATETIME,DAY_END_TIME)] = [datetime.date.strftime(Roundoffdate(MEntry.DATETIME,DAY_END_TIME),"%d-%b-%Y") + "(MMD)"]



        AttendanceRegister['SEWADARS'] = set(AttendanceRegister['SEWADARS'])

        try:
            db.tempAttendanceRegisterDetailed.drop()
        except:
            pass

        dAttendanceRegister = dworkbook.create_sheet(0)
        dAttendanceRegister.title = "AttendanceRegisterDetailed"


        row_num = 1
        for Sewadar in AttendanceRegister['SEWADARS']:
            row_num = row_num + 1
            dAttendanceRegister.cell("A"+str(row_num)).value = Sewadar
            try:
                dAttendanceRegister.cell("B"+str(row_num)).value = AttendanceRegister[Sewadar,'OldSewadarid']
                dAttendanceRegister.cell("C"+str(row_num)).value = AttendanceRegister[Sewadar,'NAME']
            except:
                pass

            for i in range((datetime.datetime.strptime(DateSelectedEnd,'%Y-%m-%d') - datetime.datetime.strptime(DateSelectedStart,'%Y-%m-%d')).days):
                dateindex = datetime.datetime.strptime(DateSelectedStart,'%Y-%m-%d') + datetime.timedelta(i)
                dAttendanceRegister.cell(get_column_letter(i+5)+"1").value = dateindex
                try:
                    CombinedDayAttendance = "\n".join(map(str, AttendanceRegister[Sewadar,dateindex]))
                    dAttendanceRegister.cell(get_column_letter(i+5)+str(row_num)).value = CombinedDayAttendance
                except:
                    dAttendanceRegister.cell(get_column_letter(i+5)+str(row_num)).value = ""


        jatha_count = {}
        for jatha in myjathalist:
            jatha_count[jatha] = 0
            #Fetch Sewadar list in the jatha
            SewadarList = db(db.MasterSheet.DEV_DTY == jatha).select(db.MasterSheet.GR_NO,db.MasterSheet.SewadarNewID,orderby=db.MasterSheet.GR_NO).as_list()
            dAttendanceRegister = dworkbook.create_sheet(0)
            dAttendanceRegister.title = jatha.replace(':','_')
            dAttendanceRegister.page_setup.orientation = dAttendanceRegister.ORIENTATION_LANDSCAPE
            dAttendanceRegister.page_setup.paperSize = dAttendanceRegister.PAPERSIZE_A4
            dAttendanceRegister.page_setup.fitToHeight = 0
            dAttendanceRegister.page_setup.fitToWidth = 1
            dAttendanceRegister.page_margins.left = 0.6/2.54
            dAttendanceRegister.page_margins.right = 0.4/2.54
            dAttendanceRegister.page_margins.bottom = 0.6/2.54
            dAttendanceRegister.page_margins.top = 0.6/2.54

            #Put headers
            dAttendanceRegister['A1'].value = "RADHA SOAMI SATSANG BEAS, DELHI\n BHATI CENTRE \n CAFETERIA DEPARTMENT"
            dAttendanceRegister['A1'].font = Font(name='Calibri',size=14,bold=True)
            dAttendanceRegister['A1'].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)

            dAttendanceRegister.merge_cells('A1:' + get_column_letter(((datetime.datetime.strptime(DateSelectedEnd,'%Y-%m-%d') - datetime.datetime.strptime(DateSelectedStart,'%Y-%m-%d')).days) + 3) + '1')
            dAttendanceRegister.row_dimensions[1].height = 60
            dAttendanceRegister.column_dimensions['A'].width = 13.14

            dAttendanceRegister['A2'].value = jatha + ': ATTENDANCE FROM ' + datetime.datetime.strftime(datetime.datetime.strptime(DateSelectedStart,'%Y-%m-%d'),'%d-%b-%Y') + ' TO ' +datetime.datetime.strftime(datetime.datetime.strptime(DateSelectedEnd,'%Y-%m-%d') - datetime.timedelta(days=1),'%d-%b-%Y')
            dAttendanceRegister['A2'].font = Font(name='Calibri',size=12,bold=True)
            dAttendanceRegister['A2'].alignment = Alignment(horizontal='center',vertical='center')

            dAttendanceRegister.merge_cells('A2:' + get_column_letter(((datetime.datetime.strptime(DateSelectedEnd,'%Y-%m-%d') - datetime.datetime.strptime(DateSelectedStart,'%Y-%m-%d')).days) + 3) + '2')
            dAttendanceRegister.row_dimensions[2].height = 22.5

            #Put column headers
            dAttendanceRegister['A3'].value = 'GR NO'
            dAttendanceRegister['A3'].font = Font(name='Calibri',size=12,bold=True)
            dAttendanceRegister['A3'].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
            dAttendanceRegister['A3'].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))

            dAttendanceRegister['B3'].value = 'NAME'
            dAttendanceRegister['B3'].font = Font(name='Calibri',size=12,bold=True)
            dAttendanceRegister['B3'].alignment = Alignment(horizontal='left',vertical='center',wrap_text=True)
            dAttendanceRegister['B3'].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
            dAttendanceRegister.column_dimensions['B'].width = 12.86


            dAttendanceRegister['C3'].value = "REQ\n BY NEXT\nVISIT"
            dAttendanceRegister['C3'].font = Font(name='Calibri',size=8,bold=True)
            dAttendanceRegister['C3'].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
            dAttendanceRegister['C3'].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
            dAttendanceRegister.row_dimensions[3].height = 49
            dAttendanceRegister.column_dimensions['C'].width = 4.14

            for i in range((datetime.datetime.strptime(DateSelectedEnd,'%Y-%m-%d') - datetime.datetime.strptime(DateSelectedStart,'%Y-%m-%d')).days):
                dateindex = (datetime.datetime.strftime(datetime.datetime.strptime(DateSelectedStart,'%Y-%m-%d') +  datetime.timedelta(days=i),'%a').upper())[:2] + '\n' + str((datetime.datetime.strptime(DateSelectedStart,'%Y-%m-%d') + datetime.timedelta(days=i)).day)
                dAttendanceRegister.cell(get_column_letter(i+4)+"3").value = dateindex
                dAttendanceRegister[get_column_letter(i+4)+"3"].font = Font(name='Calibri',size=8,bold=True)
                dAttendanceRegister[get_column_letter(i+4)+"3"].alignment = Alignment(horizontal='center',vertical='bottom',wrap_text=True)
                dAttendanceRegister[get_column_letter(i+4)+"3"].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
                dAttendanceRegister.column_dimensions[get_column_letter(i+4)].width = 3

            row_num = 4

            for Sewadar in SewadarList:
                try:
                    dAttendanceRegister.cell("B"+str(row_num)).value = AttendanceRegister[Sewadar['SewadarNewID'],'NAME']
                    dAttendanceRegister["B"+str(row_num)].font = Font(name='Calibri',size=11,bold=False)
                    dAttendanceRegister["B"+str(row_num)].alignment = Alignment(horizontal='left',vertical='center')
                    dAttendanceRegister["B"+str(row_num)].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
                    dAttendanceRegister.cell("A"+str(row_num)).value = Sewadar['GR_NO']
                    dAttendanceRegister["A"+str(row_num)].font = Font(name='Calibri',size=11,bold=False)
                    dAttendanceRegister["A"+str(row_num)].alignment = Alignment(horizontal='center',vertical='center')
                    dAttendanceRegister["A"+str(row_num)].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
                    if ((AttendanceRegister[Sewadar['SewadarNewID'],'Gender'] == 'Male')) & (AttendanceRegister[Sewadar['SewadarNewID'],'TotalCount'] < GENTS_REQUIRED):
                        if GENTS_REQUIRED - AttendanceRegister[Sewadar['SewadarNewID'],'TotalCount'] > SS_GENTS_REQUIRED:
                            dAttendanceRegister.cell("C"+str(row_num)).value = SS_GENTS_REQUIRED
                        else:
                            dAttendanceRegister.cell("C"+str(row_num)).value = GENTS_REQUIRED - AttendanceRegister[Sewadar['SewadarNewID'],'TotalCount']
                    elif ((AttendanceRegister[Sewadar['SewadarNewID'],'Gender'] == 'Female')) & (AttendanceRegister[Sewadar['SewadarNewID'],'TotalCount'] < LADIES_REQUIRED):
                        if LADIES_REQUIRED - AttendanceRegister[Sewadar['SewadarNewID'],'TotalCount'] > SS_LADIES_REQUIRED:
                            dAttendanceRegister.cell("C"+str(row_num)).value = SS_LADIES_REQUIRED
                        else:
                            dAttendanceRegister.cell("C"+str(row_num)).value = LADIES_REQUIRED - AttendanceRegister[Sewadar['SewadarNewID'],'TotalCount']
                    elif ((AttendanceRegister[Sewadar['SewadarNewID'],'Gender'] == 'Female') | (AttendanceRegister[Sewadar['SewadarNewID'],'Gender'] == 'Male')):
                        dAttendanceRegister.cell("C"+str(row_num)).value = "NIL"
                    else:
                        dAttendanceRegister.cell("C"+str(row_num)).value = SS_LADIES_REQUIRED - AttendanceRegister[Sewadar['SewadarNewID'],'TotalCount']

                    dAttendanceRegister["C"+str(row_num)].font = Font(name='Calibri',size=11,bold=True)
                    dAttendanceRegister["C"+str(row_num)].alignment = Alignment(horizontal='center',vertical='center')
                    dAttendanceRegister["C"+str(row_num)].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
                    for i in range((datetime.datetime.strptime(DateSelectedEnd,'%Y-%m-%d') - datetime.datetime.strptime(DateSelectedStart,'%Y-%m-%d')).days):
                        dateindex = datetime.datetime.strptime(DateSelectedStart,'%Y-%m-%d') + datetime.timedelta(days=i)
                        try:
                            dAttendanceRegister.cell(get_column_letter(i+4)+str(row_num)).value = AttendanceRegister[Sewadar['SewadarNewID'],dateindex,'SSTYPE']
                        except:
                            dAttendanceRegister.cell(get_column_letter(i+4)+str(row_num)).value = " "
                        dAttendanceRegister[get_column_letter(i+4)+str(row_num)].font = Font(name='Calibri',size=8,bold=True)
                        dAttendanceRegister[get_column_letter(i+4)+str(row_num)].alignment = Alignment(horizontal='center',vertical='center')
                        dAttendanceRegister[get_column_letter(i+4)+str(row_num)].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
                    jatha_count[jatha] = jatha_count[jatha] + 1
                    row_num = row_num + 1
                except:
                    print "Sewadar GRNO = " + Sewadar['GR_NO'] + ' = ' +  Sewadar['SewadarNewID'] + " does not exist in SS"


        for jatha in myjathalist:
            dAttendanceRegister = dworkbook[jatha.replace(':','_')]
            dAttendanceRegister.add_print_title(3,3,rows_or_cols='row')
            dAttendanceRegister['A' + str(jatha_count[jatha] + 5)].value = 'Note:'
            dAttendanceRegister['A' + str(jatha_count[jatha] + 6)].value = '1) PLEASE INFORM ALL THE SEWADARS.'
            dAttendanceRegister['A' + str(jatha_count[jatha] + 7)].value = '2) LAST DATE TO COMPLETE ATTENDANCE IS ' + COMPLETION_DATE
            if request.vars.Note3 == "ENDNOTE":
                pass
            else:
                dAttendanceRegister['A' + str(jatha_count[jatha] + 8)].value = '3) ' + request.vars.Note3
            if request.vars.Note4 == "ENDNOTE":
                pass
            else:
                dAttendanceRegister['A' + str(jatha_count[jatha] + 9)].value = '4) ' + request.vars.Note4
            if request.vars.Note5 == "ENDNOTE":
                pass
            else:
                dAttendanceRegister['A' + str(jatha_count[jatha] + 10)].value = '5) ' + request.vars.Note5
            if request.vars.Note6 == "ENDNOTE":
                pass
            else:
                dAttendanceRegister['A' + str(jatha_count[jatha] + 11)].value = '6) ' + request.vars.Note6
            if request.vars.Note7 == "ENDNOTE":
                pass
            else:
                dAttendanceRegister['A' + str(jatha_count[jatha] + 12)].value = '7) ' + request.vars.Note7




        dworkbook.save(dpath)
        redirect(URL(r=request, f='download_AttendanceRegisterDetailed'))
        message = "File too big probably"

    return dict(form=form,message=message)

def download_AttendanceRegisterDetailed():
    import os
    dpath = os.path.join(request.folder,'private','AttendanceRegisterDetailed.xlsx')
    return response.stream(open(dpath,'rb'), chunk_size=10**6)

@auth.requires_login()
def ShortAttendanceRegister():
    from gluon.sqlhtml import form_factory
    import datetime
    import time

    from openpyxl import Workbook
    from openpyxl.cell import get_column_letter
    import os
    response.headers['Connection'] =  'keep-alive'
    dworkbook = Workbook()
    dAttendanceRegister = dworkbook.create_sheet(0)
    dpath = os.path.join(request.folder,'private','ShortAttendanceRegister.xlsx')

    DateSelectedStart = datetime.datetime.today()
    DateSelectedEnd = datetime.datetime.today()
    Register = "EMPTY"
    DAttendanceRegisterTable = "EMPTY"
    datasource = {}
    columns = []
    orderby = []
    headers = {}
    excel_headers = {}
    DEV_DTY = ""
    TextMessage = "Jatha Wise Short Attendance Report"
    ReportDate =  0
    GENTS_REQUIRED = 30
    LADIES_REQUIRED = 36
    SS_GENTS_REQUIRED = 30
    SS_LADIES_REQUIRED = 36

    try:
        db.tempAttendanceRegisterTable.drop()
    except:
        pass

    SewadarDetails = {'Sewadars':[]}
#    form=form_factory(SQLField('JATHA','string',requires=IS_IN_DB(db,'MasterSheet.DEV_DTY','%(DEV_DTY)s')),SQLField('DateStart','date',default=datetime.datetime.today()-datetime.timedelta(days=31)),SQLField('DateEnd','date',default=datetime.datetime.today()),formname='DateSelect')
    my_list = db().select(db.MasterSheet.DEV_DTY, distinct=True).as_list()
    myjathalist = []
    for mydict in my_list:
        myjathalist.append(mydict['DEV_DTY'])

    myjathalist.sort()
    form=form_factory(SQLField('JATHA','string',requires=IS_IN_SET(myjathalist)),SQLField('DateStart','date',default=datetime.datetime.today()-datetime.timedelta(days=31)),SQLField('DateEnd','date',default=datetime.datetime.today()),SQLField('Download','string',requires=IS_IN_SET(['YES','NO']),default='NO'),formname='DateSelect')
    if form.accepts(request.vars,session,formname='DateSelect'):
        try:
            os.remove(dpath)
        except:
            pass

        download = request.vars.Download
        DateSelectedStart = request.vars.DateStart
        DateSelectedEnd = request.vars.DateEnd
        ReportDate = datetime.datetime.today()-datetime.timedelta(days=3001)

        SSDate = db((db.SSAttendanceDate.DutyDate >= DateSelectedStart) & (db.SSAttendanceDate.DutyDate <= DateSelectedEnd)).select()
        for Sewadar in SSDate:
            delta = Sewadar.DutyDate - datetime.datetime.strptime(DateSelectedStart,'%Y-%m-%d')
            SewadarDetails[Sewadar.SewadarNewID,'DAYS',delta.days] = Sewadar.DutyDate
            SewadarDetails[Sewadar.SewadarNewID,'TYPE',delta.days] = Sewadar.Duty_Type
            if ReportDate < Sewadar.DutyDate:
                ReportDate = Sewadar.DutyDate

        SSCount = db(db.SSAttendanceCount).select()
        for Sewadar in SSCount:
            SewadarDetails[Sewadar.NewID,'Gender'] = Sewadar.gender
            SewadarDetails[Sewadar.NewID,'TotalCount'] = Sewadar.Total
            SewadarDetails[Sewadar.NewID,'NAME'] = Sewadar.Name


        MasterTable = db(db.MasterSheet.GR_NO == 'SS015298').select()
        for Sewadar in MasterTable:
                 SewadarDetails[Sewadar.GR_NO,'SewadarNewID'] = Sewadar.SewadarNewID
                 SewadarDetails[Sewadar.SewadarNewID,'DEV_DTY'] = Sewadar.DEV_DTY

        DEV_DTY = request.vars.JATHA
#        if auth.user.username == 'admin':
#            DEV_DTY = request.vars.JATHA
#        else:
#            try:
#                DEV_DTY = SewadarDetails[auth.user.username,'DEV_DTY']
#            except:
#                #try:
#                    DEV_DTY = SewadarDetails[SewadarDetails[auth.user.username,'SewadarNewID'],'DEV_DTY']
#                #except:
#                #    TextMessage = 'This ID is not allocated to any Jatha'


        MasterTable = db(db.MasterSheet.DEV_DTY == DEV_DTY).select()
        for Sewadar in MasterTable:
            try:
                #Check if this ID was in SewaSamiti list
                SewadarDetails[Sewadar.SewadarNewID,'NAME']
                SewadarDetails['Sewadars'].append(Sewadar.SewadarNewID)
                SewadarDetails[Sewadar.SewadarNewID,'CANTEEN'] = Sewadar.CANTEEN
                SewadarDetails[Sewadar.SewadarNewID,'DEV_DTY'] = Sewadar.DEV_DTY
                SewadarDetails[Sewadar.SewadarNewID,'OLD_ID'] = Sewadar.GR_NO
            except:
                TextMessage = 'Sewadar ID not in list'


        Fields = [Field('D%02i' %i,'string') for i in range((datetime.datetime.strptime(DateSelectedEnd,'%Y-%m-%d') - datetime.datetime.strptime(DateSelectedStart,'%Y-%m-%d')).days)]


        #Now define the table
        db.define_table('tempAttendanceRegisterTable',
                Field('GR_NO','string'),
                Field('SewadarNewID','string'),
                Field('NAME','string'),
                Field('DEV_DTY','string'),
                Field('CANTEEN','string'),
                Field('TOTAL','integer'),
                Field('STATUS','string'),
                Field('REQD','integer'),
                Field('GENDER','string'),
                *Fields,
                migrate=True,
                redefine=True,
                format='%(SewadarNewID)s')

        db(db.tempAttendanceRegisterTable.id > 0).delete()

        mydict = {}
        mydict.clear()
        excel_headers = {1:'SewadarNewID',2:'GR_NO',3:'NAME',4:'DEV_DTY',5:'CANTEEN',6:'GENDER',7:'STATUS',8:'REQD'}
        dAttendanceRegister.append(excel_headers.values())

        SewadarNumber = 1
        for Sewadar in SewadarDetails['Sewadars']:
            if ((GENTS_REQUIRED - SewadarDetails[Sewadar,'TotalCount']) if (SewadarDetails[Sewadar,'Gender'] == 'Male') else (LADIES_REQUIRED - SewadarDetails[Sewadar,'TotalCount']) ) > 0:
                SewadarNumber = SewadarNumber + 1
                mydict['SewadarNewID'] = Sewadar
                mydict['GR_NO'] = SewadarDetails[Sewadar,'OLD_ID']
                mydict['NAME'] = SewadarDetails[Sewadar,'NAME']
                mydict['DEV_DTY'] = SewadarDetails[Sewadar,'DEV_DTY']
                mydict['CANTEEN'] = SewadarDetails[Sewadar,'CANTEEN']
                mydict['TOTAL'] = SewadarDetails[Sewadar,'TotalCount']
                mydict['GENDER'] = SewadarDetails[Sewadar,'Gender']
                mydict['STATUS'] = ''
                mydict['REQD'] = (GENTS_REQUIRED - SewadarDetails[Sewadar,'TotalCount']) if (SewadarDetails[Sewadar,'Gender'] == 'Male') else (LADIES_REQUIRED - SewadarDetails[Sewadar,'TotalCount'])
                if mydict['REQD'] < 1:
                    mydict['REQD'] = 0
                elif ((SewadarDetails[Sewadar,'Gender'] == 'Male') and mydict['REQD'] > SS_GENTS_REQUIRED):
                    mydict['REQD'] = SS_GENTS_REQUIRED
                elif ((SewadarDetails[Sewadar,'Gender'] == 'Female') and mydict['REQD'] > SS_LADIES_REQUIRED):
                    mydict['REQD'] = SS_LADIES_REQUIRED
                else:
                    pass

                #Write to excel too
                dAttendanceRegister.cell('A' + str(SewadarNumber)).value = mydict['SewadarNewID']
                dAttendanceRegister.cell('B' + str(SewadarNumber)).value = mydict['GR_NO']
                dAttendanceRegister.cell('C' + str(SewadarNumber)).value = mydict['NAME']
                dAttendanceRegister.cell('D' + str(SewadarNumber)).value = mydict['DEV_DTY']
                dAttendanceRegister.cell('E' + str(SewadarNumber)).value = mydict['CANTEEN']
                dAttendanceRegister.cell('F' + str(SewadarNumber)).value = mydict['GENDER']
                dAttendanceRegister.cell('G' + str(SewadarNumber)).value = mydict['STATUS']
                dAttendanceRegister.cell('H' + str(SewadarNumber)).value = mydict['REQD']


                columns = ['tempAttendanceRegisterTable.SewadarNewID','tempAttendanceRegisterTable.GR_NO','tempAttendanceRegisterTable.NAME','tempAttendanceRegisterTable.DEV_DTY','tempAttendanceRegisterTable.CANTEEN','tempAttendanceRegisterTable.REQD']
                orderby=columns

                headers = {'tempAttendanceRegisterTable.SewadarNewID':{'label':T('SewadarNewID'),'class':'','width':12,'truncate':12,'selected': False},
                   'tempAttendanceRegisterTable.GR_NO':{'label':T('GR_NO'),'class':'','width':10,'truncate':10,'selected': False},
                   'tempAttendanceRegisterTable.NAME':{'label':T('NAME'),'class':'','width':10,'truncate': 10,'selected': False},
                   'tempAttendanceRegisterTable.DEV_DTY':{'label':T('DEV_DTY'),'class':'','width':11,'truncate': 11,'selected': False},
                   'tempAttendanceRegisterTable.CANTEEN':{'label':T('CANTEEN'),'class':'','width':13,'truncate': 13,'selected': False},
                   'tempAttendanceRegisterTable.REQD':{'label':T('REQD'),'class':'','width':4,'truncate': 4,'selected': False}
                   }


                for i in range((datetime.datetime.strptime(DateSelectedEnd,'%Y-%m-%d') - datetime.datetime.strptime(DateSelectedStart,'%Y-%m-%d')).days):
                    dateindex = datetime.datetime.strptime(DateSelectedStart,'%Y-%m-%d') + datetime.timedelta(i)
                    columns.append('tempAttendanceRegisterTable.D%02i' %i)
                    headers['tempAttendanceRegisterTable.D%02i' %i] = {'label':datetime.date.strftime(dateindex,'%d\n%m'),'class':'','width':'3','truncate': 3,'selected': False}
                    excel_headers[i+9] = (datetime.date.strftime(dateindex,'%d-%m'))

                    try:
                        mydict['D%02i' %i] = SewadarDetails[Sewadar,'TYPE',i]
                    except:
                        mydict['D%02i' %i] = ''
                    dAttendanceRegister.cell(get_column_letter(i+9) + str(SewadarNumber)).value = mydict['D%02i' %i]

                db.tempAttendanceRegisterTable.insert(**mydict)

        for key in excel_headers.keys():
            dAttendanceRegister.cell(get_column_letter(key) + '1').value = excel_headers[key]

        Register = SQLTABLE(SSDate,headers='fieldname:capitalize')

        datasource = db(db.tempAttendanceRegisterTable).select(orderby=db.tempAttendanceRegisterTable.GR_NO)

        DAttendanceRegisterTable = SQLTABLE(datasource,columns=columns,headers=headers,orderby='GR_NO',_class='datatable')
        TextMessage = 'Attendance Report for ' + DEV_DTY +' from ' + datetime.date.strftime(datetime.datetime.strptime(DateSelectedStart, "%Y-%m-%d"),"%d-%b-%Y") + ' to ' + datetime.date.strftime(datetime.datetime.strptime(DateSelectedEnd, "%Y-%m-%d"),"%d-%b-%Y") + ' '
        ReportDate = datetime.date.strftime(ReportDate,"%d-%m-%Y")
        dworkbook.save(dpath)
        if download == "YES":
            redirect(URL(r=request, f='download_ShortAttendanceRegister'))
        #tables = plugins.powerTable
        #tables.datasource = datasource
        #tables.uitheme = 'cupertino'
        #tables.dtfeatures['sPaginationType'] = 'two button'
        #tables.keycolumn = 'tempAttendanceRegisterTable.id'
        #tables.showkeycolumn = False
        #tables.columns = ['tempAttendanceRegisterTable.GR_NO','tempAttendanceRegisterTable.DEV_DTY','tempAttendanceRegisterTable.CANTEEN']
        #tables.hiddencolumns = ['tempAttendanceRegisterTable.GR_NO','tempAttendanceRegisterTable.DEV_DTY','tempAttendanceRegisterTable.CANTEEN']
        #tables.headers = 'labels'
        #created_table = tables.create()
        #DAttendanceRegisterTable = plugin_powerTable(datasource)


    return dict(DAttendanceRegisterTable=DAttendanceRegisterTable,form=form,TextMessage=TextMessage,ReportDate=ReportDate,SewadarDetails=SewadarDetails)



@auth.requires_login()
def ShortAttendance():
    from gluon.sqlhtml import form_factory
    import datetime
    import time
    import math

    import sys
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.cell import get_column_letter
    from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
    import os
    response.headers['Connection'] =  'keep-alive'
    dworkbook = Workbook()
    dAttendanceRegisterGents = dworkbook.create_sheet(0)
    dAttendanceRegisterGents.title = "GENTS"
    dAttendanceRegisterLadies = dworkbook.create_sheet(0)
    dAttendanceRegisterLadies.title = "LADIES"
    dpath = os.path.join(request.folder,'private','ShortAttendance.xlsx')

    DateSelectedStart = datetime.datetime.today()
    DateSelectedEnd = datetime.datetime.today()
    Register = "EMPTY"
    DAttendanceRegisterTable = "EMPTY"
    datasource = {}
    columns = []
    orderby = []
    headers = {}
    excel_headers = {}
    DEV_DTY = ""
    TextMessage = "Short Attendance Report"
    ReportDate =  0
    GENTS_REQUIRED = 30
    LADIES_REQUIRED = 36
    SS_GENTS_REQUIRED = 30
    SS_LADIES_REQUIRED = 36

    try:
        db.tempAttendanceRegisterTable.drop()
    except:
        pass

    NumColumns = 1;
    SewadarDetails = {'Sewadars':[]}
    form=form_factory(SQLField('Download','string',requires=IS_IN_SET(['YES','NO']),default='YES'),SQLField('CompletionDate','date',default=datetime.datetime.strptime('15-Nov-2017','%d-%b-%Y')),formname='Download')
    if form.accepts(request.vars,session,formname='Download'):
        try:
            os.remove(dpath)
        except:
            pass

        COMPLETION_DATE = datetime.datetime.strftime(datetime.datetime.strptime(request.vars.CompletionDate,'%Y-%m-%d'),'%d-%b-%Y')
        download = request.vars.Download
        DateSelectedStart = datetime.datetime.today() - datetime.timedelta(days=365)
        DateSelectedEnd =  datetime.datetime.today()

        ReportDate = datetime.datetime.today()-datetime.timedelta(days=3001)

        SSDate = db((db.SSAttendanceDate.DutyDate >= DateSelectedStart) & (db.SSAttendanceDate.DutyDate <= DateSelectedEnd)).select()
        for Sewadar in SSDate:
            delta = Sewadar.DutyDate - DateSelectedStart
            SewadarDetails[Sewadar.SewadarNewID,'DAYS',delta.days] = Sewadar.DutyDate
            SewadarDetails[Sewadar.SewadarNewID,'TYPE',delta.days] = Sewadar.Duty_Type
            try:
                (SewadarDetails[Sewadar.SewadarNewID,'DATES']).append(Sewadar.DutyDate)
            except:
                SewadarDetails[Sewadar.SewadarNewID,'DATES'] = [Sewadar.DutyDate]

            if ReportDate < Sewadar.DutyDate:
                ReportDate = Sewadar.DutyDate

        ReportDate = datetime.date.strftime(ReportDate,"%d-%b-%Y")

        SSCount = db(db.SSAttendanceCount).select()
        for Sewadar in SSCount:
            SewadarDetails[Sewadar.NewID,'Gender'] = Sewadar.gender
            SewadarDetails[Sewadar.NewID,'TotalCount'] = Sewadar.Total
            SewadarDetails[Sewadar.NewID,'NAME'] = Sewadar.Name


        MasterTable = db(db.MasterSheet.GR_NO == 'SS015298').select()
        for Sewadar in MasterTable:
                 SewadarDetails[Sewadar.GR_NO,'SewadarNewID'] = Sewadar.SewadarNewID
                 SewadarDetails[Sewadar.SewadarNewID,'DEV_DTY'] = Sewadar.DEV_DTY


        MasterTable = db(db.MasterSheet.id).select()
        for Sewadar in MasterTable:
            try:
                #Check if this ID was in SewaSamiti list
                SewadarDetails[Sewadar.SewadarNewID,'NAME']
                SewadarDetails['Sewadars'].append(Sewadar.SewadarNewID)
                SewadarDetails[Sewadar.SewadarNewID,'CANTEEN'] = Sewadar.CANTEEN
                SewadarDetails[Sewadar.SewadarNewID,'DEV_DTY'] = Sewadar.DEV_DTY
                SewadarDetails[Sewadar.SewadarNewID,'OLD_ID'] = Sewadar.GR_NO
            except:
                TextMessage = 'Sewadar ID not in list'


        mydict = {}
        mydict.clear()
        excel_headers = {1:'S.NO',2:'GR ID',3:'NAME',4:'Attendance required',5:'Last 3 Attendance'}

        SewadarNumberGents = 3
        SewadarNumberLadies = 3
        col_num = 0
        dAttendanceRegisterGents.append(['GENTS'])
        dAttendanceRegisterGents['A1'].font = Font(name='Arial',size=72,bold=True)
        dAttendanceRegisterGents['A1'].alignment = Alignment(horizontal='center',vertical='center')

        dAttendanceRegisterGents.merge_cells('A1:E1')
        dAttendanceRegisterGents['A1'].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
        dAttendanceRegisterGents['B1'].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
        dAttendanceRegisterGents['C1'].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
        dAttendanceRegisterGents['D1'].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
        dAttendanceRegisterGents['E1'].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))

        dAttendanceRegisterGents.append(['ATTENDANCE AS ON ' + ReportDate])
        dAttendanceRegisterGents['A2'].font = Font(name='Arial',size=16,bold=True)
        dAttendanceRegisterGents['A2'].alignment = Alignment(horizontal='center',vertical='center')

        dAttendanceRegisterGents.merge_cells('A2:E2')
        dAttendanceRegisterGents['A2'].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
        dAttendanceRegisterGents['B2'].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
        dAttendanceRegisterGents['C2'].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
        dAttendanceRegisterGents['D2'].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
        dAttendanceRegisterGents['E2'].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))

        dAttendanceRegisterGents.append(excel_headers.values())
        dAttendanceRegisterGents['A3'].font = Font(name='Arial',size=10,bold=True)
        dAttendanceRegisterGents['A3'].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
        dAttendanceRegisterGents['A3'].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
        dAttendanceRegisterGents['B3'].font = Font(name='Arial',size=10,bold=True)
        dAttendanceRegisterGents['B3'].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
        dAttendanceRegisterGents['B3'].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
        dAttendanceRegisterGents['C3'].font = Font(name='Arial',size=10,bold=True)
        dAttendanceRegisterGents['C3'].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
        dAttendanceRegisterGents['C3'].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
        dAttendanceRegisterGents['D3'].font = Font(name='Arial',size=10,bold=True)
        dAttendanceRegisterGents['D3'].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
        dAttendanceRegisterGents['D3'].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
        dAttendanceRegisterGents['E3'].font = Font(name='Arial',size=10,bold=True)
        dAttendanceRegisterGents['E3'].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
        dAttendanceRegisterGents['E3'].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))


        dAttendanceRegisterLadies.append(['LADIES'])
        dAttendanceRegisterLadies['A1'].font = Font(name='Arial',size=72,bold=True)
        dAttendanceRegisterLadies['A1'].alignment = Alignment(horizontal='center',vertical='center')

        dAttendanceRegisterLadies.merge_cells('A1:E1')
        dAttendanceRegisterLadies['A1'].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
        dAttendanceRegisterLadies['B1'].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
        dAttendanceRegisterLadies['C1'].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
        dAttendanceRegisterLadies['D1'].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
        dAttendanceRegisterLadies['E1'].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))

        dAttendanceRegisterLadies.append(['ATTENDANCE AS ON ' + ReportDate])
        dAttendanceRegisterLadies['A2'].font = Font(name='Arial',size=16,bold=True)
        dAttendanceRegisterLadies['A2'].alignment = Alignment(horizontal='center',vertical='center')

        dAttendanceRegisterLadies.merge_cells('A2:E2')
        dAttendanceRegisterLadies['A2'].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
        dAttendanceRegisterLadies['B2'].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
        dAttendanceRegisterLadies['C2'].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
        dAttendanceRegisterLadies['D2'].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
        dAttendanceRegisterLadies['E2'].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))

        dAttendanceRegisterLadies.append(excel_headers.values())
        dAttendanceRegisterLadies['A3'].font = Font(name='Arial',size=10,bold=True)
        dAttendanceRegisterLadies['A3'].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
        dAttendanceRegisterLadies['A3'].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
        dAttendanceRegisterLadies['B3'].font = Font(name='Arial',size=10,bold=True)
        dAttendanceRegisterLadies['B3'].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
        dAttendanceRegisterLadies['B3'].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
        dAttendanceRegisterLadies['C3'].font = Font(name='Arial',size=10,bold=True)
        dAttendanceRegisterLadies['C3'].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
        dAttendanceRegisterLadies['C3'].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
        dAttendanceRegisterLadies['D3'].font = Font(name='Arial',size=10,bold=True)
        dAttendanceRegisterLadies['D3'].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
        dAttendanceRegisterLadies['D3'].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
        dAttendanceRegisterLadies['E3'].font = Font(name='Arial',size=10,bold=True)
        dAttendanceRegisterLadies['E3'].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
        dAttendanceRegisterLadies['E3'].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))



        for Sewadar in SewadarDetails['Sewadars']:
            mydict['SewadarNewID'] = Sewadar
            mydict['GR_NO'] = SewadarDetails[Sewadar,'OLD_ID']
            mydict['NAME'] = SewadarDetails[Sewadar,'NAME']
            mydict['DEV_DTY'] = SewadarDetails[Sewadar,'DEV_DTY']
            mydict['CANTEEN'] = SewadarDetails[Sewadar,'CANTEEN']
            mydict['TOTAL'] = SewadarDetails[Sewadar,'TotalCount']
            mydict['GENDER'] = SewadarDetails[Sewadar,'Gender']
            try:
                SewadarDetails[Sewadar,'DATES'].sort(reverse=True)
                mydict['DATES'] = SewadarDetails[Sewadar,'DATES']
            except:
                mydict['DATES'] = "EMPTY"
            mydict['STATUS'] = ''
            mydict['REQD'] = (GENTS_REQUIRED - SewadarDetails[Sewadar,'TotalCount']) if (SewadarDetails[Sewadar,'Gender'] == 'Male') else (LADIES_REQUIRED - SewadarDetails[Sewadar,'TotalCount'])
            if mydict['REQD'] < 1:
                mydict['REQD'] = 0
            elif ((SewadarDetails[Sewadar,'Gender'] == 'Male') and mydict['REQD'] > SS_GENTS_REQUIRED):
                mydict['REQD'] = SS_GENTS_REQUIRED
            elif ((SewadarDetails[Sewadar,'Gender'] == 'Female') and mydict['REQD'] > SS_LADIES_REQUIRED):
                mydict['REQD'] = SS_LADIES_REQUIRED
            else:
                pass



            #Write to excel too
            if mydict['REQD'] > 0 and mydict['GR_NO'] != "G01956" and mydict['GR_NO'] != "G13133" :
                try:
                    DateList = ",".join([datetime.date.strftime(p,"%d/%m") for p in mydict['DATES'][0:3]])
                except:
                    DateList = ""

                if mydict['GENDER'] == "Male":
                    dAttendanceRegisterGents.cell('A' + str(SewadarNumberGents+1)).value = SewadarNumberGents-2
                    dAttendanceRegisterGents.cell('B' + str(SewadarNumberGents+1)).value = mydict['GR_NO']
                    dAttendanceRegisterGents.cell('C' + str(SewadarNumberGents+1)).value = mydict['NAME']
                    dAttendanceRegisterGents.cell('D' + str(SewadarNumberGents+1)).value = mydict['REQD']
                    dAttendanceRegisterGents.cell('E' + str(SewadarNumberGents+1)).value = DateList

                    dAttendanceRegisterGents['A' + str(SewadarNumberGents+1)].font = Font(name='Calibri',size=11,bold=False)
                    dAttendanceRegisterGents['A' + str(SewadarNumberGents+1)].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
                    dAttendanceRegisterGents['A' + str(SewadarNumberGents+1)].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
                    dAttendanceRegisterGents['B' + str(SewadarNumberGents+1)].font = Font(name='Calibri',size=10,bold=False)
                    dAttendanceRegisterGents['B' + str(SewadarNumberGents+1)].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
                    dAttendanceRegisterGents['B' + str(SewadarNumberGents+1)].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
                    dAttendanceRegisterGents['C' + str(SewadarNumberGents+1)].font = Font(name='Calibri',size=10,bold=False)
                    dAttendanceRegisterGents['C' + str(SewadarNumberGents+1)].alignment = Alignment(horizontal='left',vertical='center',wrap_text=True)
                    dAttendanceRegisterGents['C' + str(SewadarNumberGents+1)].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
                    dAttendanceRegisterGents['D' + str(SewadarNumberGents+1)].font = Font(name='Calibri',size=11,bold=False)
                    dAttendanceRegisterGents['D' + str(SewadarNumberGents+1)].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
                    dAttendanceRegisterGents['D' + str(SewadarNumberGents+1)].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
                    dAttendanceRegisterGents['E' + str(SewadarNumberGents+1)].font = Font(name='Calibri',size=10,bold=False)
                    dAttendanceRegisterGents['E' + str(SewadarNumberGents+1)].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
                    dAttendanceRegisterGents['E' + str(SewadarNumberGents+1)].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))

                    dAttendanceRegisterGents.row_dimensions[SewadarNumberGents+1].height = 15

                    SewadarNumberGents = SewadarNumberGents + 1
                else:
                    dAttendanceRegisterLadies.cell('A' + str(SewadarNumberLadies+1)).value = SewadarNumberLadies-2
                    dAttendanceRegisterLadies.cell('B' + str(SewadarNumberLadies+1)).value = mydict['GR_NO']
                    dAttendanceRegisterLadies.cell('C' + str(SewadarNumberLadies+1)).value = mydict['NAME']
                    dAttendanceRegisterLadies.cell('D' + str(SewadarNumberLadies+1)).value = mydict['REQD']
                    dAttendanceRegisterLadies.cell('E' + str(SewadarNumberLadies+1)).value = DateList

                    dAttendanceRegisterLadies['A' + str(SewadarNumberLadies+1)].font = Font(name='Calibri',size=11,bold=False)
                    dAttendanceRegisterLadies['A' + str(SewadarNumberLadies+1)].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
                    dAttendanceRegisterLadies['A' + str(SewadarNumberLadies+1)].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
                    dAttendanceRegisterLadies['B' + str(SewadarNumberLadies+1)].font = Font(name='Calibri',size=10,bold=False)
                    dAttendanceRegisterLadies['B' + str(SewadarNumberLadies+1)].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
                    dAttendanceRegisterLadies['B' + str(SewadarNumberLadies+1)].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
                    dAttendanceRegisterLadies['C' + str(SewadarNumberLadies+1)].font = Font(name='Calibri',size=10,bold=False)
                    dAttendanceRegisterLadies['C' + str(SewadarNumberLadies+1)].alignment = Alignment(horizontal='left',vertical='center',wrap_text=True)
                    dAttendanceRegisterLadies['C' + str(SewadarNumberLadies+1)].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
                    dAttendanceRegisterLadies['D' + str(SewadarNumberLadies+1)].font = Font(name='Calibri',size=11,bold=False)
                    dAttendanceRegisterLadies['D' + str(SewadarNumberLadies+1)].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
                    dAttendanceRegisterLadies['D' + str(SewadarNumberLadies+1)].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
                    dAttendanceRegisterLadies['E' + str(SewadarNumberLadies+1)].font = Font(name='Calibri',size=10,bold=False)
                    dAttendanceRegisterLadies['E' + str(SewadarNumberLadies+1)].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
                    dAttendanceRegisterLadies['E' + str(SewadarNumberLadies+1)].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))

                    dAttendanceRegisterLadies.row_dimensions[SewadarNumberLadies+1].height = 15

                    SewadarNumberLadies = SewadarNumberLadies + 1



        dAttendanceRegisterGents.column_dimensions['A'].width = 6.57
        dAttendanceRegisterGents.column_dimensions['B'].width = 13
        dAttendanceRegisterGents.column_dimensions['C'].width = 23.29
        dAttendanceRegisterGents.column_dimensions['D'].width = 11
        dAttendanceRegisterGents.column_dimensions['E'].width = 16.14

        dAttendanceRegisterGents.row_dimensions[3].height = 25.5


        dAttendanceRegisterLadies.column_dimensions['A'].width = 6.57
        dAttendanceRegisterLadies.column_dimensions['B'].width = 13
        dAttendanceRegisterLadies.column_dimensions['C'].width = 23.29
        dAttendanceRegisterLadies.column_dimensions['D'].width = 11
        dAttendanceRegisterLadies.column_dimensions['E'].width = 16.14

        dAttendanceRegisterLadies.row_dimensions[3].height = 25.5

        dAttendanceRegisterGents.add_print_title(2,3,rows_or_cols='row')
        dAttendanceRegisterLadies.add_print_title(2,3,rows_or_cols='row')

        dAttendanceRegisterGents.header_footer.left_footer.text = 'LAST DATE TO COMPLETE ATTENDANCE IS ' + COMPLETION_DATE
        dAttendanceRegisterLadies.header_footer.left_footer.text = 'LAST DATE TO COMPLETE ATTENDANCE IS ' + COMPLETION_DATE


        dAttendanceRegisterGents.page_setup.paperSize = dAttendanceRegisterGents.PAPERSIZE_A4
        dAttendanceRegisterGents.page_margins.left = 0.4/2.54
        dAttendanceRegisterGents.page_margins.right = 1.8/2.54
        dAttendanceRegisterGents.page_margins.bottom = 1.4/2.54
        dAttendanceRegisterGents.page_margins.top = 0.5/2.54
        dAttendanceRegisterGents.page_margins.footer = 0.5/2.54
        dAttendanceRegisterLadies.page_setup.paperSize = dAttendanceRegisterLadies.PAPERSIZE_A4
        dAttendanceRegisterLadies.page_margins.left = 0.4/2.54
        dAttendanceRegisterLadies.page_margins.right = 1.8/2.54
        dAttendanceRegisterLadies.page_margins.bottom = 1.4/2.54
        dAttendanceRegisterLadies.page_margins.top = 0.5/2.54
        dAttendanceRegisterLadies.page_margins.footer = 0.5/2.54
        TextMessage = 'Attendance Report as on ' + ReportDate
        dworkbook.save(dpath)
        if download == "YES":
            redirect(URL(r=request, f='download_ShortAttendance'))

    return dict(DAttendanceRegisterTable=DAttendanceRegisterTable,form=form,TextMessage=TextMessage,ReportDate=ReportDate,SewadarDetails=SewadarDetails)

@auth.requires_login()
def AttendanceRegisterAll():
    from gluon.sqlhtml import form_factory
    import datetime
    import time

    db.commit()
    message = 'Schedular based'

    form=form_factory(SQLField('DateStart','date',default=datetime.datetime.today()-datetime.timedelta(days=31)),SQLField('DateEnd','date',default=datetime.datetime.today()),formname='DateSelect')

    if form.accepts(request.vars,session,formname='DateSelect'):
        DateSelectedStart = request.vars.DateStart
        DateSelectedEnd = request.vars.DateEnd
        from datetime import timedelta as timed
        scheduler.queue_task('AttendanceRegisterScheduledAll',pvars={'DateSelectedStart':DateSelectedStart,'DateSelectedEnd':DateSelectedEnd},
            start_time=request.now + timed(seconds=1),
            timeout = 6000)
    return dict(form=form)

@auth.requires_login()
def AttendanceRegister():
    from gluon.sqlhtml import form_factory
    import datetime
    import time

    from openpyxl import Workbook
    from openpyxl.cell import get_column_letter
    import os
    dpath = os.path.join(request.folder,'private','log_AttendanceRegister')
    logf = open(dpath,'wb')

    response.headers['Connection'] =  'keep-alive'
    dworkbook = Workbook()
    dAttendanceRegister = dworkbook.create_sheet(0)
    dpath = os.path.join(request.folder,'private','AttendanceRegister.xlsx')

    DateSelectedStart = datetime.datetime.today()
    DateSelectedEnd = datetime.datetime.today()
    Register = "EMPTY"
    DAttendanceRegisterTable = "EMPTY"
    datasource = {}
    columns = []
    orderby = []
    headers = {}
    excel_headers = {}
    DEV_DTY = ""
    TextMessage = "Jatha Wise Attendance Report"
    ReportDate =  0
    GENTS_REQUIRED = 30
    LADIES_REQUIRED = 36
    SS_GENTS_REQUIRED = 30
    SS_LADIES_REQUIRED = 36

    try:
        db.tempAttendanceRegisterTable.drop()
    except:
        pass

    SewadarDetails = {'Sewadars':[]}
#    form=form_factory(SQLField('JATHA','string',requires=IS_IN_DB(db,'MasterSheet.DEV_DTY','%(DEV_DTY)s')),SQLField('DateStart','date',default=datetime.datetime.today()-datetime.timedelta(days=31)),SQLField('DateEnd','date',default=datetime.datetime.today()),formname='DateSelect')
    my_list = db().select(db.MasterSheet.DEV_DTY, distinct=True).as_list()
    myjathalist = []
    for mydict in my_list:
        myjathalist.append(mydict['DEV_DTY'])

    myjathalist.sort()
    form=form_factory(SQLField('JATHA','string',requires=IS_IN_SET(myjathalist)),SQLField('DateStart','date',default=(datetime.datetime.today()-datetime.timedelta(days=31)).date()),SQLField('DateEnd','date',default=datetime.datetime.today().date()),SQLField('Download','string',requires=IS_IN_SET(['YES','NO']),default='NO'),formname='DateSelect')
    if form.accepts(request.vars,session,formname='DateSelect'):
        DateSelectedStart = request.vars.DateStart
        DateSelectedEnd = request.vars.DateEnd
        Fields = [Field('D%02i' %i,'string') for i in range((datetime.datetime.strptime(DateSelectedEnd,'%Y-%m-%d') - datetime.datetime.strptime(DateSelectedStart,'%Y-%m-%d')).days)]

        #Now define the table
        db.define_table('tempAttendanceRegisterTable',
                Field('SewadarNewID','string'),
                Field('NAME','string'),
                Field('DEV_DTY','string'),
                Field('CANTEEN','string'),
                Field('TOTAL','integer'),
                Field('STATUS','string'),
                Field('REQD','integer'),
                Field('GENDER','string'),
                Field('WWDATES','string'),
                *Fields,
                migrate=True,
                fake_migrate=True,
                redefine=True,
                format='%(SewadarNewID)s')

        db(db.tempAttendanceRegisterTable.id > 0).delete()
        try:
            os.remove(dpath)
        except:
            pass

        download = request.vars.Download
        ReportDate = datetime.datetime.today()-datetime.timedelta(days=3001)

        SSDate = db((db.SSAttendanceDate.DutyDate >= DateSelectedStart) & (db.SSAttendanceDate.DutyDate <= DateSelectedEnd)).select()
        for Sewadar in SSDate:
            delta = Sewadar.DutyDate - datetime.datetime.strptime(DateSelectedStart,'%Y-%m-%d')
            SewadarDetails[(Sewadar.SewadarNewID).replace('BH0011',''),'DAYS',delta.days] = Sewadar.DutyDate
            SewadarDetails[(Sewadar.SewadarNewID).replace('BH0011',''),'TYPE',delta.days] = Sewadar.Duty_Type
            if ReportDate < Sewadar.DutyDate:
                ReportDate = Sewadar.DutyDate

        SSCount = db(db.SSAttendanceCount).select()
        for Sewadar in SSCount:
            SewadarDetails[(Sewadar.NewID).replace('BH0011',''),'Gender'] = Sewadar.gender
            SewadarDetails[(Sewadar.NewID).replace('BH0011',''),'TotalCount'] = Sewadar.Total
            SewadarDetails[(Sewadar.NewID).replace('BH0011',''),'NAME'] = Sewadar.Name


        MasterTable = db(db.MasterSheet.SewadarNewID == 'GB4530').select()
        for Sewadar in MasterTable:
                 SewadarDetails[Sewadar.SewadarNewID,'DEV_DTY'] = Sewadar.DEV_DTY

        DEV_DTY = request.vars.JATHA
#        if auth.user.username == 'admin':
#            DEV_DTY = request.vars.JATHA
#        else:
#            try:
#                DEV_DTY = SewadarDetails[auth.user.username,'DEV_DTY']
#            except:
#                #try:
#                    DEV_DTY = SewadarDetails[SewadarDetails[auth.user.username,'SewadarNewID'],'DEV_DTY']
#                #except:
#                #    TextMessage = 'This ID is not allocated to any Jatha'


        MasterTable = db(db.MasterSheet.DEV_DTY == DEV_DTY).select()
        for Sewadar in MasterTable:
            try:
                #Check if this ID was in SewaSamiti list
                SewadarDetails[Sewadar.SewadarNewID,'NAME']
                SewadarDetails['Sewadars'].append(Sewadar.SewadarNewID)
                SewadarDetails[Sewadar.SewadarNewID,'CANTEEN'] = Sewadar.CANTEEN
                SewadarDetails[Sewadar.SewadarNewID,'DEV_DTY'] = Sewadar.DEV_DTY
            except:
                TextMessage = 'Sewadar ID not in list'



        excel_headers = {1:'SewadarNewID',2:'NAME',3:'DEV_DTY',4:'CANTEEN',5:'GENDER',6:'STATUS',7:'REQD',8:'WWDATES'}
        dAttendanceRegister.append(excel_headers.values())

        SewadarNumber = 1
        for Sewadar in SewadarDetails['Sewadars']:
            mydict = {}
            SewadarNumber = SewadarNumber + 1
            mydict['SewadarNewID'] = Sewadar
            mydict['NAME'] = SewadarDetails[Sewadar,'NAME']
            mydict['DEV_DTY'] = SewadarDetails[Sewadar,'DEV_DTY']
            mydict['CANTEEN'] = SewadarDetails[Sewadar,'CANTEEN']
            mydict['TOTAL'] = SewadarDetails[Sewadar,'TotalCount']
            mydict['GENDER'] = SewadarDetails[Sewadar,'Gender']
            mydict['STATUS'] = ''
            mydict['REQD'] = (GENTS_REQUIRED - SewadarDetails[Sewadar,'TotalCount']) if (SewadarDetails[Sewadar,'Gender'] == 'Male') else (LADIES_REQUIRED - SewadarDetails[Sewadar,'TotalCount'])
            if mydict['REQD'] < 1:
                mydict['REQD'] = 0
            elif (((SewadarDetails[Sewadar,'Gender'] == 'Male')) and mydict['REQD'] > SS_GENTS_REQUIRED):
                mydict['REQD'] = SS_GENTS_REQUIRED
            elif (((SewadarDetails[Sewadar,'Gender'] == 'Female')) and mydict['REQD'] > SS_LADIES_REQUIRED):
                mydict['REQD'] = SS_LADIES_REQUIRED
            else:
                pass



            #Write to excel too
            dAttendanceRegister.cell('A' + str(SewadarNumber)).value = mydict['SewadarNewID']
            dAttendanceRegister.cell('B' + str(SewadarNumber)).value = mydict['NAME']
            dAttendanceRegister.cell('C' + str(SewadarNumber)).value = mydict['DEV_DTY']
            dAttendanceRegister.cell('D' + str(SewadarNumber)).value = mydict['CANTEEN']
            dAttendanceRegister.cell('E' + str(SewadarNumber)).value = mydict['GENDER']
            dAttendanceRegister.cell('F' + str(SewadarNumber)).value = mydict['STATUS']
            dAttendanceRegister.cell('G' + str(SewadarNumber)).value = mydict['REQD']


            columns = ['tempAttendanceRegisterTable.SewadarNewID','tempAttendanceRegisterTable.NAME','tempAttendanceRegisterTable.DEV_DTY','tempAttendanceRegisterTable.CANTEEN','tempAttendanceRegisterTable.REQD']
            orderby=columns

            headers = {'tempAttendanceRegisterTable.SewadarNewID':{'label':T('SewadarNewID'),'class':'','width':12,'truncate':12,'selected': False},
               'tempAttendanceRegisterTable.NAME':{'label':T('NAME'),'class':'','width':10,'truncate': 10,'selected': False},
               'tempAttendanceRegisterTable.DEV_DTY':{'label':T('DEV_DTY'),'class':'','width':11,'truncate': 11,'selected': False},
               'tempAttendanceRegisterTable.CANTEEN':{'label':T('CANTEEN'),'class':'','width':13,'truncate': 13,'selected': False},
               'tempAttendanceRegisterTable.REQD':{'label':T('REQD'),'class':'','width':4,'truncate': 4,'selected': False}
               }


            for i in range((datetime.datetime.strptime(DateSelectedEnd,'%Y-%m-%d') - datetime.datetime.strptime(DateSelectedStart,'%Y-%m-%d')).days):
                dateindex = datetime.datetime.strptime(DateSelectedStart,'%Y-%m-%d') + datetime.timedelta(i)
                columns.append('tempAttendanceRegisterTable.D%02i' %i)
                headers['tempAttendanceRegisterTable.D%02i' %i] = {'label':datetime.date.strftime(dateindex,'%d\n%m'),'class':'','width':'3','truncate': 3,'selected': False}
                excel_headers[i+10] = (datetime.date.strftime(dateindex,'%d-%m'))

                try:
                    mydict['D%02i' %i] = SewadarDetails[Sewadar,'TYPE',i]
                    if mydict['D%02i' %i] == 'W':
                        try:
                            mydict["WWDATES"] = mydict["WWDATES"] + '\n' +  excel_headers[i+10]
                        except:
                            mydict["WWDATES"] = excel_headers[i+10]
                except:
                    mydict['D%02i' %i] = ''
                dAttendanceRegister.cell(get_column_letter(i+10) + str(SewadarNumber)).value = mydict['D%02i' %i]
                try:
                    dAttendanceRegister.cell(get_column_letter(9) + str(SewadarNumber)).value = mydict["WWDATES"]
                except:
                    pass

            db.tempAttendanceRegisterTable.insert(**mydict)

        for key in excel_headers.keys():
            dAttendanceRegister.cell(get_column_letter(key) + '1').value = excel_headers[key]

        Register = SQLTABLE(SSDate,headers='fieldname:capitalize')

        datasource = db(db.tempAttendanceRegisterTable).select(orderby=db.tempAttendanceRegisterTable.SewadarNewID)

        DAttendanceRegisterTable = SQLTABLE(datasource,columns=columns,headers=headers,orderby='SewadarNewID',_class='datatable')
        TextMessage = 'Attendance Report for ' + DEV_DTY +' from ' + datetime.date.strftime(datetime.datetime.strptime(DateSelectedStart, "%Y-%m-%d"),"%d-%b-%Y") + ' to ' + datetime.date.strftime(datetime.datetime.strptime(DateSelectedEnd, "%Y-%m-%d"),"%d-%b-%Y") + ' '
        ReportDate = datetime.date.strftime(ReportDate,"%d-%m-%Y")
        dworkbook.save(dpath)
        if download == "YES":
            redirect(URL(r=request, f='download_AttendanceRegister'))
        #tables = plugins.powerTable
        #tables.datasource = datasource
        #tables.uitheme = 'cupertino'
        #tables.dtfeatures['sPaginationType'] = 'two button'
        #tables.keycolumn = 'tempAttendanceRegisterTable.id'
        #tables.showkeycolumn = False
        #tables.columns = ['tempAttendanceRegisterTable.GR_NO','tempAttendanceRegisterTable.DEV_DTY','tempAttendanceRegisterTable.CANTEEN']
        #tables.hiddencolumns = ['tempAttendanceRegisterTable.GR_NO','tempAttendanceRegisterTable.DEV_DTY','tempAttendanceRegisterTable.CANTEEN']
        #tables.headers = 'labels'
        #created_table = tables.create()
        #DAttendanceRegisterTable = plugin_powerTable(datasource)


    return dict(DAttendanceRegisterTable=DAttendanceRegisterTable,form=form,TextMessage=TextMessage,ReportDate=ReportDate,SewadarDetails=SewadarDetails)


def user():
    """
    exposes:
    http://..../[app]/default/user/login
    http://..../[app]/default/user/logout
    http://..../[app]/default/user/register
    http://..../[app]/default/user/profile
    http://..../[app]/default/user/retrieve_password
    http://..../[app]/default/user/change_password
    http://..../[app]/default/user/manage_users (requires membership in
    use @auth.requires_login()
        @auth.requires_membership('group name')
        @auth.requires_permission('read','table name',record_id)
    to decorate functions that need access control
    """
    return dict(form=auth())

@cache.action()
def download():
    """
    allows downloading of uploaded files
    http://..../[app]/default/download/[filename]
    """
    return response.download(request, db)

def download_ShortAttendance():
    import os
    dpath = os.path.join(request.folder,'private','ShortAttendance.xlsx')
    return response.stream(open(dpath,'rb'), chunk_size=10**6)

def download_AttendanceRegister():
    import os
    dpath = os.path.join(request.folder,'private','AttendanceRegister.xlsx')
    return response.stream(open(dpath,'rb'), chunk_size=10**6)

def download_ShortAttendanceRegister():
    import os
    dpath = os.path.join(request.folder,'private','ShortAttendanceRegister.xlsx')
    return response.stream(open(dpath,'rb'), chunk_size=10**6)

def call():
    """
    exposes services. for example:
    http://..../[app]/default/call/jsonrpc
    decorate with @services.jsonrpc the functions to expose
    supports xml, json, xmlrpc, jsonrpc, amfrpc, rss, csv
    """
    return service()



def data():
    """
    http://..../[app]/default/data/tables
    http://..../[app]/default/data/create/[table]
    http://..../[app]/default/data/read/[table]/[id]
    http://..../[app]/default/data/update/[table]/[id]
    http://..../[app]/default/data/delete/[table]/[id]
    http://..../[app]/default/data/select/[table]
    http://..../[app]/default/data/search/[table]
    but URLs must be signed, i.e. linked with
      A('table',_href=URL('data/tables',user_signature=True))
    or with the signed load operator
      LOAD('default','data.load',args='tables',ajax=True,user_signature=True)
    """
    return dict(form=crud())

def MachineSewaSamitiDifference():
    from gluon.sqlhtml import form_factory
    import datetime
    import time

    DateSelectedStart = datetime.datetime.today()
    DateSelectedEnd = datetime.datetime.today()
    DAY_END_TIME = 19

    SSAttendance = 'Not queried'
    MachineAttendance = 'Not queried'
    MachineDifference = 'Not queried'

    form=form_factory(SQLField('DAY_END_TIME','string',default=19,requires=IS_IN_SET(range(0,25,1))),SQLField('DateStart','date',default=datetime.datetime.today()-datetime.timedelta(days=31)),SQLField('DateEnd','date',default=datetime.datetime.today()),formname='DateSelect')

    if form.accepts(request.vars,session,formname='DateSelect'):
        DateSelectedStart = request.vars.DateStart
        DateSelectedEnd = request.vars.DateEnd
        DAY_END_TIME = request.vars.DAY_END_TIME
        SSDate = db((db.SSAttendanceDate.DutyDate >= (datetime.datetime.strptime(DateSelectedStart,'%Y-%m-%d').replace(hour=0, minute=0, second=0, microsecond=0))) & (db.SSAttendanceDate.DutyDate <= (datetime.datetime.strptime(DateSelectedEnd,'%Y-%m-%d').replace(hour=23, minute=59, second=59, microsecond=999)))).select('SewadarNewID','DutyDate')
        MachineDate = db((db.MachineAttendance.DATETIME >= (datetime.datetime.strptime(DateSelectedStart,'%Y-%m-%d') - datetime.timedelta(hours=((24 - int(DAY_END_TIME)) % 24)))) & (db.MachineAttendance.DATETIME <= (datetime.datetime.strptime(DateSelectedEnd,'%Y-%m-%d') + datetime.timedelta(hours=int(DAY_END_TIME))))).select('GRNO','NewGRNO','DATETIME')

        SSAttendanceDictionary = {}
        for SSEntry in SSDate:
            SSAttendanceDictionary[SSEntry.SewadarNewID,Roundoffdate(SSEntry.DutyDate,23)] = SSEntry.DutyDate

        MachineAttendanceDictionary = {}
        MachineAttendanceAdditional = {}


        try:
            db.tempMachineAttendanceAdditional.drop()
        except:
            pass

        #Now define the table
        db.define_table('tempMachineAttendanceAdditional',
                Field('GRNO','string'),
                Field('NewGRNO','string'),
                Field('DutyDate','datetime'),
                Field('DutyDateList','list:string'),
                migrate=True,
                redefine=True,
                format='%(NewGRNO)s')

        db(db.tempMachineAttendanceAdditional.id > 0).delete()

        for MEntry in MachineDate:
            #Select earliest entry
            try:
                if MachineAttendanceDictionary[MEntry.NewGRNO,Roundoffdate(MEntry.DATETIME,DAY_END_TIME)] > MEntry.DATETIME:
                    MachineAttendanceDictionary[MEntry.NewGRNO,Roundoffdate(MEntry.DATETIME,DAY_END_TIME)] = MEntry.DATETIME
            except:
                MachineAttendanceDictionary[MEntry.NewGRNO,Roundoffdate(MEntry.DATETIME,DAY_END_TIME)] = MEntry.DATETIME

            try:
                MachineAttendanceDictionary[MEntry.NewGRNO,'LIST',Roundoffdate(MEntry.DATETIME,DAY_END_TIME)].append(MEntry.DATETIME)
            except:
                MachineAttendanceDictionary[MEntry.NewGRNO,'LIST',Roundoffdate(MEntry.DATETIME,DAY_END_TIME)] = [MEntry.DATETIME]

            #Find difference
            try:
                a = SSAttendanceDictionary[MEntry.NewGRNO,Roundoffdate(MEntry.DATETIME,DAY_END_TIME)]
            except:
                try:
                    MachineAttendanceAdditional[MEntry.NewGRNO,MEntry.GRNO,Roundoffdate(MEntry.DATETIME,DAY_END_TIME)].append(MEntry.DATETIME)
                except:
                    MachineAttendanceAdditional[MEntry.NewGRNO,MEntry.GRNO,Roundoffdate(MEntry.DATETIME,DAY_END_TIME)] = [MEntry.DATETIME]



        for key, value in MachineAttendanceAdditional.iteritems():
            MachineAttendanceAdditionalDictionary = {}
            NewGRNO, GRNO, DutyDate = key
            MachineAttendanceAdditionalDictionary['NewGRNO'] = NewGRNO
            MachineAttendanceAdditionalDictionary['GRNO'] = GRNO
            MachineAttendanceAdditionalDictionary['DutyDate'] = DutyDate
            MachineAttendanceAdditionalDictionary['DutyDateList'] = value
            db.tempMachineAttendanceAdditional.insert(**MachineAttendanceAdditionalDictionary)


        columns = ['SSAttendanceDate.SewadarNewID','SSAttendanceDate.DutyDate']
        orderby=columns

        headers = {'SSAttendanceDate.SewadarNewID':{'label':T('NewId'),'class':'','width':12,'truncate':12,'selected': False},
                   'SSAttendanceDate.DutyDate':{'label':T('Duty Date'),'class':'','width':12,'truncate':4,'selected': False}
                  }
        SSAttendance = SQLTABLE(SSDate)
        #SSAttendance = SQLTABLE(SSDate,columns=columns,headers=headers,orderby=orderby,_class='datatable')

        columns = ['MachineAttendance.GRNO','MachineAttendance.DUTY_DATE']
        orderby=columns

        headers = {'MachineAttendance.NewGRNO':{'label':T('NewId'),'class':'','width':12,'truncate':12,'selected': False},
                   'MachineAttendance.DATETIME':{'label':T('Duty Date'),'class':'','width':12,'truncate':4,'selected': False}
                  }

        MachineAttendance = SQLTABLE(MachineDate)
        #MachineAttendance = SQLTABLE(MachineDate,columns=columns,headers=headers,orderby=orderby,_class='datatable')

        #Machine Extra Attendance
        MachineDifference = SQLTABLE(db(db.tempMachineAttendanceAdditional.id > 0).select())

    return dict(form=form,MachineAttendance=MachineAttendance,SSAttendance=SSAttendance,MachineDifference=MachineDifference)

def user():
    """
    exposes:
    http://..../[app]/default/user/login
    http://..../[app]/default/user/logout
    http://..../[app]/default/user/register
    http://..../[app]/default/user/profile
    http://..../[app]/default/user/retrieve_password
    http://..../[app]/default/user/change_password
    http://..../[app]/default/user/manage_users (requires membership in
    use @auth.requires_login()
        @auth.requires_membership('group name')
        @auth.requires_permission('read','table name',record_id)
    to decorate functions that need access control
    """
    return dict(form=auth())

@cache.action()
def download():
    """
    allows downloading of uploaded files
    http://..../[app]/default/download/[filename]
    """
    return response.download(request, db)


def call():
    """
    exposes services. for example:
    http://..../[app]/default/call/jsonrpc
    decorate with @services.jsonrpc the functions to expose
    supports xml, json, xmlrpc, jsonrpc, amfrpc, rss, csv
    """
    return service()


@auth.requires_login()
def CardList():
    SSCount = db(db.SSAttendanceCount).select()
    CardList = db(db.CardList).select()
    NewGRNOs = []
    OldGRNOs = []
    OldtoNew = {}
    for Sewadar in SSCount:
        keynew = Sewadar.NewID + ':' + Sewadar.OldSewadarid + ':' + Sewadar.Name
        keyold = Sewadar.OldSewadarid + ':' + Sewadar.NewID + ':' + Sewadar.Name
        OldtoNew[Sewadar.OldSewadarid] = Sewadar.NewID
        NewGRNOs.append(keynew)
        OldGRNOs.append(keyold)

    NewGRNOs = sorted(set(NewGRNOs))
    OldGRNOs = sorted(set(OldGRNOs))

    AllCardList = []
    AllCard = db(db.AllCardList).select()
    for Card in AllCard:
        AllCardList.append(Card.PROXIMITY_CARDNUMBER)

    AllCardList = sorted(set(AllCardList))
    from gluon.sqlhtml import form_factory
    #L- form=form_factory(SQLField('SewadarNewId','string',default="",requires=IS_IN_SET(NewGRNOs)),SQLField('A-Card','string',default="",requires=[IS_MATCH('^00\d{8}','Not a valid ACARD number'),IS_LENGTH(10)]),formname='CardDetails')
    form=form_factory(SQLField('SewadarNewId','string',default="",requires=IS_IN_SET(NewGRNOs)),SQLField('A_Card','string',default="",requires=IS_MATCH('^00\d{8}$')),formname='CardDetails')
    if form.accepts(request.vars,session,keepvalues=True,formname='CardDetails'):
        SewadarNewID = re.sub(":.*","",request.vars.SewadarNewId)
        db(db.CardList.PROXIMITY_CARDNUMBER == request.vars.A_Card).delete()
        db(db.AllCardList.PROXIMITY_CARDNUMBER == request.vars.A_Card).delete()
        db(db.CardList.SewadarNewID == SewadarNewID).delete()
        db.CardList.update_or_insert(**dict(SewadarNewID=SewadarNewID,PROXIMITY_CARDNUMBER=request.vars.A_Card))
        db.AllCardList.update_or_insert(**dict(PROXIMITY_CARDNUMBER=request.vars.A_Card))

    #form2=form_factory(SQLField('SewadarOldId','string',default="",requires=IS_IN_SET(OldGRNOs)),Field('A_Card','string',default="",widget=SQLFORM.widgets.autocomplete(request,db.AllCardList.PROXIMITY_CARDNUMBER, limitby=(0,10), min_length=4)),formname='OldCardDetails')
    form2=form_factory(SQLField('SewadarOldId','string',default="",requires=IS_IN_SET(OldGRNOs)),SQLField('A_Card','string',default="",requires=IS_MATCH('^00\d{8}$')),formname='OldCardDetails')
    if form2.accepts(request.vars,session,keepvalues=True,formname='OldCardDetails'):
        match = re.search(r'.*:(.*):.*',request.vars.SewadarOldId)
        SewadarNewID = match.group(1)
        db(db.CardList.PROXIMITY_CARDNUMBER == request.vars.A_Card).delete()
        db(db.AllCardList.PROXIMITY_CARDNUMBER == request.vars.A_Card).delete()
        db(db.CardList.SewadarNewID == SewadarNewID).delete()
        db.CardList.update_or_insert(**dict(SewadarNewID=SewadarNewID,PROXIMITY_CARDNUMBER=request.vars.A_Card))
        db.AllCardList.update_or_insert(**dict(PROXIMITY_CARDNUMBER=request.vars.A_Card))

    form1=form_factory(SQLField('Generate','string',default="",requires=IS_IN_SET(['YES','NO'])),formname='Generate')
    if form1.accepts(request.vars,session,keepvalues=True,formname='Generate'):
        redirect(URL(r=request, f='download_MachineFiles'))

    return dict(form=form,form1=form1,form2=form2)


def download_MachineFiles():
    import os

    CardList = db(db.CardList).select()

    SSCountDict = {}
    SSCount = db(db.SSAttendanceCount).select()
    for Sewadar in SSCount:
        SSCountDict[Sewadar.NewID,'Gender'] = Sewadar.gender
        SSCountDict[Sewadar.NewID,'NAME'] = Sewadar.Name
        SSCountDict[Sewadar.NewID,'OldSewadarId'] = Sewadar.OldSewadarid

    MasterSheet = db(db.MasterSheet.id > 0).select('SewadarNewID','GR_NO','NAME','CANTEEN','DEV_DTY')
    for Sewadar in MasterSheet:
        SSCountDict[Sewadar.SewadarNewID,'CANTEEN'] = Sewadar.CANTEEN
        SSCountDict[Sewadar.SewadarNewID,'JATHA'] = Sewadar.DEV_DTY
        SSCountDict[Sewadar.SewadarNewID,'NAME'] = Sewadar.NAME

    dpath = os.path.join(request.folder,'private','ALL.txt')
    dpath1 = os.path.join(request.folder,'private','C1.txt')
    dpath2 = os.path.join(request.folder,'private','C2.txt')
    dpath3 = os.path.join(request.folder,'private','C3.txt')
    dpath4 = os.path.join(request.folder,'private','C4.txt')
    dpath5 = os.path.join(request.folder,'private','C5.txt')
    dpath6 = os.path.join(request.folder,'private','SHED.txt')
    dpathunknown = os.path.join(request.folder,'private','Unknown.txt')
    funknown = open(dpathunknown, 'w')
    funknown.write(str(SSCountDict))
    funknown.close()
    f = open(dpath, 'w')
    f1 = open(dpath1, 'w')
    f2 = open(dpath2, 'w')
    f3 = open(dpath3, 'w')
    f4 = open(dpath4, 'w')
    f5 = open(dpath5, 'w')
    f6 = open(dpath6, 'w')


    CARDLIST = db(db.CardList).select()
    for card in CARDLIST:
        try:
            SSCountDict[card.SewadarNewID,'CANTEEN']
            f.write(GRNOtoCARDNO(card.SewadarNewID) + "," + SSCountDict[card.SewadarNewID,'NAME'] + ',' + card.PROXIMITY_CARDNUMBER + '\n')
            if SSCountDict[card.SewadarNewID,'CANTEEN'] == 'CANTEEN NO.1' or SSCountDict[card.SewadarNewID,'CANTEEN'] == 'JAIPUR' or SSCountDict[card.SewadarNewID,'CANTEEN'] == 'KITCHEN' or SSCountDict[card.SewadarNewID,'CANTEEN'] == 'ADMIN' or SSCountDict[card.SewadarNewID,'CANTEEN'] == 'PHASE-IV CANTEEN':
                f1.write(GRNOtoCARDNO(card.SewadarNewID) + "," + SSCountDict[card.SewadarNewID,'NAME'] + ',' + card.PROXIMITY_CARDNUMBER + '\n')
            else:
                f1.write(GRNOtoCARDNO(card.SewadarNewID) + "," + SSCountDict[card.SewadarNewID,'CANTEEN'] + ',' + card.PROXIMITY_CARDNUMBER + '\n')

            if SSCountDict[card.SewadarNewID,'CANTEEN'] == 'CANTEEN NO.2' or SSCountDict[card.SewadarNewID,'CANTEEN'] == 'CANTEEN NO.6' or SSCountDict[card.SewadarNewID,'CANTEEN'] == 'CANTEEN NO.7' or SSCountDict[card.SewadarNewID,'CANTEEN'] == 'JAIPUR' or SSCountDict[card.SewadarNewID,'CANTEEN'] == 'KITCHEN' or SSCountDict[card.SewadarNewID,'CANTEEN'] == 'ADMIN' or SSCountDict[card.SewadarNewID,'CANTEEN'] == 'PHASE-IV CANTEEN':
                f2.write(GRNOtoCARDNO(card.SewadarNewID) + "," + SSCountDict[card.SewadarNewID,'NAME'] + ',' + card.PROXIMITY_CARDNUMBER + '\n')
            else:
                f2.write(GRNOtoCARDNO(card.SewadarNewID) + "," + SSCountDict[card.SewadarNewID,'CANTEEN'] + ',' + card.PROXIMITY_CARDNUMBER + '\n')

            if SSCountDict[card.SewadarNewID,'CANTEEN'] == 'CANTEEN NO.3' or SSCountDict[card.SewadarNewID,'CANTEEN'] == 'JAIPUR' or SSCountDict[card.SewadarNewID,'CANTEEN'] == 'KITCHEN' or SSCountDict[card.SewadarNewID,'CANTEEN'] == 'ADMIN' or SSCountDict[card.SewadarNewID,'CANTEEN'] == 'PHASE-IV CANTEEN':
                f3.write(GRNOtoCARDNO(card.SewadarNewID) + "," + SSCountDict[card.SewadarNewID,'NAME'] + ',' + card.PROXIMITY_CARDNUMBER + '\n')
            else:
                f3.write(GRNOtoCARDNO(card.SewadarNewID) + "," + SSCountDict[card.SewadarNewID,'CANTEEN'] + ',' + card.PROXIMITY_CARDNUMBER + '\n')

            if SSCountDict[card.SewadarNewID,'CANTEEN'] == 'CANTEEN NO.4' or SSCountDict[card.SewadarNewID,'CANTEEN'] == 'JAIPUR' or SSCountDict[card.SewadarNewID,'CANTEEN'] == 'KITCHEN' or SSCountDict[card.SewadarNewID,'CANTEEN'] == 'ADMIN' or SSCountDict[card.SewadarNewID,'CANTEEN'] == 'PHASE-IV CANTEEN':
                f4.write(GRNOtoCARDNO(card.SewadarNewID) + "," + SSCountDict[card.SewadarNewID,'NAME'] + ',' + card.PROXIMITY_CARDNUMBER + '\n')
            else:
                f4.write(GRNOtoCARDNO(card.SewadarNewID) + "," + SSCountDict[card.SewadarNewID,'CANTEEN'] + ',' + card.PROXIMITY_CARDNUMBER + '\n')

            if SSCountDict[card.SewadarNewID,'CANTEEN'] == 'CANTEEN NO.5' or SSCountDict[card.SewadarNewID,'CANTEEN'] == 'JAIPUR' or SSCountDict[card.SewadarNewID,'CANTEEN'] == 'KITCHEN' or SSCountDict[card.SewadarNewID,'CANTEEN'] == 'ADMIN' or SSCountDict[card.SewadarNewID,'CANTEEN'] == 'PHASE-IV CANTEEN':
                f5.write(GRNOtoCARDNO(card.SewadarNewID) + "," + SSCountDict[card.SewadarNewID,'NAME'] + ',' + card.PROXIMITY_CARDNUMBER + '\n')
            else:
                f5.write(GRNOtoCARDNO(card.SewadarNewID) + "," + SSCountDict[card.SewadarNewID,'CANTEEN'] + ',' + card.PROXIMITY_CARDNUMBER + '\n')

            if SSCountDict[card.SewadarNewID,'CANTEEN'] == 'SHED CANTEEN' or SSCountDict[card.SewadarNewID,'CANTEEN'] == 'JAIPUR' or SSCountDict[card.SewadarNewID,'CANTEEN'] == 'KITCHEN' or SSCountDict[card.SewadarNewID,'CANTEEN'] == 'ADMIN' or SSCountDict[card.SewadarNewID,'CANTEEN'] == 'PHASE-IV CANTEEN':
                f6.write(GRNOtoCARDNO(card.SewadarNewID) + "," + SSCountDict[card.SewadarNewID,'NAME'] + ',' + card.PROXIMITY_CARDNUMBER + '\n')
            else:
                f6.write(GRNOtoCARDNO(card.SewadarNewID) + "," + SSCountDict[card.SewadarNewID,'CANTEEN'] + ',' + card.PROXIMITY_CARDNUMBER + '\n')
        except:
            funknown = open(dpathunknown, 'a')
            funknown.write(GRNOtoCARDNO(card.SewadarNewID))
            funknown.close()
            funknown = open(dpathunknown, 'a')
            funknown.write(SSCountDict[card.SewadarNewID,'NAME'])
            funknown.close()
            funknown = open(dpathunknown, 'a')
            funknown.write(card.PROXIMITY_CARDNUMBER)
            funknown.close()
            funknown = open(dpathunknown, 'a')
            funknown.write(GRNOtoCARDNO(card.SewadarNewID) + "," + SSCountDict[card.SewadarNewID,'NAME'] + ',' + card.PROXIMITY_CARDNUMBER + '\n')
            funknown.close()
            f.write(GRNOtoCARDNO(card.SewadarNewID) + "," + SSCountDict[card.SewadarNewID,'NAME'] + ',' + card.PROXIMITY_CARDNUMBER + '\n')

    f.close()
    f1.close()
    f2.close()
    f3.close()
    f4.close()
    f5.close()
    f6.close()
    funknown.close()

    import subprocess
    os.system('echo hmm > samm')
    os.system('/home/rootcat/bitplay/bitplay ' + dpath)
    subprocess.check_call('/home/rootcat/bitplay/bitplay ' + dpath,shell=True)
    subprocess.check_call("cp -rf C_LIST MFILES/ALL",shell=True)
    subprocess.check_call('./applications/AttendanceSoftware/bitplay ' + dpath1,shell=True)
    subprocess.check_call("cp -rf C_LIST MFILES/C1",shell=True)
    subprocess.check_call('./applications/AttendanceSoftware/bitplay ' + dpath2,shell=True)
    subprocess.check_call("cp -rf C_LIST MFILES/C2",shell=True)
    subprocess.check_call('./applications/AttendanceSoftware/bitplay ' + dpath3,shell=True)
    subprocess.check_call("cp -rf C_LIST MFILES/C3",shell=True)
    subprocess.check_call('./applications/AttendanceSoftware/bitplay ' + dpath4,shell=True)
    subprocess.check_call("cp -rf C_LIST MFILES/C4",shell=True)
    subprocess.check_call('./applications/AttendanceSoftware/bitplay ' + dpath5,shell=True)
    subprocess.check_call("cp -rf C_LIST MFILES/C5",shell=True)
    subprocess.check_call('./applications/AttendanceSoftware/bitplay ' + dpath6,shell=True)
    subprocess.check_call("cp -rf C_LIST MFILES/SHED",shell=True)
    subprocess.check_call("cp -f ./applications/AttendanceSoftware/private/Unknown.txt MFILES/",shell=True)
    subprocess.check_call("tar -c MFILES -f MFILES.tar",shell=True)
    subprocess.check_call("mv MFILES.tar applications/AttendanceSoftware/private/",shell=True)
    dpath = os.path.join(request.folder,'private','MFILES.tar')
    return response.stream(open(dpath,'rb'), chunk_size=10**6)

def uploaddata_WWScheduleLadies():
    import os
    key_dict = {}
    key_dict['Sheet1'] = 'DAY'
    key_dict['Sheet1','TYPE'] = 'INDEXED'
    dict_days = CommonFunctions.XlsToDict(os.path.join(request.folder,'private','LadiesWW.xlsx'),key_dict)
    print str(dict_days['Sheet1','ARRAYCOLUMNNAMES'])
    print str(dict_days['Sheet1','ARRAYROWNAMES'])
    return locals()


def uploaddata_AllCardList():
    import xlrd
    import os
    path = os.path.join(request.folder,'private','AllCardList_xls.xlsx')
    workbook = xlrd.open_workbook(path)
    worksheet = workbook.sheet_names()
    #Attempt with any worksheet found
    worksheet = workbook.sheet_by_name(worksheet[0])
    header_row = 0
    import re

    #my headers are the headers in DB
    myheaders = ['PROXIMITY_CARDNUMBER']

    #headers are the name used in XLS
    headers = ['A_CARD']
    header_cells_in_xls = [worksheet.cell_value(0,idx) for idx in range(worksheet.ncols)]



    indexcol = [];
    for header_cell in headers:
            try:
                indexcol.append(header_cells_in_xls.index(header_cell))
            except Exception,e :
                print 'Header not found in xls: %s' %e
            else:
                print 'OK'

    print "INDEXCOL = ", indexcol
    row_dict_list = []

    try:
        db(db.AllCardList).delete()
    except:
        print "Creating CardList table"

    for row in xrange(header_row+1, worksheet.nrows):
       i=0
       row_dict = {}
       for col in indexcol:
               cell_type = worksheet.cell_type(row, col)
               if cell_type == xlrd.XL_CELL_EMPTY:
                   value = None
               elif cell_type == xlrd.XL_CELL_TEXT:
                  value = worksheet.cell_value(row, col)
               elif cell_type == xlrd.XL_CELL_NUMBER:
                  value = float(worksheet.cell_value(row,col))
               elif cell_type == xlrd.XL_CELL_DATE:
                  from datetime import datetime
                  value = datetime(*xlrd.xldate_as_tuple(worksheet.cell_value(row, col), workbook.datemode))
               elif cell_type == xlrd.XL_CELL_BOOLEAN:
                  value = bool(worksheet.cell_value(row, col))
               else:
                  value = worksheet.cell_value(row, col)
               row_dict[myheaders[i]] = value
               i=i+1
       db.AllCardList.insert(**row_dict)

    response.flash = T("Entry Successful!")
    return locals()

def uploaddata_CardList():
    import xlrd
    import os
    path = os.path.join(request.folder,'private','CardList_xls.xlsx')
    workbook = xlrd.open_workbook(path)
    worksheet = workbook.sheet_names()
    #Attempt with any worksheet found
    worksheet = workbook.sheet_by_name(worksheet[0])
    header_row = 0
    import re

    #my headers are the headers in DB
    myheaders = ['SewadarNewID','PROXIMITY_CARDNUMBER']

    #headers are the name used in XLS
    headers = ['GR_NO','A_CARD']
    header_cells_in_xls = [worksheet.cell_value(0,idx) for idx in range(worksheet.ncols)]



    indexcol = [];
    for header_cell in headers:
            try:
                indexcol.append(header_cells_in_xls.index(header_cell))
            except Exception,e :
                print 'Header not found in xls: %s' %e
            else:
                print 'OK'

    print "INDEXCOL = ", indexcol
    row_dict_list = []

    try:
        db(db.CardList).delete()
    except:
        print "Creating CardList table"

    for row in xrange(header_row+1, worksheet.nrows):
       i=0
       row_dict = {}
       for col in indexcol:
               cell_type = worksheet.cell_type(row, col)
               if cell_type == xlrd.XL_CELL_EMPTY:
                   value = None
               elif cell_type == xlrd.XL_CELL_TEXT:
                  value = worksheet.cell_value(row, col)
               elif cell_type == xlrd.XL_CELL_NUMBER:
                  value = float(worksheet.cell_value(row,col))
               elif cell_type == xlrd.XL_CELL_DATE:
                  from datetime import datetime
                  value = datetime(*xlrd.xldate_as_tuple(worksheet.cell_value(row, col), workbook.datemode))
               elif cell_type == xlrd.XL_CELL_BOOLEAN:
                  value = bool(worksheet.cell_value(row, col))
               else:
                  value = worksheet.cell_value(row, col)
               row_dict[myheaders[i]] = value
               i=i+1
       db.CardList.insert(**row_dict)

    response.flash = T("Entry Successful!")
    return locals()

import cv2
import numpy as np
import urllib2
import json

def image_dimensions():
    # Masquerade as Mozilla because some web servers may not like python bots.
    hdr = {'User-Agent': 'Mozilla/5.0'}
    # Set up the request
    req = urllib2.Request(request.vars.url, headers=hdr)
    try:
        # Obtain the content of the url
        con = urllib2.urlopen( req )
        # Read the content and convert it into an numpy array
        im_array = np.asarray(bytearray(con.read()), dtype=np.uint8)
        # Convert the numpy array into an image.
        im =  cv2.imdecode(im_array, cv2.IMREAD_GRAYSCALE)
        # Get the width and heigh of the image.
        height, width = im.shape
        # Wrap up the width and height in an object and return the encoded JSON.
        return json.dumps({"width" : width, "height" : height})
    except urllib2.HTTPError, e:
        return e.fp.read()

def timing_diagram():
    msg = "wow"
    return dict(msg=msg)

@auth.requires_login()
def MyAdmin():
    msg = "wow"
    return dict(msg=msg)

def ExcelTemplate():
    import os
    from gluon.sqlhtml import form_factory
    form=form_factory(SQLField('hop','integer',default=15),SQLField('ExcelTemplate','upload',uploadfolder='temporary'),formname='ExcelTemplate')
    if form.accepts(request.vars,session,formname='ExcelTemplate'):
        request.flash='Received: %s'%request.vars.ExcelTemplate
        path = os.path.join(request.folder,'private','ExcelTemplate.xlsx')
        import shutil
        shutil.copyfileobj(request.vars.ExcelTemplate.file,open(path, 'wb'))
        #Then redirect to the next screen (or do the processing now)
        redirect(URL(r=request, f='download_TemplateExcel' , args=['hop',request.vars.hop]))
    return dict(form=form)



def download_TemplateExcel():
    import os
    import copy
    import re
    pathlog = os.path.join(request.folder,'private','log_TemplateExcel')
    logf = open(pathlog,'w')

    hop = int(request.args[1])

    p = re.compile('{{=.*}}')

    from openpyxl import Workbook
    from openpyxl import load_workbook
    from openpyxl.cell import get_column_letter
    dpathr = os.path.join(request.folder,'private','ExcelTemplate.xlsx')
    dWorkbookr = load_workbook(filename = dpathr)
    dTemplateSheet = dWorkbookr['Template']
    dIndexSheet = dWorkbookr['Index']
    index = 1

    msg = 'hello '

    IndexSheet = {}
    row = 0
    for index_row in dIndexSheet.rows:
        if row == 0:
            pass
        else:
            col = 0
            for cellObj in index_row:
                IndexSheet[row-1,col] = cellObj.value
                col = col + 1
        row = row + 1

    i = -1
    for index_row in dIndexSheet.rows:
        if i%hop == 0:
            #Sheet name decided by first column
            mytitle = str(index_row[0].value)

            dWorkbookr.add_sheet(copy.copy(dTemplateSheet),0)
            dWorkSheetw = dWorkbookr.worksheets[0]
            dWorkSheetw.title = mytitle
        i = i + 1


    dWorkbookr.save(dpathr)
    dWorkbookr = load_workbook(filename = dpathr)
    dTemplateSheet = dWorkbookr['Template']

    i = -1
    for index_row in dIndexSheet.rows:
        if i%hop == 0:
            iterator = i
            mytitle = str(index_row[0].value)
            dWorkSheetw = dWorkbookr[mytitle]

            for row in xrange(1,dTemplateSheet.get_highest_row()+1):
                for col in xrange(1,dTemplateSheet.get_highest_column()+1):
                    value = dTemplateSheet[get_column_letter(col)+str(row)].value
                    logf.write(str(value) + ' = ')
                    try:
                        if p.match(value):
                            #msg = msg + ' ' + value + ' ' + get_column_letter(col) + str(row) + ' = ' + eval(value.replace('{{=','').replace('}}','')) + '\n'
                            try:
                                dWorkSheetw[get_column_letter(col)+str(row)].value = eval(value.replace('{{=','').replace('}}',''))
                                logf.write(eval(value.replace('{{=','').replace('}}',''))+'\n')
                                #print "evaled =" + eval(value.replace('{{=','').replace('}}',''))
                            except:
                                logf.write('\n')
                    except:
                        logf.write('\n')

        i = i + 1


    dWorkbookr.save(dpathr)
    logf.close()
    return response.stream(open(dpathr,'rb'), chunk_size=10**6)


def SplitExcel():
    import os
    from gluon.sqlhtml import form_factory
    form=form_factory(SQLField('RetainRows','integer',default=1),SQLField('BreakOnField','string',default='A'),SQLField('SplitExcel','upload',uploadfolder='temporary'),SQLField('AddSNO','string',default='YES',requires=IS_IN_SET(['YES','NO'])),SQLField('PurgeIndexColumns','string',default='YES',requires=IS_IN_SET(['YES','NO'])),SQLField('CopyIndexFormatting','string',default='YES',requires=IS_IN_SET(['YES','NO'])),SQLField('TotalFields','string',default='B;C'),formname='SplitExcel')
    if form.accepts(request.vars,session,formname='SplitExcel'):
        request.flash='Received: %s'%request.vars.SplitExcel
        path = os.path.join(request.folder,'private','SplitExcel.xlsx')
        import shutil
        shutil.copyfileobj(request.vars.SplitExcel.file,open(path, 'wb'))
        #Then redirect to the next screen (or do the processing now)
        redirect(URL(r=request, f='download_SplitExcel' , args=[request.vars.RetainRows,request.vars.BreakOnField,request.vars.AddSNO,request.vars.CopyIndexFormatting,request.vars.PurgeIndexColumns,request.vars.TotalFields]))
    return dict(form=form)



def download_SplitExcel():
    import os
    import copy
    import re
    pathlog = os.path.join(request.folder,'private','log_SplitExcel')
    logf = open(pathlog,'w')

    RetainRows = int(request.args[0])
    IndexColumns = str(request.args[1])
    AddSNO = request.args[2]
    CopyIndexFormatting = request.args[3]
    PurgeIndexColumns = request.args[4]
    TotalFields = request.args[5]
    IndexColumn = re.split('_',IndexColumns)
    TotalColumn = re.split('_',TotalFields)
    logf.write(IndexColumns + '\n')

    from openpyxl import Workbook
    from openpyxl import load_workbook
    from openpyxl.cell import get_column_letter
    from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
    dpath = os.path.join(request.folder,'private','SplitExcel.xlsx')
    dWorkbook = load_workbook(filename = dpath)
    dIndexSheet = dWorkbook['Index']
    dIndexFormatSheet = dWorkbook['IndexFormat']
    dHeaderSheet = dWorkbook['Header']
    dFooterSheet = dWorkbook['Footer']

    msg = 'start'


    IndexSheet = {}
    HeaderIndex = {}
    ListIndex = []
    row = 0
    #include header from the index sheet
    max_header_row = dHeaderSheet.get_highest_row()
    max_index_col =  dIndexSheet.get_highest_column()

    #include footer from the index sheet
    max_footer_row = dFooterSheet.get_highest_row()



    #Store index sheet in an array split into keys
    for index_row in dIndexSheet.rows:
        if row < RetainRows:
            for col in xrange(0,max_index_col):
                HeaderIndex[row,col] = dIndexSheet[get_column_letter(col+1)+str(row+1)].value
        else:
            my_IndexCol = ""
            for i_col in IndexColumn:
                my_IndexCol = my_IndexCol + ':' + str(dIndexSheet[i_col+str(row+1)].value)

            for t_col in TotalColumn:
                try:
                    IndexSheet[my_IndexCol,t_col,'TotalColumnList'].append(dIndexSheet[t_col+str(row+1)].value)
                except:
                    IndexSheet[my_IndexCol,t_col,'TotalColumnList'] = [dIndexSheet[t_col+str(row+1)].value]

                try:
                    IndexSheet[my_IndexCol,(t_col,dIndexSheet[t_col+str(row+1)].value),'TotalColumn'] = IndexSheet[my_IndexCol,(t_col,dIndexSheet[t_col+str(row+1)].value),'TotalColumn'] + 1
                except:
                    IndexSheet[my_IndexCol,(t_col,dIndexSheet[t_col+str(row+1)].value),'TotalColumn'] = 1

                IndexSheet[my_IndexCol,t_col,'TotalColumnList'] = list(set(IndexSheet[my_IndexCol,t_col,'TotalColumnList']))
                logf.write("----" + my_IndexCol + " " + str(IndexSheet[my_IndexCol,t_col,'TotalColumnList']) + "\n")

            logf.write("npw:" + my_IndexCol + '\n')
            logf.write(str(TotalColumn) + '\n')
            try:
                IndexSheet[my_IndexCol,'MAXROW'] = IndexSheet[my_IndexCol,'MAXROW'] + 1
            except:
                IndexSheet[my_IndexCol,'MAXROW'] =  1

            col = 0
            ListIndex.append(my_IndexCol)
            for cellObj in index_row:
                IndexSheet[my_IndexCol,IndexSheet[my_IndexCol,'MAXROW']-1,col] = cellObj.value
                IndexSheet[my_IndexCol,IndexSheet[my_IndexCol,'MAXROW']-1,col,'font'] = cellObj.font.copy()
                IndexSheet[my_IndexCol,IndexSheet[my_IndexCol,'MAXROW']-1,col,'alignment'] = cellObj.alignment.copy()
                IndexSheet[my_IndexCol,IndexSheet[my_IndexCol,'MAXROW']-1,col,'border'] = cellObj.border.copy()
                col = col + 1
        row = row + 1

    ListIndex = list(set(ListIndex))
    ListIndex = sorted(ListIndex,reverse=True)
    for i in ListIndex:
        logf.write('listindex= ' + str(i) + '\n')


    title_index = {}
    title_count = 0
    for index in ListIndex:
        mytitle = str(index)
        dWorkbook.add_sheet(copy.copy(dHeaderSheet),0)
        dWorkSheet = dWorkbook.worksheets[0]
        dWorkSheet.title = str(title_count)
        title_index[mytitle] = str(title_count)
        title_count = title_count + 1


    dWorkbook.save(dpath)
    dWorkbook = load_workbook(filename = dpath)


    dIndexSheet = dWorkbook['Index']

    ListIndex = list(set(ListIndex))
    max_header_row = max_header_row + 1
    for index in ListIndex:
        mytitle = str(index)
        dWorkSheet = dWorkbook[title_index[mytitle]]
        if PurgeIndexColumns == "YES":
            dWorkSheet['A' + str(max_header_row-1)].value = re.sub(r'^:','',str(mytitle)) + ' : ' + str(dWorkSheet['A' + str(max_header_row-1)].value)

        dWorkSheet['A' + str(max_header_row)].value = "Total Summary: "
        for t_col in TotalColumn:
            for vt_col in IndexSheet[index,t_col,'TotalColumnList']:
                dWorkSheet['A' + str(max_header_row)].value = str(dWorkSheet['A' + str(max_header_row)].value) + str(t_col) + '@' + str(vt_col) + ":" + str(IndexSheet[index,(t_col,vt_col),'TotalColumn']) + "   "
            dWorkSheet['A' + str(max_header_row)].value = str(dWorkSheet['A' + str(max_header_row)].value) + '@ '

        logf.write('mytitle =' + mytitle + '\n')
        logf.write('MAXROW =' + str(IndexSheet[index,'MAXROW']) + '\n')
        #copy header from Index sheet
        for hrow in xrange(0,RetainRows):
            PurgedColumnsCount = 0
            for hcol in xrange(0,max_index_col):
                if (PurgeIndexColumns == "YES") and (get_column_letter(hcol+1) in IndexColumn):
                    PurgedColumnsCount = PurgedColumnsCount + 1
                else:
                    dWorkSheet[get_column_letter(hcol-PurgedColumnsCount+1)+str(hrow+max_header_row+1)].value = HeaderIndex[hrow,hcol]
                    dWorkSheet[get_column_letter(hcol-PurgedColumnsCount+1)+str(hrow+max_header_row+1)].font = dIndexFormatSheet[get_column_letter(hcol+1)+str(hrow+1)].font.copy()
                    dWorkSheet[get_column_letter(hcol-PurgedColumnsCount+1)+str(hrow+max_header_row+1)].border = dIndexFormatSheet[get_column_letter(hcol+1)+str(hrow+1)].border.copy()
                    dWorkSheet[get_column_letter(hcol-PurgedColumnsCount+1)+str(hrow+max_header_row+1)].alignment = dIndexFormatSheet[get_column_letter(hcol+1)+str(hrow+1)].alignment.copy()

        sno = 0
        for row in xrange(0,IndexSheet[index,'MAXROW']):
            logf.write('    row =' + str(row) + '\n')
            sno = sno + 1
            if AddSNO == "YES":
                dWorkSheet['A'+str(max_header_row+RetainRows+row+1)].value = sno
                dWorkSheet['A'+str(max_header_row+RetainRows+row+1)].font = dIndexFormatSheet['A'+str(RetainRows+1)].font.copy()
                dWorkSheet['A'+str(max_header_row+RetainRows+row+1)].alignment = dIndexFormatSheet['A'+str(RetainRows+1)].alignment.copy()
                dWorkSheet['A'+str(max_header_row+RetainRows+row+1)].border = dIndexFormatSheet['A'+str(RetainRows+1)].border.copy()

            PurgedColumnsCount = 0
            for col in xrange(0,max_index_col):
                logf.write('        row =' + str(row) + '\n')
                logf.write('        ' + mytitle + ':' + str(row) + ':' + str(col))
                if AddSNO == "YES" and col == 0:
                    pass
                else:
                    if (PurgeIndexColumns == "YES") and (get_column_letter(col+1) in IndexColumn):
                        PurgedColumnsCount = PurgedColumnsCount + 1
                    else:
                        dWorkSheet[get_column_letter(col-PurgedColumnsCount+1)+str(max_header_row+RetainRows+row+1)].value = IndexSheet[index,row,col]
                        if CopyIndexFormatting=='YES':
                            dWorkSheet[get_column_letter(col-PurgedColumnsCount+1)+str(max_header_row+RetainRows+row+1)].font = IndexSheet[index,row,col,'font']
                            dWorkSheet[get_column_letter(col-PurgedColumnsCount+1)+str(max_header_row+RetainRows+row+1)].alignment = IndexSheet[index,row,col,'alignment']
                            dWorkSheet[get_column_letter(col-PurgedColumnsCount+1)+str(max_header_row+RetainRows+row+1)].border = IndexSheet[index,row,col,'border']
                        else:
                            dWorkSheet[get_column_letter(col-PurgedColumnsCount+1)+str(max_header_row+RetainRows+row+1)].font = dIndexFormatSheet[get_column_letter(col+1)+str(RetainRows+1)].font.copy()
                            dWorkSheet[get_column_letter(col-PurgedColumnsCount+1)+str(max_header_row+RetainRows+row+1)].alignment = dIndexFormatSheet[get_column_letter(col+1)+str(RetainRows+1)].alignment.copy()
                            dWorkSheet[get_column_letter(col-PurgedColumnsCount+1)+str(max_header_row+RetainRows+row+1)].border = dIndexFormatSheet[get_column_letter(col+1)+str(RetainRows+1)].border.copy()
        #dWorkSheet.add_print_title(1,RetainRows,rows_or_cols='row')
        dWorkSheet.add_print_title(max_header_row-1,max_header_row+1,rows_or_cols='row')

        footer_row_num = 1
        for footer_row in dFooterSheet.rows:
            footer_col_num = 1
            for cellObj in footer_row:
                dWorkSheet[get_column_letter(footer_col_num)+str(max_header_row+RetainRows+row+footer_row_num+1)].value = dFooterSheet[get_column_letter(footer_col_num)+str(footer_row_num)].value
                dWorkSheet[get_column_letter(footer_col_num)+str(max_header_row+RetainRows+row+footer_row_num+1)].font = dFooterSheet[get_column_letter(footer_col_num)+str(footer_row_num)].font.copy()
                dWorkSheet[get_column_letter(footer_col_num)+str(max_header_row+RetainRows+row+footer_row_num+1)].alignment = dFooterSheet[get_column_letter(footer_col_num)+str(footer_row_num)].alignment.copy()
                dWorkSheet[get_column_letter(footer_col_num)+str(max_header_row+RetainRows+row+footer_row_num+1)].border = dFooterSheet[get_column_letter(footer_col_num)+str(footer_row_num)].border.copy()
                footer_col_num = footer_col_num + 1
            footer_row_num = footer_row_num + 1

    dWorkbook.save(dpath)
    logf.close()
    return response.stream(open(dpath,'rb'), chunk_size=10**6)

def MissedMarking():
    import datetime
    import os
    from gluon.sqlhtml import form_factory
    from openpyxl import Workbook
    from openpyxl import load_workbook
    from openpyxl.cell import get_column_letter
    from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

    pathlog = os.path.join(request.folder,'private','log_MissedMarking')
    logf = open(pathlog,'w')
    dpath = os.path.join(request.folder,'private','InternalTemplateMissedMarking.xlsx')
    apath = os.path.join(request.folder,'private','MissedMarking.xlsx')
    dWorkBook = load_workbook(filename = dpath)

    MasterSheet = db(db.MasterSheet.id > 0).select('SewadarNewID','GR_NO','NAME','CANTEEN','DEV_DTY','GENDER')
    M_SHEET = {'SEWADARS':[]}
    for row in MasterSheet:
        M_SHEET['SEWADARS'].append(row.SewadarNewID)
        M_SHEET[row.SewadarNewID,'GR_NO'] = row.GR_NO
        M_SHEET[row.SewadarNewID,'NAME'] = row.NAME
        M_SHEET[row.SewadarNewID,'CANTEEN'] = row.CANTEEN
        M_SHEET[row.SewadarNewID,'DEV_DTY'] = row.DEV_DTY
        M_SHEET[row.SewadarNewID,'GENDER'] = row.GENDER

    form=form_factory(SQLField('M_StartTime','datetime',requires=IS_NOT_EMPTY()),SQLField('M_EndTime','datetime',requires=IS_NOT_EMPTY()),SQLField('E_StartTime','datetime',requires=IS_NOT_EMPTY()),SQLField('E_EndTime','datetime',requires=IS_NOT_EMPTY()),SQLField('IncludeCopy','string',default='NO',requires=IS_IN_SET(['YES','NO'])),formname='MorningEvening')
    if form.accepts(request.vars,session,formname='MorningEvening'):
        MachineDate_M = ''
        MachineDate_E = ''
        if request.vars.IncludeCopy == "YES":
            MachineDate_M = db((db.MachineAttendance.DATETIME >= (datetime.datetime.strptime(request.vars.M_StartTime,'%Y-%m-%d %H:%M:%S'))) & (db.MachineAttendance.DATETIME <= (datetime.datetime.strptime(request.vars.M_EndTime,'%Y-%m-%d %H:%M:%S')))).select('NewGRNO')
            MachineDate_E = db((db.MachineAttendance.DATETIME >= (datetime.datetime.strptime(request.vars.E_StartTime,'%Y-%m-%d %H:%M:%S'))) & (db.MachineAttendance.DATETIME <= (datetime.datetime.strptime(request.vars.E_EndTime,'%Y-%m-%d %H:%M:%S')))).select('NewGRNO')
        else:
            MachineDate_M = db((db.MachineAttendance.DATETIME >= (datetime.datetime.strptime(request.vars.M_StartTime,'%Y-%m-%d %H:%M:%S'))) & (db.MachineAttendance.DATETIME <= (datetime.datetime.strptime(request.vars.M_EndTime,'%Y-%m-%d %H:%M:%S'))) & (db.MachineAttendance.TYPE != 'COPY')).select('NewGRNO')
            MachineDate_E = db((db.MachineAttendance.DATETIME >= (datetime.datetime.strptime(request.vars.E_StartTime,'%Y-%m-%d %H:%M:%S'))) & (db.MachineAttendance.DATETIME <= (datetime.datetime.strptime(request.vars.E_EndTime,'%Y-%m-%d %H:%M:%S'))) & (db.MachineAttendance.TYPE != 'COPY')).select('NewGRNO')

        for row in MachineDate_M:
            M_SHEET[row.NewGRNO,'M'] = 'P'

        for row in MachineDate_E:
            M_SHEET[row.NewGRNO,'E'] = 'P'

        dWorkSheet = dWorkBook['Index']

        dWorkSheet['A1'] = 'S.NO'
        dWorkSheet['B1'] = 'SewadarNewID'
        dWorkSheet['C1'] = 'GR_NO'
        dWorkSheet['D1'] = 'NAME'
        dWorkSheet['E1'] = 'CANTEEN'
        dWorkSheet['F1'] = 'DEV_DTY'
        dWorkSheet['G1'] = 'GENDER'

        row_num = 1
        col_index = {'S.NO':'A','GR_NO':'C','NAME':'D','CANTEEN':'E','DEV_DTY':'F','GENDER':'G'}
        for Sewadar in M_SHEET['SEWADARS']:
            logf.write("Sewadar = " + Sewadar + "\n")
            try:
                a = M_SHEET[Sewadar,'M']
                try:
                    a = M_SHEET[Sewadar,'E']
                    logf.write("Sewadar = " + Sewadar + " = Full Present\n")
                except:
                    row_num = row_num + 1
                    logf.write("Sewadar = " + Sewadar + " = Partially Present. Row number =" + str(row_num) + "\n")
                    dWorkSheet['B'+str(row_num)] = Sewadar
                    dWorkSheet['A'+str(row_num)] = row_num-1
                    for col_name in ['GR_NO','NAME','CANTEEN','DEV_DTY','GENDER']:
                        dWorkSheet[col_index[col_name]+str(row_num)] = M_SHEET[Sewadar,col_name]
            except:
                logf.write("Sewadar = " + Sewadar + " = Absent\n")
                pass


        dWorkBook.save(apath)

        logf.close()

        return response.stream(open(apath,'rb'), chunk_size=10**6)

    return dict(form=form)

def DailyStrengthReport():
    message = "Preparing Report"
    db.commit()
    import os
    import datetime
    import time
    from openpyxl import Workbook
    from openpyxl.cell import get_column_letter
    from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
    dWorkBook = Workbook()
    dpath = os.path.join(request.folder,'private','DateReport.xlsx')
    MasterSheet = db(db.MasterSheet.id > 0).select('SewadarNewID','GR_NO','NAME','CANTEEN','DEV_DTY','GENDER')
    from gluon.sqlhtml import form_factory
    form=form_factory(SQLField('DateSelectedStart','datetime',default=datetime.datetime.now(),requires=IS_NOT_EMPTY()),SQLField('DateSelectedEnd','datetime',default=datetime.datetime.now(),requires=IS_NOT_EMPTY()),formname='MorningEvening')
    if form.accepts(request.vars,session,formname='MorningEvening'):
        DateSelectedStart = request.vars.DateSelectedStart
        DateSelectedEnd = request.vars.DateSelectedEnd
        SSDate = db((db.SSAttendanceDate.DutyDate >= datetime.datetime.strptime(DateSelectedStart,"%Y-%m-%d %H:%M:%S")) & (db.SSAttendanceDate.DutyDate <= datetime.datetime.strptime(DateSelectedEnd,"%Y-%m-%d %H:%M:%S"))).select('SewadarNewID','DutyDate','Duty_Type')
        MachineDate = db((db.MachineAttendance.DATETIME >= datetime.datetime.strptime(DateSelectedStart,"%Y-%m-%d %H:%M:%S")) & (db.MachineAttendance.DATETIME <= datetime.datetime.strptime(DateSelectedEnd,"%Y-%m-%d %H:%M:%S"))).select('NewGRNO','DATETIME','TYPE')

        CombinedAttendance = {'SEWADARS':[]}

        for row in SSDate:
            CombinedAttendance[row.SewadarNewID] = 'S'

        for row in MachineDate:
            try:
                if CombinedAttendance[row.NewGRNO] == 'S':
                    CombinedAttendance[row.NewGRNO] = 'B'
            except:
                CombinedAttendance[row.NewGRNO] = 'M'

        ReportDict = {'CANTEENS':[]}

        for row in MasterSheet:
            ReportDict['CANTEENS'].append(row.CANTEEN)
            try:
                ReportDict[row.CANTEEN,'JATHALIST'].append(row.DEV_DTY)
            except:
                ReportDict[row.CANTEEN,'JATHALIST'] = [row.DEV_DTY]

            try:
                a = CombinedAttendance[row.SewadarNewID]
                try:
                    ReportDict['PRESENT',row.GENDER] = ReportDict['PRESENT',row.GENDER] + 1
                except:
                    ReportDict['PRESENT',row.GENDER] = 1

                try:
                    ReportDict['PRESENT',row.CANTEEN,row.GENDER] = ReportDict['PRESENT',row.CANTEEN,row.GENDER] + 1
                except:
                    ReportDict['PRESENT',row.CANTEEN,row.GENDER] = 1

                try:
                    ReportDict['PRESENT',row.CANTEEN,row.DEV_DTY,row.GENDER] = ReportDict['PRESENT',row.CANTEEN,row.DEV_DTY,row.GENDER] + 1
                except:
                    ReportDict['PRESENT',row.CANTEEN,row.DEV_DTY,row.GENDER] = 1
            except:
                try:
                    ReportDict['ABSENT',row.GENDER] = ReportDict['ABSENT',row.GENDER] + 1
                except:
                    ReportDict['ABSENT',row.GENDER] = 1

                try:
                    ReportDict['ABSENT',row.CANTEEN,row.GENDER] = ReportDict['ABSENT',row.CANTEEN,row.GENDER] + 1
                except:
                    ReportDict['ABSENT',row.CANTEEN,row.GENDER] = 1

                try:
                    ReportDict['ABSENT',row.CANTEEN,row.DEV_DTY,row.GENDER] = ReportDict['ABSENT',row.CANTEEN,row.DEV_DTY,row.GENDER] + 1
                except:
                    ReportDict['ABSENT',row.CANTEEN,row.DEV_DTY,row.GENDER] = 1

        ReportDict['CANTEENS'] = sorted(list(set(ReportDict['CANTEENS'])))
        #Initialize all numbers
        for canteen in ReportDict['CANTEENS']:
            ReportDict[canteen,'JATHALIST'] = sorted(list(set(ReportDict[canteen,'JATHALIST'])))
            try:
                a = ReportDict['PRESENT',canteen,'F']
            except:
                ReportDict['PRESENT',canteen,'F'] = 0
            try:
                a = ReportDict['ABSENT',canteen,'F']
            except:
                ReportDict['ABSENT',canteen,'F'] = 0
            try:
                a = ReportDict['PRESENT',canteen,'M']
            except:
                ReportDict['PRESENT',canteen,'M'] = 0
            try:
                a = ReportDict['ABSENT',canteen,'M']
            except:
                ReportDict['ABSENT',canteen,'M'] = 0

            for jatha in ReportDict[canteen,'JATHALIST']:
                try:
                    a = ReportDict['PRESENT',canteen,jatha,'F']
                except:
                    ReportDict['PRESENT',canteen,jatha,'F'] = 0
                try:
                    a = ReportDict['ABSENT',canteen,jatha,'F']
                except:
                    ReportDict['ABSENT',canteen,jatha,'F'] = 0
                try:
                    a = ReportDict['PRESENT',canteen,jatha,'M']
                except:
                    ReportDict['PRESENT',canteen,jatha,'M'] = 0
                try:
                    a = ReportDict['ABSENT',canteen,jatha,'M']
                except:
                    ReportDict['ABSENT',canteen,jatha,'M'] = 0

        dSheet = dWorkBook.create_sheet(0)
        dSheet.page_setup.orientation = dSheet.ORIENTATION_LANDSCAPE
        dSheet.page_setup.paperSize = dSheet.PAPERSIZE_A4
        dSheet.page_setup.fitToHeight = 0
        dSheet.page_setup.fitToWidth = 1
        dSheet.title = 'COMBINED'
        dSheet['B1'] = 'LADIES'
        dSheet['E1'] = 'GENTS'
        dSheet['H1'] = 'ALL'
        dSheet.merge_cells('B1:D1')
        dSheet.merge_cells('E1:G1')
        dSheet.merge_cells('H1:J1')
        dSheet['A2'] = 'CANTEEN'
        dSheet['B2'] = 'PRESENT'
        dSheet['C2'] = 'ABSENT'
        dSheet['D2'] = 'TOTAL'
        dSheet['E2'] = 'PRESENT'
        dSheet['F2'] = 'ABSENT'
        dSheet['G2'] = 'TOTAL'
        dSheet['H2'] = 'PRESENT'
        dSheet['I2'] = 'ABSENT'
        dSheet['J2'] = 'TOTAL'
        dSheet['A2'].font = Font(name='Calibri',size=12,bold=True)
        dSheet['A2'].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
        dSheet['A2'].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
        dSheet['B2'].font = Font(name='Calibri',size=12,bold=True)
        dSheet['B2'].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
        dSheet['B2'].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
        dSheet['C2'].font = Font(name='Calibri',size=12,bold=True)
        dSheet['C2'].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
        dSheet['C2'].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
        dSheet['D2'].font = Font(name='Calibri',size=12,bold=True)
        dSheet['D2'].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
        dSheet['D2'].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
        dSheet['E2'].font = Font(name='Calibri',size=12,bold=True)
        dSheet['E2'].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
        dSheet['E2'].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
        dSheet['F2'].font = Font(name='Calibri',size=12,bold=True)
        dSheet['F2'].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
        dSheet['F2'].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
        dSheet['G2'].font = Font(name='Calibri',size=12,bold=True)
        dSheet['G2'].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
        dSheet['G2'].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
        dSheet['H2'].font = Font(name='Calibri',size=12,bold=True)
        dSheet['H2'].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
        dSheet['H2'].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
        dSheet['I2'].font = Font(name='Calibri',size=12,bold=True)
        dSheet['I2'].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
        dSheet['I2'].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
        dSheet['J2'].font = Font(name='Calibri',size=12,bold=True)
        dSheet['J2'].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
        dSheet['J2'].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
        dSheet['B1'].font = Font(name='Calibri',size=12,bold=True)
        dSheet['B1'].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
        dSheet['B1'].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
        dSheet['C1'].font = Font(name='Calibri',size=12,bold=True)
        dSheet['C1'].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
        dSheet['C1'].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
        dSheet['D1'].font = Font(name='Calibri',size=12,bold=True)
        dSheet['D1'].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
        dSheet['D1'].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
        dSheet['E1'].font = Font(name='Calibri',size=12,bold=True)
        dSheet['E1'].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
        dSheet['E1'].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
        dSheet['F1'].font = Font(name='Calibri',size=12,bold=True)
        dSheet['F1'].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
        dSheet['F1'].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
        dSheet['G1'].font = Font(name='Calibri',size=12,bold=True)
        dSheet['G1'].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
        dSheet['G1'].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
        dSheet['H1'].font = Font(name='Calibri',size=12,bold=True)
        dSheet['H1'].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
        dSheet['H1'].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
        dSheet['I1'].font = Font(name='Calibri',size=12,bold=True)
        dSheet['I1'].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
        dSheet['I1'].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
        dSheet['J1'].font = Font(name='Calibri',size=12,bold=True)
        dSheet['J1'].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
        dSheet['J1'].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
        dSheet.add_print_title(1,2,rows_or_cols='rows')
        row_num = 3
        for canteen in ReportDict['CANTEENS']:
            dSheet['A'+ str(row_num)] = canteen
            dSheet['A'+ str(row_num)].font = Font(name='Calibri',size=12,bold=True)
            dSheet['A'+ str(row_num)].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
            dSheet['A'+ str(row_num)].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
            dSheet['B'+ str(row_num)] = ReportDict['PRESENT',canteen,'F']
            dSheet['B'+ str(row_num)].font = Font(name='Calibri',size=12,bold=True)
            dSheet['B'+ str(row_num)].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
            dSheet['B'+ str(row_num)].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
            dSheet['C'+ str(row_num)] = ReportDict['ABSENT',canteen,'F']
            dSheet['C'+ str(row_num)].font = Font(name='Calibri',size=12,bold=True)
            dSheet['C'+ str(row_num)].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
            dSheet['C'+ str(row_num)].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
            dSheet['D'+ str(row_num)] = ReportDict['PRESENT',canteen,'F'] + ReportDict['ABSENT',canteen,'F']
            dSheet['D'+ str(row_num)].font = Font(name='Calibri',size=12,bold=True)
            dSheet['D'+ str(row_num)].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
            dSheet['D'+ str(row_num)].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
            dSheet['E'+ str(row_num)] = ReportDict['PRESENT',canteen,'M']
            dSheet['E'+ str(row_num)].font = Font(name='Calibri',size=12,bold=True)
            dSheet['E'+ str(row_num)].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
            dSheet['E'+ str(row_num)].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
            dSheet['F'+ str(row_num)] = ReportDict['ABSENT',canteen,'M']
            dSheet['F'+ str(row_num)].font = Font(name='Calibri',size=12,bold=True)
            dSheet['F'+ str(row_num)].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
            dSheet['F'+ str(row_num)].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
            dSheet['G'+ str(row_num)] = ReportDict['PRESENT',canteen,'M'] + ReportDict['ABSENT',canteen,'M']
            dSheet['G'+ str(row_num)].font = Font(name='Calibri',size=12,bold=True)
            dSheet['G'+ str(row_num)].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
            dSheet['G'+ str(row_num)].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
            dSheet['H'+ str(row_num)] = ReportDict['PRESENT',canteen,'F'] + ReportDict['PRESENT',canteen,'M']
            dSheet['H'+ str(row_num)].font = Font(name='Calibri',size=12,bold=True)
            dSheet['H'+ str(row_num)].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
            dSheet['H'+ str(row_num)].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
            dSheet['I'+ str(row_num)] = ReportDict['ABSENT',canteen,'F'] + ReportDict['ABSENT',canteen,'M']
            dSheet['I'+ str(row_num)].font = Font(name='Calibri',size=12,bold=True)
            dSheet['I'+ str(row_num)].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
            dSheet['I'+ str(row_num)].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
            dSheet['J'+ str(row_num)] = ReportDict['PRESENT',canteen,'F'] + ReportDict['ABSENT',canteen,'F'] + ReportDict['PRESENT',canteen,'M'] + ReportDict['ABSENT',canteen,'M']
            dSheet['J'+ str(row_num)].font = Font(name='Calibri',size=12,bold=True)
            dSheet['J'+ str(row_num)].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
            dSheet['J'+ str(row_num)].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
            row_num = row_num + 1
        dSheet['A'+str(row_num)] = 'TOTAL SUMMARY'
        dSheet['A'+ str(row_num)].font = Font(name='Calibri',size=12,bold=True)
        dSheet['A'+ str(row_num)].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
        dSheet['A'+ str(row_num)].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
        dSheet['B'+str(row_num)] = ReportDict['PRESENT','F']
        dSheet['B'+ str(row_num)].font = Font(name='Calibri',size=12,bold=True)
        dSheet['B'+ str(row_num)].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
        dSheet['B'+ str(row_num)].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
        dSheet['C'+str(row_num)] = ReportDict['ABSENT','F']
        dSheet['C'+ str(row_num)].font = Font(name='Calibri',size=12,bold=True)
        dSheet['C'+ str(row_num)].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
        dSheet['C'+ str(row_num)].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
        dSheet['D'+str(row_num)] = ReportDict['PRESENT','F'] + ReportDict['ABSENT','F']
        dSheet['D'+ str(row_num)].font = Font(name='Calibri',size=12,bold=True)
        dSheet['D'+ str(row_num)].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
        dSheet['D'+ str(row_num)].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
        dSheet['E'+str(row_num)] = ReportDict['PRESENT','M']
        dSheet['E'+ str(row_num)].font = Font(name='Calibri',size=12,bold=True)
        dSheet['E'+ str(row_num)].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
        dSheet['E'+ str(row_num)].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
        dSheet['F'+str(row_num)] = ReportDict['ABSENT','M']
        dSheet['F'+ str(row_num)].font = Font(name='Calibri',size=12,bold=True)
        dSheet['F'+ str(row_num)].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
        dSheet['F'+ str(row_num)].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
        dSheet['G'+str(row_num)] = ReportDict['PRESENT','M'] + ReportDict['ABSENT','M']
        dSheet['G'+ str(row_num)].font = Font(name='Calibri',size=12,bold=True)
        dSheet['G'+ str(row_num)].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
        dSheet['G'+ str(row_num)].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
        dSheet['H'+str(row_num)] = ReportDict['PRESENT','M'] + ReportDict['PRESENT','F']
        dSheet['H'+ str(row_num)].font = Font(name='Calibri',size=12,bold=True)
        dSheet['H'+ str(row_num)].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
        dSheet['H'+ str(row_num)].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
        dSheet['I'+str(row_num)] = ReportDict['ABSENT','M'] + ReportDict['ABSENT','F']
        dSheet['I'+ str(row_num)].font = Font(name='Calibri',size=12,bold=True)
        dSheet['I'+ str(row_num)].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
        dSheet['I'+ str(row_num)].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
        dSheet['J'+str(row_num)] = ReportDict['PRESENT','M'] + ReportDict['ABSENT','M'] + ReportDict['PRESENT','F'] + ReportDict['ABSENT','F']
        dSheet['J'+ str(row_num)].font = Font(name='Calibri',size=12,bold=True)
        dSheet['J'+ str(row_num)].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
        dSheet['J'+ str(row_num)].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))

        for canteen in ReportDict['CANTEENS']:
            dSheet = dWorkBook.create_sheet(0)
            dSheet.page_setup.orientation = dSheet.ORIENTATION_LANDSCAPE
            dSheet.page_setup.paperSize = dSheet.PAPERSIZE_A4
            dSheet.page_setup.fitToHeight = 0
            dSheet.page_setup.fitToWidth = 1
            dSheet.title = canteen
            dSheet['A1'] = canteen
            dSheet['B1'] = 'LADIES'
            dSheet['E1'] = 'GENTS'
            dSheet['H1'] = 'ALL'
            dSheet.merge_cells('B1:D1')
            dSheet.merge_cells('E1:G1')
            dSheet.merge_cells('H1:J1')
            dSheet['A2'] = 'JATHA'
            dSheet['B2'] = 'PRESENT'
            dSheet['C2'] = 'ABSENT'
            dSheet['D2'] = 'TOTAL'
            dSheet['E2'] = 'PRESENT'
            dSheet['F2'] = 'ABSENT'
            dSheet['G2'] = 'TOTAL'
            dSheet['H2'] = 'PRESENT'
            dSheet['I2'] = 'ABSENT'
            dSheet['J2'] = 'TOTAL'
            dSheet['A1'].font = Font(name='Calibri',size=12,bold=True)
            dSheet['A1'].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
            dSheet['A1'].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
            dSheet['A2'].font = Font(name='Calibri',size=12,bold=True)
            dSheet['A2'].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
            dSheet['A2'].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
            dSheet['B2'].font = Font(name='Calibri',size=12,bold=True)
            dSheet['B2'].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
            dSheet['B2'].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
            dSheet['C2'].font = Font(name='Calibri',size=12,bold=True)
            dSheet['C2'].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
            dSheet['C2'].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
            dSheet['D2'].font = Font(name='Calibri',size=12,bold=True)
            dSheet['D2'].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
            dSheet['D2'].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
            dSheet['E2'].font = Font(name='Calibri',size=12,bold=True)
            dSheet['E2'].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
            dSheet['E2'].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
            dSheet['F2'].font = Font(name='Calibri',size=12,bold=True)
            dSheet['F2'].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
            dSheet['F2'].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
            dSheet['G2'].font = Font(name='Calibri',size=12,bold=True)
            dSheet['G2'].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
            dSheet['G2'].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
            dSheet['H2'].font = Font(name='Calibri',size=12,bold=True)
            dSheet['H2'].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
            dSheet['H2'].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
            dSheet['I2'].font = Font(name='Calibri',size=12,bold=True)
            dSheet['I2'].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
            dSheet['I2'].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
            dSheet['J2'].font = Font(name='Calibri',size=12,bold=True)
            dSheet['J2'].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
            dSheet['J2'].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
            dSheet['B1'].font = Font(name='Calibri',size=12,bold=True)
            dSheet['B1'].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
            dSheet['B1'].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
            dSheet['C1'].font = Font(name='Calibri',size=12,bold=True)
            dSheet['C1'].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
            dSheet['C1'].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
            dSheet['D1'].font = Font(name='Calibri',size=12,bold=True)
            dSheet['D1'].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
            dSheet['D1'].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
            dSheet['E1'].font = Font(name='Calibri',size=12,bold=True)
            dSheet['E1'].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
            dSheet['E1'].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
            dSheet['F1'].font = Font(name='Calibri',size=12,bold=True)
            dSheet['F1'].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
            dSheet['F1'].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
            dSheet['G1'].font = Font(name='Calibri',size=12,bold=True)
            dSheet['G1'].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
            dSheet['G1'].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
            dSheet['H1'].font = Font(name='Calibri',size=12,bold=True)
            dSheet['H1'].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
            dSheet['H1'].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
            dSheet['I1'].font = Font(name='Calibri',size=12,bold=True)
            dSheet['I1'].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
            dSheet['I1'].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
            dSheet['J1'].font = Font(name='Calibri',size=12,bold=True)
            dSheet['J1'].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
            dSheet['J1'].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
            dSheet.add_print_title(1,2,rows_or_cols='rows')
            row_num = 3
            for jatha in ReportDict[canteen,'JATHALIST']:
                dSheet['A'+ str(row_num)] = jatha
                dSheet['A'+ str(row_num)].font = Font(name='Calibri',size=12,bold=True)
                dSheet['A'+ str(row_num)].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
                dSheet['A'+ str(row_num)].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
                dSheet['B'+ str(row_num)] = ReportDict['PRESENT',canteen,jatha,'F']
                dSheet['B'+ str(row_num)].font = Font(name='Calibri',size=12,bold=True)
                dSheet['B'+ str(row_num)].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
                dSheet['B'+ str(row_num)].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
                dSheet['C'+ str(row_num)] = ReportDict['ABSENT',canteen,jatha,'F']
                dSheet['C'+ str(row_num)].font = Font(name='Calibri',size=12,bold=True)
                dSheet['C'+ str(row_num)].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
                dSheet['C'+ str(row_num)].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
                dSheet['D'+ str(row_num)] = ReportDict['PRESENT',canteen,jatha,'F'] + ReportDict['ABSENT',canteen,jatha,'F']
                dSheet['D'+ str(row_num)].font = Font(name='Calibri',size=12,bold=True)
                dSheet['D'+ str(row_num)].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
                dSheet['D'+ str(row_num)].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
                dSheet['E'+ str(row_num)] = ReportDict['PRESENT',canteen,jatha,'M']
                dSheet['E'+ str(row_num)].font = Font(name='Calibri',size=12,bold=True)
                dSheet['E'+ str(row_num)].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
                dSheet['E'+ str(row_num)].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
                dSheet['F'+ str(row_num)] = ReportDict['ABSENT',canteen,jatha,'M']
                dSheet['F'+ str(row_num)].font = Font(name='Calibri',size=12,bold=True)
                dSheet['F'+ str(row_num)].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
                dSheet['F'+ str(row_num)].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
                dSheet['G'+ str(row_num)] = ReportDict['PRESENT',canteen,jatha,'M'] + ReportDict['ABSENT',canteen,jatha,'M']
                dSheet['G'+ str(row_num)].font = Font(name='Calibri',size=12,bold=True)
                dSheet['G'+ str(row_num)].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
                dSheet['G'+ str(row_num)].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
                dSheet['H'+ str(row_num)] = ReportDict['PRESENT',canteen,jatha,'F'] + ReportDict['PRESENT',canteen,jatha,'M']
                dSheet['H'+ str(row_num)].font = Font(name='Calibri',size=12,bold=True)
                dSheet['H'+ str(row_num)].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
                dSheet['H'+ str(row_num)].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
                dSheet['I'+ str(row_num)] = ReportDict['ABSENT',canteen,jatha,'F'] + ReportDict['ABSENT',canteen,jatha,'M']
                dSheet['I'+ str(row_num)].font = Font(name='Calibri',size=12,bold=True)
                dSheet['I'+ str(row_num)].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
                dSheet['I'+ str(row_num)].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
                dSheet['J'+ str(row_num)] = ReportDict['PRESENT',canteen,jatha,'F'] + ReportDict['ABSENT',canteen,jatha,'F'] + ReportDict['PRESENT',canteen,jatha,'M'] + ReportDict['ABSENT',canteen,jatha,'M']
                dSheet['J'+ str(row_num)].font = Font(name='Calibri',size=12,bold=True)
                dSheet['J'+ str(row_num)].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
                dSheet['J'+ str(row_num)].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
                row_num = row_num + 1
            dSheet['A'+str(row_num)] = 'TOTAL SUMMARY'
            dSheet['A'+ str(row_num)].font = Font(name='Calibri',size=12,bold=True)
            dSheet['A'+ str(row_num)].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
            dSheet['A'+ str(row_num)].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
            dSheet['B'+str(row_num)] = ReportDict['PRESENT',canteen,'F']
            dSheet['B'+ str(row_num)].font = Font(name='Calibri',size=12,bold=True)
            dSheet['B'+ str(row_num)].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
            dSheet['B'+ str(row_num)].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
            dSheet['C'+str(row_num)] = ReportDict['ABSENT',canteen,'F']
            dSheet['C'+ str(row_num)].font = Font(name='Calibri',size=12,bold=True)
            dSheet['C'+ str(row_num)].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
            dSheet['C'+ str(row_num)].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
            dSheet['D'+str(row_num)] = ReportDict['PRESENT',canteen,'F'] + ReportDict['ABSENT',canteen,'F']
            dSheet['D'+ str(row_num)].font = Font(name='Calibri',size=12,bold=True)
            dSheet['D'+ str(row_num)].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
            dSheet['D'+ str(row_num)].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
            dSheet['E'+str(row_num)] = ReportDict['PRESENT',canteen,'M']
            dSheet['E'+ str(row_num)].font = Font(name='Calibri',size=12,bold=True)
            dSheet['E'+ str(row_num)].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
            dSheet['E'+ str(row_num)].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
            dSheet['F'+str(row_num)] = ReportDict['ABSENT',canteen,'M']
            dSheet['F'+ str(row_num)].font = Font(name='Calibri',size=12,bold=True)
            dSheet['F'+ str(row_num)].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
            dSheet['F'+ str(row_num)].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
            dSheet['G'+str(row_num)] = ReportDict['PRESENT',canteen,'M'] + ReportDict['ABSENT',canteen,'M']
            dSheet['G'+ str(row_num)].font = Font(name='Calibri',size=12,bold=True)
            dSheet['G'+ str(row_num)].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
            dSheet['G'+ str(row_num)].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
            dSheet['H'+str(row_num)] = ReportDict['PRESENT',canteen,'M'] + ReportDict['PRESENT',canteen,'F']
            dSheet['H'+ str(row_num)].font = Font(name='Calibri',size=12,bold=True)
            dSheet['H'+ str(row_num)].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
            dSheet['H'+ str(row_num)].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
            dSheet['I'+str(row_num)] = ReportDict['ABSENT',canteen,'M'] + ReportDict['ABSENT',canteen,'F']
            dSheet['I'+ str(row_num)].font = Font(name='Calibri',size=12,bold=True)
            dSheet['I'+ str(row_num)].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
            dSheet['I'+ str(row_num)].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
            dSheet['J'+str(row_num)] = ReportDict['PRESENT',canteen,'M'] + ReportDict['ABSENT',canteen,'M'] + ReportDict['PRESENT',canteen,'F'] + ReportDict['ABSENT',canteen,'F']
            dSheet['J'+ str(row_num)].font = Font(name='Calibri',size=12,bold=True)
            dSheet['J'+ str(row_num)].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
            dSheet['J'+ str(row_num)].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))

        dWorkBook.save(dpath)

        return response.stream(open(dpath,'rb'), chunk_size=10**6)
    return dict(form=form)

def schedule_report_mail():
    import datetime
    scheduler.queue_task(AttendanceFetch,
                        #start_time=datetime.datetime.strptime('11-July-2018 00:27:00','%d-%B-%Y %H:%M:%S'),  # datetime
                        start_time=datetime.datetime.now() + datetime.timedelta(seconds=3),  # datetime
                        retry_failed=-1,
                        stop_time=None,  # datetime
                        timeout = 5000,  # seconds
                        prevent_drift=True,
                        period=86400,  # seconds
                        immediate=False,
                        repeats=1)
    return 0

def schedule_sms_mail():
    import datetime
    scheduler.queue_task(SMSReport,
                        #start_time=datetime.datetime.strptime('11-July-2018 00:27:00','%d-%B-%Y %H:%M:%S'),  # datetime
                        start_time=datetime.datetime.now() + datetime.timedelta(seconds=3),  # datetime
                        retry_failed=-1,
                        stop_time=None,  # datetime
                        timeout = 5000,  # seconds
                        prevent_drift=True,
                        period=86400,  # seconds
                        immediate=False,
                        repeats=1)
    return 0

def schedule_auto_sms():
    import datetime
    scheduler.queue_task(SendSMSWarning,
                        #start_time=datetime.datetime.strptime('11-July-2018 00:27:00','%d-%B-%Y %H:%M:%S'),  # datetime
                        start_time=datetime.datetime.now() + datetime.timedelta(seconds=3),  # datetime
                        retry_failed=-1,
                        stop_time=None,  # datetime
                        timeout = 5000,  # seconds
                        prevent_drift=True,
                        period=86400,  # seconds
                        immediate=False,
                        repeats=1)
    return 0



@auth.requires(auth.user_id == 3)
def update_master_from_google_sheets():
    import os
    pathlog = os.path.join(request.folder,'private','log_update_master_from_google_sheet')
    logf = open(pathlog,'wb')
    from apiclient.discovery import build
    from httplib2 import Http
    from oauth2client import file as oauth_file, client, tools
    import pandas as pd
    import pprint

    # If modifying these scopes, delete the file token.json.
    SCOPES = 'https://www.googleapis.com/auth/spreadsheets.readonly'

    # The ID and range of a sample spreadsheet.
    #SAMPLE_SPREADSHEET_ID = '1BxiMVs0XRA5nFMdKvBdBZjgmUUqptlbs74OgvE2upms'
    SAMPLE_SPREADSHEET_ID = '18Hm5KEHTsEDBdFFdpVuoWNexjL7ch89BKV-8edpEKiE'
    SAMPLE_RANGE_NAME = 'MasterSheet!A:H'


    """Shows basic usage of the Sheets API.

    Prints values from a sample spreadsheet.
    """
    store = oauth_file.Storage('/home/rootcat/new_web2py/web2py/applications/AttendanceSoftware/private/token.json')
    creds = store.get()
    if not creds or creds.invalid:
        flow = client.flow_from_clientsecrets('/home/rootcat/new_web2py/web2py/applications/AttendanceSoftware/private/credentials.json', SCOPES)
        creds = tools.run_flow(flow, store)
    service = build('sheets', 'v4', http=creds.authorize(Http()))

    # Call the Sheets API
    result = service.spreadsheets().values().get(spreadsheetId=SAMPLE_SPREADSHEET_ID,range=SAMPLE_RANGE_NAME).execute()
    values = result.get('values', [])

    db.MasterSheet.drop()
    db.define_table('MasterSheet',
            Field('SewadarNewID','string'),
            Field('CANTEEN','string'),
            Field('DEV_DTY','string'),
            Field('MOBILE','string'),
            Field('AGE','integer'),
            migrate=True,
            redefine=True,
            format='%(SewadarNewID)s')


    if not values:
        print('No data found.')
    else:
        MasterSheet_df = pd.DataFrame.from_records(values)
        MasterSheet_df.columns = MasterSheet_df.iloc[0]
        MasterSheet_df.drop(MasterSheet_df.index[0],inplace=True)
        for row in MasterSheet_df.iterrows():
            row_dict = {}
            pprint.pprint(row[1],stream=logf)
            row_dict['SewadarNewID']=row[1]['GR ID']
            row_dict['CANTEEN']=row[1]['Canteen No'].upper()
            row_dict['DEV_DTY']=row[1]['New Jatha'].upper()
            row_dict['MOBILE']=row[1]['MOBILE']
            row_dict['AGE']=row[1]['AGE']
            db.MasterSheet.insert(**row_dict)


    pprint.pprint(MasterSheet_df.columns,stream=logf)
    pprint.pprint(MasterSheet_df,stream=logf)

    logf.close()
    return dict(MasterSheet_df="Successfully updated")
