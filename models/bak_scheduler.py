#db2 = DAL('mysql://rootcat:7133783@rootcat.mysql.pythonanywhere-services.com/rootcat$AttendanceDB',pool_size=1,check_reserved=['all'])
#db2 = DAL('mysql://sql695965:qT6!xZ4!@sql6.freesqldatabase.com:3306/sql695965',pool_size=1,check_reserved=['all'])
#hmm
db2 = DAL('sqlite://storage4.db')

from gluon.tools import Mail
mail = Mail()
mail.settings.server = 'smtp.gmail.com:25'
mail.settings.sender = 'softwareattendance@gmail.com'
mail.settings.login = 'softwareattendance:hajari123'

def Roundoffdate(DATETIME, CUTOFF):
    import datetime
    RoundedDate = DATETIME
    if (DATETIME.hour < int(CUTOFF)) or (DATETIME.hour == int(CUTOFF) and DATETIME.minute == 0) or (CUTOFF == 24):
       RoundedDate = RoundedDate.replace(hour=0, minute=0, second=0, microsecond=0)
    else:
       RoundedDate = (RoundedDate + datetime.timedelta(days=1)).replace(hour=0, minute=0, second=0, microsecond=0) 
    return RoundedDate

def uploaddata_SSAttendance():
    import xlrd
    import os
    import re, time
    import datetime
    import logging
    logger = logging.getLogger("web2py.app.AttendanceSoftware")
    logger.setLevel(logging.DEBUG)

    logger.debug("uploaddata_SSAttendance got invoked")

    db.commit()

    #SSDates = db(db.SSAttendanceDate).select(db.SSAttendanceDate.SewadarNewID,db.SSAttendanceDate.DutyDate).as_list()
    #FastAccess = {}
    #for dates in SSDates:
    #    FastAccess[dates['SewadarNewID'] + datetime.datetime.strftime(dates['DutyDate'],'%Y-%m-%d')] = 1


    logger.debug("indexed SSAttendance Dates")
    path = os.path.join(request.folder,'private','SSAttendanceDates_xls.xlsx')
    pathlog = os.path.join(request.folder,'private','log_SSAttendance')
    logf = open(pathlog,'w')

    import commands
    logf.write(commands.getoutput('date'))
    logf.write("indexed SSAttendance Dates\n")
    logf.close()
    logf = open(pathlog,'a')
    workbook = xlrd.open_workbook(path)

    worksheet = workbook.sheet_names()
    #Attempt with any worksheet found
    worksheet = workbook.sheet_by_name(worksheet[0])
    header_row = 0

    #my headers are the headers in DB
    myheaders = ['OldSewadarID','SewadarNewID','Name','Gender','DepartmentID','DutyDate','Duty_Type']

    #headers are the name used in XLS
    headers = ['OldSewadarID','SewadarNewID','Name','Gender','DepartmentID','DutyDate','Duty Type']
    header_cells_in_xls = [worksheet.cell_value(0,idx) for idx in range(worksheet.ncols)]

    indexcol = [];
    for header_cell in headers:
            try:
                indexcol.append(header_cells_in_xls.index(header_cell))
            except Exception,e :
                print 'Header not found in xls: %s' %e
            else:
                print 'OK'

    row_dict_list = []

    for row in xrange(header_row+1, worksheet.nrows):
       i=0
       row_dict = {}
       logf.write(str(row) + "\n")
       logf.close()
       logf = open(pathlog,'a')

       for col in indexcol:
           cell_type = worksheet.cell_type(row, col)
           if cell_type == xlrd.XL_CELL_EMPTY:
               value = None
           elif cell_type == xlrd.XL_CELL_TEXT:
              value = worksheet.cell_value(row, col)
           elif cell_type == xlrd.XL_CELL_DATE:
              value = datetime.datetime(*xlrd.xldate_as_tuple(worksheet.cell_value(row, col), workbook.datemode))
           else:
              value = worksheet.cell_value(row, col)

           if i<=1:
               try:
                   value = re.sub("\s*","",value)
               except:
                   pass

           row_dict[myheaders[i]] = value
           i=i+1

       #print row_dict['SewadarNewID']
       try:
           db.SSAttendanceDate.insert(**row_dict)
       except:
           db((db.SSAttendanceDate.SewadarNewID == row_dict['SewadarNewID']) & (db.SSAttendanceDate.DutyDate == row_dict['DutyDate'])).update(Duty_Type=row_dict['Duty_Type'])
    #my headers are the headers in DB
       #db.SSAttendanceDate.update_or_insert(**row_dict)
       #db.SSAttendanceDate.update_or_insert(**dict(OldSewadarID=row_dict['OldSewadarID'],SewadarNewID=row_dict['SewadarNewID'],Name=row_dict['Name'],Gender=row_dict['Gender'],DepartmentID=row_dict['DepartmentID'],DutyDate=row_dict['DutyDate'],Duty_Type=row_dict['Duty_Type']))
                                             
       #db.SSAttendanceDate.update_or_insert((db.SSAttendanceDate.SewadarNewID==row_dict['SewadarNewID']) & (db.SSAttendanceDate.DutyDate==row_dict['DutyDate']),**row_dict)
    logger.debug("Sent to database engine")
    logf.write("Send to database engine\n")

    db.commit()

    logger.debug("Commited")
    logf.write("Commited\n")
    logf.close()
    mail.send('softwareattendance@gmail.com',
        'Comitted SSAttendance to database',
        'Success',
        attachments = mail.Attachment('/home/rootcat/new_web2py/web2py/applications/AttendanceSoftware/private/log_SSAttendance', content_id='text'))
    return 0


def ParshadListScheduled(DateSelectedStart,DateSelectedEnd,LALastLadiesNewGRNO,LBLastLadiesNewGRNO,GALastGentsNewGRNO,GBLastGentsNewGRNO,LastOSS,SSCountCutOffGents,SSCountCutOffLadies,CVCutOff,VisitCountCutOff,WWCutOff,DAY_END_TIME,DumpMachineAttendance,DumpSSAttendance):
    db.commit()
    import os
    pathlog = os.path.join(request.folder,'private','log_ParshadListScheduled')
    logf = open(pathlog,'w')
    import os
    from openpyxl import Workbook
    from openpyxl.cell import get_column_letter
    from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
    response.headers['Connection'] =  'keep-alive'
    dworkbook = Workbook()
    dworkbookAttendance = Workbook()
    dpath = os.path.join(request.folder,'private','TentativeParshadList.xlsx')
    dpathAttendance = os.path.join(request.folder,'private','DumpAttendance.xlsx')

    
    from gluon.sqlhtml import form_factory
    import datetime
    import time



    db.commit()
    SSCount = db(db.SSAttendanceCount).select()
    SSCountDict = {}
    for Sewadar in SSCount:
        SSCountDict[Sewadar.NewID,'Gender'] = Sewadar.gender
        SSCountDict[Sewadar.NewID,'TotalVisit'] = Sewadar.TotalVisit
        SSCountDict[Sewadar.NewID,'TotalCount'] = Sewadar.Total
        SSCountDict[Sewadar.NewID,'NAME'] = Sewadar.Name
        SSCountDict[Sewadar.NewID,'OldSewadarid'] = Sewadar.OldSewadarid
        SSCountDict[Sewadar.NewID,'status'] = Sewadar.status
    #Next setup: Change current visit dates(move to old visit dates)
    #Change total visits
    

    SSCURRENT_VISIT_MORNING_START = {}
    SSCURRENT_VISIT_MORNING_START['D0'] = datetime.datetime.strptime('28-September-2016 00:00:00','%d-%B-%Y %H:%M:%S')
    SSCURRENT_VISIT_MORNING_START['D1'] = datetime.datetime.strptime('29-September-2016 00:00:00','%d-%B-%Y %H:%M:%S')
    SSCURRENT_VISIT_MORNING_START['D2'] = datetime.datetime.strptime('30-September-2016 00:00:00','%d-%B-%Y %H:%M:%S')
    SSCURRENT_VISIT_MORNING_START['D3'] = datetime.datetime.strptime('01-October-2016 00:00:00','%d-%B-%Y %H:%M:%S')
    SSCURRENT_VISIT_MORNING_START['D4'] = datetime.datetime.strptime('02-October-2016 00:00:00','%d-%B-%Y %H:%M:%S')
    SSCURRENT_VISIT_MORNING_START['D5'] = datetime.datetime.strptime('03-October-2016 00:00:00','%d-%B-%Y %H:%M:%S')

    SSCURRENT_VISIT_MORNING_END = {}
    SSCURRENT_VISIT_MORNING_END['D0'] = datetime.datetime.strptime('28-September-2016 14:30:00','%d-%B-%Y %H:%M:%S')
    SSCURRENT_VISIT_MORNING_END['D1'] = datetime.datetime.strptime('29-September-2016 14:30:00','%d-%B-%Y %H:%M:%S')
    SSCURRENT_VISIT_MORNING_END['D2'] = datetime.datetime.strptime('30-September-2016 14:30:00','%d-%B-%Y %H:%M:%S')
    SSCURRENT_VISIT_MORNING_END['D3'] = datetime.datetime.strptime('01-October-2016 14:30:00','%d-%B-%Y %H:%M:%S')
    SSCURRENT_VISIT_MORNING_END['D4'] = datetime.datetime.strptime('02-October-2016 14:30:00','%d-%B-%Y %H:%M:%S')
    SSCURRENT_VISIT_MORNING_END['D5'] = datetime.datetime.strptime('03-October-2016 14:30:00','%d-%B-%Y %H:%M:%S')

    SSCURRENT_VISIT_EVENING_START = {}
    SSCURRENT_VISIT_EVENING_START['D0'] = datetime.datetime.strptime('28-September-2016 14:30:01','%d-%B-%Y %H:%M:%S')
    SSCURRENT_VISIT_EVENING_START['D1'] = datetime.datetime.strptime('29-September-2016 14:30:01','%d-%B-%Y %H:%M:%S')
    SSCURRENT_VISIT_EVENING_START['D2'] = datetime.datetime.strptime('30-September-2016 14:30:01','%d-%B-%Y %H:%M:%S')
    SSCURRENT_VISIT_EVENING_START['D3'] = datetime.datetime.strptime('01-October-2016 14:30:01','%d-%B-%Y %H:%M:%S')
    SSCURRENT_VISIT_EVENING_START['D4'] = datetime.datetime.strptime('02-October-2016 14:30:01','%d-%B-%Y %H:%M:%S')
    SSCURRENT_VISIT_EVENING_START['D5'] = datetime.datetime.strptime('03-October-2016 14:00:01','%d-%B-%Y %H:%M:%S')

    SSCURRENT_VISIT_EVENING_END = {}
    SSCURRENT_VISIT_EVENING_END['D0'] = datetime.datetime.strptime('28-September-2016 11:59:59','%d-%B-%Y %H:%M:%S')
    SSCURRENT_VISIT_EVENING_END['D1'] = datetime.datetime.strptime('29-September-2016 11:59:59','%d-%B-%Y %H:%M:%S')
    SSCURRENT_VISIT_EVENING_END['D2'] = datetime.datetime.strptime('30-September-2016 11:59:59','%d-%B-%Y %H:%M:%S')
    SSCURRENT_VISIT_EVENING_END['D3'] = datetime.datetime.strptime('01-October-2016 11:59:59','%d-%B-%Y %H:%M:%S')
    SSCURRENT_VISIT_EVENING_END['D4'] = datetime.datetime.strptime('02-October-2016 11:59:59','%d-%B-%Y %H:%M:%S')
    SSCURRENT_VISIT_EVENING_END['D5'] = datetime.datetime.strptime('03-October-2016 11:59:59','%d-%B-%Y %H:%M:%S')


    SSCURRENT_VISIT = {}
    SSCURRENT_VISIT['COUNT'] = 5
    SSCURRENT_VISIT['D0'] = datetime.datetime.strptime('28-September-2016 00:00:00','%d-%B-%Y %H:%M:%S')
    SSCURRENT_VISIT['D1'] = datetime.datetime.strptime('28-September-2016 12:00:00','%d-%B-%Y %H:%M:%S')
    SSCURRENT_VISIT['D2'] = datetime.datetime.strptime('29-September-2016 13:00:01','%d-%B-%Y %H:%M:%S')
    SSCURRENT_VISIT['D3'] = datetime.datetime.strptime('30-September-2016 09:00:01','%d-%B-%Y %H:%M:%S')
    SSCURRENT_VISIT['D4'] = datetime.datetime.strptime('01-October-2016 09:00:01','%d-%B-%Y %H:%M:%S')
    SSCURRENT_VISIT['D5'] = datetime.datetime.strptime('02-October-2016 09:00:01','%d-%B-%Y %H:%M:%S')

    VISIT_DATES = {}
    TOTAL_VISIT = 13
    VISIT_DATES['V0','COUNT'] = 5
    VISIT_DATES['V0','D0'] = datetime.datetime.strptime('05-October-2011 00:00:00','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V0','D1'] = datetime.datetime.strptime('05-October-2011 14:30:01','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V0','D2'] = datetime.datetime.strptime('06-October-2011 12:00:00','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V0','D3'] = datetime.datetime.strptime('07-October-2011 12:00:00','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V0','D4'] = datetime.datetime.strptime('08-October-2011 12:00:00','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V0','D5'] = datetime.datetime.strptime('09-October-2011 12:00:00','%d-%B-%Y %H:%M:%S')

    VISIT_DATES['V1','COUNT'] = 5
    VISIT_DATES['V1','D0'] = datetime.datetime.strptime('23-November-2011 00:00:00','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V1','D1'] = datetime.datetime.strptime('23-November-2011 14:30:01','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V1','D2'] = datetime.datetime.strptime('24-November-2011 12:00:00','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V1','D3'] = datetime.datetime.strptime('25-November-2011 12:00:00','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V1','D4'] = datetime.datetime.strptime('26-November-2011 12:00:00','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V1','D5'] = datetime.datetime.strptime('27-November-2011 12:00:00','%d-%B-%Y %H:%M:%S')

    VISIT_DATES['V2','COUNT'] = 5
    VISIT_DATES['V2','D0'] = datetime.datetime.strptime('07-March-2012 00:00:00','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V2','D1'] = datetime.datetime.strptime('07-March-2012 14:30:01','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V2','D2'] = datetime.datetime.strptime('08-March-2012 12:00:00','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V2','D3'] = datetime.datetime.strptime('09-March-2012 12:00:00','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V2','D4'] = datetime.datetime.strptime('10-March-2012 12:00:00','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V2','D5'] = datetime.datetime.strptime('11-March-2012 12:00:00','%d-%B-%Y %H:%M:%S')

    VISIT_DATES['V3','COUNT'] = 5
    VISIT_DATES['V3','D0'] = datetime.datetime.strptime('10-October-2012 00:00:00','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V3','D1'] = datetime.datetime.strptime('10-October-2012 14:30:01','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V3','D2'] = datetime.datetime.strptime('11-October-2012 12:00:00','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V3','D3'] = datetime.datetime.strptime('12-October-2012 12:00:00','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V3','D4'] = datetime.datetime.strptime('13-October-2012 12:00:00','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V3','D5'] = datetime.datetime.strptime('14-October-2012 12:00:00','%d-%B-%Y %H:%M:%S')

    VISIT_DATES['V4','COUNT'] = 5
    VISIT_DATES['V4','D0'] = datetime.datetime.strptime('28-November-2012 00:00:00','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V4','D1'] = datetime.datetime.strptime('28-November-2012 14:30:01','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V4','D2'] = datetime.datetime.strptime('29-November-2012 12:00:00','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V4','D3'] = datetime.datetime.strptime('30-November-2012 12:00:00','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V4','D4'] = datetime.datetime.strptime('01-December-2012 12:00:00','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V4','D5'] = datetime.datetime.strptime('02-December-2012 12:00:00','%d-%B-%Y %H:%M:%S')

    VISIT_DATES['V5','COUNT'] = 5
    VISIT_DATES['V5','D0'] = datetime.datetime.strptime('13-March-2013 00:00:00','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V5','D1'] = datetime.datetime.strptime('13-March-2013 14:30:01','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V5','D2'] = datetime.datetime.strptime('14-March-2013 12:00:00','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V5','D3'] = datetime.datetime.strptime('15-March-2013 12:00:00','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V5','D4'] = datetime.datetime.strptime('16-March-2013 12:00:00','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V5','D5'] = datetime.datetime.strptime('17-March-2013 12:00:00','%d-%B-%Y %H:%M:%S')

    VISIT_DATES['V6','COUNT'] = 5
    VISIT_DATES['V6','D0'] = datetime.datetime.strptime('09-October-2013 00:00:00','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V6','D1'] = datetime.datetime.strptime('09-October-2013 14:30:01','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V6','D2'] = datetime.datetime.strptime('10-October-2013 12:00:00','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V6','D3'] = datetime.datetime.strptime('11-October-2013 12:00:00','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V6','D4'] = datetime.datetime.strptime('12-October-2013 12:00:00','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V6','D5'] = datetime.datetime.strptime('13-October-2013 12:00:00','%d-%B-%Y %H:%M:%S')

    VISIT_DATES['V7','COUNT'] = 3
    VISIT_DATES['V7','D0'] = datetime.datetime.strptime('10-January-2014 00:00:00','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V7','D1'] = datetime.datetime.strptime('10-January-2014 14:30:01','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V7','D2'] = datetime.datetime.strptime('11-January-2014 12:00:00','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V7','D3'] = datetime.datetime.strptime('12-January-2014 12:00:00','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V7','D4'] = datetime.datetime.strptime('13-January-2014 12:00:00','%d-%B-%Y %H:%M:%S')

    VISIT_DATES['V8','COUNT'] = 5
    VISIT_DATES['V8','D0'] = datetime.datetime.strptime('01-October-2014 00:00:00','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V8','D1'] = datetime.datetime.strptime('01-October-2014 14:30:01','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V8','D2'] = datetime.datetime.strptime('02-October-2014 12:00:00','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V8','D3'] = datetime.datetime.strptime('03-October-2014 12:00:00','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V8','D4'] = datetime.datetime.strptime('04-October-2014 12:00:00','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V8','D5'] = datetime.datetime.strptime('05-October-2014 12:00:00','%d-%B-%Y %H:%M:%S')

    VISIT_DATES['V9','COUNT'] = 5
    VISIT_DATES['V9','D0'] = datetime.datetime.strptime('26-November-2014 00:00:00','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V9','D1'] = datetime.datetime.strptime('26-November-2014 14:30:01','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V9','D2'] = datetime.datetime.strptime('27-November-2014 12:00:00','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V9','D3'] = datetime.datetime.strptime('28-November-2014 12:00:00','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V9','D4'] = datetime.datetime.strptime('29-November-2014 12:00:00','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V9','D5'] = datetime.datetime.strptime('30-November-2014 12:00:00','%d-%B-%Y %H:%M:%S')

    VISIT_DATES['V10','COUNT'] = 5
    VISIT_DATES['V10','D0'] = datetime.datetime.strptime('11-March-2015 00:00:00','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V10','D1'] = datetime.datetime.strptime('11-March-2015 14:30:01','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V10','D2'] = datetime.datetime.strptime('12-March-2015 12:00:00','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V10','D3'] = datetime.datetime.strptime('13-March-2015 12:00:00','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V10','D4'] = datetime.datetime.strptime('14-March-2015 12:00:00','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V10','D5'] = datetime.datetime.strptime('15-March-2015 12:00:00','%d-%B-%Y %H:%M:%S')
    
    VISIT_DATES['V11','COUNT'] = 5
    VISIT_DATES['V11','D0'] = datetime.datetime.strptime('07-October-2015 00:00:00','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V11','D1'] = datetime.datetime.strptime('07-October-2015 14:30:01','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V11','D2'] = datetime.datetime.strptime('08-October-2015 12:00:00','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V11','D3'] = datetime.datetime.strptime('09-October-2015 12:00:00','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V11','D4'] = datetime.datetime.strptime('10-October-2015 12:00:00','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V11','D5'] = datetime.datetime.strptime('11-October-2015 12:00:00','%d-%B-%Y %H:%M:%S')


    VISIT_DATES['V12','COUNT'] = 5
    VISIT_DATES['V12','D0'] = datetime.datetime.strptime('25-November-2015 00:00:00','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V12','D1'] = datetime.datetime.strptime('25-November-2015 14:30:01','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V12','D2'] = datetime.datetime.strptime('26-November-2015 12:00:00','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V12','D3'] = datetime.datetime.strptime('27-November-2015 12:00:00','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V12','D4'] = datetime.datetime.strptime('28-November-2015 12:00:00','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V12','D5'] = datetime.datetime.strptime('29-November-2015 12:00:00','%d-%B-%Y %H:%M:%S')

    VISIT_DATES['V13','COUNT'] = 5
    VISIT_DATES['V13','D0'] = datetime.datetime.strptime('9-March-2016 00:00:00','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V13','D1'] = datetime.datetime.strptime('9-March-2016 14:30:01','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V13','D2'] = datetime.datetime.strptime('10-March-2016 12:00:00','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V13','D3'] = datetime.datetime.strptime('11-March-2016 12:00:00','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V13','D4'] = datetime.datetime.strptime('12-March-2016 12:00:00','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V13','D5'] = datetime.datetime.strptime('13-March-2016 12:00:00','%d-%B-%Y %H:%M:%S')

    message = "ALL OK "
    ParshadList = {}


    try:
        os.remove(dpath)
    except:
        pass

    try:
        os.remove(dpathAttendance)
    except:
        pass

    SSDate = db((db.SSAttendanceDate.DutyDate >= (datetime.datetime.strptime(DateSelectedStart,'%Y-%m-%d').replace(hour=0, minute=0, second=0, microsecond=0))) & (db.SSAttendanceDate.DutyDate <= (datetime.datetime.strptime(DateSelectedEnd,'%Y-%m-%d').replace(hour=23, minute=59, second=59, microsecond=999)))).select('SewadarNewID','DutyDate','Duty_Type')
    MachineDate = db((db.MachineAttendance.DATETIME >= (datetime.datetime.strptime(DateSelectedStart,'%Y-%m-%d') - datetime.timedelta(hours=((24 - int(DAY_END_TIME)) % 24)))) & (db.MachineAttendance.DATETIME <= (datetime.datetime.strptime(DateSelectedEnd,'%Y-%m-%d') + datetime.timedelta(hours=int(DAY_END_TIME))))).select('GRNO','NewGRNO','DATETIME','TYPE')
    dExceptionMail = db(db.ParshadMailException).select()
    print "Machine Attendance between date"
    print "length of MachineDate =" + str(len(MachineDate))

    ExceptionMail = {}
    #Keywords for exception
    #ALL , Visits Count,Current Visit
    for row in dExceptionMail:
        ExceptionMail[row.NewGRNO] = row.Status
        ExceptionMail[row.NewGRNO,'ExceptionField'] = row.ExceptionField


    ParshadList['SEWADARS'] = []
    SSAttendanceDictionary = {}
    print "Collecting SSDate"
    for SSEntry in SSDate:
        ParshadList['SEWADARS'].append(SSEntry.SewadarNewID)
        if SSEntry.Duty_Type == 'W':
            now = datetime.datetime.now()
            SSAttendanceDictionary[SSEntry.SewadarNewID,Roundoffdate(SSEntry.DutyDate,23),'W'] = SSEntry.DutyDate
            try:
                ParshadList[SSEntry.SewadarNewID,'SSCount'] = ParshadList[SSEntry.SewadarNewID,'SSCount'] + 2
            except:
                ParshadList[SSEntry.SewadarNewID,'SSCount'] = 2
            try:
                if SSEntry.DutyDate.year == now.year:
                    ParshadList[SSEntry.SewadarNewID,'SSWWCount'] = ParshadList[SSEntry.SewadarNewID,'SSWWCount'] + 1
            except:
                if SSEntry.DutyDate.year == now.year:
                    ParshadList[SSEntry.SewadarNewID,'SSWWCount'] = 1
        else:
            SSAttendanceDictionary[SSEntry.SewadarNewID,Roundoffdate(SSEntry.DutyDate,23),'D'] = SSEntry.DutyDate
            try:
                ParshadList[SSEntry.SewadarNewID,'SSCount'] = ParshadList[SSEntry.SewadarNewID,'SSCount'] + 1
            except:
                ParshadList[SSEntry.SewadarNewID,'SSCount'] = 1

        ParshadList[SSEntry.SewadarNewID,'SSCountOld'] = ParshadList[SSEntry.SewadarNewID,'SSCount']

    MachineAttendanceAdditional = {}


    try:
        db.tempMachineAttendanceAdditional.drop()
    except:
        pass

    #Now define the table
    db.define_table('tempMachineAttendanceAdditional',
            Field('GRNO','string'),
            Field('NewGRNO','string'),
            Field('Duty_Type','string'),
            Field('DutyDate','datetime'),
            Field('DutyDateList','list:string'),
            migrate=True,
            redefine=True,
            format='%(NewGRNO)s')

    try:
        db(db.tempMachineAttendanceAdditional.id > 0).delete()
    except:
        pass

    print "Collecting Machine Date"

    for MEntry in MachineDate:
        if ((MEntry.DATETIME < VISIT_DATES['V12','D5']) and (MEntry.DATETIME > VISIT_DATES['V12','D0'])) or ((MEntry.DATETIME < VISIT_DATES['V11','D5']) and (MEntry.DATETIME > VISIT_DATES['V11','D0'])) or ((MEntry.DATETIME < VISIT_DATES['V10','D5']) and (MEntry.DATETIME > VISIT_DATES['V10','D0'])) or ((MEntry.DATETIME < VISIT_DATES['V9','D5']) and (MEntry.DATETIME > VISIT_DATES['V9','D0'])) or  ((MEntry.DATETIME < VISIT_DATES['V8','D5']) and (MEntry.DATETIME > VISIT_DATES['V8','D0'])):
            pass
        else:
            #Select earliest entry but give preference to WW attendance
            if MEntry.TYPE == 'WMANUAL':
                try:
                    a = SSAttendanceDictionary[MEntry.NewGRNO,Roundoffdate(MEntry.DATETIME,DAY_END_TIME),'W']
                except:
                    try:
                        MachineAttendanceAdditional[MEntry.NewGRNO,MEntry.GRNO,Roundoffdate(MEntry.DATETIME,DAY_END_TIME),'W'].append(MEntry.DATETIME) 
                    except:
                        MachineAttendanceAdditional[MEntry.NewGRNO,MEntry.GRNO,Roundoffdate(MEntry.DATETIME,DAY_END_TIME),'W'] = [MEntry.DATETIME]


            else:
                try:
                    a = SSAttendanceDictionary[MEntry.NewGRNO,Roundoffdate(MEntry.DATETIME,DAY_END_TIME),'D']
                except:
                    try:
                        a = SSAttendanceDictionary[MEntry.NewGRNO,Roundoffdate(MEntry.DATETIME,DAY_END_TIME),'W']
                    except:
                        try:
                            MachineAttendanceAdditional[MEntry.NewGRNO,MEntry.GRNO,Roundoffdate(MEntry.DATETIME,DAY_END_TIME),'D'].append(MEntry.DATETIME) 
                        except:
                            MachineAttendanceAdditional[MEntry.NewGRNO,MEntry.GRNO,Roundoffdate(MEntry.DATETIME,DAY_END_TIME),'D'] = [MEntry.DATETIME]


    del SSAttendanceDictionary

    print "Preparing SS Count"

    for key, value in MachineAttendanceAdditional.iteritems():
        NewGRNO, GRNO, DutyDate , DutyType= key
        ParshadList['SEWADARS'].append(NewGRNO)
        if DutyType == 'W':
            try:
                ParshadList[NewGRNO,'SSCount'] = ParshadList[NewGRNO,'SSCount'] + 2
            except:
                ParshadList[NewGRNO,'SSCount'] = 2
        else:
            try:
                a = MachineAttendanceAdditional[NewGRNO,GRNO,DutyDate,'W']
            except:
                try:
                    ParshadList[NewGRNO,'SSCount'] = ParshadList[NewGRNO,'SSCount'] + 1
                except:
                    ParshadList[NewGRNO,'SSCount'] = 1


    for key, value in MachineAttendanceAdditional.iteritems():
        MachineAttendanceAdditionalDictionary = {}
        NewGRNO, GRNO, DutyDate , DutyType= key
        ParshadList['SEWADARS'].append(NewGRNO)
        MachineAttendanceAdditionalDictionary['NewGRNO'] = NewGRNO 
        MachineAttendanceAdditionalDictionary['GRNO'] = GRNO
        MachineAttendanceAdditionalDictionary['Duty_Type'] = DutyType
        MachineAttendanceAdditionalDictionary['DutyDate'] = DutyDate
        MachineAttendanceAdditionalDictionary['DutyDateList'] = value
        try:
            ParshadList[NewGRNO,'Before Visit Machine Additional'].append(DutyDate)
        except:
            ParshadList[NewGRNO,'Before Visit Machine Additional'] = [DutyDate]
        db.tempMachineAttendanceAdditional.insert(**MachineAttendanceAdditionalDictionary)


    del MachineAttendanceAdditional

    print "Creating Attendance Sheets"

    if DumpSSAttendance == 'YES': 
        dSSdate = dworkbookAttendance.create_sheet(0)
        dSSdate.title = "SSDate"
        dSSdate.append(['SewadarNewID','DutyDate','Duty_Type'])
        for row in SSDate:
            dSSdate.append([row.SewadarNewID , row.DutyDate , row.Duty_Type])

    if DumpMachineAttendance == 'YES': 
        dMachineAttendance = dworkbookAttendance.create_sheet(0)
        dMachineAttendance.title = "MachineAttendance"
        dMachineAttendance.append(['GRNO','NewGRNO','DATETIME','TYPE'])
        for row in MachineDate:
            dMachineAttendance.append([row.GRNO,row.NewGRNO,row.DATETIME,row.TYPE])

    MachineDifference = db(db.tempMachineAttendanceAdditional.id > 0).select()

    dMachineDifference = dworkbook.create_sheet(0)
    dMachineDifference.title = "MachineDifference"
    dMachineDifference.append(['NewGRNO','GRNO','Duty_Type','DutyDate','DutyDateList','TimeDifference'])
    for row in MachineDifference:
        DutyDateList = ", ".join(map(str, row.DutyDateList))
        n = len(row.DutyDateList)
        TimeDifference = max(map(datetime.datetime.strptime,row.DutyDateList,['%Y-%m-%d %H:%M:%S']*n)) - min(map(datetime.datetime.strptime,row.DutyDateList,['%Y-%m-%d %H:%M:%S']*n))
        dMachineDifference.append([row.NewGRNO,row.GRNO,row.Duty_Type,row.DutyDate,DutyDateList,str(TimeDifference)])


    dworkbookAttendance.save(dpathAttendance)
    

    throttle = 1

    for visit in xrange(0,TOTAL_VISIT):
        VCOUNT = 0
        VCOUNT = VISIT_DATES['V'+str(visit),'COUNT']

        vc = 'V'+str(visit)
        
        for day in xrange(0,VCOUNT):
            dc = 'D'+str(day)
            dc1 = 'D'+str(day+1)
            print "Analyzing " + vc + dc

            SSDate = db((db.SSAttendanceDate.DutyDate >= datetime.datetime(VISIT_DATES[vc,dc].year,VISIT_DATES[vc,dc].month,VISIT_DATES[vc,dc].day,0,0,0)) & (db.SSAttendanceDate.DutyDate <= datetime.datetime((VISIT_DATES[vc,dc]+datetime.timedelta(hours=24)).year,(VISIT_DATES[vc,dc]+datetime.timedelta(hours=24)).month,(VISIT_DATES[vc,dc]+datetime.timedelta(hours=24)).day,0,0,0))).select('SewadarNewID')
            MachineDate = db((db.MachineAttendance.DATETIME >= VISIT_DATES[vc,dc]) & (db.MachineAttendance.DATETIME <= VISIT_DATES[vc,dc1])).select('NewGRNO')
            print "length of SSDate =" + str(len(SSDate))
            print "length of MachineDate =" + str(len(MachineDate))

            for row in SSDate:
                ParshadList['SEWADARS'].append(row.SewadarNewID)
                try:
                    a = ParshadList[row.SewadarNewID,vc]
                except:
                    ParshadList[row.SewadarNewID,vc] = VCOUNT - day

            for row in MachineDate:
                ParshadList['SEWADARS'].append(row.NewGRNO)
                try:
                    if ParshadList[row.NewGRNO,vc] < VCOUNT - day :
                        ParshadList[row.NewGRNO,vc] = VCOUNT - day
                        try:
                            ParshadList[row.NewGRNO,'Additional Machine Visit Days'].append(vc + dc + ' onwards')
                        except:
                            ParshadList[row.NewGRNO,'Additional Machine Visit Days'] = [vc + dc + ' onwards']
                except:
                    ParshadList[row.NewGRNO,vc] = (VCOUNT - day)
                    try:
                        ParshadList[row.NewGRNO,'Additional Machine Visit Days'].append(vc + dc + ' onwards')
                    except:
                        ParshadList[row.NewGRNO,'Additional Machine Visit Days'] = [vc + dc + ' onwards']
            del SSDate
            del MachineDate


    print "reading tentative parshad list"
    SSTentativeParshadList = db(db.SSTentativeParshadList.id > 0).select('NewGRNO','Status')

    for row in SSTentativeParshadList:
        ParshadList['SEWADARS'].append(row.NewGRNO)
        ParshadList[row.NewGRNO,'SS Tentative Parshad Status'] = row.Status


    print "reading initiated list"
    InitiatedList = db(db.InitiatedList.id > 0).select('NewGRNO','Status')

    for row in InitiatedList:
        ParshadList['SEWADARS'].append(row.NewGRNO)
        ParshadList[row.NewGRNO,'Initiation Status'] = row.Status


    print "reading previous parshad list"
    PreviousParshadList = db(db.PreviousParshadList.id > 0).select('NewGRNO','Status')

    for row in PreviousParshadList:
        ParshadList['SEWADARS'].append(row.NewGRNO)
        ParshadList[row.NewGRNO,'Previous Visit Parshad Status'] = row.Status

    print "Analyzing Current Visit"
    logf.write("Analkysinz current visit")
    
    #time.sleep (1);
    for day in xrange(0,SSCURRENT_VISIT['COUNT']):
        dc = 'D'+str(day)
        dc1 = 'D'+str(day+1)
        print "Analyzing Current Visit day " + str(day)

        SSDate = db((db.SSAttendanceDate.DutyDate >= SSCURRENT_VISIT['D'+str(day)]) & (db.SSAttendanceDate.DutyDate <= SSCURRENT_VISIT['D'+str(day+1)])).select('SewadarNewID')
        print "Collecting SSDate"
        time.sleep (1);
        MachineDate = db((db.MachineAttendance.DATETIME >= SSCURRENT_VISIT['D'+str(day)]) & (db.MachineAttendance.DATETIME <= SSCURRENT_VISIT['D'+str(day+1)])).select('NewGRNO')
        MachineDateMorning = db((db.MachineAttendance.DATETIME >= SSCURRENT_VISIT_MORNING_START['D'+str(day)]) & (db.MachineAttendance.DATETIME <= SSCURRENT_VISIT_MORNING_END['D'+str(day+1)])).select('NewGRNO')
        MachineDateEvening = db((db.MachineAttendance.DATETIME >= SSCURRENT_VISIT_EVENING_START['D'+str(day)]) & (db.MachineAttendance.DATETIME <= SSCURRENT_VISIT_EVENING_END['D'+str(day+1)])).select('NewGRNO')
        print "Collected Machine and SSDate"

        for row in SSDate:
            ParshadList['SEWADARS'].append(row.SewadarNewID)
            try:
                a = ParshadList[row.SewadarNewID,'CV']
            except:
                ParshadList[row.SewadarNewID,'CV'] = SSCURRENT_VISIT['COUNT'] - day
                ParshadList[row.SewadarNewID,'CVOld'] = SSCURRENT_VISIT['COUNT'] - day

            ParshadList[row.SewadarNewID,'CV '+ dc] = 'P'


        print "Analyzed SSDate for current visit"

        for row in MachineDate:
            ParshadList['SEWADARS'].append(row.NewGRNO)
            try:
                a = ParshadList[row.NewGRNO,'CV']
            except:
                ParshadList[row.NewGRNO,'CV'] = SSCURRENT_VISIT['COUNT'] - day

            try:
                a = ParshadList[row.NewGRNO,'CV '+ dc]
            except:
                ParshadList[row.NewGRNO,'CV '+ dc] = 'C'

        for row in MachineDateMorning:
            ParshadList['SEWADARS'].append(row.NewGRNO)
            ParshadList[row.NewGRNO,'CVMORNING',day] = 'P'

        for row in MachineDateEvening:
            ParshadList['SEWADARS'].append(row.NewGRNO)
            ParshadList[row.NewGRNO,'CVMORNING',day] = 'P'

        print "Analyzed Machine Dates for current visit"

    ParshadList['SEWADARS'] = set(ParshadList['SEWADARS'])

    print "Creating Super Set Parshad Status"
    logf.write("Creating Super set")

    dParshadList = dworkbook.create_sheet(0)
    dParshadList.title = "ParshadList"
    dParshadList.append(['NewGRNO','SSCount','SSCountOld','CVCount','CVCountOld','MachineBeforeVisitAdditional','WW Count','CV D1','CV D2','CV D3','CV D4','CV D5','MachinePreviousVisitAddition','InitiationStatus','SS Tentative Parshad Status','Previous Visit Parshad Status','VISIT_COUNTS','V0','V1','V2','V3','V4','V5','V6','V7','V8','V9','V10','V11','V12'])
    row_num = 1
    for Sewadar in ParshadList['SEWADARS']:
        #print "SuperSet Sewadar " + Sewadar
        try:
            SSCount = ParshadList[Sewadar,'SSCount'] 
        except:
            SSCount = 0

        try:
            SSWWCount = ParshadList[Sewadar,'SSWWCount'] 
        except:
            SSWWCount = 0


        try:
            SSCountOld = ParshadList[Sewadar,'SSCountOld'] 
        except:
            SSCountOld = 0
        try:
            CVCount = ParshadList[Sewadar,'CV'] 
        except:
            CVCount = 0
        try:
            CVCountOld = ParshadList[Sewadar,'CVOld'] 
        except:
            CVCountOld = 0
        try:
            CVD0 = ParshadList[Sewadar,'CV D0']
        except:
            CVD0 = 'A'
        try:
            CVD1 = ParshadList[Sewadar,'CV D1']
        except:
            CVD1 = 'A'
        try:
            CVD2 = ParshadList[Sewadar,'CV D2']
        except:
            CVD2 = 'A'
        try:
            CVD3 = ParshadList[Sewadar,'CV D3']
        except:
            CVD3 = 'A'
        try:
            CVD4 = ParshadList[Sewadar,'CV D4']
        except:
            CVD4 = 'A'
        try:
            MachineBeforeVisitAdditional = ", ".join(map(str, ParshadList[Sewadar,'Before Visit Machine Additional']))
        except:
            MachineBeforeVisitAdditional = ""
        try:
            MachinePreviousVisitAddition = ", ".join(map(str, ParshadList[Sewadar,'Additional Machine Visit Days']))
        except:
            MachinePreviousVisitAddition = ""
        try:
            InitiationStatus = ParshadList[Sewadar,'Initiation Status']
        except:
            InitiationStatus = "Not Found"
        try:
            PreviousVisitParshadStatus = ParshadList[Sewadar,'Previous Visit Parshad Status']
        except:
            PreviousVisitParshadStatus = "Not Found"
        try:
            SSTentativeParshadStatus = ParshadList[Sewadar,'SS Tentative Parshad Status']
        except:
            SSTentativeParshadStatus = "Not Found"

        dParshadList.append([Sewadar,SSCount,SSCountOld,CVCount,CVCountOld,MachineBeforeVisitAdditional,SSWWCount,CVD0,CVD1,CVD2,CVD3,CVD4,MachinePreviousVisitAddition,InitiationStatus,SSTentativeParshadStatus,PreviousVisitParshadStatus])
       
        row_num = row_num + 1
        visit_counts = 0
        for visit in xrange(0,TOTAL_VISIT):
            try:
                dParshadList.cell(get_column_letter(visit+18)+str(row_num)).value = ParshadList[Sewadar,'V'+str(visit)]
                if ParshadList[Sewadar,'V'+str(visit)] >= 3:
                    visit_counts = visit_counts + 1 
            except:
                dParshadList.cell(get_column_letter(visit+18)+str(row_num)).value = 0
                
        dParshadList.cell(get_column_letter(17)+str(row_num)).value = visit_counts


    MasterSheet = db(db.MasterSheet.id > 0).select('SewadarNewID','GR_NO','NAME','CANTEEN','DEV_DTY')
    dWorkingParshad = dworkbook.create_sheet(0)
    dWorkingParshad.title = "WorkingParshad"
    dWorkingParshad.append(['SewadarNewID','GR_NO','NAME','CANTEEN','DEV_DTY','InitiationStatus','SS Tentative Parshad Status','CANTEEN Parshad Status','CANTEEN Parshad Remarks','SSCount','SSCountNew','CVCount','CVCountNew','MachineBeforeVisitAdditional','WW Count','CV D1','CV D2','CV D3','CV D4','CV D5','MachinePreviousVisitAddition','Previous Visit Parshad Status','VISIT_COUNTS','V0','V1','V2','V3','V4','V5','V6','V7','V8','V9','V10','V11','V12'])

    print "Creating Master Specific Parshad Status"
    logf.write("Creating Master Specific Parshad Status")

    row_num = 2
    for row in MasterSheet:
        Sewadar = row.SewadarNewID
        #print "Master Sewadar " + Sewadar

        dWorkingParshad.cell(get_column_letter(1) + str(row_num)).value = row.SewadarNewID
        dWorkingParshad.cell(get_column_letter(2) + str(row_num)).value = row.GR_NO
        dWorkingParshad.cell(get_column_letter(3) + str(row_num)).value = row.NAME
        dWorkingParshad.cell(get_column_letter(4) + str(row_num)).value = row.CANTEEN
        dWorkingParshad.cell(get_column_letter(5) + str(row_num)).value = row.DEV_DTY




        try:
            SSCount = ParshadList[Sewadar,'SSCount'] 
        except:
            SSCount = 0
        try:
            SSCountOld = ParshadList[Sewadar,'SSCountOld'] 
        except:
            SSCountOld = 0
        try:
            SSWWCount = ParshadList[Sewadar,'SSWWCount'] 
        except:
            SSWWCount = 0
        try:
            CVCount = ParshadList[Sewadar,'CV'] 
        except:
            CVCount = 0
        try:
            CVCountOld = ParshadList[Sewadar,'CVOld'] 
        except:
            CVCountOld = 0
        try:
            CVD0 = ParshadList[Sewadar,'CV D0']
        except:
            CVD0 = 'A'
        try:
            CVD1 = ParshadList[Sewadar,'CV D1']
        except:
            CVD1 = 'A'
        try:
            CVD2 = ParshadList[Sewadar,'CV D2']
        except:
            CVD2 = 'A'
        try:
            CVD3 = ParshadList[Sewadar,'CV D3']
        except:
            CVD3 = 'A'
        try:
            CVD4 = ParshadList[Sewadar,'CV D4']
        except:
            CVD4 = 'A'
        try:
            MachineBeforeVisitAdditional = ", ".join(map(str, ParshadList[Sewadar,'Before Visit Machine Additional']))
        except:
            MachineBeforeVisitAdditional = ""
        try:
            MachinePreviousVisitAddition = ", ".join(map(str, ParshadList[Sewadar,'Additional Machine Visit Days']))
        except:
            MachinePreviousVisitAddition = ""
        try:
            InitiationStatus = ParshadList[Sewadar,'Initiation Status']
        except:
            InitiationStatus = "Not Found"
        try:
            PreviousVisitParshadStatus = ParshadList[Sewadar,'Previous Visit Parshad Status']
        except:
            PreviousVisitParshadStatus = "Not Found"
        try:
            SSTentativeParshadStatus = ParshadList[Sewadar,'SS Tentative Parshad Status']
        except:
            SSTentativeParshadStatus = "Not Found"


        visit_counts = 0
        for visit in xrange(0,TOTAL_VISIT):
            try:
                dWorkingParshad.cell(get_column_letter(visit+24)+str(row_num)).value = ParshadList[Sewadar,'V'+str(visit)]
                if ParshadList[Sewadar,'V'+str(visit)] >= 3:
                    visit_counts = visit_counts + 1 
            except:
                dWorkingParshad.cell(get_column_letter(visit+24)+str(row_num)).value = 0
                
        try:
            if visit_counts < SSCountDict[Sewadar,'TotalVisit']:
                visit_counts = SSCountDict[Sewadar,'TotalVisit']
        except:
            test = 1

        dWorkingParshad.cell(get_column_letter(23)+str(row_num)).value = visit_counts

        CanteenParshadStatus = "OK"
        CanteenParshadRemark = []


        try:
            if SSCountDict[row.SewadarNewID,'status'] == 'Slip':
                try:
                    if ExceptionMail[row.SewadarNewID,'ExceptionField'] == "Visits Count":
                        pass
                    else:
                        if (int(visit_counts) < int(VisitCountCutOff)):
                            CanteenParshadRemark.append(str(visit_counts) + " Visits Attended")
                            CanteenParshadStatus = "Not OK"
                except:
                    if (int(visit_counts) < int(VisitCountCutOff)):
                        CanteenParshadRemark.append(str(visit_counts) + " Visits Attended")
                        CanteenParshadStatus = "Not OK"
        except:
            CanteenParshadRemark.append("SS Missing")
            CanteenParshadStatus = "SS Missing"

        try:
            if ExceptionMail[row.SewadarNewID,'ExceptionField'] == "Current Visit":
                pass
            else:
                if (CVCount < int(CVCutOff)):
                    CanteenParshadRemark.append("Current Visit Short")
                    CanteenParshadStatus = "Not OK"
        except:
            if (CVCount < int(CVCutOff)):
                CanteenParshadRemark.append("Current Visit Short")
                CanteenParshadStatus = "Not OK"

        if (row.SewadarNewID.find("G") > -1):
            try:
                if ExceptionMail[row.SewadarNewID,'ExceptionField'] == "WW Count":
                    pass
                else:
                    if SSCount >= 70:
                        if (SSWWCount < int(WWCutOff) - 1):
                            #CanteenParshadRemark.append("WW short by " + str(int(WWCutOff) - SSWWCount - 1))
                            CanteenParshadRemark.append(str(SSWWCount) + " WW done")
                            CanteenParshadStatus = "Not OK"
                    else:
                        if (SSWWCount < int(WWCutOff)):
                            #CanteenParshadRemark.append("WW short by " + str(int(WWCutOff) - SSWWCount))
                            CanteenParshadRemark.append(str(SSWWCount) + " WW done")
                            CanteenParshadStatus = "Not OK"
            except:
                if SSCount >= 70:
                    if (SSWWCount < int(WWCutOff) - 1):
                        #CanteenParshadRemark.append("WW short by " + str(int(WWCutOff) - SSWWCount - 1))
                        CanteenParshadRemark.append(str(SSWWCount) + " WW done")
                        CanteenParshadStatus = "Not OK"
                else:
                    if (SSWWCount < int(WWCutOff)):
                        #CanteenParshadRemark.append("WW short by " + str(int(WWCutOff) - SSWWCount))
                        CanteenParshadRemark.append(str(SSWWCount) + " WW done")
                        CanteenParshadStatus = "Not OK"

            try:
                if ExceptionMail[row.SewadarNewID,'ExceptionField'] == "SS Count":
                    pass
                else:
                    if (SSCount < int(SSCountCutOffGents)):
                        CanteenParshadRemark.append(str(int(SSCountCutOffGents) - SSCount) + " Short Before Visit")
                        CanteenParshadStatus = "Not OK"
            except:
                if (SSCount < int(SSCountCutOffGents)):
                    CanteenParshadRemark.append(str(int(SSCountCutOffGents) - SSCount) + " Short Before Visit")
                    CanteenParshadStatus = "Not OK"

        if (row.SewadarNewID.find("L") > -1):
            try:
                if ExceptionMail[row.SewadarNewID,'ExceptionField'] == "SS Count":
                    pass
                else:
                    if (SSCount < int(SSCountCutOffLadies)):
                        CanteenParshadRemark.append(str(int(SSCountCutOffLadies) - SSCount) + " Short Before Visit")
                        CanteenParshadStatus = "Not OK"
            except:
                if (SSCount < int(SSCountCutOffLadies)):
                    CanteenParshadRemark.append(str(int(SSCountCutOffLadies) - SSCount) + " Short Before Visit")
                    CanteenParshadStatus = "Not OK"

        try:
            if ExceptionMail[row.SewadarNewID,'ExceptionField'] == "Initiation":
                pass
            else:
                if InitiationStatus.find("N") > -1:
                    CanteenParshadRemark = ["Non Initiated"]
                    CanteenParshadStatus = "Not OK"
        except:
            if InitiationStatus.find("N") > -1:
                CanteenParshadRemark = ["Non Initiated"]
                CanteenParshadStatus = "Not OK"

        #if (row.GR_NO.find("SS0") > -1):
        #    SewadarId = (row.GR_NO)[2:]
        #    if (int(SewadarId) > int(LastOSS)):
        #        CanteenParshadRemark.append("Slip Greater than SS" + str(LastOSS))
        #        CanteenParshadStatus = "Not OK"

        #try:
        #    if SSCountDict[row.SewadarNewID,'status'] == 'Slip':
        #        if (row.SewadarNewID.find("BH0011GA") > -1):
        #            SewadarId = (row.SewadarNewID)[8:]
        #            if int(SewadarId) > int(GALastGentsNewGRNO):
        #                CanteenParshadRemark.append("Parshad Slip Greater than BH0011GA" + str(GALastGentsNewGRNO))
        #                CanteenParshadStatus = "Not OK"

        #        if (row.SewadarNewID.find("BH0011LA") > -1):
        #            SewadarId = (row.SewadarNewID)[8:]
        #            if int(SewadarId) > int(LALastLadiesNewGRNO):
        #                CanteenParshadRemark.append("Parshad Slip Greater than BH0011LA" + str(LALastLadiesNewGRNO))
        #                CanteenParshadStatus = "Not OK"

        #        if (row.SewadarNewID.find("BH0011GB") > -1):
        #            SewadarId = (row.SewadarNewID)[8:]
        #            if int(SewadarId) > int(GBLastGentsNewGRNO):
        #                CanteenParshadRemark.append("Parshad Slip Greater than BH0011GB" + str(GBLastGentsNewGRNO))
        #                CanteenParshadStatus = "Not OK"

        #        if (row.SewadarNewID.find("BH0011LB") > -1):
        #            SewadarId = (row.SewadarNewID)[8:]
        #            if int(SewadarId) > int(LBLastLadiesNewGRNO):
        #                CanteenParshadRemark.append("Parshad Slip Greater than BH0011LB" + str(LBLastLadiesNewGRNO))
        #                CanteenParshadStatus = "Not OK"

        #except:
        #    CanteenParshadRemark.append("SS Missing")
        #    CanteenParshadStatus = "SS Missing" 



        #This visit criteria for Parshad Packet is that the Sewadar should be present
        if (CVCount >= int(CVCutOff)) and CanteenParshadStatus == "Not OK":
            #CanteenParshadRemark = "PARSHAD PACKET"
            CanteenParshadStatus = "PACKET OK"

        try:
            if ExceptionMail[row.SewadarNewID,'ExceptionField'] == 'ALL':
                CanteenParshadStatus = ExceptionMail[row.SewadarNewID]
                CanteenParshadRemark = [ExceptionMail[row.SewadarNewID]]
        except:
            pass


        dWorkingParshad.cell(get_column_letter(6 ) + str(row_num)).value = InitiationStatus
        dWorkingParshad.cell(get_column_letter(7 ) + str(row_num)).value = SSTentativeParshadStatus
        dWorkingParshad.cell(get_column_letter(8 ) + str(row_num)).value = CanteenParshadStatus
        if len(CanteenParshadRemark) == 0:
            dWorkingParshad.cell(get_column_letter(9 ) + str(row_num)).value = "OK"
        else:
            dWorkingParshad.cell(get_column_letter(9 ) + str(row_num)).value = "\n".join(CanteenParshadRemark)
        dWorkingParshad.cell(get_column_letter(10) + str(row_num)).value = SSCountOld
        dWorkingParshad.cell(get_column_letter(11) + str(row_num)).value = SSCount
        dWorkingParshad.cell(get_column_letter(12) + str(row_num)).value = CVCountOld
        dWorkingParshad.cell(get_column_letter(13) + str(row_num)).value = CVCount
        dWorkingParshad.cell(get_column_letter(14) + str(row_num)).value = MachineBeforeVisitAdditional
        dWorkingParshad.cell(get_column_letter(15) + str(row_num)).value = SSWWCount
        dWorkingParshad.cell(get_column_letter(16) + str(row_num)).value = CVD0
        dWorkingParshad.cell(get_column_letter(17) + str(row_num)).value = CVD1
        dWorkingParshad.cell(get_column_letter(18) + str(row_num)).value = CVD2
        dWorkingParshad.cell(get_column_letter(19) + str(row_num)).value = CVD3
        dWorkingParshad.cell(get_column_letter(20) + str(row_num)).value = CVD4
        dWorkingParshad.cell(get_column_letter(21) + str(row_num)).value = MachinePreviousVisitAddition
        dWorkingParshad.cell(get_column_letter(22) + str(row_num)).value = PreviousVisitParshadStatus


        row_num = row_num + 1

    print "Almost Done!!"
    logf.write("Almost done!\n")

    del ParshadList
    dworkbook.save(dpath)
    mail.send('softwareattendance@gmail.com',
        'Tentative Parshad List',
        'Tentative Parshad List',
        attachments = mail.Attachment('/home/rootcat/new_web2py/web2py/applications/AttendanceSoftware/private/TentativeParshadList.xlsx', content_id='text'))

    logf.write("Mail sent!\n")
    logf.close()

    return dict(message=message)

def DevParshadListScheduled(DateSelectedStart,DateSelectedEnd,LALastLadiesNewGRNO,LBLastLadiesNewGRNO,GALastGentsNewGRNO,GBLastGentsNewGRNO,LastOSS,SSCountCutOffGents,SSCountCutOffLadies,CVCutOff,VisitCountCutOff,WWCutOff,DAY_END_TIME,DumpMachineAttendance,DumpSSAttendance,CANTEENWISE_REPORT):
    db.commit()
    import os
    pathlog = os.path.join(request.folder,'private','log_DevParshadListScheduled')
    logf = open(pathlog,'w')
    from openpyxl import Workbook
    from openpyxl.cell import get_column_letter
    from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
    response.headers['Connection'] =  'keep-alive'
    dworkbook = Workbook()
    dworkbookAttendance = Workbook()
    dpath = os.path.join(request.folder,'private','JathaTentativeParshadList.xlsx')
    dpathAttendance = os.path.join(request.folder,'private','DumpAttendance.xlsx')

    
    from gluon.sqlhtml import form_factory
    import datetime
    import time



    db.commit()
    SSCount = db(db.SSAttendanceCount).select()
    SSCountDict = {}
    for Sewadar in SSCount:
        SSCountDict[Sewadar.NewID,'Gender'] = Sewadar.gender
        SSCountDict[Sewadar.NewID,'TotalVisit'] = Sewadar.TotalVisit
        SSCountDict[Sewadar.NewID,'TotalCount'] = Sewadar.Total
        SSCountDict[Sewadar.NewID,'NAME'] = Sewadar.Name
        SSCountDict[Sewadar.NewID,'OldSewadarid'] = Sewadar.OldSewadarid
        SSCountDict[Sewadar.NewID,'status'] = Sewadar.status
    #Next setup: Change current visit dates(move to old visit dates)
    #Change total visits
    



    SSCURRENT_VISIT = {}
    SSCURRENT_VISIT['COUNT'] = 5
    SSCURRENT_VISIT['D0'] = datetime.datetime.strptime('9-March-2016 00:00:00','%d-%B-%Y %H:%M:%S')
    SSCURRENT_VISIT['D1'] = datetime.datetime.strptime('9-March-2016 12:00:00','%d-%B-%Y %H:%M:%S')
    SSCURRENT_VISIT['D2'] = datetime.datetime.strptime('10-March-2016 13:00:01','%d-%B-%Y %H:%M:%S')
    SSCURRENT_VISIT['D3'] = datetime.datetime.strptime('11-March-2016 09:00:01','%d-%B-%Y %H:%M:%S')
    SSCURRENT_VISIT['D4'] = datetime.datetime.strptime('12-March-2016 09:00:01','%d-%B-%Y %H:%M:%S')
    SSCURRENT_VISIT['D5'] = datetime.datetime.strptime('13-March-2016 09:00:01','%d-%B-%Y %H:%M:%S')

    VISIT_DATES = {}
    TOTAL_VISIT = 13
    VISIT_DATES['V0','COUNT'] = 5
    VISIT_DATES['V0','D0'] = datetime.datetime.strptime('05-October-2011 00:00:00','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V0','D1'] = datetime.datetime.strptime('05-October-2011 14:30:01','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V0','D2'] = datetime.datetime.strptime('06-October-2011 12:00:00','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V0','D3'] = datetime.datetime.strptime('07-October-2011 12:00:00','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V0','D4'] = datetime.datetime.strptime('08-October-2011 12:00:00','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V0','D5'] = datetime.datetime.strptime('09-October-2011 12:00:00','%d-%B-%Y %H:%M:%S')

    VISIT_DATES['V1','COUNT'] = 5
    VISIT_DATES['V1','D0'] = datetime.datetime.strptime('23-November-2011 00:00:00','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V1','D1'] = datetime.datetime.strptime('23-November-2011 14:30:01','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V1','D2'] = datetime.datetime.strptime('24-November-2011 12:00:00','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V1','D3'] = datetime.datetime.strptime('25-November-2011 12:00:00','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V1','D4'] = datetime.datetime.strptime('26-November-2011 12:00:00','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V1','D5'] = datetime.datetime.strptime('27-November-2011 12:00:00','%d-%B-%Y %H:%M:%S')

    VISIT_DATES['V2','COUNT'] = 5
    VISIT_DATES['V2','D0'] = datetime.datetime.strptime('07-March-2012 00:00:00','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V2','D1'] = datetime.datetime.strptime('07-March-2012 14:30:01','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V2','D2'] = datetime.datetime.strptime('08-March-2012 12:00:00','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V2','D3'] = datetime.datetime.strptime('09-March-2012 12:00:00','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V2','D4'] = datetime.datetime.strptime('10-March-2012 12:00:00','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V2','D5'] = datetime.datetime.strptime('11-March-2012 12:00:00','%d-%B-%Y %H:%M:%S')

    VISIT_DATES['V3','COUNT'] = 5
    VISIT_DATES['V3','D0'] = datetime.datetime.strptime('10-October-2012 00:00:00','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V3','D1'] = datetime.datetime.strptime('10-October-2012 14:30:01','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V3','D2'] = datetime.datetime.strptime('11-October-2012 12:00:00','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V3','D3'] = datetime.datetime.strptime('12-October-2012 12:00:00','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V3','D4'] = datetime.datetime.strptime('13-October-2012 12:00:00','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V3','D5'] = datetime.datetime.strptime('14-October-2012 12:00:00','%d-%B-%Y %H:%M:%S')

    VISIT_DATES['V4','COUNT'] = 5
    VISIT_DATES['V4','D0'] = datetime.datetime.strptime('28-November-2012 00:00:00','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V4','D1'] = datetime.datetime.strptime('28-November-2012 14:30:01','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V4','D2'] = datetime.datetime.strptime('29-November-2012 12:00:00','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V4','D3'] = datetime.datetime.strptime('30-November-2012 12:00:00','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V4','D4'] = datetime.datetime.strptime('01-December-2012 12:00:00','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V4','D5'] = datetime.datetime.strptime('02-December-2012 12:00:00','%d-%B-%Y %H:%M:%S')

    VISIT_DATES['V5','COUNT'] = 5
    VISIT_DATES['V5','D0'] = datetime.datetime.strptime('13-March-2013 00:00:00','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V5','D1'] = datetime.datetime.strptime('13-March-2013 14:30:01','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V5','D2'] = datetime.datetime.strptime('14-March-2013 12:00:00','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V5','D3'] = datetime.datetime.strptime('15-March-2013 12:00:00','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V5','D4'] = datetime.datetime.strptime('16-March-2013 12:00:00','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V5','D5'] = datetime.datetime.strptime('17-March-2013 12:00:00','%d-%B-%Y %H:%M:%S')

    VISIT_DATES['V6','COUNT'] = 5
    VISIT_DATES['V6','D0'] = datetime.datetime.strptime('09-October-2013 00:00:00','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V6','D1'] = datetime.datetime.strptime('09-October-2013 14:30:01','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V6','D2'] = datetime.datetime.strptime('10-October-2013 12:00:00','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V6','D3'] = datetime.datetime.strptime('11-October-2013 12:00:00','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V6','D4'] = datetime.datetime.strptime('12-October-2013 12:00:00','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V6','D5'] = datetime.datetime.strptime('13-October-2013 12:00:00','%d-%B-%Y %H:%M:%S')

    VISIT_DATES['V7','COUNT'] = 3
    VISIT_DATES['V7','D0'] = datetime.datetime.strptime('10-January-2014 00:00:00','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V7','D1'] = datetime.datetime.strptime('10-January-2014 14:30:01','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V7','D2'] = datetime.datetime.strptime('11-January-2014 12:00:00','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V7','D3'] = datetime.datetime.strptime('12-January-2014 12:00:00','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V7','D4'] = datetime.datetime.strptime('13-January-2014 12:00:00','%d-%B-%Y %H:%M:%S')

    VISIT_DATES['V8','COUNT'] = 5
    VISIT_DATES['V8','D0'] = datetime.datetime.strptime('01-October-2014 00:00:00','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V8','D1'] = datetime.datetime.strptime('01-October-2014 14:30:01','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V8','D2'] = datetime.datetime.strptime('02-October-2014 12:00:00','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V8','D3'] = datetime.datetime.strptime('03-October-2014 12:00:00','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V8','D4'] = datetime.datetime.strptime('04-October-2014 12:00:00','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V8','D5'] = datetime.datetime.strptime('05-October-2014 12:00:00','%d-%B-%Y %H:%M:%S')

    VISIT_DATES['V9','COUNT'] = 5
    VISIT_DATES['V9','D0'] = datetime.datetime.strptime('26-November-2014 00:00:00','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V9','D1'] = datetime.datetime.strptime('26-November-2014 14:30:01','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V9','D2'] = datetime.datetime.strptime('27-November-2014 12:00:00','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V9','D3'] = datetime.datetime.strptime('28-November-2014 12:00:00','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V9','D4'] = datetime.datetime.strptime('29-November-2014 12:00:00','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V9','D5'] = datetime.datetime.strptime('30-November-2014 12:00:00','%d-%B-%Y %H:%M:%S')

    VISIT_DATES['V10','COUNT'] = 5
    VISIT_DATES['V10','D0'] = datetime.datetime.strptime('11-March-2015 00:00:00','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V10','D1'] = datetime.datetime.strptime('11-March-2015 14:30:01','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V10','D2'] = datetime.datetime.strptime('12-March-2015 12:00:00','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V10','D3'] = datetime.datetime.strptime('13-March-2015 12:00:00','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V10','D4'] = datetime.datetime.strptime('14-March-2015 12:00:00','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V10','D5'] = datetime.datetime.strptime('15-March-2015 12:00:00','%d-%B-%Y %H:%M:%S')
    
    VISIT_DATES['V11','COUNT'] = 5
    VISIT_DATES['V11','D0'] = datetime.datetime.strptime('07-October-2015 00:00:00','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V11','D1'] = datetime.datetime.strptime('07-October-2015 14:30:01','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V11','D2'] = datetime.datetime.strptime('08-October-2015 12:00:00','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V11','D3'] = datetime.datetime.strptime('09-October-2015 12:00:00','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V11','D4'] = datetime.datetime.strptime('10-October-2015 12:00:00','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V11','D5'] = datetime.datetime.strptime('11-October-2015 12:00:00','%d-%B-%Y %H:%M:%S')


    VISIT_DATES['V12','COUNT'] = 5
    VISIT_DATES['V12','D0'] = datetime.datetime.strptime('25-November-2015 00:00:00','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V12','D1'] = datetime.datetime.strptime('25-November-2015 14:30:01','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V12','D2'] = datetime.datetime.strptime('26-November-2015 12:00:00','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V12','D3'] = datetime.datetime.strptime('27-November-2015 12:00:00','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V12','D4'] = datetime.datetime.strptime('28-November-2015 12:00:00','%d-%B-%Y %H:%M:%S')
    VISIT_DATES['V12','D5'] = datetime.datetime.strptime('29-November-2015 12:00:00','%d-%B-%Y %H:%M:%S')

    message = "ALL OK "
    ParshadList = {}


    try:
        os.remove(dpath)
    except:
        pass

    try:
        os.remove(dpathAttendance)
    except:
        pass

    SSDate = db((db.SSAttendanceDate.DutyDate >= (datetime.datetime.strptime(DateSelectedStart,'%Y-%m-%d').replace(hour=0, minute=0, second=0, microsecond=0))) & (db.SSAttendanceDate.DutyDate <= (datetime.datetime.strptime(DateSelectedEnd,'%Y-%m-%d').replace(hour=23, minute=59, second=59, microsecond=999)))).select('SewadarNewID','DutyDate','Duty_Type')
    MachineDate = db((db.MachineAttendance.DATETIME >= (datetime.datetime.strptime(DateSelectedStart,'%Y-%m-%d') - datetime.timedelta(hours=((24 - int(DAY_END_TIME)) % 24)))) & (db.MachineAttendance.DATETIME <= (datetime.datetime.strptime(DateSelectedEnd,'%Y-%m-%d') + datetime.timedelta(hours=int(DAY_END_TIME))))).select('GRNO','NewGRNO','DATETIME','TYPE')
    print "Machine Attendance between date"
    print "length of MachineDate =" + str(len(MachineDate))
    dExceptionMail = db(db.ParshadMailException.id>0).select()

    ExceptionMail = {}
    for row in dExceptionMail:
        ExceptionMail[row.NewGRNO] = row.Status

    ParshadList['SEWADARS'] = []
    SSAttendanceDictionary = {}
    print "Collecting SSDate"
    logf.write("Collecting SSDate")
    for SSEntry in SSDate:
        ParshadList['SEWADARS'].append(SSEntry.SewadarNewID)
        if SSEntry.Duty_Type == 'W':
            now = datetime.datetime.now()
            SSAttendanceDictionary[SSEntry.SewadarNewID,Roundoffdate(SSEntry.DutyDate,23),'W'] = SSEntry.DutyDate
            try:
                ParshadList[SSEntry.SewadarNewID,'SSCount'] = ParshadList[SSEntry.SewadarNewID,'SSCount'] + 2
            except:
                ParshadList[SSEntry.SewadarNewID,'SSCount'] = 2
            try:
                if SSEntry.DutyDate.year == now.year:
                    ParshadList[SSEntry.SewadarNewID,'SSWWCount'] = ParshadList[SSEntry.SewadarNewID,'SSWWCount'] + 1
            except:
                if SSEntry.DutyDate.year == now.year:
                    ParshadList[SSEntry.SewadarNewID,'SSWWCount'] = 1
        else:
            SSAttendanceDictionary[SSEntry.SewadarNewID,Roundoffdate(SSEntry.DutyDate,23),'D'] = SSEntry.DutyDate
            try:
                ParshadList[SSEntry.SewadarNewID,'SSCount'] = ParshadList[SSEntry.SewadarNewID,'SSCount'] + 1
            except:
                ParshadList[SSEntry.SewadarNewID,'SSCount'] = 1

        ParshadList[SSEntry.SewadarNewID,'SSCountOld'] = ParshadList[SSEntry.SewadarNewID,'SSCount']

    MachineAttendanceAdditional = {}


    try:
        db.tempMachineAttendanceAdditional.drop()
    except:
        pass

    #Now define the table
    db.define_table('tempMachineAttendanceAdditional',
            Field('GRNO','string'),
            Field('NewGRNO','string'),
            Field('Duty_Type','string'),
            Field('DutyDate','datetime'),
            Field('DutyDateList','list:string'),
            migrate=True,
            redefine=True,
            format='%(NewGRNO)s')

    try:
        db(db.tempMachineAttendanceAdditional.id > 0).delete()
    except:
        pass

    print "Collecting Machine Date"
    logf.write("Collecting Machine date")

    for MEntry in MachineDate:
        #check if visit dates
        if ((MEntry.DATETIME < VISIT_DATES['V12','D5']) and (MEntry.DATETIME > VISIT_DATES['V12','D0'])) or ((MEntry.DATETIME < VISIT_DATES['V11','D5']) and (MEntry.DATETIME > VISIT_DATES['V11','D0'])) or ((MEntry.DATETIME < VISIT_DATES['V10','D5']) and (MEntry.DATETIME > VISIT_DATES['V10','D0'])) or ((MEntry.DATETIME < VISIT_DATES['V9','D5']) and (MEntry.DATETIME > VISIT_DATES['V9','D0'])) or  ((MEntry.DATETIME < VISIT_DATES['V8','D5']) and (MEntry.DATETIME > VISIT_DATES['V8','D0'])):
            pass
        else:
            #Select earliest entry but give preference to WW attendance
            if MEntry.TYPE == 'WMANUAL':
                try:
                    a = SSAttendanceDictionary[MEntry.NewGRNO,Roundoffdate(MEntry.DATETIME,DAY_END_TIME),'W']
                except:
                    try:
                        MachineAttendanceAdditional[MEntry.NewGRNO,MEntry.GRNO,Roundoffdate(MEntry.DATETIME,DAY_END_TIME),'W'].append(MEntry.DATETIME) 
                    except:
                        MachineAttendanceAdditional[MEntry.NewGRNO,MEntry.GRNO,Roundoffdate(MEntry.DATETIME,DAY_END_TIME),'W'] = [MEntry.DATETIME]


            else:
                try:
                    a = SSAttendanceDictionary[MEntry.NewGRNO,Roundoffdate(MEntry.DATETIME,DAY_END_TIME),'D']
                except:
                    try:
                        a = SSAttendanceDictionary[MEntry.NewGRNO,Roundoffdate(MEntry.DATETIME,DAY_END_TIME),'W']
                    except:
                        try:
                            MachineAttendanceAdditional[MEntry.NewGRNO,MEntry.GRNO,Roundoffdate(MEntry.DATETIME,DAY_END_TIME),'D'].append(MEntry.DATETIME) 
                        except:
                            MachineAttendanceAdditional[MEntry.NewGRNO,MEntry.GRNO,Roundoffdate(MEntry.DATETIME,DAY_END_TIME),'D'] = [MEntry.DATETIME]


    del SSAttendanceDictionary

    print "Preparing SS Count"
    logf.write("Preparing SS Count")

    for key, value in MachineAttendanceAdditional.iteritems():
        NewGRNO, GRNO, DutyDate , DutyType= key
        ParshadList['SEWADARS'].append(NewGRNO)
        if DutyType == 'W':
            try:
                ParshadList[NewGRNO,'SSCount'] = ParshadList[NewGRNO,'SSCount'] + 2
            except:
                ParshadList[NewGRNO,'SSCount'] = 2
        else:
            try:
                a = MachineAttendanceAdditional[NewGRNO,GRNO,DutyDate,'W']
            except:
                try:
                    ParshadList[NewGRNO,'SSCount'] = ParshadList[NewGRNO,'SSCount'] + 1
                except:
                    ParshadList[NewGRNO,'SSCount'] = 1


    for key, value in MachineAttendanceAdditional.iteritems():
        MachineAttendanceAdditionalDictionary = {}
        NewGRNO, GRNO, DutyDate , DutyType= key
        ParshadList['SEWADARS'].append(NewGRNO)
        MachineAttendanceAdditionalDictionary['NewGRNO'] = NewGRNO 
        MachineAttendanceAdditionalDictionary['GRNO'] = GRNO
        MachineAttendanceAdditionalDictionary['Duty_Type'] = DutyType
        MachineAttendanceAdditionalDictionary['DutyDate'] = DutyDate
        MachineAttendanceAdditionalDictionary['DutyDateList'] = value
        try:
            ParshadList[NewGRNO,'Before Visit Machine Additional'].append(DutyDate)
        except:
            ParshadList[NewGRNO,'Before Visit Machine Additional'] = [DutyDate]
        db.tempMachineAttendanceAdditional.insert(**MachineAttendanceAdditionalDictionary)


    del MachineAttendanceAdditional

    print "Creating Attendance Sheets"
    logf.write("Creating Attendance Sheets")

    if DumpSSAttendance == 'YES': 
        dSSdate = dworkbookAttendance.create_sheet(0)
        dSSdate.title = "SSDate"
        dSSdate.append(['SewadarNewID','DutyDate','Duty_Type'])
        for row in SSDate:
            dSSdate.append([row.SewadarNewID , row.DutyDate , row.Duty_Type])

    if DumpMachineAttendance == 'YES': 
        dMachineAttendance = dworkbookAttendance.create_sheet(0)
        dMachineAttendance.title = "MachineAttendance"
        dMachineAttendance.append(['GRNO','NewGRNO','DATETIME','TYPE'])
        for row in MachineDate:
            dMachineAttendance.append([row.GRNO,row.NewGRNO,row.DATETIME,row.TYPE])

    MachineDifference = db(db.tempMachineAttendanceAdditional.id > 0).select()

    dMachineDifference = dworkbook.create_sheet(0)
    dMachineDifference.title = "MachineDifference"
    dMachineDifference.append(['NewGRNO','GRNO','Duty_Type','DutyDate','DutyDateList','TimeDifference'])
    for row in MachineDifference:
        DutyDateList = ", ".join(map(str, row.DutyDateList))
        n = len(row.DutyDateList)
        TimeDifference = max(map(datetime.datetime.strptime,row.DutyDateList,['%Y-%m-%d %H:%M:%S']*n)) - min(map(datetime.datetime.strptime,row.DutyDateList,['%Y-%m-%d %H:%M:%S']*n))
        dMachineDifference.append([row.NewGRNO,row.GRNO,row.Duty_Type,row.DutyDate,DutyDateList,str(TimeDifference)])


    dworkbookAttendance.save(dpathAttendance)
    

    throttle = 1

    for visit in xrange(0,TOTAL_VISIT):
        VCOUNT = 0
        VCOUNT = VISIT_DATES['V'+str(visit),'COUNT']

        vc = 'V'+str(visit)
        
        for day in xrange(0,VCOUNT):
            dc = 'D'+str(day)
            dc1 = 'D'+str(day+1)
            print "Analyzing " + vc + dc

            SSDate = db((db.SSAttendanceDate.DutyDate >= datetime.datetime(VISIT_DATES[vc,dc].year,VISIT_DATES[vc,dc].month,VISIT_DATES[vc,dc].day,0,0,0)) & (db.SSAttendanceDate.DutyDate <= datetime.datetime((VISIT_DATES[vc,dc]+datetime.timedelta(hours=24)).year,(VISIT_DATES[vc,dc]+datetime.timedelta(hours=24)).month,(VISIT_DATES[vc,dc]+datetime.timedelta(hours=24)).day,0,0,0))).select('SewadarNewID')
            MachineDate = db((db.MachineAttendance.DATETIME >= VISIT_DATES[vc,dc]) & (db.MachineAttendance.DATETIME <= VISIT_DATES[vc,dc1])).select('NewGRNO')
            print "length of SSDate =" + str(len(SSDate))
            print "length of MachineDate =" + str(len(MachineDate))

            for row in SSDate:
                ParshadList['SEWADARS'].append(row.SewadarNewID)
                try:
                    a = ParshadList[row.SewadarNewID,vc]
                except:
                    ParshadList[row.SewadarNewID,vc] = VCOUNT - day

            for row in MachineDate:
                ParshadList['SEWADARS'].append(row.NewGRNO)
                try:
                    if ParshadList[row.NewGRNO,vc] < VCOUNT - day :
                        ParshadList[row.NewGRNO,vc] = VCOUNT - day
                        try:
                            ParshadList[row.NewGRNO,'Additional Machine Visit Days'].append(vc + dc + ' onwards')
                        except:
                            ParshadList[row.NewGRNO,'Additional Machine Visit Days'] = [vc + dc + ' onwards']
                except:
                    ParshadList[row.NewGRNO,vc] = (VCOUNT - day)
                    try:
                        ParshadList[row.NewGRNO,'Additional Machine Visit Days'].append(vc + dc + ' onwards')
                    except:
                        ParshadList[row.NewGRNO,'Additional Machine Visit Days'] = [vc + dc + ' onwards']
            del SSDate
            del MachineDate


    print "reading tentative parshad list"
    logf.write("reading tentative parshad list")
    SSTentativeParshadList = db(db.SSTentativeParshadList.id > 0).select('NewGRNO','Status')

    for row in SSTentativeParshadList:
        ParshadList['SEWADARS'].append(row.NewGRNO)
        ParshadList[row.NewGRNO,'SS Tentative Parshad Status'] = row.Status


    print "reading initiated list"
    InitiatedList = db(db.InitiatedList.id > 0).select('NewGRNO','Status')

    for row in InitiatedList:
        ParshadList['SEWADARS'].append(row.NewGRNO)
        ParshadList[row.NewGRNO,'Initiation Status'] = row.Status


    print "reading previous parshad list"
    PreviousParshadList = db(db.PreviousParshadList.id > 0).select('NewGRNO','Status')

    for row in PreviousParshadList:
        ParshadList['SEWADARS'].append(row.NewGRNO)
        ParshadList[row.NewGRNO,'Previous Visit Parshad Status'] = row.Status

    print "Analyzing Current Visit"
    logf.write("Analyzing Current visit")
    
    #time.sleep (1);
    for day in xrange(0,SSCURRENT_VISIT['COUNT']):
        dc = 'D'+str(day)
        dc1 = 'D'+str(day+1)
        print "Analyzing Current Visit day " + str(day)

        SSDate = db((db.SSAttendanceDate.DutyDate >= SSCURRENT_VISIT['D'+str(day)]) & (db.SSAttendanceDate.DutyDate <= SSCURRENT_VISIT['D'+str(day+1)])).select('SewadarNewID')
        print "Collecting SSDate"
        time.sleep (1);
        MachineDate = db((db.MachineAttendance.DATETIME >= SSCURRENT_VISIT['D'+str(day)]) & (db.MachineAttendance.DATETIME <= SSCURRENT_VISIT['D'+str(day+1)])).select('NewGRNO')
        print "Collected Machine and SSDate"

        for row in SSDate:
            ParshadList['SEWADARS'].append(row.SewadarNewID)
            try:
                a = ParshadList[row.SewadarNewID,'CV']
            except:
                ParshadList[row.SewadarNewID,'CV'] = SSCURRENT_VISIT['COUNT'] - day
                ParshadList[row.SewadarNewID,'CVOld'] = SSCURRENT_VISIT['COUNT'] - day

            ParshadList[row.SewadarNewID,'CV '+ dc] = 'P'


        print "Analyzed SSDate for current visit"

        for row in MachineDate:
            ParshadList['SEWADARS'].append(row.NewGRNO)
            try:
                a = ParshadList[row.NewGRNO,'CV']
            except:
                ParshadList[row.NewGRNO,'CV'] = SSCURRENT_VISIT['COUNT'] - day

            try:
                a = ParshadList[row.NewGRNO,'CV '+ dc]
            except:
                ParshadList[row.NewGRNO,'CV '+ dc] = 'C'



        print "Analyzed Machine Dates for current visit"

    ParshadList['SEWADARS'] = set(ParshadList['SEWADARS'])

    print "Creating Super Set Parshad Status"

    dParshadList = dworkbook.create_sheet(0)
    dParshadList.title = "ParshadList"
    dParshadList.append(['NewGRNO','SSCount','SSCountOld','CVCount','CVCountOld','MachineBeforeVisitAdditional','WW Count','CV D1','CV D2','CV D3','CV D4','CV D5','MachinePreviousVisitAddition','InitiationStatus','SS Tentative Parshad Status','Previous Visit Parshad Status','VISIT_COUNTS','V0','V1','V2','V3','V4','V5','V6','V7','V8','V9','V10','V11','V12'])
    row_num = 1
    for Sewadar in ParshadList['SEWADARS']:
        #print "SuperSet Sewadar " + Sewadar
        try:
            SSCount = ParshadList[Sewadar,'SSCount'] 
        except:
            SSCount = 0

        try:
            SSWWCount = ParshadList[Sewadar,'SSWWCount'] 
        except:
            SSWWCount = 0


        try:
            SSCountOld = ParshadList[Sewadar,'SSCountOld'] 
        except:
            SSCountOld = 0
        try:
            CVCount = ParshadList[Sewadar,'CV'] 
        except:
            CVCount = 0
        try:
            CVCountOld = ParshadList[Sewadar,'CVOld'] 
        except:
            CVCountOld = 0
        try:
            CVD0 = ParshadList[Sewadar,'CV D0']
        except:
            CVD0 = 'A'
        try:
            CVD1 = ParshadList[Sewadar,'CV D1']
        except:
            CVD1 = 'A'
        try:
            CVD2 = ParshadList[Sewadar,'CV D2']
        except:
            CVD2 = 'A'
        try:
            CVD3 = ParshadList[Sewadar,'CV D3']
        except:
            CVD3 = 'A'
        try:
            CVD4 = ParshadList[Sewadar,'CV D4']
        except:
            CVD4 = 'A'
        try:
            MachineBeforeVisitAdditional = ", ".join(map(str, ParshadList[Sewadar,'Before Visit Machine Additional']))
        except:
            MachineBeforeVisitAdditional = ""
        try:
            MachinePreviousVisitAddition = ", ".join(map(str, ParshadList[Sewadar,'Additional Machine Visit Days']))
        except:
            MachinePreviousVisitAddition = ""
        try:
            InitiationStatus = ParshadList[Sewadar,'Initiation Status']
        except:
            InitiationStatus = "Not Found"
        try:
            PreviousVisitParshadStatus = ParshadList[Sewadar,'Previous Visit Parshad Status']
        except:
            PreviousVisitParshadStatus = "Not Found"
        try:
            SSTentativeParshadStatus = ParshadList[Sewadar,'SS Tentative Parshad Status']
        except:
            SSTentativeParshadStatus = "Not Found"

        dParshadList.append([Sewadar,SSCount,SSCountOld,CVCount,CVCountOld,MachineBeforeVisitAdditional,SSWWCount,CVD0,CVD1,CVD2,CVD3,CVD4,MachinePreviousVisitAddition,InitiationStatus,SSTentativeParshadStatus,PreviousVisitParshadStatus])
       
        row_num = row_num + 1
        visit_counts = 0
        for visit in xrange(0,TOTAL_VISIT):
            try:
                dParshadList.cell(get_column_letter(visit+18)+str(row_num)).value = ParshadList[Sewadar,'V'+str(visit)]
                if ParshadList[Sewadar,'V'+str(visit)] >= 3:
                    visit_counts = visit_counts + 1 
            except:
                dParshadList.cell(get_column_letter(visit+18)+str(row_num)).value = 0
                
        dParshadList.cell(get_column_letter(17)+str(row_num)).value = visit_counts


    my_list = db().select(db.MasterSheet.DEV_DTY, distinct=True).as_list()
    myjathalist = []
    for mydict in my_list:
        myjathalist.append(mydict['DEV_DTY'])

    myjathalist.sort()

    mycanteenlist = []
    if CANTEENWISE_REPORT == "YES":
        my_list = db().select(db.MasterSheet.CANTEEN, distinct=True).as_list()
        for mydict in my_list:
            mycanteenlist.append(mydict['CANTEEN'])

        mycanteenlist.sort()
    elif CANTEENWISE_REPORT == "NO":
        mycanteenlist = [" "]
    else:
        mycanteenlist = [" "]
        myjathalist = [" "]



    for Canteen in mycanteenlist:
        logf.write("Creating sheet for " + Canteen + "\n")
        for Jatha in myjathalist:
            logf.write("  Creating sheet for " + Jatha + "\n")
            for Gender in ['M','F']:
                if CANTEENWISE_REPORT == "YES":
                    MasterSheet = db((db.MasterSheet.CANTEEN  == Canteen) & (db.MasterSheet.DEV_DTY  == Jatha) & (db.MasterSheet.Gender  == Gender)).select('SewadarNewID','GR_NO','NAME','CANTEEN','DEV_DTY',orderby=db.MasterSheet.GR_NO)
                elif CANTEENWISE_REPORT == "NO":
                    MasterSheet = db((db.MasterSheet.DEV_DTY  == Jatha) & (db.MasterSheet.Gender  == Gender)).select('SewadarNewID','GR_NO','NAME','CANTEEN','DEV_DTY',orderby=db.MasterSheet.GR_NO)
                    logf.write("Fetched query\n")
                else:
                    MasterSheet = db(db.MasterSheet.Gender == Gender).select('SewadarNewID','GR_NO','NAME','CANTEEN','DEV_DTY','Gender',orderby=db.MasterSheet.GR_NO)
                    logf.write("Fetched query\n")

                if (len(MasterSheet) == 0):
                    continue

                Mytitle = ""

                if CANTEENWISE_REPORT == "FLAT":
                    Mytitle = "ALL"
                else:
                    Mytitle = Gender + '_' + Canteen + '_' + (Jatha.replace(':','_'))[0:10]
                dWorkingParshad = dworkbook.create_sheet(title=Mytitle)
                dWorkingParshad.page_setup.orientation = dWorkingParshad.ORIENTATION_PORTRAIT
                dWorkingParshad.page_setup.paperSize = dWorkingParshad.PAPERSIZE_A4
                dWorkingParshad.page_setup.fitToHeight = 0
                dWorkingParshad.page_setup.fitToWidth = 1
                dWorkingParshad.page_margins.left = 0.6/2.54
                dWorkingParshad.page_margins.right = 0.4/2.54
                dWorkingParshad.page_margins.bottom = 0.6/2.54
                dWorkingParshad.page_margins.top = 0.6/2.54

                print "Creating Master Specific Parshad Status"

                row_num = 4
                for row in MasterSheet:
                    Sewadar = row.SewadarNewID
                    #print "Master Sewadar " + Sewadar

                    dWorkingParshad.cell(get_column_letter(1) + str(row_num)).value = row.SewadarNewID
                    dWorkingParshad.cell(get_column_letter(2) + str(row_num)).value = row.GR_NO
                    dWorkingParshad.cell(get_column_letter(3) + str(row_num)).value = row.NAME
                    dWorkingParshad.cell(get_column_letter(10) + str(row_num)).value = row.Gender
                    dWorkingParshad.cell(get_column_letter(11) + str(row_num)).value = row.CANTEEN
                    dWorkingParshad.cell(get_column_letter(12) + str(row_num)).value = row.DEV_DTY


                    try:
                        SSCount = ParshadList[Sewadar,'SSCount'] 
                    except:
                        SSCount = 0
                    try:
                        SSCountOld = ParshadList[Sewadar,'SSCountOld'] 
                    except:
                        SSCountOld = 0
                    try:
                        SSWWCount = ParshadList[Sewadar,'SSWWCount'] 
                    except:
                        SSWWCount = 0
                    try:
                        CVCount = ParshadList[Sewadar,'CV'] 
                    except:
                        CVCount = 0
                    try:
                        CVCountOld = ParshadList[Sewadar,'CVOld'] 
                    except:
                        CVCountOld = 0
                    try:
                        CVD0 = ParshadList[Sewadar,'CV D0']
                    except:
                        CVD0 = 'A'
                    try:
                        CVD1 = ParshadList[Sewadar,'CV D1']
                    except:
                        CVD1 = 'A'
                    try:
                        CVD2 = ParshadList[Sewadar,'CV D2']
                    except:
                        CVD2 = 'A'
                    try:
                        CVD3 = ParshadList[Sewadar,'CV D3']
                    except:
                        CVD3 = 'A'
                    try:
                        CVD4 = ParshadList[Sewadar,'CV D4']
                    except:
                        CVD4 = 'A'
                    try:
                        MachineBeforeVisitAdditional = ", ".join(map(str, ParshadList[Sewadar,'Before Visit Machine Additional']))
                    except:
                        MachineBeforeVisitAdditional = ""
                    try:
                        MachinePreviousVisitAddition = ", ".join(map(str, ParshadList[Sewadar,'Additional Machine Visit Days']))
                    except:
                        MachinePreviousVisitAddition = ""
                    try:
                        InitiationStatus = ParshadList[Sewadar,'Initiation Status']
                    except:
                        InitiationStatus = "Not Found"
                    try:
                        PreviousVisitParshadStatus = ParshadList[Sewadar,'Previous Visit Parshad Status']
                    except:
                        PreviousVisitParshadStatus = "Not Found"
                    try:
                        SSTentativeParshadStatus = ParshadList[Sewadar,'SS Tentative Parshad Status']
                    except:
                        SSTentativeParshadStatus = "Not Found"


                    visit_counts = 0
                    for visit in xrange(0,TOTAL_VISIT):
                        try:
                            if ParshadList[Sewadar,'V'+str(visit)] >= 3:
                                visit_counts = visit_counts + 1 
                        except:
                            pass
                            
                    try:
                        if visit_counts < SSCountDict[Sewadar,'TotalVisit']:
                            visit_counts = SSCountDict[Sewadar,'TotalVisit']
                    except:
                        test = 1

                    dWorkingParshad.cell(get_column_letter(4)+str(row_num)).value = visit_counts

                    CanteenParshadStatus = "OK"
                    CanteenParshadRemark = "OK"


                    #TBD: Deduce based on 'status' whether the Sewadar is permanent or temporary
                    try:
                        if SSCountDict[row.SewadarNewID,'status'] == 'Slip':
                            if (int(visit_counts) < int(VisitCountCutOff)):
                                #CanteenParshadRemark = VisitCountCutOff + " Visit Short"
                                CanteenParshadRemark = str(visit_counts) + " Visits Attended"
                                CanteenParshadStatus = "Not OK"
                    except:
                        CanteenParshadRemark = "Not in Sewa Samiti Count Sheet"
                        CanteenParshadStatus = "Not in Sewa Samiti Count Sheet"

                    if (CVCount < int(CVCutOff)):
                        CanteenParshadRemark = "Current Visit Short"
                        CanteenParshadStatus = "Not OK"

                    if (row.SewadarNewID.find("G") > -1):
                        if (SSWWCount < int(WWCutOff)):
                            CanteenParshadRemark = "WW short by " + str(int(WWCutOff) - SSWWCount)
                            CanteenParshadStatus = "Not OK"
                        if (SSCount < int(SSCountCutOffGents)):
                            CanteenParshadRemark = "Short Before Visit"
                            CanteenParshadStatus = "Not OK"

                    if (row.SewadarNewID.find("L") > -1):
                        if (SSCount < int(SSCountCutOffLadies)):
                            CanteenParshadRemark = "Short Before Visit"
                            CanteenParshadStatus = "Not OK"

                    if InitiationStatus.find("N") > -1:
                        CanteenParshadRemark  = "NI"
                        CanteenParshadStatus = "Not OK"

                    if (row.GR_NO.find("SS0") > -1):
                        SewadarId = (row.GR_NO)[2:]
                        if (int(SewadarId) > int(LastOSS)):
                            CanteenParshadRemark = "Slip Greater than SS" + str(LastOSS)
                            CanteenParshadStatus = "Not OK"

                    try:
                        if SSCountDict[row.SewadarNewID,'status'] == 'Slip':
                            if (row.SewadarNewID.find("BH0011GA") > -1):
                                SewadarId = (row.SewadarNewID)[8:]
                                if int(SewadarId) > int(GALastGentsNewGRNO):
                                    CanteenParshadRemark = "Parshad Slip Greater than BH0011GA" + str(GALastGentsNewGRNO)
                                    CanteenParshadStatus = "Not OK"

                            if (row.SewadarNewID.find("BH0011LA") > -1):
                                SewadarId = (row.SewadarNewID)[8:]
                                if int(SewadarId) > int(LALastLadiesNewGRNO):
                                    CanteenParshadRemark = "Parshad Slip Greater than BH0011LA" + str(LALastLadiesNewGRNO)
                                    CanteenParshadStatus = "Not OK"

                            if (row.SewadarNewID.find("BH0011GB") > -1):
                                SewadarId = (row.SewadarNewID)[8:]
                                if int(SewadarId) > int(GBLastGentsNewGRNO):
                                    CanteenParshadRemark = "Parshad Slip Greater than BH0011GB" + str(GBLastGentsNewGRNO)
                                    CanteenParshadStatus = "Not OK"

                            if (row.SewadarNewID.find("BH0011LB") > -1):
                                SewadarId = (row.SewadarNewID)[8:]
                                if int(SewadarId) > int(LBLastLadiesNewGRNO):
                                    CanteenParshadRemark = "Parshad Slip Greater than BH0011LB" + str(LBLastLadiesNewGRNO)
                                    CanteenParshadStatus = "Not OK"

                    except:
                        CanteenParshadRemark = "Not in SSCOunt sheet"
                        CanteenParshadStatus = "Not in SSCount sheet" 


                    #This visit criteria for Parshad Packet is that the Sewadar should be present
                    if (CVCount >= int(CVCutOff)) and CanteenParshadStatus == "Not OK":
                        #CanteenParshadRemark = "PARSHAD PACKET"
                        CanteenParshadStatus = "PACKET OK"

                    try:
                        CanteenParshadStatus = ExceptionMail[row.SewadarNewID]
                        CanteenParshadRemark = ExceptionMail[row.SewadarNewID]
                    except:
                        pass

                    if (SSTentativeParshadStatus == "Not Found"  and CanteenParshadStatus ==  "OK"):
                        dWorkingParshad.cell(get_column_letter(8) + str(row_num)).value = "Tentative"
                        dWorkingParshad.cell(get_column_letter(9) + str(row_num)).value = "Tentative"
                    else:
                        dWorkingParshad.cell(get_column_letter(8) + str(row_num)).value = CanteenParshadStatus
                        dWorkingParshad.cell(get_column_letter(9) + str(row_num)).value = CanteenParshadRemark





                    dWorkingParshad.cell(get_column_letter(5) + str(row_num)).value = SSCountOld
                    dWorkingParshad.cell(get_column_letter(6) + str(row_num)).value = SSWWCount
                    dWorkingParshad.cell(get_column_letter(7) + str(row_num)).value = CVCount

                    dWorkingParshad[get_column_letter(1) + str(row_num)].font = Font(name='Calibri',size=8,bold=False)
                    dWorkingParshad[get_column_letter(1) + str(row_num)].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
                    dWorkingParshad[get_column_letter(1) + str(row_num)].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
                    dWorkingParshad[get_column_letter(2) + str(row_num)].font = Font(name='Calibri',size=8,bold=False)
                    dWorkingParshad[get_column_letter(2) + str(row_num)].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
                    dWorkingParshad[get_column_letter(2) + str(row_num)].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
                    dWorkingParshad[get_column_letter(3) + str(row_num)].font = Font(name='Calibri',size=8,bold=False)
                    dWorkingParshad[get_column_letter(3) + str(row_num)].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
                    dWorkingParshad[get_column_letter(3) + str(row_num)].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
                    dWorkingParshad[get_column_letter(4) + str(row_num)].font = Font(name='Calibri',size=8,bold=False)
                    dWorkingParshad[get_column_letter(4) + str(row_num)].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
                    dWorkingParshad[get_column_letter(4) + str(row_num)].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
                    dWorkingParshad[get_column_letter(5) + str(row_num)].font = Font(name='Calibri',size=8,bold=False)
                    dWorkingParshad[get_column_letter(5) + str(row_num)].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
                    dWorkingParshad[get_column_letter(5) + str(row_num)].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
                    dWorkingParshad[get_column_letter(6) + str(row_num)].font = Font(name='Calibri',size=8,bold=False)
                    dWorkingParshad[get_column_letter(6) + str(row_num)].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
                    dWorkingParshad[get_column_letter(6) + str(row_num)].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
                    dWorkingParshad[get_column_letter(7) + str(row_num)].font = Font(name='Calibri',size=8,bold=False)
                    dWorkingParshad[get_column_letter(7) + str(row_num)].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
                    dWorkingParshad[get_column_letter(7) + str(row_num)].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
                    dWorkingParshad[get_column_letter(8) + str(row_num)].font = Font(name='Calibri',size=8,bold=False)
                    dWorkingParshad[get_column_letter(8) + str(row_num)].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
                    dWorkingParshad[get_column_letter(8) + str(row_num)].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
                    dWorkingParshad[get_column_letter(9) + str(row_num)].font = Font(name='Calibri',size=8,bold=False)
                    dWorkingParshad[get_column_letter(9) + str(row_num)].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
                    dWorkingParshad[get_column_letter(9) + str(row_num)].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
                    dWorkingParshad[get_column_letter(10) + str(row_num)].font = Font(name='Calibri',size=8,bold=False)
                    dWorkingParshad[get_column_letter(10) + str(row_num)].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
                    dWorkingParshad[get_column_letter(10) + str(row_num)].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
                    dWorkingParshad[get_column_letter(11) + str(row_num)].font = Font(name='Calibri',size=8,bold=False)
                    dWorkingParshad[get_column_letter(11) + str(row_num)].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
                    dWorkingParshad[get_column_letter(11) + str(row_num)].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
                    dWorkingParshad[get_column_letter(12) + str(row_num)].font = Font(name='Calibri',size=8,bold=False)
                    dWorkingParshad[get_column_letter(12) + str(row_num)].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
                    dWorkingParshad[get_column_letter(12) + str(row_num)].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
                    dWorkingParshad.column_dimensions['A'].width = 10.86
                    dWorkingParshad.column_dimensions['B'].width = 6.86
                    dWorkingParshad.column_dimensions['C'].width = 10.57
                    dWorkingParshad.column_dimensions['D'].width = 5.86
                    dWorkingParshad.column_dimensions['E'].width = 7.14
                    dWorkingParshad.column_dimensions['F'].width = 7.14
                    dWorkingParshad.column_dimensions['G'].width = 7.14
                    dWorkingParshad.column_dimensions['H'].width = 8.71
                    dWorkingParshad.column_dimensions['I'].width = 11.57
                    dWorkingParshad.column_dimensions['J'].width = 11.57
                    dWorkingParshad.column_dimensions['K'].width = 11.57
                    dWorkingParshad.column_dimensions['L'].width = 11.57
                    dWorkingParshad.add_print_title(3,3,rows_or_cols='row')

                    row_num = row_num + 1

                dWorkingParshad['A3'] = 'NEW ID'
                dWorkingParshad['B3'] = 'GR NO.'
                dWorkingParshad['C3'] = 'NAME'
                dWorkingParshad['D3'] = 'Total\nVisits'
                dWorkingParshad['E3'] = 'Total\nCount'
                dWorkingParshad['F3'] = 'WW\nCount'
                dWorkingParshad['G3'] = 'Visit\nCount'
                dWorkingParshad['H3'] = 'Parshad\nStatus'
                dWorkingParshad['I3'] = 'PARSHAD\nREMARKS'
                dWorkingParshad['J3'] = 'Sex'
                dWorkingParshad['K3'] = 'CANTEEN'
                dWorkingParshad['L3'] = 'JATHA'
                dWorkingParshad[get_column_letter(1) + str(3)].font = Font(name='Calibri',size=11,bold=True)
                dWorkingParshad[get_column_letter(1) + str(3)].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
                dWorkingParshad[get_column_letter(1) + str(3)].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
                dWorkingParshad[get_column_letter(2) + str(3)].font = Font(name='Calibri',size=11,bold=True)
                dWorkingParshad[get_column_letter(2) + str(3)].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
                dWorkingParshad[get_column_letter(2) + str(3)].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
                dWorkingParshad[get_column_letter(3) + str(3)].font = Font(name='Calibri',size=11,bold=True)
                dWorkingParshad[get_column_letter(3) + str(3)].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
                dWorkingParshad[get_column_letter(3) + str(3)].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
                dWorkingParshad[get_column_letter(4) + str(3)].font = Font(name='Calibri',size=11,bold=True)
                dWorkingParshad[get_column_letter(4) + str(3)].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
                dWorkingParshad[get_column_letter(4) + str(3)].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
                dWorkingParshad[get_column_letter(5) + str(3)].font = Font(name='Calibri',size=11,bold=True)
                dWorkingParshad[get_column_letter(5) + str(3)].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
                dWorkingParshad[get_column_letter(5) + str(3)].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
                dWorkingParshad[get_column_letter(6) + str(3)].font = Font(name='Calibri',size=11,bold=True)
                dWorkingParshad[get_column_letter(6) + str(3)].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
                dWorkingParshad[get_column_letter(6) + str(3)].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
                dWorkingParshad[get_column_letter(7) + str(3)].font = Font(name='Calibri',size=11,bold=True)
                dWorkingParshad[get_column_letter(7) + str(3)].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
                dWorkingParshad[get_column_letter(7) + str(3)].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
                dWorkingParshad[get_column_letter(8) + str(3)].font = Font(name='Calibri',size=11,bold=True)
                dWorkingParshad[get_column_letter(8) + str(3)].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
                dWorkingParshad[get_column_letter(8) + str(3)].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
                dWorkingParshad[get_column_letter(9) + str(3)].font = Font(name='Calibri',size=11,bold=True)
                dWorkingParshad[get_column_letter(9) + str(3)].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
                dWorkingParshad[get_column_letter(9) + str(3)].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
                dWorkingParshad[get_column_letter(10) + str(3)].font = Font(name='Calibri',size=11,bold=True)
                dWorkingParshad[get_column_letter(10) + str(3)].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
                dWorkingParshad[get_column_letter(10) + str(3)].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
                dWorkingParshad[get_column_letter(11) + str(3)].font = Font(name='Calibri',size=11,bold=True)
                dWorkingParshad[get_column_letter(11) + str(3)].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
                dWorkingParshad[get_column_letter(11) + str(3)].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
                dWorkingParshad[get_column_letter(12) + str(3)].font = Font(name='Calibri',size=11,bold=True)
                dWorkingParshad[get_column_letter(12) + str(3)].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
                dWorkingParshad[get_column_letter(12) + str(3)].border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))


                dWorkingParshad['A1'].value = "RADHA SOAMI SATSANG BEAS, DELHI\n BHATI CENTRE \n CAFETERIA DEPARTMENT"
                dWorkingParshad['A1'].font = Font(name='Calibri',size=14,bold=True)
                dWorkingParshad['A1'].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)

                dWorkingParshad.merge_cells('A1:I1')
                dWorkingParshad.row_dimensions[1].height = 60
                dWorkingParshad.column_dimensions['A'].width = 13.14


                if (CANTEENWISE_REPORT == "YES"):
                    dWorkingParshad['A2'].value =  'TENTATIVE PARSHAD STATUS LIST FOR ' + Canteen + ':' + Jatha + '(' + Gender + ')'
                else:
                    dWorkingParshad['A2'].value =  'TENTATIVE PARSHAD STATUS LIST FOR ' + Jatha + '(' + Gender + ')'
                dWorkingParshad['A2'].font = Font(name='Calibri',size=12,bold=True)
                dWorkingParshad['A2'].alignment = Alignment(horizontal='center',vertical='center')

                dWorkingParshad.merge_cells('A2:I2')
                dWorkingParshad.row_dimensions[2].height = 22.5

    logf.write("Almost done!\n")
    print "Almost Done!!"

    del ParshadList
    dworkbook.save(dpath)
    logf.write("Saved XLS!\n")
    mail.send('softwareattendance@gmail.com',
        'JATHAWISE Tentative Parshad List',
        'JATHAWISE Tentative Parshad List',
        attachments = mail.Attachment('/home/rootcat/new_web2py/web2py/applications/AttendanceSoftware/private/JathaTentativeParshadList.xlsx', content_id='text'))

    logf.write("Mail sent!\n")
    logf.close()
    return dict(message=message)

#@auth.requires_login()
def AttendanceRegisterScheduledAll(DateSelectedStart,DateSelectedEnd):
    import os
    pathlog = os.path.join(request.folder,'private','log_AttendanceRegisterAll')
    logf = open(pathlog,'w')

    from gluon.sqlhtml import form_factory
    import datetime
    import time

    from openpyxl import Workbook
    from openpyxl.cell import get_column_letter
    import os
    response.headers['Connection'] =  'keep-alive'
    dworkbook = Workbook()
    dAttendanceRegister = dworkbook.create_sheet(0)
    dpath = os.path.join(request.folder,'private','AttendanceRegister.xlsx')

    Register = "EMPTY"
    DAttendanceRegisterTable = "EMPTY"
    datasource = {}
    columns = []
    orderby = []
    headers = {}
    excel_headers = {}
    DEV_DTY = ""
    TextMessage = "Jatha Wise Attendance Report"
    ReportDate =  0
    GENTS_REQUIRED = 30
    LADIES_REQUIRED = 36
    SS_GENTS_REQUIRED = 30
    SS_LADIES_REQUIRED = 36

    try:
        db.tempAttendanceRegisterTable.drop()
    except:
        pass

    SewadarDetails = {'Sewadars':[]}
#    form=form_factory(SQLField('JATHA','string',requires=IS_IN_DB(db,'MasterSheet.DEV_DTY','%(DEV_DTY)s')),SQLField('DateStart','date',default=datetime.datetime.today()-datetime.timedelta(days=31)),SQLField('DateEnd','date',default=datetime.datetime.today()),formname='DateSelect')
    my_list = db().select(db.MasterSheet.DEV_DTY, distinct=True).as_list()
    myjathalist = []
    for mydict in my_list:
        myjathalist.append(mydict['DEV_DTY'])

    myjathalist.sort()
    form=form_factory(SQLField('JATHA','string',requires=IS_IN_SET(myjathalist)),SQLField('DateStart','date',default=datetime.datetime.today()-datetime.timedelta(days=31)),SQLField('DateEnd','date',default=datetime.datetime.today()),SQLField('Download','string',requires=IS_IN_SET(['YES','NO']),default='NO'),formname='DateSelect')
    try:
        os.remove(dpath)
    except:
        pass

    download = request.vars.Download
    ReportDate = datetime.datetime.today()-datetime.timedelta(days=3001)

    #SSDate = db((db.SSAttendanceDate.DutyDate >= DateSelectedStart) & (db.SSAttendanceDate.DutyDate <= DateSelectedEnd) & (db.SSAttendanceDate.Duty_Type == "W")).select()
    SSDate = db((db.SSAttendanceDate.DutyDate >= DateSelectedStart) & (db.SSAttendanceDate.DutyDate <= DateSelectedEnd)).select()
    for Sewadar in SSDate:
        delta = Sewadar.DutyDate - datetime.datetime.strptime(DateSelectedStart,'%Y-%m-%d')
        SewadarDetails[Sewadar.SewadarNewID,'DAYS',delta.days] = Sewadar.DutyDate
        SewadarDetails[Sewadar.SewadarNewID,'TYPE',delta.days] = Sewadar.Duty_Type
        if ReportDate < Sewadar.DutyDate:
            ReportDate = Sewadar.DutyDate

    SSCount = db(db.SSAttendanceCount).select()
    for Sewadar in SSCount:
        SewadarDetails[Sewadar.NewID,'Gender'] = Sewadar.gender
        SewadarDetails[Sewadar.NewID,'TotalCount'] = Sewadar.Total
        SewadarDetails[Sewadar.NewID,'NAME'] = Sewadar.Name


    MasterTable = db(db.MasterSheet.GR_NO == 'SS015298').select()
    for Sewadar in MasterTable:
             SewadarDetails[Sewadar.GR_NO,'SewadarNewID'] = Sewadar.SewadarNewID
             SewadarDetails[Sewadar.SewadarNewID,'DEV_DTY'] = Sewadar.DEV_DTY

    DEV_DTY = request.vars.JATHA
#        if auth.user.username == 'admin':
#            DEV_DTY = request.vars.JATHA
#        else:
#            try:
#                DEV_DTY = SewadarDetails[auth.user.username,'DEV_DTY']
#            except:
#                #try:
#                    DEV_DTY = SewadarDetails[SewadarDetails[auth.user.username,'SewadarNewID'],'DEV_DTY']
#                #except:
#                #    TextMessage = 'This ID is not allocated to any Jatha'


    MasterTable = db(db.MasterSheet.id > 0).select()
    for Sewadar in MasterTable:
        try:
            #Check if this ID was in SewaSamiti list
            SewadarDetails[Sewadar.SewadarNewID,'NAME']
            SewadarDetails['Sewadars'].append(Sewadar.SewadarNewID)
            SewadarDetails[Sewadar.SewadarNewID,'CANTEEN'] = Sewadar.CANTEEN
            SewadarDetails[Sewadar.SewadarNewID,'DEV_DTY'] = Sewadar.DEV_DTY
            SewadarDetails[Sewadar.SewadarNewID,'OLD_ID'] = Sewadar.GR_NO
        except:
            TextMessage = 'Sewadar ID not in list'


    Fields = [Field('D%02i' %i,'string') for i in range((datetime.datetime.strptime(DateSelectedEnd,'%Y-%m-%d') - datetime.datetime.strptime(DateSelectedStart,'%Y-%m-%d')).days)]


    #Now define the table
    db.define_table('tempAttendanceRegisterTable',
            Field('GR_NO','string'),
            Field('SewadarNewID','string'),
            Field('NAME','string'),
            Field('DEV_DTY','string'),
            Field('CANTEEN','string'),
            Field('TOTAL','integer'),
            Field('STATUS','string'),
            Field('REQD','integer'),
            Field('GENDER','string'),
            Field('WWDATES','string'),
            *Fields,
            migrate=True,
            redefine=True,
            format='%(SewadarNewID)s')

    db(db.tempAttendanceRegisterTable.id > 0).delete()

    excel_headers = {1:'SewadarNewID',2:'GR_NO',3:'NAME',4:'DEV_DTY',5:'CANTEEN',6:'GENDER',7:'STATUS',8:'REQD',9:'WWDATES'}
    dAttendanceRegister.append(excel_headers.values())

    SewadarNumber = 1
    for Sewadar in SewadarDetails['Sewadars']:
        logf.write(Sewadar)
        mydict = {}
        SewadarNumber = SewadarNumber + 1
        mydict['SewadarNewID'] = Sewadar
        mydict['GR_NO'] = SewadarDetails[Sewadar,'OLD_ID']
        mydict['NAME'] = SewadarDetails[Sewadar,'NAME']
        mydict['DEV_DTY'] = SewadarDetails[Sewadar,'DEV_DTY']
        mydict['CANTEEN'] = SewadarDetails[Sewadar,'CANTEEN']
        mydict['TOTAL'] = SewadarDetails[Sewadar,'TotalCount']
        mydict['GENDER'] = SewadarDetails[Sewadar,'Gender']
        mydict['STATUS'] = ''
        mydict['REQD'] = (GENTS_REQUIRED - SewadarDetails[Sewadar,'TotalCount']) if (SewadarDetails[Sewadar,'Gender'] == 'Male') else (LADIES_REQUIRED - SewadarDetails[Sewadar,'TotalCount'])
        if mydict['REQD'] < 1:
            mydict['REQD'] = 0
        elif (((SewadarDetails[Sewadar,'Gender'] == 'Male')) and mydict['REQD'] > SS_GENTS_REQUIRED):
            mydict['REQD'] = SS_GENTS_REQUIRED
        elif (((SewadarDetails[Sewadar,'Gender'] == 'Female')) and mydict['REQD'] > SS_LADIES_REQUIRED):
            mydict['REQD'] = SS_LADIES_REQUIRED
        else:
            pass



        #Write to excel too
        dAttendanceRegister.cell('A' + str(SewadarNumber)).value = mydict['SewadarNewID']
        dAttendanceRegister.cell('B' + str(SewadarNumber)).value = mydict['GR_NO']
        dAttendanceRegister.cell('C' + str(SewadarNumber)).value = mydict['NAME']
        dAttendanceRegister.cell('D' + str(SewadarNumber)).value = mydict['DEV_DTY']
        dAttendanceRegister.cell('E' + str(SewadarNumber)).value = mydict['CANTEEN']
        dAttendanceRegister.cell('F' + str(SewadarNumber)).value = mydict['GENDER']
        dAttendanceRegister.cell('G' + str(SewadarNumber)).value = mydict['STATUS']
        dAttendanceRegister.cell('H' + str(SewadarNumber)).value = mydict['REQD']


        columns = ['tempAttendanceRegisterTable.SewadarNewID','tempAttendanceRegisterTable.GR_NO','tempAttendanceRegisterTable.NAME','tempAttendanceRegisterTable.DEV_DTY','tempAttendanceRegisterTable.CANTEEN','tempAttendanceRegisterTable.REQD']
        orderby=columns

        headers = {'tempAttendanceRegisterTable.SewadarNewID':{'label':T('SewadarNewID'),'class':'','width':12,'truncate':12,'selected': False},
           'tempAttendanceRegisterTable.GR_NO':{'label':T('GR_NO'),'class':'','width':10,'truncate':10,'selected': False},
           'tempAttendanceRegisterTable.NAME':{'label':T('NAME'),'class':'','width':10,'truncate': 10,'selected': False},
           'tempAttendanceRegisterTable.DEV_DTY':{'label':T('DEV_DTY'),'class':'','width':11,'truncate': 11,'selected': False},
           'tempAttendanceRegisterTable.CANTEEN':{'label':T('CANTEEN'),'class':'','width':13,'truncate': 13,'selected': False},
           'tempAttendanceRegisterTable.REQD':{'label':T('REQD'),'class':'','width':4,'truncate': 4,'selected': False}
           }


        for i in range((datetime.datetime.strptime(DateSelectedEnd,'%Y-%m-%d') - datetime.datetime.strptime(DateSelectedStart,'%Y-%m-%d')).days):
            dateindex = datetime.datetime.strptime(DateSelectedStart,'%Y-%m-%d') + datetime.timedelta(i)
            columns.append('tempAttendanceRegisterTable.D%02i' %i)
            headers['tempAttendanceRegisterTable.D%02i' %i] = {'label':datetime.date.strftime(dateindex,'%d\n%m'),'class':'','width':'3','truncate': 3,'selected': False}
            excel_headers[i+10] = (datetime.date.strftime(dateindex,'%d-%m'))
            logf.write(excel_headers[i+10])
            logf.write("\n")

            try:
                mydict['D%02i' %i] = SewadarDetails[Sewadar,'TYPE',i]
                if mydict['D%02i' %i] == 'W':
                    try:
                        mydict["WWDATES"] = mydict["WWDATES"] + '\n' +  excel_headers[i+10]
                    except:
                        mydict["WWDATES"] = excel_headers[i+10]
            except:
                mydict['D%02i' %i] = ''
            dAttendanceRegister.cell(get_column_letter(i+10) + str(SewadarNumber)).value = mydict['D%02i' %i]
            try:
                dAttendanceRegister.cell(get_column_letter(9) + str(SewadarNumber)).value = mydict["WWDATES"]
            except:
                pass

        db.tempAttendanceRegisterTable.insert(**mydict)

    for key in excel_headers.keys():
        dAttendanceRegister.cell(get_column_letter(key) + '1').value = excel_headers[key]

    Register = SQLTABLE(SSDate,headers='fieldname:capitalize')

    datasource = db(db.tempAttendanceRegisterTable).select(orderby=db.tempAttendanceRegisterTable.GR_NO)

    DAttendanceRegisterTable = SQLTABLE(datasource,columns=columns,headers=headers,orderby='GR_NO',_class='datatable')
    TextMessage = ""
    ReportDate = datetime.date.strftime(ReportDate,"%d-%m-%Y")
    dworkbook.save(dpath)
    if download == "YES":
        redirect(URL(r=request, f='download_AttendanceRegister'))
    #tables = plugins.powerTable
    #tables.datasource = datasource
    #tables.uitheme = 'cupertino'
    #tables.dtfeatures['sPaginationType'] = 'two button'
    #tables.keycolumn = 'tempAttendanceRegisterTable.id'
    #tables.showkeycolumn = False
    #tables.columns = ['tempAttendanceRegisterTable.GR_NO','tempAttendanceRegisterTable.DEV_DTY','tempAttendanceRegisterTable.CANTEEN']
    #tables.hiddencolumns = ['tempAttendanceRegisterTable.GR_NO','tempAttendanceRegisterTable.DEV_DTY','tempAttendanceRegisterTable.CANTEEN']
    #tables.headers = 'labels'
    #created_table = tables.create()
    #DAttendanceRegisterTable = plugin_powerTable(datasource)

    mail.send('softwareattendance@gmail.com',
        'All Attendance register',
        'All attendance register',
        attachments = mail.Attachment('/home/rootcat/new_web2py/web2py/applications/AttendanceSoftware/private/AttendanceRegister.xlsx', content_id='text'))

    logf.close()
    return 0



from gluon.scheduler import Scheduler
scheduler = Scheduler(db2)
