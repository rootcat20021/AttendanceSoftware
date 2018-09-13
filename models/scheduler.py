#Other database support
#db2 = DAL('mysql://rootcat:7133783@rootcat.mysql.pythonanywhere-services.com/rootcat$AttendanceDB',pool_size=1,check_reserved=['all'])
#db2 = DAL('mysql://sql695965:qT6!xZ4!@sql6.freesqldatabase.com:3306/sql695965',pool_size=1,check_reserved=['all'])
db2 = DAL('sqlite://storage8.db')

Incharge_Email_Ids = {}
Incharge_Email_Ids['1'] = 'badgecanteen1@gmail.com'
Incharge_Email_Ids['2'] = 'badgecanteen2@gmail.com'
Incharge_Email_Ids['3'] = 'badgecanteen3@gmail.com'
Incharge_Email_Ids['4'] = 'badgecanteen4@gmail.com'
Incharge_Email_Ids['5'] = 'badgecanteen5@gmail.com'
#Incharge_Email_Ids['1'] = 'pushkar.sareen@gmail.com'
#Incharge_Email_Ids['2'] = 'pushkar.sareen@gmail.com'
#Incharge_Email_Ids['3'] = 'pushkar.sareen@gmail.com'
#Incharge_Email_Ids['4'] = 'pushkar.sareen@gmail.com'
#Incharge_Email_Ids['5'] = 'pushkar.sareen@gmail.com'


from gluon.tools import Mail
mail = Mail()
mail.settings.server = 'smtp.gmail.com:25'
mail.settings.sender = 'acknowledgesynchronization@gmail.com'
mail.settings.login = 'acknowledgesynchronization:synchronizationacknowledge'

def Roundoffdate(DATETIME, CUTOFF):
    import datetime
    RoundedDate = DATETIME
    if (DATETIME.hour < int(CUTOFF)) or (DATETIME.hour == int(CUTOFF) and DATETIME.minute == 0) or (CUTOFF == 24):
       RoundedDate = RoundedDate.replace(hour=0, minute=0, second=0, microsecond=0)
    else:
       RoundedDate = (RoundedDate + datetime.timedelta(days=1)).replace(hour=0, minute=0, second=0, microsecond=0)
    return RoundedDate

def uploaddata_SSAttendance():
    import xlrd
    import os
    import re, time
    import datetime
    import logging
    import pandas as pd
    db.commit()

    path = os.path.join(request.folder,'private','SSAttendanceDates_xls.xlsx')
    pathlog = os.path.join(request.folder,'private','log_SSAttendance')
    logf = open(pathlog,'w')

    import commands
    logf.write(commands.getoutput('date'))
    logf.write("indexed SSAttendance Dates\n")
    logf.close()
    logf = open(pathlog,'a')

    LastUpdated = datetime.datetime(2000, 1, 1)
    datasource  = db(db.LocalVariables.id > 0).select()
    for data in datasource:
        LastUpdated = data['LastUpdated']

    logf.write('Attendance was last updated on ' + datetime.datetime.strftime(LastUpdated,'%d-%b-%Y') + '\n')

    df = pd.read_excel(path,header=0,index_col=0)

    for row in df.iterrows():
       row_dict = {}
       #logf.write(str(row[1]) + "\n")
       #logf.close()
       #logf = open(pathlog,'a')

       row_dict = {}
       row_dict['SewadarNewID'] = row[1]['SewadarNewID']
       if type(row[1]['DutyDateTime']) == unicode:
           row_dict['DutyDate'] = datetime.datetime.strptime(row[1]['DutyDateTime'],'%d/%m/%Y %H:%M:%S')
           logf.write("unicode=")
       else:
           row_dict['DutyDate'] = row[1]['DutyDateTime'].replace(day=row[1]['DutyDateTime'].month,month=row[1]['DutyDateTime'].day)
           logf.write("datetime=")

       logf.write(datetime.datetime.strftime(row_dict['DutyDate'],'%d-%b-%Y') + '  ')
       logf.write("Diff = " + str(row_dict['DutyDate'].replace(hour=0, minute=0, second=0, microsecond=0) - row[1]['DutyDate'].replace(hour=0, minute=0, second=0, microsecond=0)) + '\n')
       row_dict['Duty_Type'] = row[1]['Duty Type']

       #print row_dict['SewadarNewID']
       if LastUpdated < row_dict['DutyDate']:
           LastUpdated = row_dict['DutyDate']

       try:
           db.SSAttendanceDate.insert(**row_dict)
       except:
           db((db.SSAttendanceDate.SewadarNewID == row_dict['SewadarNewID']) & (db.SSAttendanceDate.DutyDate == row_dict['DutyDate'])).update(Duty_Type=row_dict['Duty_Type'])

    db(db.LocalVariables.id>0).update(LastUpdated=LastUpdated)
    logf.write("Send to database engine\n")

    db.commit()

    logf.write("Commited\n")
    logf.close()
    mail.send('acknowledgesynchronization@gmail.com',
        'Comitted SSAttendance to database',
        'Success',
        attachments = mail.Attachment('/home/rootcat/new_web2py/web2py/applications/AttendanceSoftware/private/log_SSAttendance', content_id='text'))
    return 0


def ParshadListScheduled(DateSelectedStart,DateSelectedEnd,MandatoryDaysDateStart,MandatoryDaysDateEnd,MandatoryDaysCountCutoff,SSCountCutOffGents,SSCountCutOffLadies,CVCutOff,VisitCountCutOff,WWCutOff,WWWaiver,WWAgeWaiver,DAY_END_TIME,MailSubject):
    import os
    import pandas as pd
    import pprint
    CURRENT_VISIT_START = {}
    CURRENT_VISIT_START[0] = datetime.datetime.strptime('26-September-2018 00:00:00','%d-%B-%Y %H:%M:%S')
    CURRENT_VISIT_START[1] = datetime.datetime.strptime('27-September-2018 00:00:00','%d-%B-%Y %H:%M:%S')
    CURRENT_VISIT_START[2] = datetime.datetime.strptime('28-September-2018 00:00:00','%d-%B-%Y %H:%M:%S')
    CURRENT_VISIT_START[3] = datetime.datetime.strptime('29-September-2018 00:00:00','%d-%B-%Y %H:%M:%S')
    CURRENT_VISIT_START[4] = datetime.datetime.strptime('30-September-2018 00:00:00','%d-%B-%Y %H:%M:%S')
    CURRENT_VISIT_START['COUNT'] = 5

    pathlog = os.path.join(request.folder,'private','log_ParshadListScheduled')
    logf = open(pathlog,'w')
    logf.write(str(datetime.datetime.now()) + "\n")
    db.commit()
    logf.write("Committed\n")
    df_MasterSheet = pd.DataFrame.from_records(db(db.MasterSheet.id > 0).select().as_list())
    df_MasterSheet.to_excel(os.path.join(request.folder,'private','df_MasterSheet.xlsx'))
    df_CountSheet = pd.DataFrame.from_records(db(db.SSAttendanceCount.id > 0).select().as_list())
    df_CountSheet.drop(columns=['w','B','V1','V2','V3','V4','Father_Husband_Name'],inplace=True)
    df_CountSheet['SewadarNewID'] = df_CountSheet['NewID'].str.replace('BH0011','')
    df_CountSheet = df_CountSheet.set_index(['SewadarNewID'])
    df_CountSheet.drop(columns=['id','areaname','OldSewadarid'],inplace=True)
    df_CountSheet.drop_duplicates(inplace=True)
    df_CountSheet.to_excel(os.path.join(request.folder,'private','df_CountSheet.xlsx'))

    df_DateSheet_visit_days = pd.DataFrame.from_records(db((db.SSAttendanceDate.DutyDate >= CURRENT_VISIT_START[0]) & (db.SSAttendanceDate.DutyDate <= CURRENT_VISIT_START[CURRENT_VISIT_START['COUNT']-1])).select().as_list())
    df_DateSheet_visit_days.reset_index(level=0,inplace=True)
    df_DateSheet_visit_days.to_excel(os.path.join(request.folder,'private','df_visit_days.xlsx'))

    df_DateSheet_visit_pivot = ""
    try:
        df_DateSheet_visit_pivot = pd.pivot_table(df_DateSheet_visit_days,index=['SewadarNewID'],columns=['Duty_Type'],aggfunc='count',margins=False,fill_value=0)
        df_DateSheet_visit_pivot = df_DateSheet_visit_pivot.xs('DutyDate', axis=1, drop_level=True)
    except:
        df_DateSheet_visit_pivot = pd.DataFrame(df_CountSheet.xs('Name',axis=1))
        df_DateSheet_visit_pivot.rename_axis("SewadarNewID", axis='index', inplace=True)
        df_DateSheet_visit_pivot.loc[:, 'B'] = 0
        df_DateSheet_visit_pivot.loc[:, 'D'] = 0
        df_DateSheet_visit_pivot.loc[:, 'V'] = 0
        df_DateSheet_visit_pivot.loc[:, 'W'] = 0



    df_DateSheet_visit_pivot.loc[:, 'VISIT_COUNT'] = 0
    try:
        df_DateSheet_visit_pivot['VISIT_COUNT'] = df_DateSheet_visit_pivot['W']*2
    except:
        pass
    try:
        df_DateSheet_visit_pivot['VISIT_COUNT'] = df_DateSheet_visit_pivot['VISIT_COUNT'] + df_DateSheet_visit_pivot['D']
    except:
        pass
    try:
        df_DateSheet_visit_pivot['VISIT_COUNT'] = df_DateSheet_visit_pivot['VISIT_COUNT'] + df_DateSheet_visit_pivot['B']
    except:
        pass
    try:
        df_DateSheet_visit_pivot['VISIT_COUNT'] = df_DateSheet_visit_pivot['VISIT_COUNT'] + df_DateSheet_visit_pivot['V']
    except:
        pass

    try:
        df_DateSheet_visit_pivot.drop(columns='B',inplace=True)
    except:
        pass
    try:
        df_DateSheet_visit_pivot.drop(columns='D',inplace=True)
    except:
        pass
    try:
        df_DateSheet_visit_pivot.drop(columns='V',inplace=True)
    except:
        pass
    try:
        df_DateSheet_visit_pivot.drop(columns='W',inplace=True)
    except:
        pass


    df_DateSheet_visit_pivot.to_excel(os.path.join(request.folder,'private','df_visit_days_pivot.xlsx'))

    df_DateSheet_WW  = pd.DataFrame.from_records(db((db.SSAttendanceDate.DutyDate >= (datetime.datetime.strptime(DateSelectedEnd,'%Y-%m-%d %H:%M:%S').replace(day=1,month=1,hour=0, minute=0, second=0, microsecond=0))) & (db.SSAttendanceDate.DutyDate <= (datetime.datetime.strptime(DateSelectedEnd,'%Y-%m-%d %H:%M:%S').replace(hour=23, minute=59, second=59, microsecond=999))) & (db.SSAttendanceDate.Duty_Type == 'W')).select().as_list())
    df_DateSheet_WW['SewadarNewID'] = df_DateSheet_WW['SewadarNewID'].str.replace('BH0011','')
    df_DateSheet_WW_pivot = pd.pivot_table(df_DateSheet_WW,index=['SewadarNewID'],columns=['Duty_Type'],aggfunc='count',margins=False,fill_value=0)
    df_DateSheet_WW_pivot = df_DateSheet_WW_pivot.xs('DutyDate', axis=1, drop_level=True)
    try:
        df_DateSheet_WW_pivot.drop(columns='D',inplace=True)
    except:
        pass
    try:
        df_DateSheet_WW_pivot.drop(columns='B',inplace=True)
    except:
        pass
    try:
        df_DateSheet_WW_pivot.drop(columns='V',inplace=True)
    except:
        pass

    df_DateSheet_WW_pivot.rename(columns={'W':'WW_Count'},inplace=True)
    df_DateSheet_WW_pivot.to_excel(os.path.join(request.folder,'private','df_dates_WW_pivot.xlsx'))

    df_DateSheet_mandatory_days = pd.DataFrame.from_records(db((db.SSAttendanceDate.DutyDate >= (datetime.datetime.strptime(MandatoryDaysDateStart,'%Y-%m-%d %H:%M:%S'))) & (db.SSAttendanceDate.DutyDate <= (datetime.datetime.strptime(MandatoryDaysDateEnd,'%Y-%m-%d %H:%M:%S')))).select().as_list())
    df_DateSheet_mandatory_days['SewadarNewID'] = df_DateSheet_mandatory_days['SewadarNewID'].str.replace('BH0011','')
    df_DateSheet_mandatory_days.to_excel(os.path.join(request.folder,'private','df_DateSheet_mandatory_days.xlsx'))

    df_DateSheet_mandatory_pivot = pd.pivot_table(df_DateSheet_mandatory_days,index=['SewadarNewID'],columns=['Duty_Type'],aggfunc='count',margins=False,fill_value=0)
    df_DateSheet_mandatory_pivot = df_DateSheet_mandatory_pivot.xs('DutyDate', axis=1, drop_level=True)
    df_DateSheet_mandatory_pivot.loc[: ,'MANDATORY_COUNT'] = 0
    try:
        df_DateSheet_mandatory_pivot.loc[:, 'MANDATORY_COUNT'] = df_DateSheet_mandatory_pivot.loc[:, 'W']*2
        df_DateSheet_mandatory_pivot.drop(columns='W',inplace=True)
    except:
        pass
    try:
        df_DateSheet_mandatory_pivot.loc[:, 'MANDATORY_COUNT'] = df_DateSheet_mandatory_pivot.loc[:, 'D'] + df_DateSheet_mandatory_pivot.loc[:, 'MANDATORY_COUNT']
        df_DateSheet_mandatory_pivot.drop(columns='D',inplace=True)
    except:
        pass

    try:
        df_DateSheet_mandatory_pivot[:, 'MANDATORY_COUNT'] = df_DateSheet_mandatory_pivot.loc[:, 'MANDATORY_COUNT'] + df_DateSheet_mandatory_pivot.loc[:, 'V']
        df_DateSheet_mandatory_pivot.drop(columns='V',inplace=True)
    except:
        pass
    try:
        df_DateSheet_mandatory_pivot[:, 'MANDATORY_COUNT'] = df_DateSheet_mandatory_pivot.loc[:, 'MANDATORY_COUNT'] + df_DateSheet_mandatory_pivot.loc[:, 'B']
        df_DateSheet_mandatory_pivot.drop(columns='B',inplace=True)
    except:
        pass

    df_DateSheet_mandatory_pivot.to_excel(os.path.join(request.folder,'private','df_mandatory_dates_pivot.xlsx'))

    df_MergedMasterSheet = df_CountSheet.merge(df_MasterSheet,on=['SewadarNewID'],how='left')
    #df_MergedMasterSheet = df_MasterSheet.merge(df_CountSheet,on=['SewadarNewID'],how='left')
    #Add count of visit days and mandatory days
    df_MergedMasterSheet = df_MergedMasterSheet.merge(df_DateSheet_visit_pivot,on=['SewadarNewID'],how='left')
    df_MergedMasterSheet = df_MergedMasterSheet.merge(df_DateSheet_mandatory_pivot,on=['SewadarNewID'],how='left')
    try:
        df_MergedMasterSheet = df_MergedMasterSheet.merge(df_DateSheet_WW_pivot,on=['SewadarNewID'],how='left')
    except:
        df_MergedMasterSheet.loc[:,'WW_Count'] = 0

    df_MergedMasterSheet.to_excel(os.path.join(request.folder,'private','df_MergedMaster_all_input.xlsx'))
    #Keywords for exception
    #ALL , Visits Count,Current Visit
    df_ExceptionMail = pd.DataFrame.from_records(db(db.ParshadMailException).select().as_list())
    df_MergedMasterSheet = df_MergedMasterSheet.merge(df_ExceptionMail,on=['SewadarNewID'],how='left')
    df_SSTentativeParshadList = pd.DataFrame.from_records(db(db.SSTentativeParshadList.id > 0).select().as_list())
    df_MergedMasterSheet = df_MergedMasterSheet.merge(df_SSTentativeParshadList,on=['SewadarNewID'],how='left')
    logf.write("fetched records\n")
    message = "ALL OK "

    df_MergedMasterSheet = df_MergedMasterSheet.set_index(['SewadarNewID'])
    #time.sleep (1);
    for row in df_DateSheet_visit_days.iterrows():
        df_MergedMasterSheet.at[row[1]['SewadarNewID'],row[1]['DutyDate'].replace(hour=0, minute=0, second=0, microsecond=0)] = 'P'

    for row in df_DateSheet_mandatory_days.iterrows():
        df_MergedMasterSheet.at[row[1]['SewadarNewID'],row[1]['DutyDate'].replace(hour=0, minute=0, second=0, microsecond=0)] = 'M'

    #Fill Empty WW_COunt and Total and TotalVisit with 0
    df_MergedMasterSheet['Total'].fillna(0,inplace=True)
    df_MergedMasterSheet['TotalVisit'].fillna(0,inplace=True)
    df_MergedMasterSheet['WW_Count'].fillna(0,inplace=True)
    #Start the checks
    df_MergedMasterSheet.loc[:,'ParshadStatus'] = 'OK'
    df_MergedMasterSheet.loc[:,'ParshadRemark'] = 'OK'
    for row in df_MergedMasterSheet.iterrows():
        if row[1]['ExceptionField'] == 'ALL':
            df_MergedMasterSheet.at[row[0],'ParshadStatus'] = 'OK'
            df_MergedMasterSheet.at[row[0],'ParshadRemark'] = 'CO'
            continue
        #Check initiation

        if not(row[1]['ExceptionField'] == 'Initiation'):
            if row[1]['Initiated_Status'] == 'Y':
                df_MergedMasterSheet.at[row[0],'ParshadStatus'] = 'No'
                try:
                    df_MergedMasterSheet.at[row[0],'ParshadRemark'] =  df_MergedMasterSheet.at[row[0],'ParshadRemark'] + '\n' + "Non Initiated"
                except:
                    df_MergedMasterSheet.at[row[0],'ParshadRemark'] = "Non Initiated"

        #Check 9 visit
        if not(row[1]['ExceptionField'] == 'Visits Count'):
            try:
                if row[1]['status'].upper() != 'PERMANENT':
                    if int(row[1]['TotalVisit']) < int(VisitCountCutOff):
                        df_MergedMasterSheet.at[row[0],'ParshadStatus'] = 'No'
                        try:
                            df_MergedMasterSheet.at[row[0],'ParshadRemark'] =  df_MergedMasterSheet.at[row[0],'ParshadRemark'] + '\n' + str(int(VisitCountCutOff) - int(df_MergedMasterSheet.at[row[0],'TotalVisit'])) + ' Visits Short'
                        except:
                            df_MergedMasterSheet.at[row[0],'ParshadRemark'] = str(int(VisitCountCutOff) - int(df_MergedMasterSheet.at[row[0],'TotalVisit'])) + ' Visits'
            except:
                df_MergedMasterSheet.at[row[0],'ParshadStatus'] = 'MissingEntry:Status'
                df_MergedMasterSheet.at[row[0],'ParshadRemark'] = 'MissingEntry:Status'


        #Check mandatory attendance
        if not(row[1]['ExceptionField'] == 'Mandatory Days'):
            if not(pd.isnull(row[1]['MANDATORY_COUNT'])):
                if int(row[1]['MANDATORY_COUNT']) < int(MandatoryDaysCountCutoff):
                    df_MergedMasterSheet.at[row[0],'ParshadStatus'] = 'No'
                    try:
                        df_MergedMasterSheet.at[row[0],'ParshadRemark'] =  df_MergedMasterSheet.at[row[0],'ParshadRemark'] + '\n' + "Missi Roti Attendance Short by " + str(int(MandatoryDaysCountCutoff) - int(row[1]['MANDATORY_COUNT']))
                    except:
                        df_MergedMasterSheet.at[row[0],'ParshadRemark'] = "Missi Roti Attendance Short by " + str(int(MandatoryDaysCountCutoff) - int(row[1]['MANDATORY_COUNT']))
            else:
                df_MergedMasterSheet.at[row[0],'ParshadStatus'] = 'MissingEntry:MANDATORY_COUNT'
                df_MergedMasterSheet.at[row[0],'ParshadRemark'] = 'MissingEntry:MANDATORY_COUNT'
        #Check WW
        if not(row[1]['ExceptionField'] == 'WW Count'):
            if not(pd.isnull(row[1]['gender'])):
                if row[1]['gender'].upper() == "MALE":
                    if int(row[1]['WW_Count']) < int(WWCutOff):
                        df_MergedMasterSheet.at[row[0],'ParshadStatus'] = 'No'
                        try:
                            df_MergedMasterSheet.at[row[0],'ParshadRemark'] =  df_MergedMasterSheet.at[row[0],'ParshadRemark'] + '\n' + "WW Short by " + str(int(WWCutOff) - int(row[1]['WW_Count']))
                        except:
                            df_MergedMasterSheet.at[row[0],'ParshadRemark'] = " WW Short by " + str(int(WWCutOff) - int(row[1]['WW_Count']))
            else:
                df_MergedMasterSheet.at[row[0],'ParshadStatus'] = 'MissingEntry:WW'
                df_MergedMasterSheet.at[row[0],'ParshadRemark'] = 'MissingEntry:WW'

        #Check Before visit
        if not(row[1]['ExceptionField'] == 'SS Count'):
            if not(pd.isnull(row[1]['gender'])):
                if row[1]['gender'].upper() == "MALE":
                    if int(row[1]['Total']) < int(SSCountCutOffGents):
                        df_MergedMasterSheet.at[row[0],'ParshadStatus'] = 'No'
                        try:
                            df_MergedMasterSheet.at[row[0],'ParshadRemark'] =  df_MergedMasterSheet.at[row[0],'ParshadRemark'] + '\n' + "Before Visit Attendance Short by " + str(int(SSCountCutOffGents) - int(row[1]['Total']))
                        except:
                            df_MergedMasterSheet.at[row[0],'ParshadRemark'] = "Before Visit Attendance Short by " + str(int(SSCountCutOffGents) - int(row[1]['Total']))
                else:
                    if int(row[1]['Total']) < int(SSCountCutOffLadies):
                        df_MergedMasterSheet.at[row[0],'ParshadStatus'] = 'No'
                        try:
                            df_MergedMasterSheet.at[row[0],'ParshadRemark'] =  df_MergedMasterSheet.at[row[0],'ParshadRemark'] + '\n' + "Before Visit Attendance Short by " + str(int(SSCountCutOffLadies) - int(row[1]['Total']))
                        except:
                            df_MergedMasterSheet.at[row[0],'ParshadRemark'] = "Before Visit Attendance Short by " + str(int(SSCountCutOffLadies) - int(row[1]['Total']))

            else:
                df_MergedMasterSheet.at[row[0],'ParshadStatus'] = 'MissingEntry:Total'
                df_MergedMasterSheet.at[row[0],'ParshadRemark'] = 'MissingEntry:Total'

        #Check visit attendance
        if not(row[1]['ExceptionField'] == 'Current Visit'):
            if not(pd.isnull(row[1]['VISIT_COUNT'])):
                if int(row[1]['VISIT_COUNT']) < int(CVCutOff):
                    df_MergedMasterSheet.at[row[0],'ParshadStatus'] = 'No'
                    try:
                        df_MergedMasterSheet.at[row[0],'ParshadRemark'] =  df_MergedMasterSheet.at[row[0],'ParshadRemark'] + '\n' + "Current Visit Short by " + str(int(CVCutOff) - int(row[1]['VISIT_COUNT']))
                    except:
                        df_MergedMasterSheet.at[row[0],'ParshadRemark'] = "Current Visit Short by " + str(int(CVCutOff) - int(row[1]['VISIT_COUNT']))
                elif (df_MergedMasterSheet.at[row[0],'ParshadStatus'] == 'No'):
                    df_MergedMasterSheet.at[row[0],'ParshadStatus'] = 'Packet'

            else:
                df_MergedMasterSheet.at[row[0],'ParshadStatus'] = 'MissingEntry:VisitCount'
                df_MergedMasterSheet.at[row[0],'ParshadRemark'] = 'MissingEntry:VisitCount'

    df_MergedMasterSheet.to_excel(os.path.join(request.folder,'private','df_ParshadStatus.xlsx'))
    df_MergedMasterSheet.to_csv(os.path.join(request.folder,'private','df_ParshadStatus.csv'))

    print "Marked visit days and mandatory days attendance"
    mail.send('acknowledgesynchronization@gmail.com',
        MailSubject,
        'Tentative Parshad List' + '\n DateSelectedStart=' + str(DateSelectedStart) + '\n DateSelectedEnd=' + str(DateSelectedEnd) + '\n SSCountCutOffLadies=' + str(SSCountCutOffLadies) + '\n SSCountCutOffGents=' + str(SSCountCutOffGents) + '\n MandatoryDaysDateStart=' + str(MandatoryDaysDateStart) + '\n MandatoryDaysDateEnd=' + str(MandatoryDaysDateEnd) + '\n MandatoryDaysCountCutoff=' + str(MandatoryDaysCountCutoff) + '\n VisitCountCutOff=' + str(VisitCountCutOff) + '\n CVCutOff=' + str(CVCutOff) + '\n WWCutOff=' + str(WWCutOff)  + '\n WWWaiver=' + str(WWWaiver) + '\n WWAgeWaiver =' + str(WWAgeWaiver),
        attachments = mail.Attachment('/home/rootcat/new_web2py/web2py/applications/AttendanceSoftware/private/df_ParshadStatus.xlsx', content_id='text'))

    mail.send('acknowledgesynchronization@gmail.com',
        "SENDING TENTATIVE PARSHAD STATUS",
        'Tentative Parshad List' + '\n DateSelectedStart=' + str(DateSelectedStart) + '\n DateSelectedEnd=' + str(DateSelectedEnd) + '\n SSCountCutOffLadies=' + str(SSCountCutOffLadies) + '\n SSCountCutOffGents=' + str(SSCountCutOffGents) + '\n MandatoryDaysDateStart=' + str(MandatoryDaysDateStart) + '\n MandatoryDaysDateEnd=' + str(MandatoryDaysDateEnd) + '\n MandatoryDaysCountCutoff=' + str(MandatoryDaysCountCutoff) + '\n VisitCountCutOff=' + str(VisitCountCutOff) + '\n CVCutOff=' + str(CVCutOff) + '\n WWCutOff=' + str(WWCutOff)  + '\n WWWaiver=' + str(WWWaiver) + '\n WWAgeWaiver =' + str(WWAgeWaiver),
        attachments = mail.Attachment('/home/rootcat/new_web2py/web2py/applications/AttendanceSoftware/private/df_ParshadStatus.csv', content_id='text'))

    logf.write("Mail sent!\n")
    logf.close()

    return dict(message=message)


#@auth.requires_login()
def AttendanceRegisterScheduledAll(DateSelectedStart,DateSelectedEnd):
    import os
    pathlog = os.path.join(request.folder,'private','log_AttendanceRegisterAll')
    logf = open(pathlog,'w')

    from gluon.sqlhtml import form_factory
    import datetime
    import time

    from openpyxl import Workbook
    from openpyxl.cell import get_column_letter
    import os
    response.headers['Connection'] =  'keep-alive'
    dworkbook = Workbook()
    dAttendanceRegister = dworkbook.create_sheet(0)
    dpath = os.path.join(request.folder,'private','AttendanceRegister.xlsx')

    Register = "EMPTY"
    DAttendanceRegisterTable = "EMPTY"
    datasource = {}
    columns = []
    orderby = []
    headers = {}
    excel_headers = {}
    DEV_DTY = ""
    TextMessage = "Jatha Wise Attendance Report"
    ReportDate =  0
    GENTS_REQUIRED = 30
    LADIES_REQUIRED = 36
    SS_GENTS_REQUIRED = 30
    SS_LADIES_REQUIRED = 36

    try:
        db.tempAttendanceRegisterTable.drop()
    except:
        pass

    SewadarDetails = {'Sewadars':[]}
#    form=form_factory(SQLField('JATHA','string',requires=IS_IN_DB(db,'MasterSheet.DEV_DTY','%(DEV_DTY)s')),SQLField('DateStart','date',default=datetime.datetime.today()-datetime.timedelta(days=31)),SQLField('DateEnd','date',default=datetime.datetime.today()),formname='DateSelect')
    my_list = db().select(db.MasterSheet.DEV_DTY, distinct=True).as_list()
    myjathalist = []
    for mydict in my_list:
        myjathalist.append(mydict['DEV_DTY'])

    myjathalist.sort()
    form=form_factory(SQLField('JATHA','string',requires=IS_IN_SET(myjathalist)),SQLField('DateStart','date',default=datetime.datetime.today()-datetime.timedelta(days=31)),SQLField('DateEnd','date',default=datetime.datetime.today()),SQLField('Download','string',requires=IS_IN_SET(['YES','NO']),default='NO'),formname='DateSelect')
    try:
        os.remove(dpath)
    except:
        pass

    download = request.vars.Download
    ReportDate = datetime.datetime.today()-datetime.timedelta(days=3001)

    #SSDate = db((db.SSAttendanceDate.DutyDate >= DateSelectedStart) & (db.SSAttendanceDate.DutyDate <= DateSelectedEnd) & (db.SSAttendanceDate.Duty_Type == "W")).select()
    SSDate = db((db.SSAttendanceDate.DutyDate >= DateSelectedStart) & (db.SSAttendanceDate.DutyDate <= DateSelectedEnd)).select()
    for Sewadar in SSDate:
        delta = Sewadar.DutyDate - datetime.datetime.strptime(DateSelectedStart,'%Y-%m-%d')
        SewadarDetails[Sewadar.SewadarNewID,'DAYS',delta.days] = Sewadar.DutyDate
        SewadarDetails[Sewadar.SewadarNewID,'TYPE',delta.days] = Sewadar.Duty_Type
        if ReportDate < Sewadar.DutyDate:
            ReportDate = Sewadar.DutyDate

    SSCount = db(db.SSAttendanceCount).select()
    for Sewadar in SSCount:
        SewadarDetails[Sewadar.NewID,'Gender'] = Sewadar.gender
        SewadarDetails[Sewadar.NewID,'TotalCount'] = Sewadar.Total
        SewadarDetails[Sewadar.NewID,'NAME'] = Sewadar.Name


    MasterTable = db(db.MasterSheet.GR_NO == 'SS015298').select()
    for Sewadar in MasterTable:
             SewadarDetails[Sewadar.GR_NO,'SewadarNewID'] = Sewadar.SewadarNewID
             SewadarDetails[Sewadar.SewadarNewID,'DEV_DTY'] = Sewadar.DEV_DTY

    DEV_DTY = request.vars.JATHA
#        if auth.user.username == 'admin':
#            DEV_DTY = request.vars.JATHA
#        else:
#            try:
#                DEV_DTY = SewadarDetails[auth.user.username,'DEV_DTY']
#            except:
#                #try:
#                    DEV_DTY = SewadarDetails[SewadarDetails[auth.user.username,'SewadarNewID'],'DEV_DTY']
#                #except:
#                #    TextMessage = 'This ID is not allocated to any Jatha'


    MasterTable = db(db.MasterSheet.id > 0).select()
    for Sewadar in MasterTable:
        try:
            #Check if this ID was in SewaSamiti list
            SewadarDetails[Sewadar.SewadarNewID,'NAME']
            SewadarDetails['Sewadars'].append(Sewadar.SewadarNewID)
            SewadarDetails[Sewadar.SewadarNewID,'CANTEEN'] = Sewadar.CANTEEN
            SewadarDetails[Sewadar.SewadarNewID,'DEV_DTY'] = Sewadar.DEV_DTY
            SewadarDetails[Sewadar.SewadarNewID,'OLD_ID'] = Sewadar.GR_NO
        except:
            TextMessage = 'Sewadar ID not in list'


    Fields = [Field('D%02i' %i,'string') for i in range((datetime.datetime.strptime(DateSelectedEnd,'%Y-%m-%d') - datetime.datetime.strptime(DateSelectedStart,'%Y-%m-%d')).days)]


    #Now define the table
    db.define_table('tempAttendanceRegisterTable',
            Field('GR_NO','string'),
            Field('SewadarNewID','string'),
            Field('NAME','string'),
            Field('DEV_DTY','string'),
            Field('CANTEEN','string'),
            Field('TOTAL','integer'),
            Field('STATUS','string'),
            Field('REQD','integer'),
            Field('GENDER','string'),
            Field('WWDATES','string'),
            *Fields,
            migrate=True,
            redefine=True,
            format='%(SewadarNewID)s')

    db(db.tempAttendanceRegisterTable.id > 0).delete()

    excel_headers = {1:'SewadarNewID',2:'GR_NO',3:'NAME',4:'DEV_DTY',5:'CANTEEN',6:'GENDER',7:'STATUS',8:'REQD',9:'WWDATES'}
    dAttendanceRegister.append(excel_headers.values())

    SewadarNumber = 1
    for Sewadar in SewadarDetails['Sewadars']:
        logf.write(Sewadar)
        mydict = {}
        SewadarNumber = SewadarNumber + 1
        mydict['SewadarNewID'] = Sewadar
        mydict['GR_NO'] = SewadarDetails[Sewadar,'OLD_ID']
        mydict['NAME'] = SewadarDetails[Sewadar,'NAME']
        mydict['DEV_DTY'] = SewadarDetails[Sewadar,'DEV_DTY']
        mydict['CANTEEN'] = SewadarDetails[Sewadar,'CANTEEN']
        mydict['TOTAL'] = SewadarDetails[Sewadar,'TotalCount']
        mydict['GENDER'] = SewadarDetails[Sewadar,'Gender']
        mydict['STATUS'] = ''
        mydict['REQD'] = (GENTS_REQUIRED - SewadarDetails[Sewadar,'TotalCount']) if (SewadarDetails[Sewadar,'Gender'] == 'Male') else (LADIES_REQUIRED - SewadarDetails[Sewadar,'TotalCount'])
        if mydict['REQD'] < 1:
            mydict['REQD'] = 0
        elif (((SewadarDetails[Sewadar,'Gender'] == 'Male')) and mydict['REQD'] > SS_GENTS_REQUIRED):
            mydict['REQD'] = SS_GENTS_REQUIRED
        elif (((SewadarDetails[Sewadar,'Gender'] == 'Female')) and mydict['REQD'] > SS_LADIES_REQUIRED):
            mydict['REQD'] = SS_LADIES_REQUIRED
        else:
            pass



        #Write to excel too
        dAttendanceRegister.cell('A' + str(SewadarNumber)).value = mydict['SewadarNewID']
        dAttendanceRegister.cell('B' + str(SewadarNumber)).value = mydict['GR_NO']
        dAttendanceRegister.cell('C' + str(SewadarNumber)).value = mydict['NAME']
        dAttendanceRegister.cell('D' + str(SewadarNumber)).value = mydict['DEV_DTY']
        dAttendanceRegister.cell('E' + str(SewadarNumber)).value = mydict['CANTEEN']
        dAttendanceRegister.cell('F' + str(SewadarNumber)).value = mydict['GENDER']
        dAttendanceRegister.cell('G' + str(SewadarNumber)).value = mydict['STATUS']
        dAttendanceRegister.cell('H' + str(SewadarNumber)).value = mydict['REQD']


        columns = ['tempAttendanceRegisterTable.SewadarNewID','tempAttendanceRegisterTable.GR_NO','tempAttendanceRegisterTable.NAME','tempAttendanceRegisterTable.DEV_DTY','tempAttendanceRegisterTable.CANTEEN','tempAttendanceRegisterTable.REQD']
        orderby=columns

        headers = {'tempAttendanceRegisterTable.SewadarNewID':{'label':T('SewadarNewID'),'class':'','width':12,'truncate':12,'selected': False},
           'tempAttendanceRegisterTable.GR_NO':{'label':T('GR_NO'),'class':'','width':10,'truncate':10,'selected': False},
           'tempAttendanceRegisterTable.NAME':{'label':T('NAME'),'class':'','width':10,'truncate': 10,'selected': False},
           'tempAttendanceRegisterTable.DEV_DTY':{'label':T('DEV_DTY'),'class':'','width':11,'truncate': 11,'selected': False},
           'tempAttendanceRegisterTable.CANTEEN':{'label':T('CANTEEN'),'class':'','width':13,'truncate': 13,'selected': False},
           'tempAttendanceRegisterTable.REQD':{'label':T('REQD'),'class':'','width':4,'truncate': 4,'selected': False}
           }


        for i in range((datetime.datetime.strptime(DateSelectedEnd,'%Y-%m-%d') - datetime.datetime.strptime(DateSelectedStart,'%Y-%m-%d')).days):
            dateindex = datetime.datetime.strptime(DateSelectedStart,'%Y-%m-%d') + datetime.timedelta(i)
            columns.append('tempAttendanceRegisterTable.D%02i' %i)
            headers['tempAttendanceRegisterTable.D%02i' %i] = {'label':datetime.date.strftime(dateindex,'%d\n%m'),'class':'','width':'3','truncate': 3,'selected': False}
            excel_headers[i+10] = (datetime.date.strftime(dateindex,'%d-%m'))
            logf.write(excel_headers[i+10])
            logf.write("\n")

            try:
                mydict['D%02i' %i] = SewadarDetails[Sewadar,'TYPE',i]
                if mydict['D%02i' %i] == 'W':
                    try:
                        mydict["WWDATES"] = mydict["WWDATES"] + '\n' +  excel_headers[i+10]
                    except:
                        mydict["WWDATES"] = excel_headers[i+10]
            except:
                mydict['D%02i' %i] = ''
            dAttendanceRegister.cell(get_column_letter(i+10) + str(SewadarNumber)).value = mydict['D%02i' %i]
            try:
                dAttendanceRegister.cell(get_column_letter(9) + str(SewadarNumber)).value = mydict["WWDATES"]
            except:
                pass

        db.tempAttendanceRegisterTable.insert(**mydict)

    for key in excel_headers.keys():
        dAttendanceRegister.cell(get_column_letter(key) + '1').value = excel_headers[key]

    Register = SQLTABLE(SSDate,headers='fieldname:capitalize')

    datasource = db(db.tempAttendanceRegisterTable).select(orderby=db.tempAttendanceRegisterTable.GR_NO)

    DAttendanceRegisterTable = SQLTABLE(datasource,columns=columns,headers=headers,orderby='GR_NO',_class='datatable')
    TextMessage = ""
    ReportDate = datetime.date.strftime(ReportDate,"%d-%m-%Y")
    dworkbook.save(dpath)
    if download == "YES":
        redirect(URL(r=request, f='download_AttendanceRegister'))
    #tables = plugins.powerTable
    #tables.datasource = datasource
    #tables.uitheme = 'cupertino'
    #tables.dtfeatures['sPaginationType'] = 'two button'
    #tables.keycolumn = 'tempAttendanceRegisterTable.id'
    #tables.showkeycolumn = False
    #tables.columns = ['tempAttendanceRegisterTable.GR_NO','tempAttendanceRegisterTable.DEV_DTY','tempAttendanceRegisterTable.CANTEEN']
    #tables.hiddencolumns = ['tempAttendanceRegisterTable.GR_NO','tempAttendanceRegisterTable.DEV_DTY','tempAttendanceRegisterTable.CANTEEN']
    #tables.headers = 'labels'
    #created_table = tables.create()
    #DAttendanceRegisterTable = plugin_powerTable(datasource)

    mail.send('acknowledgesynchronization@gmail.com',
        'All Attendance register',
        'All attendance register',
        attachments = mail.Attachment('/home/rootcat/new_web2py/web2py/applications/AttendanceSoftware/private/AttendanceRegister.xlsx', content_id='text'))

    logf.close()
    return 0

def AttendanceFetch():
    import os,time
    import getpass, imaplib, email
    import pandas as pd
    import numpy as np
    import pickle
    import pprint
    from StyleFrame import StyleFrame, Styler, utils

    sorting_order = {}
    sorting_order[0,1] = ['3','3','3']
    sorting_order[0,2] = ['3','3','3']
    sorting_order[0,3] = ['3','3','3']
    sorting_order[0,4] = ['3','3','3']
    sorting_order[0,5] = ['3','3','3']
    sorting_order[1,1] = ['2','2','2']
    sorting_order[1,2] = ['2','2','2']
    sorting_order[1,3] = ['2','2','2']
    sorting_order[1,4] = ['2','2','2']
    sorting_order[1,5] = ['2','2','2']
    sorting_order[2,1] = ['1','1','1']
    sorting_order[2,2] = ['2','2','2']
    sorting_order[2,3] = ['5','5','5']
    sorting_order[2,4] = ['4','4','4']
    sorting_order[2,5] = ['3','3','3']
    sorting_order[3,1] = ['4','4','4']
    sorting_order[3,2] = ['4','4','4']
    sorting_order[3,3] = ['4','4','4']
    sorting_order[3,4] = ['4','4','4']
    sorting_order[3,5] = ['4','4','4']
    sorting_order[4,1] = ['1','1','1']
    sorting_order[4,2] = ['1','1','1']
    sorting_order[4,3] = ['1','1','1']
    sorting_order[4,4] = ['1','1','1']
    sorting_order[4,5] = ['1','1','1']
    sorting_order[5,1] = ['5','5','5']
    sorting_order[5,2] = ['5','5','5']
    sorting_order[5,3] = ['5','5','5']
    sorting_order[5,4] = ['5','5','5']
    sorting_order[5,5] = ['5','5','5']
    sorting_order[6,1] = ['3','3','3']
    sorting_order[6,2] = ['4','4','4']
    sorting_order[6,3] = ['1','1','1']
    sorting_order[6,4] = ['2','2','2']
    sorting_order[6,5] = ['1','2','4']
    revolve = len(sorting_order[0,1])


    pathlog = os.path.join(request.folder,'private','log_auto_SSAttendance')
    logf = open(pathlog,'w')
    logf.write("starting\n")
    import commands
    logf.write(commands.getoutput('date'))

    #db.SSAttendanceDate.drop()
    data_email_id = 'acknowledgesynchronization'
    data_email_password = 'synchronizationacknowledge'
    M = ''
    try:
        M = imaplib.IMAP4_SSL("imap.gmail.com", 993)
    except:
        logf.write("Cannot connect to gmail")

    try:
        M.login(data_email_id+'@gmail.com',data_email_password)
    except:
        logf.write("Unable to login to " + data_email_id + "@gmail.\n You probably need to open the email from your browser once or try to open:\n https://www.google.com/accounts/DisplayUnlockCaptcha")

    M.select('inbox')
    result, data = M.uid('search', None, '(SUBJECT "SENDING_INCREMENTAL_UPDATE:DATE: ")')
    uids = data[0].split()

    try:
        result, data = M.uid('fetch', uids[-1], '(RFC822)')
    except:
        logf.write(" client didn't acknowledged yet: ")

    m = email.message_from_string(data[0][1])
    if m.get_content_maintype() == 'multipart': #multipart messages only
        for part in m.walk():
            if part.get_content_maintype() == 'multipart': continue
            if part.get('Content-Disposition') is None: continue

            #save the attachment in the program directory
            filename = part.get_filename()
            fp = open(os.path.join(request.folder,'private',filename), 'wb')
            fp.write(part.get_payload(decode=True))
            fp.close()
            logf.write('%s saved!\n' % filename)


    result, data = M.uid('search', None, '(SUBJECT "SENDING_INCREMENTAL_UPDATE:COUNT: ")')
    uids = data[0].split()

    try:
        result, data = M.uid('fetch', uids[-1], '(RFC822)')
    except:
        logf.write(" Count client didn't acknowledged yet: ")

    m = email.message_from_string(data[0][1])
    if m.get_content_maintype() == 'multipart': #multipart messages only
        for part in m.walk():
            if part.get_content_maintype() == 'multipart': continue
            if part.get('Content-Disposition') is None: continue

            #save the attachment in the program directory
            filename = part.get_filename()
            fp = open(os.path.join(request.folder,'private',filename), 'wb')
            fp.write(part.get_payload(decode=True))
            fp.close()
            logf.write('%s saved!\n' % filename)




    #Update Count
    filename = 'attendance_count.pkl'
    db.commit()
    logf.write("Commited!\n")
    logf.write("loading attendance count pkl!\n")
    pkl_file = ""
    try:
        logf.write("loading pkl now!\n")
        pkl_file = open(os.path.join(request.folder,'private',filename), 'rb')
        logf.write("loaded pkl!\n")
    except:
        logf.write("Unable to open pkl\n")
    AttendanceCount = pd.read_pickle(os.path.join(request.folder,'private',filename))
    AttendanceCount.drop_duplicates(inplace=True)
    logf.write("read pickle!\n")
    db(db.SSAttendanceCount.id > 0).delete()


    #my headers are the headers in DB
    header_map_dict = {}
    header_map_dict['NewID'] =  'NewID'
    header_map_dict['OldSewadarid']        =  'OldSewadarid'
    header_map_dict['Name']                =  'Name'
    header_map_dict['Father_Husband_Name'] =  'Father_Husband_Name'
    header_map_dict['status']              =  'status'
    header_map_dict['Gender']              =  'gender'
    header_map_dict['B']                   =  'B'
    header_map_dict['w']                   =  'w'
    header_map_dict['V1']                  =  'V1'
    header_map_dict['V2']                  =  'V2'
    header_map_dict['V3']                  =  'V3'
    header_map_dict['V4']                  =  'V4'
    header_map_dict['Initiated_Status']    =  'Initiated_Status'
    header_map_dict['TotalVisit']          =  'TotalVisit'
    header_map_dict['Total']               =  'Total'
    header_map_dict['areaname']            =  'areaname'

    for row in AttendanceCount.iterrows():
        i=0
        row_dict = {}
        logf.write(str(row) + "\n")
        logf.close()
        logf = open(pathlog,'a')

        for col in header_map_dict.keys():
            if(col == 'OldSewadarid'):
                if row[1][col] == 'NA' or row[1][col] == '' or pd.isnull(row[1][col]):
                    row_dict[header_map_dict[col]] = row[1]['NewID']
                else:
                    row_dict[header_map_dict[col]] = row[1][col]
            else:
               row_dict[header_map_dict[col]] = row[1][col]
           #logf.write(col + ' = ' + str(row[1][col]) + "\n")

        logf.write("abou to insert\n")
        try:
            db.SSAttendanceCount.insert(**row_dict)
            logf.write("insert succesfull\n")
        except:
            logf.write("Unable to update count\n")

    db.commit()

    ######################################################################
    #Merge Attendance
    ######################################################################
    filename = 'todays_attendance.pkl'
    db.commit()
    logf.write("Commited!\n")
    logf.write("loading pkl!\n")
    pkl_file = ""
    try:
        logf.write("loading pkl now!\n")
        pkl_file = open(os.path.join(request.folder,'private',filename), 'rb')
        logf.write("loaded pkl!\n")
    except:
        logf.write("Unable to open pkl\n")
    AttendanceTodays = pd.read_pickle(os.path.join(request.folder,'private',filename))
    logf.write("read pickle!\n")
    logf.write("indexed SSAttendance Dates\n")
    logf.close()
    logf = open(pathlog,'a')

    LastUpdated = datetime.datetime(2000, 1, 1)
    datasource  = db(db.LocalVariables.id > 0).select()
    for data in datasource:
        LastUpdated = data['LastUpdated']

    logf.write('Attendance was last updated on ' + datetime.datetime.strftime(LastUpdated,'%d-%b-%Y') + '\n')

    #my headers are the headers in DB
    header_map_dict = {}
    header_map_dict['SewadarNewID'] =  'SewadarNewID'
    header_map_dict['DutyDateTime'] =  'DutyDate'
    header_map_dict['Duty Type'] =  'Duty_Type'

    for row in AttendanceTodays.iterrows():
       i=0
       row_dict = {}
       logf.write(str(row) + "\n")
       #logf.close()
       #logf = open(pathlog,'a')

       for col in header_map_dict.keys():
           if(col == 'DutyDateTime'):
               row_dict['DutyDate'] = datetime.datetime.strptime(row[1]['DutyDateTime'],'%d/%m/%Y %H:%M:%S')
           else:
               row_dict[header_map_dict[col]] = row[1][col]
           #logf.write(col + ' = ' + str(row[1][col]) + "\n")

       logf.write("abou to insert\n")
       #print row_dict['SewadarNewID']
       if LastUpdated < row[1]['DutyDate']:
           LastUpdated = row[1]['DutyDate']

       logf.write("about to insert\n")
       try:
           db.SSAttendanceDate.insert(**row_dict)
           logf.write("insert succesfull\n")
       except:
           db((db.SSAttendanceDate.SewadarNewID == row_dict['SewadarNewID']) & (db.SSAttendanceDate.DutyDate == row_dict['DutyDate'])).update(Duty_Type=row_dict['Duty_Type'])
           logf.write("trying duty type updaate\n")

    db(db.LocalVariables.id>0).update(LastUpdated=LastUpdated)
    logf.write("Send to database engine\n")

    db.commit()

    logf.write("Commited\n")
    SSDates = {}
    if datetime.datetime.now().hour >= 19:
        SSDates = db((db.SSAttendanceDate.DutyDate >= (datetime.datetime.now().replace(hour=19, minute=0, second=0, microsecond=0)))).select().as_list()
    else:
        #SSDates = db((db.SSAttendanceDate.DutyDate >= ((datetime.datetime.now() - datetime.timedelta(days=1)).replace(hour=19, minute=0, second=0, microsecond=0)))).select().as_list()
        SSDates = db((db.SSAttendanceDate.DutyDate >= ((datetime.datetime.now() - datetime.timedelta(days=0)).replace(hour=0, minute=0, second=0, microsecond=0)))).select().as_list()

    #SSDates = db(db.SSAttendanceDate).select().as_list()
    df_dates = pd.DataFrame.from_records(SSDates)
    df_dates.to_excel(os.path.join(request.folder,'private','df_dates.xlsx'))
    if len(df_dates.index) == 0:
        mail.send('canteenattendance@gmail.com',
            'Attendance Report :' + datetime.datetime.strftime(datetime.datetime.now(),'%d-%b-%Y'),
            'No Attendance marked in BIMS for today!')
        return 0
    df_Master = pd.DataFrame.from_records(db(db.MasterSheet).select().as_list())
    df_Master['SewadarNewID'] = df_Master['SewadarNewID'].apply(lambda x: "{}{}".format('BH0011',x))
    pprint.pprint(df_Master,stream=logf)
    df_count = pd.DataFrame.from_records(db(db.SSAttendanceCount).select().as_list())
    df_count.rename(columns={'NewID':'SewadarNewID'},inplace=True)
    df_SMSReport = df_Master.merge(df_count,on=['SewadarNewID'],how='left')
    for row in df_SMSReport.iterrows():
        try:
            if row[1]['gender'] == "Male":
                if row[1]['Total'] <= 30:
                    df_SMSReport.set_value(row[0],'MESSAGE','ID:' + row[1]['SewadarNewID'].replace('BH0011','') + ':Dear ' + row[1]['Name'] + ' ji, Attendance short by ' + str(30-row[1]['Total']))
                else:
                    df_SMSReport.set_value(row[0],'MESSAGE','ID:' + row[1]['SewadarNewID'].replace('BH0011','') + ':Dear ' + row[1]['Name'] + ' ji, Attendance complete')
            else:
                if row[1]['Total'] <= 36:
                    df_SMSReport.set_value(row[0],'MESSAGE','ID:' + row[1]['SewadarNewID'].replace('BH0011','') + ':Dear ' + row[1]['Name'] + ' ji, Attendance short by ' + str(36-row[1]['Total']))
                else:
                    logf.write('\n---------\n' + str(row[0]))
                    logf.write('\n---------\n' + str(row[1]))
                    df_SMSReport.set_value(row[0],'MESSAGE','ID:' + row[1]['SewadarNewID'].replace('BH0011','') + ':Dear ' + row[1]['Name'] + ' ji, Attendance complete')
        except:
            df_SMSReport.set_value(row[0],'MESSAGE','ID:' + row[1]['SewadarNewID'].replace('BH0011','') + 'Attendance short by ' + str(36-row[1]['Total']) + '. Records missing from SewaSamiti. Please contact G.P.Alagh: 8800298700' )

    df_SMSReport.to_csv(os.path.join(request.folder,'private','SendSMS.csv'))

    logf.write('Going to join now\n')
    temp_df_daily_report = df_dates.merge(df_Master,on=['SewadarNewID'],how='left')
    temp_df_daily_report.to_excel(os.path.join(request.folder,'private','temp_df_daily_report.xlsx'))
    logf.write('Going to join again now\n')
    df_daily_report = temp_df_daily_report.merge(df_count,on=['SewadarNewID'],how='left')
    df_daily_report.rename(columns={'DEV_DTY':'JATHA'},inplace=True)
    df_daily_report.rename(columns={'SewadarNewID':'ID'},inplace=True)
    df_daily_report.sort_values(by=['CANTEEN','JATHA','ID'],axis=0,inplace=True)
    df_daily_report = df_daily_report.reset_index()
    pprint.pprint(df_daily_report,stream=logf)
    df_daily_report['ID'] = df_daily_report['ID'].str.replace('BH0011','')
    writer = pd.ExcelWriter(os.path.join(request.folder,'private','df_daily_report.xlsx'),engine='xlsxwriter')
    df_daily_report.to_excel(os.path.join(request.folder,'private','temppp_df_daily_report.xlsx'))
    df_daily_report_pivot = pd.pivot_table(df_daily_report,index=['CANTEEN'],columns=['gender'],values='ID',aggfunc='count',margins=True)
    weekday = datetime.datetime.today().weekday()
    day = datetime.datetime.today().day
    weeknumber = int((day-1)/7)+1

    ReportsLocalVariables = pd.DataFrame.from_records(db(db.ReportsLocalVariables.id > 0).select().as_list())
    ReportsLocalVariables = ReportsLocalVariables.set_index(['Weekday','Weeknumber'])

    if (ReportsLocalVariables.at[(weekday,weeknumber),'LastDate'].replace(hour=0, minute=0, second=0, microsecond=0) == datetime.datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)):
        ordernumber = ReportsLocalVariables.at[(weekday,weeknumber),'LastCanteenInchargeIndex']
    else:
        ordernumber = (ReportsLocalVariables.at[(weekday,weeknumber),'LastCanteenInchargeIndex'] + 1) % revolve

    db(db.ReportsLocalVariables.Weekday == weekday).update(LastCanteenInchargeIndex=ordernumber,LastDate=datetime.datetime.now(),Weekday=weekday)

    canteen_incharge = sorting_order[weekday,weeknumber][ordernumber]
    header_df = pd.DataFrame(columns=['CANTEEN DAILY REPORT: ' + datetime.datetime.today().strftime('%d-%b-%Y') ])
    header_df.set_value('INCHARGE','CANTEEN DAILY REPORT: ' + datetime.datetime.today().strftime('%d-%b-%Y'),'CANTEEN: ' + str(canteen_incharge))
    #df_daily_report_pivot = df_daily_report_pivot.append(df_daily_report_pivot.sum(numeric_only=True), ignore_index=True)
    #logf.write("rename:" + str(df_daily_report_pivot.index.values.tolist()[-1]) + '\n')
    #df_daily_report_pivot.rename(index={df_daily_report_pivot.index.values.tolist()[-1]:'TOTAL'},inplace=True)
    df_daily_report_pivot.to_excel(writer,sheet_name='Summary',startrow=3,startcol=0)
    header_df.to_excel(writer,sheet_name='Summary',startrow=0,startcol=0,index=True)
    df_daily_report_canteen_incharge = df_daily_report[df_daily_report['CANTEEN'] == str(canteen_incharge)]
    df_daily_report_canteen_incharge.index = range(1,len(df_daily_report_canteen_incharge)+1)
    df_daily_report_others = df_daily_report[df_daily_report['CANTEEN'] != str(canteen_incharge)]
    df_daily_report_others.index = range(1,len(df_daily_report_others)+1)
    df_daily_report_canteen_incharge.to_excel(writer,columns=['ID','Name','JATHA','CANTEEN','DutyDate'],freeze_panes=(1,1),sheet_name='Incharge',startrow=0,startcol=0)
    df_daily_report_others.to_excel(writer,columns=['ID','Name','JATHA','CANTEEN','DutyDate'],freeze_panes=(1,1),sheet_name='Others',startrow=0,startcol=0)
    workbook  = writer.book
    cell_format = workbook.add_format()
    cell_format.set_align('center')
    cell_format.set_border(1)

    cell_format1 = workbook.add_format()
    cell_format1.set_align('left')
    cell_format1.set_border(1)

    worksheet = writer.sheets['Summary']
    worksheet.set_column('B:B', 8, cell_format)
    worksheet.set_column('A:A', 8, cell_format)
    worksheet.set_column('C:C', 8, cell_format)
    worksheet.set_column('D:D', 8, cell_format)
    worksheet = writer.sheets['Incharge']
    worksheet.set_column('B:B', 8, cell_format)
    worksheet.set_column('A:A', 5, cell_format)
    worksheet.set_column('C:C', 28, cell_format1)
    worksheet.set_column('D:D', 28, cell_format1)
    worksheet.set_column('E:E', None, cell_format)
    worksheet.set_column('F:F', 18, cell_format)
    worksheet = writer.sheets['Others']
    worksheet.set_column('B:B', 8, cell_format)
    worksheet.set_column('A:A', 5, cell_format)
    worksheet.set_column('C:C', 28, cell_format1)
    worksheet.set_column('D:D', 28, cell_format1)
    worksheet.set_column('E:E', None, cell_format)
    worksheet.set_column('F:F', 18, cell_format)
    logf.write("Saving xls")
    writer.save()
    logf.write("Saved xls!")

    #writer = StyleFrame.ExcelWriter(os.path.join(request.folder,'private',"stylepandas.xlsx"))
    #sf=StyleFrame(df_daily_report)
    #sf.apply_column_style(cols_to_style=df_daily_report.columns, styler_obj=Styler(bg_color=utils.colors.white, bold=True, font=utils.fonts.arial,font_size=8),style_header=True)
    #sf.apply_headers_style(styler_obj=Styler(bg_color=utils.colors.blue, bold=True, font_size=8, font_color=utils.colors.white,number_format=utils.number_formats.general, protection=False))
    #writer.save()
    #logf.write(df_dates)
    logf.write("Tadaaaa!!")
    logf.close()
    mail.send('canteenattendance@gmail.com',
        'Attendance Report :' + datetime.datetime.strftime(datetime.datetime.now(),'%d-%b-%Y'),
        'Success',
        attachments = [mail.Attachment('/home/rootcat/new_web2py/web2py/applications/AttendanceSoftware/private/df_daily_report.xlsx', content_id='excel')])
    mail.send(Incharge_Email_Ids[canteen_incharge],
        'Attendance Report :' + datetime.datetime.strftime(datetime.datetime.now(),'%d-%b-%Y'),
        'Success',
        attachments = [mail.Attachment('/home/rootcat/new_web2py/web2py/applications/AttendanceSoftware/private/df_daily_report.xlsx', content_id='excel')])


    return 0

def SMSReport():
    mail.send('acknowledgesynchronization@gmail.com',
        'SMSReport',
        'Success',
        attachments = [mail.Attachment('/home/rootcat/new_web2py/web2py/applications/AttendanceSoftware/private/SendSMS.csv', content_id='excel')])
    return 0

def SendSMSWarning():
    import os,time
    import pandas as pd
    import numpy as np
    import pprint


    df_MasterSheet = pd.DataFrame.from_records(db(db.MasterSheet.id > 0).select().as_list())
    df_CountSheet = pd.DataFrame.from_records(db(db.SSAttendanceCount.id > 0).select().as_list())
    df_CountSheet['SewadarNewID'] = df_CountSheet['NewID'].str.replace('BH0011','')
    df_MergedMasterSheet = df_MasterSheet.merge(df_CountSheet,on=['SewadarNewID'],how='left')
    for row in df_MergedMasterSheet.iterrows():
        if row[1]['gender'] == "Male":
            if row[1]['Total'] <= 30:
                df_MergedMasterSheet.set_value(row[0],'MESSAGE','ID:' + row[1]['SewadarNewID'].replace('BH0011','') + ':Dear ' + row[1]['Name'] + ' ji, Attendance short by ' + str(30-row[1]['Total']))
            else:
                df_MergedMasterSheet.set_value(row[0],'MESSAGE','ID:' + row[1]['SewadarNewID'].replace('BH0011','') + ':Dear ' + row[1]['Name'] + ' ji, Attendance complete')
        else:
            if row[1]['Total'] <= 36:
                df_MergedMasterSheet.set_value(row[0],'MESSAGE','ID:' + row[1]['SewadarNewID'].replace('BH0011','') + ':Dear ' + row[1]['Name'] + ' ji, Attendance short by ' + str(36-row[1]['Total']))
            else:
                df_MergedMasterSheet.set_value(row[0],'MESSAGE','ID:' + row[1]['SewadarNewID'].replace('BH0011','') + ':Dear ' + row[1]['Name'] + ' ji, Attendance complete')

        if not(pd.isnull(df_MergedMasterSheet.at[row[0],'MOBILE'])):
            sms_url_api = 'http://enterprise.easyserve.me/http-api.php?username=vajoff&password=vajoff@123&senderid=BAJOFF&route=1&number=' + df_MergedMasterSheet.at[row[0],'MOBILE'] + '&message=' + df_MergedMasterSheet.at[row[0],'MESSAGE'];
            df_MergedMasterSheet.set_value(row[0],'API_URL',sms_url_api)

    df_MergedMasterSheet.to_csv(os.path.join(request.folder,'private','CronSendSMS.csv'))
    return 0


from gluon.scheduler import Scheduler
scheduler = Scheduler(db2)
import datetime
#scheduler.queue_task(AttendanceFetch,
#                    start_time=datetime.datetime.strptime('09-July-2018 11:00:00','%d-%B-%Y %H:%M:%S'),  # datetime
#                    stop_time=None,  # datetime
#                    timeout = 3600,  # seconds
#                    prevent_drift=True,
#                    period=3600*24,  # seconds
#                    immediate=False,
#                    repeats=0)
