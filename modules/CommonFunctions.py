import os
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.cell import get_column_letter
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

def DictToXls(Dict,dpath,index_names):
    wb = Workbook()
    wb.remove_sheet(wb.get_sheet_by_name('Sheet'))
    for sheet_name in Dict['SHEETS']:
        sheet = wb.create_sheet(0)
        sheet.title = sheet_name
        print "Creating sheet = " + sheet_name

        #Put header
        i=1
        try:
            for col_name in list(index_names[sheet_name]):
                sheet[get_column_letter(i)+'1'].value = col_name
                i = i + 1
            #Save index length
            h = i

            for col_name in Dict[sheet_name,'ARRAYCOLUMNNAMES']:
                sheet[get_column_letter(i)+'1'].value = col_name
                i = i + 1


            j = 2
            for row_name in Dict[sheet_name,'ARRAYROWNAMES']:
                i = 1
                for index_name in list(row_name):
                    sheet[get_column_letter(i)+str(j)].value = index_name
                    i = i +1

                for col_name in Dict[sheet_name,'ARRAYCOLUMNNAMES']:
                    sheet[get_column_letter(i)+str(j)].value = Dict[sheet_name,row_name,col_name]
                    i = i + 1
               
                j = j + 1
        except:
            print "Skipping Sheet: " + sheet_name


    wb.save(dpath) 
    return 0


def XlsToDict(dpath,key_dict):
    Dict = {}
    wb = load_workbook(dpath)
    Dict['SHEETS'] = wb.get_sheet_names()    
    for sheet_name in Dict['SHEETS']:
        Dict[(sheet_name,'ARRAYROWNAMES')] = []
        Dict[(sheet_name,'ARRAYCOLUMNNAMES')] = []
        sheet = wb[sheet_name]
        ColNumberToColNameMapping = {}
        ColNameToColNumberMapping = {}
        max_Y = sheet.max_row
        max_X = sheet.max_column
        for i in xrange(max_X):
            i = i + 1
            ColNumberToColNameMapping[i] = sheet[get_column_letter(i)+'1'].value
            if ColNumberToColNameMapping[i] == None:
                ColNumberToColNameMapping[i] = 'EMPTY_HEADER'

            ColNameToColNumberMapping[ColNumberToColNameMapping[i]] = i

            try:
                if (ColNumberToColNameMapping[i] in key_dict[sheet_name]) or ((i in key_dict[sheet_name]) and key_dict[sheet_name,'TYPE']):
                    pass
                else:
                    Dict[(sheet_name,'ARRAYCOLUMNNAMES')].append(ColNumberToColNameMapping[i])
            except:
                print ColNameToColNumberMapping
                print ColNumberToColNameMapping
                print "Skipping something in " + sheet_name

        for j in xrange(max_Y-1):
            j = j+2
            #find index name
            index_name = ()

            try:
                for key in key_dict[sheet_name]:
                    if key_dict[sheet_name,'TYPE'] == 'ROW_NUMBER':
                        if (sheet[get_column_letter(key)+str(j)].value,) == None:
                            index_name = index_name + ('EMPTY_INDEX',)
                        else:
                            index_name = index_name + (sheet[get_column_letter(key)+str(j)].value,)

                    else:
                        index_name = index_name + (sheet[get_column_letter(ColNameToColNumberMapping[key])+str(j)].value,)

                #print sheet_name + ": index_name = " + str(index_name)
                Dict[(sheet_name,'ARRAYROWNAMES')].append(index_name)
                try:
                    for col in key_dict[sheet_name,'select_columns']:
                        if (col in key_dict[sheet_name]):
                            pass
                        else:
                            Dict[(sheet_name,index_name,col)] = sheet[get_column_letter(ColNameToColNumberMapping[col])+str(j)].value
                except:
                    print "Using non-optimum path"
                    for i in xrange(max_X):
                        i = i + 1
                        if (ColNumberToColNameMapping[i] in key_dict[sheet_name]) or (i in key_dict[sheet_name]):
                            pass
                        else:
                            Dict[(sheet_name,index_name,ColNumberToColNameMapping[i])] = sheet[get_column_letter(i)+str(j)].value
                    


            except:
                print "Skipping something in " + sheet_name



    return Dict


def CsvToDict(dpath,key_dict,csv_name):
    import csv
    Dict = {}
    reader = csv.DictReader(open(dpath),delimiter=',')
    Dict['SHEETS'] = [csv_name]
    Dict[(csv_name,'ARRAYROWNAMES')] = []
    Dict[(csv_name,'ARRAYCOLUMNNAMES')] = []
    for row in reader:
        index_name = ()
        for key in key_dict[csv_name]:
            index_name = (row[key],) + index_name

        Dict[(csv_name,'ARRAYROWNAMES')].append(index_name)
        try:
            for col in key_dict[csv_name,'select_columns']:
                if col in key_dict[csv_name]:
                    pass
                else:
                    Dict[(csv_name,index_name,col)] = row[col]
                    Dict[(csv_name,'ARRAYCOLUMNNAMES')].append(col)
        except:
            print "Using non-optimum path"
            for key, value in row.iteritems():
                if key in key_dict[csv_name]:
                    pass
                else:
                    Dict[(csv_name,index_name,key)] = value
                



    Dict[(csv_name,'ARRAYCOLUMNNAMES')] = list(set(Dict[(csv_name,'ARRAYCOLUMNNAMES')]))

    return Dict




import numpy
def smooth(x,window_len=11,window='hanning'):
    """smooth the data using a window with requested size.

    This method is based on the convolution of a scaled window with the signal.
    The signal is prepared by introducing reflected copies of the signal
    (with the window size) in both ends so that transient parts are minimized
    in the begining and end part of the output signal.

    input:
        x: the input signal
        window_len: the dimension of the smoothing window; should be an odd integer
        window: the type of window from 'flat', 'hanning', 'hamming', 'bartlett', 'blackman'
            flat window will produce a moving average smoothing.

    output:
        the smoothed signal

    example:

    t=linspace(-2,2,0.1)
    x=sin(t)+randn(len(t))*0.1
    y=smooth(x)

    see also:

    numpy.hanning, numpy.hamming, numpy.bartlett, numpy.blackman, numpy.convolve
    scipy.signal.lfilter

    TODO: the window parameter could be the window itself if an array instead of a string
    NOTE: length(output) != length(input), to correct this: return y[(window_len/2-1):-(window_len/2)] instead of just y.
    """

    if x.ndim != 1:
        raise ValueError, "smooth only accepts 1 dimension arrays."

    if x.size < window_len:
        raise ValueError, "Input vector needs to be bigger than window size."


    if window_len<3:
        return x


    if not window in ['flat', 'hanning', 'hamming', 'bartlett', 'blackman']:
        raise ValueError, "Window is on of 'flat', 'hanning', 'hamming', 'bartlett', 'blackman'"


    s=numpy.r_[x[window_len-1:0:-1],x,x[-1:-window_len:-1]]
    #print(len(s))
    if window == 'flat': #moving average
        w=numpy.ones(window_len,'d')
    else:
        w=eval('numpy.'+window+'(window_len)')

    y=numpy.convolve(w/w.sum(),s,mode='valid')
    return y


#Example
#dpath = 'AllRunIndividual/IMPO.xlsx'
#dpath1 = 'AllRunIndividual/IMPO1.xlsx'
#key_dict = {}
#key_dict['IMPO'] = [1]
#key_dict['IMPO','TYPE'] = 'ROW_NUMBER'
#MyDict = XlsToDict(dpath,key_dict)
#print MyDict
#for key,value in MyDict.items():
#    print 'MyDict[' + str(key) + '] = ' + str(value)
#
#header_dict = {}
#header_dict['Sheet'] = ()
#header_dict['IMPO'] = ('hello',)
#DictToXls(MyDict,dpath1,header_dict)

